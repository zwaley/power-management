import os
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

from fastapi import FastAPI, Request, Depends, Form, UploadFile, File, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.encoders import jsonable_encoder
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session, aliased
from sqlalchemy import func, or_
import pandas as pd
from typing import List, Optional
from urllib.parse import quote
import io
import traceback # 导入 traceback 用于打印详细的错误堆栈
from datetime import datetime, timedelta, date
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pydantic import BaseModel
import re
from sqlalchemy import and_
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.encoders import jsonable_encoder
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session, aliased
from sqlalchemy import func, or_
import pandas as pd
from typing import List, Optional
from urllib.parse import quote
import io
import traceback # 导入 traceback 用于打印详细的错误堆栈
from datetime import datetime, timedelta, date
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pydantic import BaseModel
import re
from sqlalchemy import and_

# Import configuration
from config import ADMIN_PASSWORD, PORT

# Fixed imports, using correct function names and models
from models import SessionLocal, Device, Connection, LifecycleRule, create_db_and_tables
from device_types import STANDARD_DEVICE_TYPES, validate_device_type, get_device_type_suggestions, STANDARD_DEVICE_TYPES


# Import error tracking system
from topology_error_tracker import topology_error_tracker, ErrorCategory, ErrorLevel


# --- Port Statistics Service Class ---

class PortStatisticsService:
    """端口统计服务类，用于处理设备端口使用情况的统计分析"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def _get_device_port_summary(self) -> dict:
        """获取设备端口总览 - 采用集合统计逻辑，统计所有有连接的端口"""
        try:
            # Count total devices
            total_devices = self.db.query(Device).count()
            
            # Use sets to avoid duplicate counting of the same port
            all_ports = set()
            connected_ports = set()
            
            # Get all connection records
            connections = self.db.query(Connection).all()
            
            for conn in connections:
                # Count source ports (A side)
                if conn.source_fuse_number and conn.source_device_id:
                    port_key = f"device_{conn.source_device_id}_fuse_{conn.source_fuse_number}"
                    all_ports.add(port_key)
                    # Determine port status by checking if connection_type field is empty
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
                        
                if conn.source_breaker_number and conn.source_device_id:
                    port_key = f"device_{conn.source_device_id}_breaker_{conn.source_breaker_number}"
                    all_ports.add(port_key)
                    # 通过连接类型字段是否为空判断端口使用状态
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
                
                # Count target ports (B side) - meets the design document requirement of "one connection occupies two ports"
                if conn.target_fuse_number and conn.target_device_id:
                    port_key = f"device_{conn.target_device_id}_fuse_{conn.target_fuse_number}"
                    all_ports.add(port_key)
                    # 通过连接类型字段是否为空判断端口使用状态
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
                        
                if conn.target_breaker_number and conn.target_device_id:
                    port_key = f"device_{conn.target_device_id}_breaker_{conn.target_breaker_number}"
                    all_ports.add(port_key)
                    # 通过连接类型字段是否为空判断端口使用状态
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
            
            total_ports = len(all_ports)
            connected_count = len(connected_ports)
            idle_ports = total_ports - connected_count
            utilization_rate = (connected_count / total_ports * 100) if total_ports > 0 else 0
            
            return {
                "total_devices": total_devices,
                "total_ports": total_ports,
                "connected_ports": connected_count,
                "idle_ports": idle_ports,
                "utilization_rate": round(utilization_rate, 2)
            }
        except Exception as e:
            print(f"获取设备端口总览时出错: {e}")
            return {
                "total_devices": 0,
                "total_ports": 0,
                "connected_ports": 0,
                "idle_ports": 0,
                "utilization_rate": 0
            }
    
    def get_device_port_details(self, device_id: int) -> dict:
        """获取指定设备的端口详情 - 基于连接表中该设备的实际端口数据"""
        try:
            # 获取设备信息
            device = self.db.query(Device).filter(Device.id == device_id).first()
            if not device:
                raise HTTPException(status_code=404, detail="设备不存在")
            
            # 获取该设备作为A端设备的所有连接记录，从中提取实际端口信息
            connections = self.db.query(Connection).filter(
                Connection.source_device_id == device_id
            ).all()
            
            # 收集该设备的所有端口信息（基于连接表中的实际数据）
            ports = []
            port_usage_map = {}
            
            for conn in connections:
                # 处理熔丝端口
                if conn.source_fuse_number:
                    port_key = f"熔丝-{conn.source_fuse_number}"
                    if port_key not in port_usage_map:
                        port_info = {
                            "port_name": port_key,
                            "port_type": "熔丝",
                            "port_number": conn.source_fuse_number,
                            "specification": conn.source_fuse_spec or "未知规格",
                            "rating": self._extract_rating_from_spec(conn.source_fuse_spec or ""),
                            "status": "已连接" if conn.connection_type else "空闲",
                            "connected_device": conn.target_device.name if conn.target_device and conn.connection_type else None,
                            "connection_id": conn.id if conn.connection_type else None
                        }
                        ports.append(port_info)
                        port_usage_map[port_key] = port_info
                
                # 处理空开端口
                if conn.source_breaker_number:
                    port_key = f"空开-{conn.source_breaker_number}"
                    if port_key not in port_usage_map:
                        port_info = {
                            "port_name": port_key,
                            "port_type": "空开",
                            "port_number": conn.source_breaker_number,
                            "specification": conn.source_breaker_spec or "未知规格",
                            "rating": self._extract_rating_from_spec(conn.source_breaker_spec or ""),
                            "status": "已连接" if conn.connection_type else "空闲",
                            "connected_device": conn.target_device.name if conn.target_device and conn.connection_type else None,
                            "connection_id": conn.id if conn.connection_type else None
                        }
                        ports.append(port_info)
                        port_usage_map[port_key] = port_info
            
            # If no ports are found, return an empty list (indicating the device has no port configuration or connection records)
            if not ports:
                return {
                    "device_info": {
                        "id": device.id,
                        "name": device.name,
                        "device_type": device.device_type or "未知",
                        "station": device.station or "未知",
                        "location": device.location or "未知"
                    },
                    "port_summary": {
                        "total_ports": 0,
                        "connected_ports": 0,
                        "idle_ports": 0,
                        "utilization_rate": 0.0
                    },
                    "ports": []
                }
            
            # Statistics
            total_ports = len(ports)
            connected_ports = len([p for p in ports if p["status"] == "已连接"])
            idle_ports = total_ports - connected_ports
            utilization_rate = (connected_ports / total_ports * 100) if total_ports > 0 else 0
            
            return {
                "device_info": {
                    "id": device.id,
                    "name": device.name,
                    "device_type": device.device_type or "未知",
                    "station": device.station or "未知",
                    "location": device.location or "未知"
                },
                "port_summary": {
                    "total_ports": total_ports,
                    "connected_ports": connected_ports,
                    "idle_ports": idle_ports,
                    "utilization_rate": round(utilization_rate, 2)
                },
                "ports": sorted(ports, key=lambda x: (x["port_type"], x["port_number"]))
            }
        except HTTPException:
            raise
        except Exception as e:
            print(f"获取设备端口详情时出错: {e}")
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"获取设备端口详情失败: {str(e)}")
    
    def _extract_rating_from_spec(self, spec: str) -> str:
        """从规格字符串中提取额定值"""
        if not spec:
            return "未知"
        
        # 尝试提取数字和单位（如：63A, 100A, 2.5mm²等）
        import re
        pattern = r'(\d+(?:\.\d+)?)\s*([A-Za-z²]+)'
        match = re.search(pattern, spec)
        if match:
            return f"{match.group(1)}{match.group(2)}"
        else:
            return "未知"

# --- Port topology service class ---
class PortTopologyService:
    """Port topology service class for generating port topology data"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def get_port_service_data(self, device_id: int, mode: str = "all") -> dict:
        """
        获取端口拓扑图数据 - 服务层方法
        
        Args:
            device_id: 设备ID
            mode: 显示模式，可选 'all' 或 'used'
        
        Returns:
            包含节点和边的拓扑图数据
        """
        try:
            # Get the center device
            logger.info(f"查询中心设备，设备ID: {device_id}")
            center_device = self.db.query(Device).filter(Device.id == device_id).first()
            if not center_device:
                logger.error(f"设备不存在，设备ID: {device_id}")
                raise HTTPException(status_code=404, detail="设备不存在")
            
            # Get all connections related to this device
            logger.info(f"查询与设备ID {device_id} 相关的所有连接")
            connections = self.db.query(Connection).filter(
                (Connection.source_device_id == device_id) |
                (Connection.target_device_id == device_id)
            ).all()
            
            if not connections:
                logger.info(f"设备ID {device_id} 没有相关连接")
                return {
                    "nodes": [],
                    "edges": []
                }
            
            # Build nodes and edges
            nodes = []
            edges = []
            
            # Add center device node
            center_node = {
                "id": f"device_{device_id}",
                "label": center_device.name,
                "type": "device",
                "x": 0,
                "y": 0,
                "size": 20,
                "color": "#848484"
            }
            nodes.append(center_node)
            
            # Process connections
            left_ports = []
            right_ports = []
            port_count = 0
            
            for conn in connections:
                # Check if display conditions are met
                if mode == "used" and not conn.connection_type:
                    continue  # 跳过空闲端口
                
                # Process source port (local)
                if conn.source_device_id == device_id:
                    source_port = conn.source_fuse_number or conn.source_breaker_number
                    if source_port:
                        port_count += 1
                        port_id = f"port_{device_id}_{source_port}"
                        port_label = f"{source_port}"
                        
                        # Distribute ports left/right based on count
                        if port_count % 2 == 1:
                            port_x = -100
                            port_y = 50 + (port_count // 2) * 80
                            port_side = "left"
                        else:
                            port_x = 100
                            port_y = 50 + ((port_count - 1) // 2) * 80
                            port_side = "right"
                        
                        port_node = {
                            "id": port_id,
                            "label": port_label,
                            "type": "port",
                            "x": port_x,
                            "y": port_y,
                            "size": 10,
                            "color": "#848484"
                        }
                        nodes.append(port_node)
                        
                        # 添加连接线
                        edge_id = f"edge_{device_id}_{source_port}"
                        edge_data = {
                            "id": edge_id,
                            "from": f"device_{device_id}",
                            "to": port_id,
                            "width": 2,
                            "color": "#848484",
                            "cable_model": conn.cable_model,
                            "remark": conn.remark,
                            "arrows": {"to": {"enabled": True, "type": "arrow"}}
                        }
                        edges.append(edge_data)
                        
                        # 处理对端设备
                        if conn.target_device:
                            target_device = conn.target_device
                            target_port = conn.target_fuse_number or conn.target_breaker_number
                            
                            # 创建对端设备节点
                            target_node_id = f"target_{conn.target_device_id}"
                            target_node = {
                                "id": target_node_id,
                                "label": f"{target_device.name} {target_port}",
                                "type": "device",
                                "x": port_x + (-150 if port_side == "left" else 150),
                                "y": port_y,
                                "size": 15,
                                "color": "#848484"
                            }
                            nodes.append(target_node)
                            
                            # 添加连接线到对端设备
                            edge_target_id = f"edge_{device_id}_{source_port}_target"
                            edge_target_data = {
                                "id": edge_target_id,
                                "from": port_id,
                                "to": target_node_id,
                                "width": 2,
                                "color": "#848484",
                                "cable_model": conn.cable_model,
                                "remark": conn.remark,
                                "arrows": {"to": {"enabled": True, "type": "arrow"}}
                            }
                            edges.append(edge_target_data)
                
                # 处理目标端口（本端）
                elif conn.target_device_id == device_id:
                    target_port = conn.target_fuse_number or conn.target_breaker_number
                    if target_port:
                        port_count += 1
                        port_id = f"port_{device_id}_{target_port}"
                        port_label = f"{target_port}"
                        
                        # 根据端口数量决定左右分布
                        if port_count % 2 == 1:
                            port_x = -100
                            port_y = 50 + (port_count // 2) * 80
                            port_side = "left"
                        else:
                            port_x = 100
                            port_y = 50 + ((port_count - 1) // 2) * 80
                            port_side = "right"
                        
                        port_node = {
                            "id": port_id,
                            "label": port_label,
                            "type": "port",
                            "x": port_x,
                            "y": port_y,
                            "size": 10,
                            "color": "#848484"
                        }
                        nodes.append(port_node)
                        
                        # 添加连接线
                        edge_id = f"edge_{device_id}_{target_port}"
                        edge_data = {
                            "id": edge_id,
                            "from": f"device_{device_id}",
                            "to": port_id,
                            "width": 2,
                            "color": "#848484",
                            "cable_model": conn.cable_model,
                            "remark": conn.remark,
                            "arrows": {"to": {"enabled": True, "type": "arrow"}}
                        }
                        edges.append(edge_data)
                        
                        # 处理对端设备
                        if conn.source_device:
                            source_device = conn.source_device
                            source_port = conn.source_fuse_number or conn.source_breaker_number
                            
                            # 创建对端设备节点
                            source_node_id = f"source_{conn.source_device_id}"
                            source_node = {
                                "id": source_node_id,
                                "label": f"{source_device.name} {source_port}",
                                "type": "device",
                                "x": port_x + (-150 if port_side == "left" else 150),
                                "y": port_y,
                                "size": 15,
                                "color": "#848484"
                            }
                            nodes.append(source_node)
                            
                            # 添加连接线到对端设备
                            edge_source_id = f"edge_{device_id}_{target_port}_source"
                            edge_source_data = {
                                "id": edge_source_id,
                                "from": port_id,
                                "to": source_node_id,
                                "width": 2,
                                "color": "#848484",
                                "cable_model": conn.cable_model,
                                "remark": conn.remark,
                                "arrows": {"to": {"enabled": True, "type": "arrow"}}
                            }
                            edges.append(edge_source_data)
            
            logger.info(f"成功生成端口拓扑图数据，节点数: {len(nodes)}, 边数: {len(edges)}")
            return {
                "nodes": nodes,
                "edges": edges
            }
        except Exception as e:
            logger.error(f"获取端口拓扑图数据时出错: {str(e)}")
            traceback.print_exc()
            # 返回标准格式的空数据，避免前端解析失败
            return {
                "nodes": [],
                "edges": []
            }

# Create FastAPI application instance
app = FastAPI(title="DC Asset Manager", description="Power Resource Management System")

# Database session dependency
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Health check endpoints
@app.get("/health")
@app.get("/api/health")
async def health_check(db: Session = Depends(get_db)):
    """健康检查：验证服务与数据库可用性"""
    try:
        # 简单数据库连通性检查
        device_count = db.query(func.count(Device.id)).scalar() or 0
        return {
            "status": "ok",
            "db": "ok",
            "device_count": int(device_count),
            "port": PORT
        }
    except Exception as e:
        topology_error_tracker.log_error(
            category=ErrorCategory.DATABASE_ERROR,
            level=ErrorLevel.ERROR,
            message=f"健康检查失败: {str(e)}",
            exception=e
        )
        raise HTTPException(status_code=500, detail="healthcheck failed")

# Register port topology API endpoint
@app.get("/api/port-topology/{device_id}")
async def get_port_topology_data(device_id: int, mode: str = "detailed", db: Session = Depends(get_db)):
    """
    获取设备的端口拓扑图数据 - 严格按照设计规范实现
    """
    try:
        # 获取设备信息
        device = db.query(Device).filter(Device.id == device_id).first()
        if not device:
            raise HTTPException(status_code=404, detail="设备不存在")
        
        # 获取设备的所有连接
        connections = db.query(Connection).filter(
            or_(Connection.source_device_id == device_id, 
                Connection.target_device_id == device_id)
        ).all()
        
        nodes = []
        edges = []
        
        # 1. 创建中心设备节点 - 按设计规范显示"机房名称 + 设备名称"
        device_label = device.name
        if (device.station and 
            device.station != "未知站点" and 
            device.station.lower() != 'nan' and
            device.station.strip() != '' and
            device.station.lower() != 'none'):
            device_label = f"{device.station}\n{device.name}"
            
        center_device_node = {
            "id": f"device_{device_id}",
            "label": device_label,
            "type": "center_device",
            "nodeType": "device",  # 添加nodeType字段供前端识别
            "x": 0,
            "y": 0,
            "size": 30,
            "color": {"background": "#3b82f6", "border": "#1e40af"},
            "shape": "box",
            "font": {"size": 14, "color": "#ffffff"}
        }
        nodes.append(center_device_node)
        
        # 2. 收集所有端口连接信息
        all_port_connections = []
        for conn in connections:
            if conn.source_device_id == device_id:
                # 本设备是源设备 - 使用原始port_name
                local_port = conn.source_port or f"端口{conn.id}"
                remote_device_id = conn.target_device_id
                remote_port = conn.target_port or f"入线{conn.id}"
                connection_direction = "outgoing"
            else:
                # 本设备是目标设备 - 使用原始port_name
                local_port = conn.target_port or f"端口{conn.id}"
                remote_device_id = conn.source_device_id
                remote_port = conn.source_port or f"出线{conn.id}"
                connection_direction = "incoming"
            
            all_port_connections.append({
                'conn': conn,
                'local_port': local_port,
                'remote_device_id': remote_device_id,
                'remote_port': remote_port,
                'direction': connection_direction
            })
        
        # 3. 按设计规范实现左右分区 - 平均分配端口数量
        total_ports = len(all_port_connections)
        left_port_count = (total_ports + 1) // 2  # 奇数时左侧多一个
        
        left_ports = all_port_connections[:left_port_count]
        right_ports = all_port_connections[left_port_count:]
        
        # 4. 处理左侧端口
        for i, port_info in enumerate(left_ports):
            conn = port_info['conn']
            local_port = port_info['local_port']
            remote_device_id = port_info['remote_device_id']
            remote_port = port_info['remote_port']
            direction = port_info['direction']
            
            # 获取对端设备信息
            remote_device = db.query(Device).filter(Device.id == remote_device_id).first()
            
            # 左侧端口位置计算 - 自上而下排列，整体居中
            port_x = -280
            if left_port_count > 1:
                total_height = (left_port_count - 1) * 100
                start_y = -total_height / 2
                port_y = start_y + i * 100
            else:
                port_y = 0
            
            # 创建端口节点
            port_node_id = f"port_{device_id}_{local_port}_{conn.id}"
            port_node = {
                "id": port_node_id,
                "label": local_port,  # 严格使用原始port_name
                "type": "port",
                "nodeType": "port",  # 添加nodeType字段供前端识别
                "x": port_x,
                "y": port_y,
                "size": 15,
                "color": {"background": "#10b981", "border": "#059669"},
                "shape": "circle",
                "font": {"size": 12}
            }
            nodes.append(port_node)
            
            # 连接中心设备到端口
            center_to_port_edge = {
                "id": f"center_to_{port_node_id}",
                "from": f"device_{device_id}",
                "to": port_node_id,
                "width": 2,
                "color": {"color": "#6b7280"},
                "arrows": {"to": {"enabled": False}}
            }
            edges.append(center_to_port_edge)
            
            # 如果有对端设备，创建对端设备节点
            if remote_device:
                remote_node_id = f"remote_{remote_device_id}_{remote_port}_{conn.id}"
                remote_x = port_x - 200
                
                # 对端设备标签格式：设备名称 + 端口名称
                remote_label = f"{remote_device.name}\n{remote_port}"
                
                remote_node = {
                    "id": remote_node_id,
                    "label": remote_label,
                    "type": "remote_device",
                    "nodeType": "remote_device",  # 添加nodeType字段供前端识别
                    "x": remote_x,
                    "y": port_y,
                    "size": 20,
                    "color": {"background": "#f59e0b", "border": "#d97706"},
                    "shape": "box",
                    "font": {"size": 10}
                }
                nodes.append(remote_node)
                
                # 根据连接类型设置颜色 - 按设计规范
                connection_type = conn.connection_type or ""
                if "交流" in connection_type:
                    edge_color = "#f59e0b"  # 黄色
                elif "直流" in connection_type:
                    edge_color = "#ef4444"  # 红色
                else:
                    edge_color = "#6b7280"  # 灰色
                
                # 根据上下游字段控制箭头方向 - 按设计规范
                upstream_downstream = conn.upstream_downstream or ""
                if direction == "outgoing":
                    # 本设备是源设备
                    if upstream_downstream == "上游":
                        # 本端是上游，箭头从本端指向对端
                        arrow_from = port_node_id
                        arrow_to = remote_node_id
                    else:
                        # 本端是下游，箭头从对端指向本端
                        arrow_from = remote_node_id
                        arrow_to = port_node_id
                else:
                    # 本设备是目标设备
                    if upstream_downstream == "下游":
                        # 本端是下游，箭头从对端指向本端
                        arrow_from = remote_node_id
                        arrow_to = port_node_id
                    else:
                        # 本端是上游，箭头从本端指向对端
                        arrow_from = port_node_id
                        arrow_to = remote_node_id
                
                # 创建连接边
                port_to_remote_edge = {
                    "id": f"{port_node_id}_to_{remote_node_id}",
                    "from": arrow_from,
                    "to": arrow_to,
                    "width": 3,
                    "color": {"color": edge_color},
                    "arrows": {"to": {"enabled": True}},
                    "label": conn.cable_model or conn.connection_type or "",
                    "font": {"size": 10}
                }
                edges.append(port_to_remote_edge)
        
        # 5. 处理右侧端口（逻辑类似左侧）
        for i, port_info in enumerate(right_ports):
            conn = port_info['conn']
            local_port = port_info['local_port']
            remote_device_id = port_info['remote_device_id']
            remote_port = port_info['remote_port']
            direction = port_info['direction']
            
            # 获取对端设备信息
            remote_device = db.query(Device).filter(Device.id == remote_device_id).first()
            
            # 右侧端口位置计算
            port_x = 280
            if len(right_ports) > 1:
                total_height = (len(right_ports) - 1) * 100
                start_y = -total_height / 2
                port_y = start_y + i * 100
            else:
                port_y = 0
            
            # 创建端口节点
            port_node_id = f"port_{device_id}_{local_port}_{conn.id}"
            port_node = {
                "id": port_node_id,
                "label": local_port,
                "type": "port",
                "nodeType": "port",  # 添加nodeType字段供前端识别
                "x": port_x,
                "y": port_y,
                "size": 15,
                "color": {"background": "#10b981", "border": "#059669"},
                "shape": "circle",
                "font": {"size": 12}
            }
            nodes.append(port_node)
            
            # 连接中心设备到端口
            center_to_port_edge = {
                "id": f"center_to_{port_node_id}",
                "from": f"device_{device_id}",
                "to": port_node_id,
                "width": 2,
                "color": {"color": "#6b7280"},
                "arrows": {"to": {"enabled": False}}
            }
            edges.append(center_to_port_edge)
            
            # 如果有对端设备，创建对端设备节点
            if remote_device:
                remote_node_id = f"remote_{remote_device_id}_{remote_port}_{conn.id}"
                remote_x = port_x + 200
                
                remote_label = f"{remote_device.name}\n{remote_port}"
                
                remote_node = {
                    "id": remote_node_id,
                    "label": remote_label,
                    "type": "remote_device",
                    "nodeType": "remote_device",  # 添加nodeType字段供前端识别
                    "x": remote_x,
                    "y": port_y,
                    "size": 20,
                    "color": {"background": "#f59e0b", "border": "#d97706"},
                    "shape": "box",
                    "font": {"size": 10}
                }
                nodes.append(remote_node)
                
                # 连接类型颜色
                connection_type = conn.connection_type or ""
                if "交流" in connection_type:
                    edge_color = "#f59e0b"  # 黄色
                elif "直流" in connection_type:
                    edge_color = "#ef4444"  # 红色
                else:
                    edge_color = "#6b7280"  # 灰色
                
                # 箭头方向控制
                upstream_downstream = conn.upstream_downstream or ""
                if direction == "outgoing":
                    if upstream_downstream == "上游":
                        arrow_from = port_node_id
                        arrow_to = remote_node_id
                    else:
                        arrow_from = remote_node_id
                        arrow_to = port_node_id
                else:
                    if upstream_downstream == "下游":
                        arrow_from = remote_node_id
                        arrow_to = port_node_id
                    else:
                        arrow_from = port_node_id
                        arrow_to = remote_node_id
                
                port_to_remote_edge = {
                    "id": f"{port_node_id}_to_{remote_node_id}",
                    "from": arrow_from,
                    "to": arrow_to,
                    "width": 3,
                    "color": {"color": edge_color},
                    "arrows": {"to": {"enabled": True}},
                    "label": conn.cable_model or conn.connection_type or "",
                    "font": {"size": 10}
                }
                edges.append(port_to_remote_edge)
        
        return {"nodes": nodes, "edges": edges}
        
    except Exception as e:
        print(f"端口拓扑图API调用失败: {str(e)}")
        traceback.print_exc()
        return {"nodes": [], "edges": []}

# 端口拓扑图数据生成函数
def generate_port_topology_data(device_id: int):
    """生成端口拓扑图数据"""
    try:
        with SessionLocal() as db:
            device = db.query(Device).filter(Device.id == device_id).first()
            if not device:
                return {"nodes": [], "edges": []}
            
            nodes = []
            edges = []
            
            # 获取设备的所有端口连接
            connections = db.query(Connection).filter(
                or_(Connection.source_device_id == device_id,
                    Connection.target_device_id == device_id)
            ).all()
            
            # 为每个端口创建节点
            for conn in connections:
                if conn.source_device_id == device_id:
                    port_info = f"{conn.source_port}"
                    connected_device = db.query(Device).filter(Device.id == conn.target_device_id).first()
                    connected_port = conn.target_port
                else:
                    port_info = f"{conn.target_port}"
                    connected_device = db.query(Device).filter(Device.id == conn.source_device_id).first()
                    connected_port = conn.source_port
                
                # 创建端口节点
                port_node = {
                    "id": f"port_{conn.id}",
                    "label": port_info,
                    "type": "port",
                    "x": 0,
                    "y": len(nodes) * 100,
                    "size": 15,
                    "color": "#3b82f6",
                    "shape": "circle"
                }
                nodes.append(port_node)
                
                # 创建连接的设备节点
                if connected_device:
                    device_node = {
                        "id": f"device_{connected_device.id}",
                        "label": f"{connected_device.name}\
{connected_port}",
                        "type": "connected_device",
                        "x": 200,
                        "y": len(nodes) * 100,
                        "size": 20,
                        "color": "#f59e0b",
                        "shape": "box"
                    }
                    nodes.append(device_node)
                    
                    # 创建连接边
                    edge = {
                        "id": f"edge_{conn.id}",
                        "from": f"port_{conn.id}",
                        "to": f"device_{connected_device.id}",
                        "width": 2,
                        "color": "#6b7280",
                        "label": conn.cable_model or ""
                    }
                    edges.append(edge)
            
            return {"nodes": nodes, "edges": edges}
            
    except Exception as e:
        topology_error_tracker.log_error(
            category=ErrorCategory.API_ERROR,
            level=ErrorLevel.ERROR,
            message=f"端口拓扑图数据生成失败: {str(e)}",
            context={"device_id": device_id},
            exception=e
        )
        return {"nodes": [], "edges": []}
        
        # 3. 按左右分区原则布局端口
        port_names = list(port_info.keys())
        total_ports = len(port_names)
        left_ports = port_names[:total_ports//2 + total_ports%2]  # 左侧端口（奇数时多一个）
        right_ports = port_names[total_ports//2 + total_ports%2:]  # 右侧端口
        
        # 4. 创建端口节点和连接
        for i, port_name in enumerate(left_ports):
            # 左侧端口
            port_node_id = f"port_left_{i}"
            port_x = -200
            port_y = -100 + i * 80
            
            port_node = {
                "id": port_node_id,
                "label": port_name,
                "type": "port",
                "x": port_x,
                "y": port_y,
                "size": 15,
                "color": "#10b981",
                "shape": "ellipse",
                "font": {"size": 12}
            }
            nodes.append(port_node)
            
            # 连接到中心设备
            edge_to_center = {
                "id": f"edge_center_{port_node_id}",
                "from": f"device_{device_id}",
                "to": port_node_id,
                "width": 2,
                "color": "#6b7280"
            }
            edges.append(edge_to_center)
            
            # 处理该端口的所有连接
            for conn_info in port_info[port_name]:
                conn = conn_info["connection"]
                remote_device = conn_info["remote_device"]
                remote_port = conn_info["remote_port"]
                
                if remote_device:
                    # 创建对端设备节点
                    remote_node_id = f"remote_{remote_device.id}_{remote_port}"
                    remote_node = {
                        "id": remote_node_id,
                        "label": f"{remote_device.name}\
{remote_port}",
                        "type": "remote_device",
                        "x": port_x - 150,
                        "y": port_y,
                        "size": 20,
                        "color": "#f59e0b",
                        "shape": "box",
                        "font": {"size": 10}
                    }
                    nodes.append(remote_node)
                    
                    # 连接线颜色根据连接类型
                    cable_color = "#fbbf24" if conn.connection_type == "交流" else "#ef4444" if conn.connection_type == "直流" else "#6b7280"
                    
                    # 箭头方向根据上下游关系
                    arrows = {}
                    if conn.upstream_downstream == "上游":
                        arrows = {"to": {"enabled": True}} if conn_info["is_source"] else {"from": {"enabled": True}}
                    elif conn.upstream_downstream == "下游":
                        arrows = {"from": {"enabled": True}} if conn_info["is_source"] else {"to": {"enabled": True}}
                    
                    edge_to_remote = {
                        "id": f"edge_{port_node_id}_{remote_node_id}",
                        "from": port_node_id,
                        "to": remote_node_id,
                        "width": 3,
                        "color": cable_color,
                        "arrows": arrows,
                        "label": conn.cable_model or "",
                        "title": f"电缆型号: {conn.cable_model or 'N/A'}\
备注: {conn.remark or 'N/A'}"
                    }
                    edges.append(edge_to_remote)
        
        # 右侧端口处理
        for i, port_name in enumerate(right_ports):
            port_node_id = f"port_right_{i}"
            port_x = 200
            port_y = -100 + i * 80
            
            port_node = {
                "id": port_node_id,
                "label": port_name,
                "type": "port",
                "x": port_x,
                "y": port_y,
                "size": 15,
                "color": "#10b981",
                "shape": "ellipse",
                "font": {"size": 12}
            }
            nodes.append(port_node)
            
            # 连接到中心设备
            edge_to_center = {
                "id": f"edge_center_{port_node_id}",
                "from": f"device_{device_id}",
                "to": port_node_id,
                "width": 2,
                "color": "#6b7280"
            }
            edges.append(edge_to_center)
            
            # 处理该端口的所有连接
            for conn_info in port_info[port_name]:
                conn = conn_info["connection"]
                remote_device = conn_info["remote_device"]
                remote_port = conn_info["remote_port"]
                
                if remote_device:
                    # 创建对端设备节点
                    remote_node_id = f"remote_{remote_device.id}_{remote_port}"
                    remote_node = {
                        "id": remote_node_id,
                        "label": f"{remote_device.name}\
{remote_port}",
                        "type": "remote_device",
                        "x": port_x + 150,
                        "y": port_y,
                        "size": 20,
                        "color": "#f59e0b",
                        "shape": "box",
                        "font": {"size": 10}
                    }
                    nodes.append(remote_node)
                    
                    # 连接线颜色根据连接类型
                    cable_color = "#fbbf24" if conn.connection_type == "交流" else "#ef4444" if conn.connection_type == "直流" else "#6b7280"
                    
                    # 箭头方向根据上下游关系
                    arrows = {}
                    if conn.upstream_downstream == "上游":
                        arrows = {"to": {"enabled": True}} if conn_info["is_source"] else {"from": {"enabled": True}}
                    elif conn.upstream_downstream == "下游":
                        arrows = {"from": {"enabled": True}} if conn_info["is_source"] else {"to": {"enabled": True}}
                    
                    edge_to_remote = {
                        "id": f"edge_{port_node_id}_{remote_node_id}",
                        "from": port_node_id,
                        "to": remote_node_id,
                        "width": 3,
                        "color": cable_color,
                        "arrows": arrows,
                        "label": conn.cable_model or "",
                        "title": f"电缆型号: {conn.cable_model or 'N/A'}\
备注: {conn.remark or 'N/A'}"
                    }
                    edges.append(edge_to_remote)
        
        topology_error_tracker.log_error(
            category=ErrorCategory.API_SUCCESS,
            level=ErrorLevel.INFO,
            message=f"端口拓扑图数据生成成功",
            context={
                "device_id": device_id,
                "device_name": device.name,
                "total_ports": total_ports,
                "nodes_count": len(nodes),
                "edges_count": len(edges)
            }
        )
        
        return {"nodes": nodes, "edges": edges}
        
    except Exception as e:
        topology_error_tracker.log_error(
            category=ErrorCategory.API_ERROR,
            level=ErrorLevel.ERROR,
            message=f"端口拓扑图API调用失败: {str(e)}",
            context={"device_id": device_id},
            exception=e
        )
        return {"nodes": [], "edges": []}

# 端口拓扑图数据生成函数
def generate_port_topology_data(device_id: int):
    """生成端口拓扑图数据"""
    try:
        with SessionLocal() as db:
            device = db.query(Device).filter(Device.id == device_id).first()
            if not device:
                return {"nodes": [], "edges": []}
            
            nodes = []
            edges = []
            
            # 获取设备的所有端口连接
            connections = db.query(Connection).filter(
                or_(Connection.source_device_id == device_id,
                    Connection.target_device_id == device_id)
            ).all()
            
            # 为每个端口创建节点
            for conn in connections:
                if conn.source_device_id == device_id:
                    port_info = f"{conn.source_port}"
                    connected_device = db.query(Device).filter(Device.id == conn.target_device_id).first()
                    connected_port = conn.target_port
                else:
                    port_info = f"{conn.target_port}"
                    connected_device = db.query(Device).filter(Device.id == conn.source_device_id).first()
                    connected_port = conn.source_port
                
                # 创建端口节点
                port_node = {
                    "id": f"port_{conn.id}",
                    "label": port_info,
                    "type": "port",
                    "x": 0,
                    "y": len(nodes) * 100,
                    "size": 15,
                    "color": "#3b82f6",
                    "shape": "circle"
                }
                nodes.append(port_node)
                
                # 创建连接的设备节点
                if connected_device:
                    device_node = {
                        "id": f"device_{connected_device.id}",
                        "label": f"{connected_device.name}\
{connected_port}",
                        "type": "connected_device",
                        "x": 200,
                        "y": len(nodes) * 100,
                        "size": 20,
                        "color": "#f59e0b",
                        "shape": "box"
                    }
                    nodes.append(device_node)
                    
                    # 创建连接边
                    edge = {
                        "id": f"edge_{conn.id}",
                        "from": f"port_{conn.id}",
                        "to": f"device_{connected_device.id}",
                        "width": 2,
                        "color": "#6b7280",
                        "label": conn.cable_model or ""
                    }
                    edges.append(edge)
            
            return {"nodes": nodes, "edges": edges}
            
    except Exception as e:
        topology_error_tracker.log_error(
            category=ErrorCategory.API_ERROR,
            level=ErrorLevel.ERROR,
            message=f"端口拓扑图数据生成失败: {str(e)}",
            context={"device_id": device_id},
            exception=e
        )
        return {"nodes": [], "edges": []}

        # 添加连接边
        for conn in topology_data.get("connections", []):
            edge = {
                "id": conn["id"],
                "from": conn["from_port_id"],
                "to": conn["to_port_id"],
                "label": conn.get("connection_type", ""),
                "title": "连接类型: " + str(conn.get('connection_type', 'Unknown')) + ", 带宽: " + str(conn.get('bandwidth', 'N/A')),
                "color": get_connection_edge_color(conn.get("status", "active")),
                "width": 2,
                "connectionData": conn
            }
            edges.append(edge)
        
        return {
            "nodes": nodes,
            "edges": edges,
            "metadata": topology_data.get("metadata", {}),
            "device": topology_data.get("device", {})
        }
    except HTTPException as e:
        logger.error(f"HTTP异常: {e.detail}")
        raise e
    except Exception as e:
        logger.error(f"获取端口拓扑图数据时出错: {str(e)}")
        traceback.print_exc()
        # 返回标准格式的空数据，避免前端解析失败
        return JSONResponse(content={"nodes": [], "edges": []}, status_code=500)

        station_utilization.sort(key=lambda x: x["idle_rate"], reverse=True)
        return station_utilization
    
    def _check_idle_rate_alerts(self) -> list:
        """Check idle rate alerts"""
        alerts = []
        
        # 检查总体空闲率
        overall_idle = self._calculate_overall_idle_rate()
        if overall_idle["idle_rate"] < 10:  # 空闲率低于10%预警
            alerts.append({
                "type": "overall",
                "level": "warning",
                "message": f"系统总体空闲率仅为 {overall_idle['idle_rate']}%，资源紧张",
                "idle_rate": overall_idle["idle_rate"]
            })
        
        # 检查设备类型空闲率
        device_type_idle = self._calculate_device_type_idle_rate()
        for item in device_type_idle:
            if item["idle_rate"] < 5:  # 设备类型空闲率低于5%预警
                alerts.append({
                    "type": "device_type",
                    "level": "critical",
                    "message": f"{item['device_type']} 类型设备空闲率仅为 {item['idle_rate']}%，急需扩容",
                    "device_type": item["device_type"],
                    "idle_rate": item["idle_rate"]
                })
        
        # 检查站点空闲率
        station_idle = self._calculate_station_idle_rate()
        for item in station_idle:
            if item["idle_rate"] < 5:  # 站点空闲率低于5%预警
                alerts.append({
                    "type": "station",
                    "level": "critical",
                    "message": f"{item['station']} 站点空闲率仅为 {item['idle_rate']}%，急需扩容",
                    "station": item["station"],
                    "idle_rate": item["idle_rate"]
                })
        
        return alerts
    

    
    def _calculate_port_capacity_distribution(self) -> dict:
        """Calculate port capacity distribution"""
        try:
            # 统计不同规格端口的使用分布
            connections = self.db.query(Connection).all()
            
            fuse_specs = {}
            breaker_specs = {}
            
            for conn in connections:
                # 统计熔断器规格分布
                if conn.source_fuse_spec:
                    spec = conn.source_fuse_spec
                    if spec not in fuse_specs:
                        fuse_specs[spec] = {"total": 0, "connected": 0}
                    fuse_specs[spec]["total"] += 1
                    if conn.connection_type and conn.connection_type.strip():
                        fuse_specs[spec]["connected"] += 1
                
                if conn.target_fuse_spec:
                    spec = conn.target_fuse_spec
                    if spec not in fuse_specs:
                        fuse_specs[spec] = {"total": 0, "connected": 0}
                    fuse_specs[spec]["total"] += 1
                    if conn.connection_type and conn.connection_type.strip():
                        fuse_specs[spec]["connected"] += 1
                
                # 统计空开规格分布
                if conn.source_breaker_spec:
                    spec = conn.source_breaker_spec
                    if spec not in breaker_specs:
                        breaker_specs[spec] = {"total": 0, "connected": 0}
                    breaker_specs[spec]["total"] += 1
                    if conn.connection_type and conn.connection_type.strip():
                        breaker_specs[spec]["connected"] += 1
                
                if conn.target_breaker_spec:
                    spec = conn.target_breaker_spec
                    if spec not in breaker_specs:
                        breaker_specs[spec] = {"total": 0, "connected": 0}
                    breaker_specs[spec]["total"] += 1
                    if conn.connection_type and conn.connection_type.strip():
                        breaker_specs[spec]["connected"] += 1
            
            # 计算各规格的使用率
            for spec_data in fuse_specs.values():
                spec_data["utilization_rate"] = round(
                    (spec_data["connected"] / spec_data["total"] * 100) if spec_data["total"] > 0 else 0, 2
                )
            
            for spec_data in breaker_specs.values():
                spec_data["utilization_rate"] = round(
                    (spec_data["connected"] / spec_data["total"] * 100) if spec_data["total"] > 0 else 0, 2
                )
            
            return {
                "fuse_specifications": fuse_specs,
                "breaker_specifications": breaker_specs
            }
            
        except Exception as e:
            print(f"计算端口容量分布时出错: {e}")
            return {
                "fuse_specifications": {},
                "breaker_specifications": {}
            }
    
    def _calculate_load_balance_analysis(self) -> dict:
        """计算负载均衡分析"""
        try:
            # 获取所有设备的使用率
            devices = self.db.query(Device).all()
            device_utilizations = []
            
            for device in devices:
                utilization_rate = self._get_device_utilization_rate(device.id)
                device_utilizations.append({
                    "device_id": device.id,
                    "device_name": device.name,
                    "device_type": device.device_type or "未知",
                    "station": device.station or "未知",
                    "utilization_rate": utilization_rate
                })
            
            if not device_utilizations:
                return {
                    "balance_score": 0,
                    "average_utilization": 0,
                    "utilization_variance": 0,
                    "overloaded_devices": [],
                    "underutilized_devices": []
                }
            
            # 计算平均使用率和方差
            utilization_rates = [d["utilization_rate"] for d in device_utilizations]
            average_utilization = sum(utilization_rates) / len(utilization_rates)
            variance = sum((rate - average_utilization) ** 2 for rate in utilization_rates) / len(utilization_rates)
            
            # 负载均衡评分（方差越小，均衡性越好）
            balance_score = max(0, 100 - variance)  # 简化的评分算法
            
            # 识别过载和低利用率设备
            overloaded_devices = [d for d in device_utilizations if d["utilization_rate"] > 90]
            underutilized_devices = [d for d in device_utilizations if d["utilization_rate"] < 20]
            
            return {
                "balance_score": round(balance_score, 2),
                "average_utilization": round(average_utilization, 2),
                "utilization_variance": round(variance, 2),
                "overloaded_devices": overloaded_devices,
                "underutilized_devices": underutilized_devices
            }
            
        except Exception as e:
            print(f"计算负载均衡分析时出错: {e}")
            return {
                "balance_score": 0,
                "average_utilization": 0,
                "utilization_variance": 0,
                "overloaded_devices": [],
                "underutilized_devices": []
            }
    
    def _get_top_utilized_devices(self, limit: int = 10) -> list:
        """获取使用率最高的设备"""
        try:
            devices = self.db.query(Device).all()
            device_utilizations = []
            
            for device in devices:
                utilization_rate = self._get_device_utilization_rate(device.id)
                device_utilizations.append({
                    "device_id": device.id,
                    "device_name": device.name,
                    "device_type": device.device_type or "未知",
                    "station": device.station or "未知",
                    "utilization_rate": utilization_rate
                })
            
            # 按使用率降序排序并返回前N个
            device_utilizations.sort(key=lambda x: x["utilization_rate"], reverse=True)
            return device_utilizations[:limit]
            
        except Exception as e:
            print(f"获取使用率最高设备时出错: {e}")
            return []
    
    def _get_device_utilization_rate(self, device_id: int) -> float:
        """获取单个设备的使用率"""
        try:
            # 获取设备的所有端口
            connections = self.db.query(Connection).filter(
                (Connection.source_device_id == device_id) |
                (Connection.target_device_id == device_id)
            ).all()
            
            all_ports = set()
            connected_ports = set()
            
            for conn in connections:
                if conn.source_device_id == device_id:
                    if conn.source_fuse_number:
                        port_key = f"fuse_{conn.source_fuse_number}"
                        all_ports.add(port_key)
                        if conn.connection_type and conn.connection_type.strip():
                            connected_ports.add(port_key)
                    if conn.source_breaker_number:
                        port_key = f"breaker_{conn.source_breaker_number}"
                        all_ports.add(port_key)
                        if conn.connection_type and conn.connection_type.strip():
                            connected_ports.add(port_key)
                
                if conn.target_device_id == device_id:
                    if conn.target_fuse_number:
                        port_key = f"fuse_{conn.target_fuse_number}"
                        all_ports.add(port_key)
                        if conn.connection_type and conn.connection_type.strip():
                            connected_ports.add(port_key)
                    if conn.target_breaker_number:
                        port_key = f"breaker_{conn.target_breaker_number}"
                        all_ports.add(port_key)
                        if conn.connection_type and conn.connection_type.strip():
                            connected_ports.add(port_key)
            
            total_ports = len(all_ports)
            connected_count = len(connected_ports)
            
            return (connected_count / total_ports * 100) if total_ports > 0 else 0
            
        except Exception as e:
            print(f"获取设备 {device_id} 使用率时出错: {e}")
            return 0
    

    
    def get_port_statistics(self) -> dict:
        """获取全局端口统计信息"""
        try:
            # 1. 设备端口总览
            device_port_summary = self._get_device_port_summary()
            
            # 2. 端口类型统计
            port_type_statistics = self._get_port_type_statistics()
            
            # 3. 容量统计
            capacity_statistics = self._get_capacity_statistics()
            
            # 4. 设备端口详情
            device_port_details = self._get_device_port_details()
            
            return {
                "device_port_summary": device_port_summary,
                "port_type_statistics": port_type_statistics,
                "capacity_statistics": capacity_statistics,
                "device_port_details": device_port_details
            }
        except Exception as e:
            print(f"获取端口统计信息时出错: {e}")
            raise HTTPException(status_code=500, detail=f"获取端口统计信息失败: {str(e)}")
    
    def _get_device_port_summary(self) -> dict:
        """获取设备端口总览 - 采用集合统计逻辑，统计所有有连接的端口"""
        try:
            # 统计总设备数
            total_devices = self.db.query(Device).count()
            
            # 使用集合来避免重复计算同一个端口
            all_ports = set()
            connected_ports = set()
            
            # 获取所有连接记录
            connections = self.db.query(Connection).all()
            
            for conn in connections:
                # 统计源端口（A端）
                if conn.source_fuse_number and conn.source_device_id:
                    port_key = f"device_{conn.source_device_id}_fuse_{conn.source_fuse_number}"
                    all_ports.add(port_key)
                    # 通过连接类型字段是否为空判断端口使用状态
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
                        
                if conn.source_breaker_number and conn.source_device_id:
                    port_key = f"device_{conn.source_device_id}_breaker_{conn.source_breaker_number}"
                    all_ports.add(port_key)
                    # 通过连接类型字段是否为空判断端口使用状态
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
                
                # 统计目标端口（B端）- 符合设计文档中"一个连接占用两个端口"的要求
                if conn.target_fuse_number and conn.target_device_id:
                    port_key = f"device_{conn.target_device_id}_fuse_{conn.target_fuse_number}"
                    all_ports.add(port_key)
                    # 通过连接类型字段是否为空判断端口使用状态
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
                        
                if conn.target_breaker_number and conn.target_device_id:
                    port_key = f"device_{conn.target_device_id}_breaker_{conn.target_breaker_number}"
                    all_ports.add(port_key)
                    # 通过连接类型字段是否为空判断端口使用状态
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
            
            total_ports = len(all_ports)
            connected_count = len(connected_ports)
            idle_ports = total_ports - connected_count
            utilization_rate = (connected_count / total_ports * 100) if total_ports > 0 else 0
            
            return {
                "total_devices": total_devices,
                "total_ports": total_ports,
                "connected_ports": connected_count,
                "idle_ports": idle_ports,
                "utilization_rate": round(utilization_rate, 2)
            }
        except Exception as e:
            print(f"获取设备端口总览时出错: {e}")
            return {
                "total_devices": 0,
                "total_ports": 0,
                "connected_ports": 0,
                "idle_ports": 0,
                "utilization_rate": 0
            }
    
    def _get_port_type_statistics(self) -> dict:
        """获取端口类型统计 - 基于A端设备统计"""
        try:
            # 获取所有连接记录
            connections = self.db.query(Connection).all()
            
            # 使用集合来避免重复计算端口
            fuse_ports = set()
            breaker_ports = set()
            connected_fuse_ports = set()
            connected_breaker_ports = set()
            
            for conn in connections:
                # 只统计A端（源端）设备的端口
                if conn.source_fuse_number and conn.source_device_id:
                    port_key = f"{conn.source_device_id}_fuse_{conn.source_fuse_number}"
                    fuse_ports.add(port_key)
                    # 通过连接类型字段是否为空判断端口使用状态
                    if conn.connection_type and conn.connection_type.strip():
                        connected_fuse_ports.add(port_key)
                
                if conn.source_breaker_number and conn.source_device_id:
                    port_key = f"{conn.source_device_id}_breaker_{conn.source_breaker_number}"
                    breaker_ports.add(port_key)
                    # 通过连接类型字段是否为空判断端口使用状态
                    if conn.connection_type and conn.connection_type.strip():
                        connected_breaker_ports.add(port_key)
            
            fuse_total = len(fuse_ports)
            fuse_connected = len(connected_fuse_ports)
            breaker_total = len(breaker_ports)
            breaker_connected = len(connected_breaker_ports)
            
            return {
                "fuse_ports": {
                    "total": fuse_total,
                    "connected": fuse_connected,
                    "idle": fuse_total - fuse_connected,
                    "utilization_rate": round((fuse_connected / fuse_total * 100) if fuse_total > 0 else 0, 2)
                },
                "breaker_ports": {
                    "total": breaker_total,
                    "connected": breaker_connected,
                    "idle": breaker_total - breaker_connected,
                    "utilization_rate": round((breaker_connected / breaker_total * 100) if breaker_total > 0 else 0, 2)
                }
            }
        except Exception as e:
            print(f"获取端口类型统计时出错: {e}")
            return {
                "fuse_ports": {"total": 0, "connected": 0, "idle": 0, "utilization_rate": 0},
                "breaker_ports": {"total": 0, "connected": 0, "idle": 0, "utilization_rate": 0}
            }
    
    def _get_capacity_statistics(self) -> dict:
        """获取容量统计"""
        try:
            # 获取所有连接的规格信息
            connections = self.db.query(Connection).all()
            
            capacity_stats = {}
            high_capacity_available = {"630A_above": 0, "400A_above": 0, "250A_above": 0}
            
            for conn in connections:
                # 处理各种规格字段
                for spec_field in [conn.source_fuse_spec, conn.source_breaker_spec, 
                                 conn.target_fuse_spec, conn.target_breaker_spec]:
                    if spec_field:
                        rating = self._extract_rating_from_spec(spec_field)
                        if rating and rating != "未知":
                            if rating not in capacity_stats:
                                capacity_stats[rating] = {"total": 0, "connected": 0, "idle": 0}
                            
                            capacity_stats[rating]["total"] += 1
                            
                            # 判断是否已连接
                            if conn.source_device_id and conn.target_device_id:
                                capacity_stats[rating]["connected"] += 1
                            else:
                                capacity_stats[rating]["idle"] += 1
                                
                                # 统计大容量可用端口
                                try:
                                    rating_value = int(rating.replace('A', ''))
                                    if rating_value >= 630:
                                        high_capacity_available["630A_above"] += 1
                                    if rating_value >= 400:
                                        high_capacity_available["400A_above"] += 1
                                    if rating_value >= 250:
                                        high_capacity_available["250A_above"] += 1
                                except ValueError:
                                    pass  # 忽略无法转换的容量值
            
            return {
                "by_rating": capacity_stats,
                "high_capacity_available": high_capacity_available
            }
        except Exception as e:
            print(f"获取容量统计时出错: {e}")
            return {
                "by_rating": {},
                "high_capacity_available": {"630A_above": 0, "400A_above": 0, "250A_above": 0}
            }
    
    def _get_device_port_details(self) -> list:
        """获取设备端口详情 - 基于A端设备统计"""
        try:
            # 获取所有设备及其端口使用情况
            devices = self.db.query(Device).all()

            device_details = []
            
            for device in devices:
                # 只统计该设备作为A端（源端）的连接记录
                device_connections = self.db.query(Connection).filter(
                    Connection.source_device_id == device.id
                ).all()

                # 使用集合来避免重复计算端口
                all_ports = set()
                connected_ports_set = set()
                fuse_ports = set()
                breaker_ports = set()
                
                for conn in device_connections:
                    # 只统计该设备作为A端的端口
                    if conn.source_fuse_number:
                        port_key = f"fuse_{conn.source_fuse_number}"
                        all_ports.add(port_key)
                        fuse_ports.add(port_key)
                        # 通过连接类型字段是否为空判断端口使用状态
                        if conn.connection_type and conn.connection_type.strip():
                            connected_ports_set.add(port_key)

                    if conn.source_breaker_number:
                        port_key = f"breaker_{conn.source_breaker_number}"
                        all_ports.add(port_key)
                        breaker_ports.add(port_key)
                        # 通过连接类型字段是否为空判断端口使用状态
                        if conn.connection_type and conn.connection_type.strip():
                            connected_ports_set.add(port_key)

                
                total_ports = len(all_ports)
                connected_ports = len(connected_ports_set)
                
                idle_ports = total_ports - connected_ports
                utilization_rate = (connected_ports / total_ports * 100) if total_ports > 0 else 0
                
                device_details.append({
                    "device_id": device.id,
                    "device_name": device.name,
                    "device_type": device.device_type or "未知",
                    "station": device.station or "未知",
                    "total_ports": total_ports,
                    "connected_ports": connected_ports,
                    "idle_ports": idle_ports,
                    "utilization_rate": round(utilization_rate, 2),
                    "fuse_ports": len(fuse_ports),
                    "breaker_ports": len(breaker_ports)
                })
            
            # 按利用率降序排序
            device_details.sort(key=lambda x: x["utilization_rate"], reverse=True)
            
            return device_details
        except Exception as e:
            print(f"获取设备端口详情时出错: {e}")
            return []
    
    def _extract_rating_from_spec(self, spec_string: str) -> str:
        """从规格字符串中提取电流等级"""
        if not spec_string:
            return "未知"
        
        try:
            # 匹配括号内的电流值，如 "NT4(500A)" -> "500A"
            match = re.search(r'\((\d+)A\)', spec_string)
            if match:
                return f"{match.group(1)}A"
            
            # 匹配直接的电流值，如 "500A" -> "500A"
            match = re.search(r'(\d+)A', spec_string)
            if match:
                return f"{match.group(1)}A"
            
            return "未知"
        except Exception as e:
            print(f"提取电流等级时出错: {e}")
            return "未知"
    



def verify_admin_password(password: str) -> bool:
    """
    验证管理员密码
    Args:
        password: 用户输入的密码
    Returns:
        bool: 密码是否正确
    """
    return password == ADMIN_PASSWORD

# --- FastAPI 应用设置 ---


# 挂载静态文件目录
app.mount("/static", StaticFiles(directory="static"), name="static")
# 设置模板目录
templates = Jinja2Templates(directory="templates")

# --- 数据库会话管理 ---

def get_db():
    """
    数据库会话管理函数
    增加了详细的日志记录来跟踪数据库连接的创建和关闭过程
    """
    print("\n--- 创建数据库会话 ---")
    db = None
    try:
        db = SessionLocal()
        print(f"数据库会话创建成功: {id(db)}")
        yield db
    except Exception as e:
        print(f"数据库会话创建失败: {e}")
        if db:
            print("正在回滚数据库事务...")
            db.rollback()
        raise
    finally:
        if db:
            print(f"正在关闭数据库会话: {id(db)}")
            db.close()
            print("数据库会话已关闭")
        print("--- 数据库会话管理结束 ---\n")

# --- 应用启动事件 ---

@app.on_event("startup")
def on_startup():
    """
    应用启动事件处理函数
    增加了详细的日志记录来跟踪应用启动过程
    """
    print("\n" + "=" * 60)
    print("🚀 动力资源资产管理系统启动中...")
    print("=" * 60)
    
    try:
        # 检查并创建数据库目录
        db_dir = './database'
        if not os.path.exists(db_dir):
            print(f"📁 创建数据库目录: {db_dir}")
            os.makedirs(db_dir)
        else:
            print(f"📁 数据库目录已存在: {db_dir}")
        
        # 初始化数据库
        print("🗄️ 正在初始化数据库...")
        create_db_and_tables()
        
        print("✅ 应用启动完成！")
        print(f"🌐 服务器地址: http://localhost:{PORT}")
        print("=" * 60 + "\n")
        
    except Exception as e:
        print(f"\n❌ 应用启动失败!")
        print(f"错误类型: {type(e).__name__}")
        print(f"错误信息: {e}")
        print("\n完整错误堆栈:")
        traceback.print_exc()
        print("=" * 60)
        raise  # 重新抛出异常，停止应用启动

# --- 路由和视图函数 ---

@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request, db: Session = Depends(get_db)):
    """
    首页路由 - 显示所有设备列表
    增加了详细的日志记录来跟踪数据获取过程
    """
    print("\n=== 首页数据获取开始 ===")
    
    try:
        # 获取设备数据
        print("正在从数据库查询设备数据...")
        devices = db.query(Device).order_by(Device.id).all()
        device_count = len(devices)
        print(f"查询到 {device_count} 个设备")
        
        # 获取生命周期规则
        lifecycle_rules = db.query(LifecycleRule).filter(LifecycleRule.is_active == 'true').all()
        rules_dict = {rule.device_type: rule for rule in lifecycle_rules}
        print(f"加载了 {len(rules_dict)} 个生命周期规则")
        
        # 为每个设备计算生命周期状态
        for device in devices:
            lifecycle_status = "unknown"
            lifecycle_status_text = "未配置规则"
            
            if device.device_type and device.device_type in rules_dict:
                rule = rules_dict[device.device_type]
                if device.commission_date:
                    try:
                        # 解析投产日期
                        commission_date = None
                        date_str = str(device.commission_date).strip()
                        
                        # 处理特殊格式：YYYYMM (如 202312)
                        if re.match(r'^\d{6}$', date_str):
                            try:
                                year = int(date_str[:4])
                                month = int(date_str[4:6])
                                commission_date = datetime(year, month, 1)
                            except ValueError:
                                pass
                        
                        # 如果特殊格式解析失败，尝试标准格式
                        if not commission_date:
                            date_formats = [
                                "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
                                "%Y-%m", "%Y/%m", "%Y.%m",
                                "%Y年%m月%d日", "%Y年%m月"
                            ]
                            
                            for fmt in date_formats:
                                try:
                                    commission_date = datetime.strptime(date_str, fmt)
                                    break
                                except ValueError:
                                    continue
                        
                        if commission_date:
                            # 计算服役时间
                            today = datetime.now()
                            service_years = (today - commission_date).days / 365.25
                            
                            # 判断状态
                            if service_years >= rule.lifecycle_years:
                                lifecycle_status = "expired"
                                lifecycle_status_text = "已超期"
                            elif service_years >= (rule.lifecycle_years - rule.warning_months / 12):
                                lifecycle_status = "warning"
                                lifecycle_status_text = "临近超限"
                            else:
                                lifecycle_status = "normal"
                                lifecycle_status_text = "正常"
                        else:
                            lifecycle_status = "unknown"
                            lifecycle_status_text = "投产日期格式无法识别"
                    except Exception as e:
                        lifecycle_status = "unknown"
                        lifecycle_status_text = "投产日期格式无法识别"
                else:
                    lifecycle_status = "unknown"
                    lifecycle_status_text = "投产日期未填写"
            
            # 将状态信息添加到设备对象
            device.lifecycle_status = lifecycle_status
            device.lifecycle_status_text = lifecycle_status_text
        
        # 显示前几个设备的信息用于调试
        if device_count > 0:
            print("\n前3个设备信息:")
            for i, device in enumerate(devices[:3]):
                print(f"  设备{i+1}: ID={device.id}, 资产编号={device.asset_id}, 名称={device.name}, 生命周期状态={device.lifecycle_status}")
        else:
            print("警告: 数据库中没有设备数据！")
        
        # 获取连接数据用于统计
        connections = db.query(Connection).all()
        connection_count = len(connections)
        print(f"数据库中共有 {connection_count} 个连接")
        
        # 获取所有不重复的局站列表，用于筛选下拉框
        print("正在获取局站列表...")
        stations = db.query(Device.station).filter(Device.station.isnot(None)).filter(Device.station != '').distinct().all()
        station_list = [station[0] for station in stations if station[0]]  # 提取局站名称并过滤空值
        station_list.sort()  # 按字母顺序排序
        print(f"找到 {len(station_list)} 个不同的局站: {station_list}")
        
        # 使用预定义的标准设备类型列表
        print("正在加载标准设备类型列表...")
        device_type_list = sorted(STANDARD_DEVICE_TYPES)
        print(f"加载了 {len(device_type_list)} 个标准设备类型: {device_type_list}")
        
        # 获取所有不重复的厂家列表，用于筛选下拉框
        print("正在获取厂家列表...")
        vendors = db.query(Device.vendor).filter(Device.vendor.isnot(None)).filter(Device.vendor != '').distinct().all()
        vendor_list = [vendor[0] for vendor in vendors if vendor[0]]  # 提取厂家名称并过滤空值
        vendor_list.sort()  # 按字母顺序排序
        print(f"找到 {len(vendor_list)} 个不同的厂家: {vendor_list}")
        
        # 检查是否有上传错误信息
        upload_error = request.query_params.get("error")
        if upload_error:
            print(f"检测到上传错误信息: {upload_error}")
        else:
            print("没有上传错误信息")
        
        # 检查是否有成功信息
        success_message = request.query_params.get("success")
        if success_message:
            print(f"检测到成功信息: {success_message}")
        else:
            print("没有成功信息")
        
        print("=== 首页数据获取完成 ===")
        
        return templates.TemplateResponse("index.html", {
            "request": request, 
            "devices": devices, 
            "stations": station_list,
            "device_types": device_type_list,
            "vendors": vendor_list,
            "upload_error": upload_error,
            "success_message": success_message
        })
        
    except Exception as e:
        print(f"\n!!! 首页数据获取失败 !!!")
        print(f"错误类型: {type(e).__name__}")
        print(f"错误信息: {e}")
        print("\n完整错误堆栈:")
        traceback.print_exc()
        print("=" * 50)
        
        # 返回错误页面或空设备列表
        return templates.TemplateResponse("index.html", {
            "request": request, 
            "devices": [], 
            "stations": [],
            "device_types": [],
            "vendors": [],
            "upload_error": f"获取设备数据时出错: {e}"
        })

@app.post("/upload")
async def upload_excel(file: UploadFile = File(...), password: str = Form(...), db: Session = Depends(get_db)):
    """
    处理 Excel 文件上传。
    如果失败，则重定向回主页并附带详细错误信息。
    增加了详细的日志记录来跟踪处理过程。
    """
    print("\n=== 开始处理上传的Excel文件 ===")
    print(f"上传文件名: {file.filename}")
    print(f"文件类型: {file.content_type}")
    
    # 验证管理员密码
    if not verify_admin_password(password):
        error_message = "密码错误，无权限执行此操作。"
        print(f"权限验证失败: {error_message}")
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
    
    print("管理员密码验证通过")
    
    try:
        # 步骤 1: 增量更新模式 - 保留手工添加的设备，只更新Excel中的设备
        print("\n步骤 1: 采用增量更新模式，保留现有手工添加的设备...")
        
        # 记录当前数据量
        current_connections_count = db.query(Connection).count()
        current_devices_count = db.query(Device).count()
        print(f"当前数据库状态: {current_connections_count} 个连接, {current_devices_count} 个设备")
        print("步骤 1: 完成。将采用增量更新模式处理Excel数据。")

        contents = await file.read()
        print(f"文件大小: {len(contents)} 字节")
        buffer = io.BytesIO(contents)
        
        # 步骤 2: 读取Excel文件
        print("\n步骤 2: 使用 pandas 读取Excel文件...")
        # 通过 dtype 参数指定列以字符串形式读取，避免自动转换格式
        # 重要：假设"上级设备"列现在包含的是父设备的资产编号
        df = pd.read_excel(buffer, dtype={
            '资产编号': str,
            '设备投产时间': str,
            '上级设备': str 
        })
        df = df.where(pd.notna(df), None) # 将 NaN 替换为 None
        print(f"步骤 2: 完成。读取到 {len(df)} 行数据。")
        print(f"Excel 文件列名: {df.columns.tolist()}")
        
        # 验证必要的列是否存在
        required_columns = ['资产编号', '设备名称']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            error_msg = f"Excel文件缺少必要的列: {missing_columns}"
            print(f"错误: {error_msg}")
            return RedirectResponse(url=f"/?error={quote(error_msg)}", status_code=303)
        
        # 显示前几行数据样本用于调试
        print("\n前3行数据样本:")
        for i in range(min(3, len(df))):
            print(f"第{i+1}行: 资产编号={df.iloc[i].get('资产编号')}, 设备名称={df.iloc[i].get('设备名称')}")

        devices_map = {} # 这个映射将以 资产编号 为键
        devices_created_count = 0
        devices_updated_count = 0
        skipped_rows = []

        # 步骤 3: 增量更新设备（创建或更新）
        print("\n步骤 3: 开始第一遍处理 - 增量更新设备（创建新设备或更新现有设备）...")
        for index, row in df.iterrows():
            # 新增：获取并校验资产编号
            asset_id = row.get("资产编号")
            if isinstance(asset_id, str):
                asset_id = asset_id.strip()

            if not asset_id or asset_id == 'nan' or asset_id.lower() == 'none':
                skip_reason = f"资产编号为空或无效: '{asset_id}'"
                print(f"  - 第 {index+2} 行：跳过，{skip_reason}")
                skipped_rows.append((index+2, skip_reason))
                continue
            
            device_name = row.get("设备名称")
            if isinstance(device_name, str):
                device_name = device_name.strip()

            if not device_name or device_name == 'nan' or device_name.lower() == 'none':
                skip_reason = f"设备名称为空或无效: '{device_name}'"
                print(f"  - 第 {index+2} 行：跳过，{skip_reason}")
                skipped_rows.append((index+2, skip_reason))
                continue
            
            # 检查资产编号是否已在本次上传中重复
            if asset_id in devices_map:
                skip_reason = f"资产编号 '{asset_id}' 在Excel文件中重复"
                print(f"  - 第 {index+2} 行：跳过，{skip_reason}")
                skipped_rows.append((index+2, skip_reason))
                continue

            try:
                # 检查数据库中是否已存在该资产编号的设备
                existing_device = db.query(Device).filter(Device.asset_id == asset_id).first()
                
                # 获取局站信息
                station = row.get("局站")
                if isinstance(station, str):
                    station = station.strip()
                if not station or station == 'nan' or station.lower() == 'none':
                    skip_reason = f"局站信息为空或无效: '{station}'"
                    print(f"  - 第 {index+2} 行：跳过，{skip_reason}")
                    skipped_rows.append((index+2, skip_reason))
                    continue
                
                # 获取并验证设备类型
                device_type = row.get("设备类型")
                if isinstance(device_type, str):
                    device_type = device_type.strip()
                
                # 验证设备类型是否在标准列表中
                if device_type and device_type != 'nan' and device_type.lower() != 'none':
                    if not validate_device_type(device_type):
                        # 提供建议的设备类型
                        suggestions = get_device_type_suggestions(device_type)
                        if suggestions:
                            suggestion_text = f"，建议使用: {', '.join(suggestions[:3])}"
                        else:
                            suggestion_text = ""
                        skip_reason = f"设备类型 '{device_type}' 不在标准列表中{suggestion_text}"
                        print(f"  - 第 {index+2} 行：跳过，{skip_reason}")
                        skipped_rows.append((index+2, skip_reason))
                        continue
                else:
                    # 如果设备类型为空，设置为"待确认"
                    device_type = "待确认"
                
                if existing_device:
                    # 更新现有设备
                    existing_device.name = device_name
                    existing_device.station = station
                    existing_device.model = row.get("设备型号")
                    existing_device.device_type = device_type  # 使用验证后的设备类型
                    existing_device.location = row.get("机房内空间位置")
                    existing_device.power_rating = row.get("设备额定容量")
                    existing_device.vendor = row.get("设备生产厂家")
                    existing_device.commission_date = row.get("设备投产时间")
                    existing_device.remark = row.get("备注")
                    
                    # 注意：以下机房相关字段被忽略（根据用户要求）：
                    # - 机房名称
                    # - 资源系统机房名称  
                    # - 资源系统机房编码
                    # - 机房等级
                    
                    devices_map[asset_id] = existing_device
                    devices_updated_count += 1
                    print(f"  - 第 {index+2} 行：准备更新现有设备 '{device_name}' (资产编号: {asset_id}, 局站: {station})")
                else:
                    # 创建新设备
                    device = Device(
                        asset_id=asset_id,
                        name=device_name,
                        station=station,
                        model=row.get("设备型号"),
                        device_type=device_type,  # 使用验证后的设备类型
                        location=row.get("机房内空间位置"),
                        power_rating=row.get("设备额定容量"),
                        vendor=row.get("设备生产厂家"),
                        commission_date=row.get("设备投产时间"),
                        remark=row.get("备注")
                        # 注意：以下机房相关字段被忽略（根据用户要求）：
                        # - 机房名称
                        # - 资源系统机房名称  
                        # - 资源系统机房编码
                        # - 机房等级
                    )
                    db.add(device)
                    devices_map[asset_id] = device
                    devices_created_count += 1
                    print(f"  - 第 {index+2} 行：准备创建新设备 '{device_name}' (资产编号: {asset_id}, 局站: {station})")
                    
            except Exception as device_error:
                skip_reason = f"处理设备失败: {device_error}"
                print(f"  - 第 {index+2} 行：跳过，{skip_reason}")
                skipped_rows.append((index+2, skip_reason))
                continue
        
        print(f"\n准备提交设备更改到数据库（新建: {devices_created_count}, 更新: {devices_updated_count}）...")
        try:
            db.commit() # 提交事务以生成设备ID
            print("设备提交成功！")
        except Exception as commit_error:
            print(f"设备提交失败: {commit_error}")
            db.rollback()
            raise commit_error
            
        # 验证设备数量
        actual_device_count = db.query(Device).count()
        print(f"步骤 3: 完成。新建 {devices_created_count} 个设备，更新 {devices_updated_count} 个设备，数据库中总共有 {actual_device_count} 个设备。")
        
        if skipped_rows:
            print(f"\n跳过的行数统计: {len(skipped_rows)} 行")
            for row_num, reason in skipped_rows[:5]:  # 只显示前5个
                print(f"  第{row_num}行: {reason}")
            if len(skipped_rows) > 5:
                print(f"  ... 还有 {len(skipped_rows) - 5} 行被跳过")

        # 刷新映射，确保对象包含数据库生成的ID
        print("\n刷新设备对象以获取数据库生成的ID...")
        for asset_id_key in list(devices_map.keys()):
            try:
                db.refresh(devices_map[asset_id_key])
                print(f"  设备 {asset_id_key} ID: {devices_map[asset_id_key].id}")
            except Exception as refresh_error:
                print(f"  刷新设备 {asset_id_key} 失败: {refresh_error}")

        # 步骤 4: 清理涉及Excel设备的旧连接
        print("\n步骤 4: 清理涉及Excel中设备的旧连接...")
        excel_device_ids = [device.id for device in devices_map.values()]
        if excel_device_ids:
            # 删除涉及这些设备的所有连接（作为源设备或目标设备）
            old_connections_deleted = db.query(Connection).filter(
                (Connection.source_device_id.in_(excel_device_ids)) |
                (Connection.target_device_id.in_(excel_device_ids))
            ).delete(synchronize_session=False)
            db.commit()
            print(f"删除了 {old_connections_deleted} 个涉及Excel设备的旧连接")
        else:
            print("没有Excel设备，跳过连接清理")
            
        connections_created_count = 0
        connection_skipped_rows = []
        
        # 步骤 5: 创建新连接
        print("\n步骤 5: 开始第二遍处理 - 创建新连接...")
        for index, row in df.iterrows():
            # 使用资产编号来查找设备
            source_asset_id = row.get("上级设备")
            target_asset_id = row.get("资产编号")

            if isinstance(source_asset_id, str):
                source_asset_id = source_asset_id.strip()
            if isinstance(target_asset_id, str):
                target_asset_id = target_asset_id.strip()
            
            # 检查是否有上级设备信息
            if not source_asset_id or source_asset_id == 'nan' or source_asset_id.lower() == 'none':
                print(f"  - 第 {index+2} 行：跳过连接创建，无上级设备信息")
                continue
                
            # 确保源和目标设备都存在于映射中
            if target_asset_id and source_asset_id:
                if source_asset_id not in devices_map:
                    skip_reason = f"上级设备 '{source_asset_id}' 不存在"
                    print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                    connection_skipped_rows.append((index+2, skip_reason))
                    continue
                    
                if target_asset_id not in devices_map:
                    skip_reason = f"目标设备 '{target_asset_id}' 不存在"
                    print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                    connection_skipped_rows.append((index+2, skip_reason))
                    continue
                
                source_device = devices_map[source_asset_id]
                target_device = devices_map[target_asset_id]
                
                try:
                    connection = Connection(
                        source_device_id=source_device.id,
                        source_port=row.get("上级端口"),
                        target_device_id=target_device.id,
                        target_port=row.get("本端端口"),
                        cable_type=row.get("线缆类型")
                    )
                    db.add(connection)
                    connections_created_count += 1
                    print(f"  - 第 {index+2} 行：准备创建从 '{source_device.name}' 到 '{target_device.name}' 的连接")
                except Exception as conn_error:
                    skip_reason = f"创建连接对象失败: {conn_error}"
                    print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                    connection_skipped_rows.append((index+2, skip_reason))
                    continue
        
        print(f"\n准备提交 {connections_created_count} 个连接到数据库...")
        try:
            db.commit()
            print("连接提交成功！")
        except Exception as commit_error:
            print(f"连接提交失败: {commit_error}")
            db.rollback()
            raise commit_error
            
        # 验证连接是否真的被创建
        actual_connection_count = db.query(Connection).count()
        print(f"步骤 5: 完成。预期创建 {connections_created_count} 个连接，实际数据库中有 {actual_connection_count} 个连接。")
        
        if connection_skipped_rows:
            print(f"\n连接跳过的行数统计: {len(connection_skipped_rows)} 行")
            for row_num, reason in connection_skipped_rows[:5]:  # 只显示前5个
                print(f"  第{row_num}行: {reason}")
            if len(connection_skipped_rows) > 5:
                print(f"  ... 还有 {len(connection_skipped_rows) - 5} 行连接被跳过")
        
        # 步骤 6: 处理Sheet2连接数据
        sheet2_connections_count = 0
        sheet2_skipped_rows = []
        
        try:
            print("\n步骤 6: 开始处理Sheet2连接数据...")
            
            # 尝试读取Sheet2（连接表）
            try:
                # 重置buffer位置到开头，因为之前读取Sheet1时已经移动了位置
                buffer.seek(0)
                df_connections = pd.read_excel(buffer, sheet_name='连接')
                print(f"成功读取Sheet2，共 {len(df_connections)} 行连接数据")
            except Exception as sheet_error:
                print(f"无法读取Sheet2（连接表）: {sheet_error}")
                print("跳过Sheet2处理，继续完成导入")
                df_connections = None
            
            if df_connections is not None and len(df_connections) > 0:
                # 连接类型映射 - 扩展映射表以包含更多可能的空值表示
                CONNECTION_TYPE_MAPPING = {
                    # 标准连接类型
                    '电缆': 'cable',
                    '铜排': 'busbar', 
                    '母线': 'busway',
                    'cable': 'cable',
                    'busbar': 'busbar',
                    'busway': 'busway',
                    # 电气连接类型 - 根据实际Excel数据添加
                    '直流': 'DC',
                    '交流': 'AC',
                    'DC': 'DC',
                    'AC': 'AC',
                    'dc': 'DC',
                    'ac': 'AC',
                    # 空值的各种表示方式 - 统一映射为None表示空闲端口
                    '无': None,
                    '空': None,
                    '空闲': None,
                    '未连接': None,
                    'N/A': None,
                    'n/a': None,
                    'NA': None,
                    'na': None,
                    '无连接': None,
                    '待连接': None,
                    '预留': None,
                    'None': None,
                    'null': None,
                    'NULL': None,
                    '': None,  # 空字符串
                    ' ': None,  # 空格
                }
                
                # 辅助函数：获取或创建设备
                def get_or_create_device(device_name: str, default_station: str = "未知站点"):
                    """获取设备，如果不存在则自动创建"""
                    if not device_name:
                        return None
                    
                    device = db.query(Device).filter(Device.name == device_name).first()
                    if not device:
                        # 自动创建设备
                        device = Device(
                            name=device_name,
                            asset_id=f"AUTO_{len(device_name)}_{hash(device_name) % 10000:04d}",  # 生成唯一资产编号
                            station=default_station,
                            device_type="待确认",
                            location="待确认",
                            remark="通过Excel导入时自动创建，请完善设备信息"
                        )
                        db.add(device)
                        db.flush()  # 获取ID但不提交
                        print(f"  * 自动创建设备: {device_name} (ID: {device.id})")
                    return device
                
                # 统计信息
                created_devices = []
                warnings = []
                
                for index, row in df_connections.iterrows():
                    try:
                        # 获取设备名称
                        source_device_name = str(row.get('A端设备名称', '')).strip()
                        target_device_name = str(row.get('B端设备名称', '')).strip()
                        
                        # 处理空设备名称的情况
                        if not source_device_name and not target_device_name:
                            skip_reason = "A端和B端设备名称都为空"
                            print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                            sheet2_skipped_rows.append((index+2, skip_reason))
                            continue
                        elif not source_device_name:
                            skip_reason = "A端设备名称为空"
                            print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                            sheet2_skipped_rows.append((index+2, skip_reason))
                            continue
                        elif not target_device_name:
                            skip_reason = "B端设备名称为空"
                            print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                            sheet2_skipped_rows.append((index+2, skip_reason))
                            continue
                        
                        # 获取或创建设备
                        source_device = get_or_create_device(source_device_name)
                        target_device = get_or_create_device(target_device_name)
                        
                        if not source_device or not target_device:
                            skip_reason = "设备创建失败"
                            print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                            sheet2_skipped_rows.append((index+2, skip_reason))
                            continue
                        
                        # 记录新创建的设备
                        if source_device.remark and "通过Excel导入时自动创建" in source_device.remark:
                            if source_device_name not in created_devices:
                                created_devices.append(source_device_name)
                        if target_device.remark and "通过Excel导入时自动创建" in target_device.remark:
                            if target_device_name not in created_devices:
                                created_devices.append(target_device_name)
                        
                        # 处理端口逻辑
                        def build_port_info(fuse_number, fuse_spec, breaker_number, breaker_spec):
                            """构建端口信息，优先使用熔丝，其次使用空开"""
                            fuse_num = str(fuse_number).strip() if pd.notna(fuse_number) else ''
                            fuse_sp = str(fuse_spec).strip() if pd.notna(fuse_spec) else ''
                            breaker_num = str(breaker_number).strip() if pd.notna(breaker_number) else ''
                            breaker_sp = str(breaker_spec).strip() if pd.notna(breaker_spec) else ''
                            
                            if fuse_num and fuse_num != 'nan':
                                return f"{fuse_num} ({fuse_sp})" if fuse_sp and fuse_sp != 'nan' else fuse_num
                            elif breaker_num and breaker_num != 'nan':
                                return f"{breaker_num} ({breaker_sp})" if breaker_sp and breaker_sp != 'nan' else breaker_num
                            else:
                                return None
                        
                        # 构建A端和B端端口信息
                        source_port = build_port_info(
                            row.get('A端熔丝编号'), row.get('A端熔丝规格'),
                            row.get('A端空开编号'), row.get('A端空开规格')
                        )
                        target_port = build_port_info(
                            row.get('B端熔丝编号'), row.get('B端熔丝规格'),
                            row.get('B端空开编号'), row.get('空开规格')
                        )
                        
                        # 处理连接类型 - 修复空闲端口被错误归类为电缆的问题
                        connection_type_raw = row.get('连接类型（交流/直流）')  # 修正：使用实际Excel列名
                        # 规则：空值 -> None；非空则按映射表转换，无法映射也置 None 并记录警告
                        connection_type = None
                        if pd.isna(connection_type_raw) or str(connection_type_raw).strip() == '':
                            connection_type = None
                        else:
                            raw = str(connection_type_raw).strip()
                            connection_type = CONNECTION_TYPE_MAPPING.get(raw, None)
                            if raw not in CONNECTION_TYPE_MAPPING:
                                print(f"  * 警告：第 {index+2} 行连接类型 '{raw}' 无法识别，设置为空闲端口")
                                warnings.append(f"第 {index+2} 行：连接类型 '{raw}' 无法识别")

                        # 检查是否已存在相同连接（按设备与端口维度去重）
                        existing_connection = db.query(Connection).filter(
                            Connection.source_device_id == source_device.id,
                            Connection.target_device_id == target_device.id,
                            Connection.source_port == source_port,
                            Connection.target_port == target_port
                        ).first()

                        if existing_connection:
                            skip_reason = "连接已存在"
                            print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                            sheet2_skipped_rows.append((index+2, skip_reason))
                            continue

                        # 创建连接记录（保留现有字段，未设置 cable_type，避免误用）
                        connection = Connection(
                            source_device_id=source_device.id,
                            target_device_id=target_device.id,
                            source_port=source_port,
                            target_port=target_port,
                            # A端信息
                            source_fuse_number=str(row.get('A端熔丝编号', '')).strip() if pd.notna(row.get('A端熔丝编号')) else None,
                            source_fuse_spec=str(row.get('A端熔丝规格', '')).strip() if pd.notna(row.get('A端熔丝规格')) else None,
                            source_breaker_number=str(row.get('A端空开编号', '')).strip() if pd.notna(row.get('A端空开编号')) else None,
                            source_breaker_spec=str(row.get('A端空开规格', '')).strip() if pd.notna(row.get('A端空开规格')) else None,
                            # B端信息
                            target_fuse_number=str(row.get('B端熔丝编号', '')).strip() if pd.notna(row.get('B端熔丝编号')) else None,
                            target_fuse_spec=str(row.get('B端熔丝规格', '')).strip() if pd.notna(row.get('B端熔丝规格')) else None,
                            target_breaker_number=str(row.get('B端空开编号', '')).strip() if pd.notna(row.get('B端空开编号')) else None,
                            target_breaker_spec=str(row.get('空开规格', '')).strip() if pd.notna(row.get('空开规格')) else None,
                            target_device_location=str(row.get('B端设备位置（非动力设备）', '')).strip() if pd.notna(row.get('B端设备位置（非动力设备）')) else None,
                            # 额定电流信息
                            a_rated_current=str(row.get('A端额定电流', '')).strip() if pd.notna(row.get('A端额定电流')) else None,
                            b_rated_current=str(row.get('B端额定电流', '')).strip() if pd.notna(row.get('B端额定电流')) else None,
                            # 连接信息
                            hierarchy_relation=str(row.get('上下级', '')).strip() if pd.notna(row.get('上下级')) else None,
                            upstream_downstream=str(row.get('上下游', '')).strip() if pd.notna(row.get('上下游')) else None,
                            connection_type=connection_type,
                            cable_model=str(row.get('电缆型号', '')).strip() if pd.notna(row.get('电缆型号')) else None,
                            # 附加信息
                            source_device_photo=str(row.get('A端设备照片', '')).strip() if pd.notna(row.get('A端设备照片')) else None,
                            target_device_photo=str(row.get('B端设备照片', '')).strip() if pd.notna(row.get('B端设备照片')) else None,
                            remark=str(row.get('备注', '')).strip() if pd.notna(row.get('备注')) else None,
                            # 安装日期（Excel中没有此字段，设置为None）
                            installation_date=None
                        )

                        db.add(connection)
                        sheet2_connections_count += 1
                        print(f"  - 第 {index+2} 行：准备创建从 '{source_device_name}' 到 '{target_device_name}' 的连接")
                        print(f"    源端口: {source_port}, 目标端口: {target_port}, 连接类型: {connection_type}")
                        
                    except Exception as conn_error:
                        skip_reason = f"处理连接失败: {conn_error}"
                        print(f"  - 第 {index+2} 行：跳过连接，{skip_reason}")
                        sheet2_skipped_rows.append((index+2, skip_reason))
                        continue
                
                # 提交Sheet2连接
                if sheet2_connections_count > 0:
                    print(f"\n准备提交 {sheet2_connections_count} 个Sheet2连接到数据库...")
                    try:
                        db.commit()
                        print("Sheet2连接提交成功！")
                    except Exception as commit_error:
                        print(f"Sheet2连接提交失败: {commit_error}")
                        db.rollback()
                        raise commit_error
                
                # 生成详细的导入报告
                print(f"\n=== Sheet2连接导入报告 ===")
                print(f"总连接数: {len(df_connections)} 行")
                print(f"成功导入: {sheet2_connections_count} 个连接")
                print(f"跳过连接: {len(sheet2_skipped_rows)} 行")
                
                if created_devices:
                    print(f"\n自动创建的设备 ({len(created_devices)} 个):")
                    for device_name in created_devices:
                        print(f"  + {device_name}")
                    print("\n注意: 自动创建的设备信息不完整，请在设备管理页面完善相关信息。")
                
                if sheet2_skipped_rows:
                    print(f"\n跳过的连接详情:")
                    skip_reasons = {}
                    for row_num, reason in sheet2_skipped_rows:
                        if reason not in skip_reasons:
                            skip_reasons[reason] = []
                        skip_reasons[reason].append(row_num)
                    
                    for reason, rows in skip_reasons.items():
                        print(f"  {reason}: {len(rows)} 行 (第{', '.join(map(str, rows[:3]))}行{'...' if len(rows) > 3 else ''})")
                
                # 计算导入成功率
                success_rate = (sheet2_connections_count / len(df_connections)) * 100 if len(df_connections) > 0 else 0
                print(f"\n导入成功率: {success_rate:.1f}% ({sheet2_connections_count}/{len(df_connections)})")
            
            print(f"步骤 6: 完成。从Sheet2创建了 {sheet2_connections_count} 个连接")
            
        except Exception as sheet2_error:
            print(f"处理Sheet2时出错: {sheet2_error}")
            print("继续完成导入，忽略Sheet2错误")
        
        # 最终统计
        final_connection_count = db.query(Connection).count()
        total_connections_created = connections_created_count + sheet2_connections_count
        
        print("\n=== Excel文件增量更新处理成功 ===")
        print(f"处理结果: 新建 {devices_created_count} 个设备, 更新 {devices_updated_count} 个设备")
        print(f"连接创建: Sheet1创建 {connections_created_count} 个, Sheet2创建 {sheet2_connections_count} 个, 总计 {total_connections_created} 个")
        print(f"数据库最终状态: {actual_device_count} 个设备, {final_connection_count} 个连接")

    except Exception as e:
        print(f"\n!!! 发生异常，开始回滚事务 !!!")
        try:
            db.rollback()
            print("事务回滚成功")
        except Exception as rollback_error:
            print(f"事务回滚失败: {rollback_error}")
            
        error_message = f"处理Excel文件时出错: {e}"
        print(f"\n=== Excel文件处理失败 ===")
        print(f"错误类型: {type(e).__name__}")
        print(f"错误信息: {error_message}")
        print("\n完整错误堆栈:")
        traceback.print_exc()
        print("=" * 50)
        
        # 检查数据库状态
        try:
            final_device_count = db.query(Device).count()
            final_connection_count = db.query(Connection).count()
            print(f"\n错误后数据库状态: {final_device_count} 个设备, {final_connection_count} 个连接")
        except Exception as db_check_error:
            print(f"无法检查数据库状态: {db_check_error}")
            
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)

    print(f"\n上传处理完成，重定向到首页...")
    return RedirectResponse(url="/", status_code=303)

# 更新设备信息
@app.post("/devices/{device_id}")
async def update_device(
    device_id: int,
    asset_id: str = Form(...),
    name: str = Form(...),
    station: str = Form(...),
    model: str = Form(None),
    device_type: str = Form(None),
    location: str = Form(None),
    power_rating: str = Form(None),
    vendor: str = Form(None),
    commission_date: str = Form(None),
    remark: str = Form(None),
    db: Session = Depends(get_db)
):
    """更新设备信息（编辑功能不需要密码验证，因为在进入编辑页面时已验证）"""
    try:
        # 获取要更新的设备
        device = db.query(Device).filter(Device.id == device_id).first()
        if not device:
            error_message = "设备不存在。"
            return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
        
        # 检查资产编号唯一性（排除当前设备）
        existing_device = db.query(Device).filter(
            Device.asset_id == asset_id,
            Device.id != device_id
        ).first()
        if existing_device:
            error_message = f"资产编号 {asset_id} 已存在，请使用其他编号。"
            return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
        
        # 更新设备信息
        device.asset_id = asset_id
        device.name = name
        device.station = station
        device.model = model if model else None
        device.device_type = device_type if device_type else None
        device.location = location if location else None
        device.power_rating = power_rating if power_rating else None
        device.vendor = vendor if vendor else None
        device.commission_date = commission_date if commission_date else None
        device.remark = remark if remark else None
        
        db.commit()
        
        success_message = f"设备 {name} 更新成功。"
        return RedirectResponse(url=f"/?success={quote(success_message)}", status_code=303)
        
    except Exception as e:
        db.rollback()
        error_message = f"更新设备失败：{str(e)}"
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)

# 编辑设备页面
@app.get("/edit/{device_id}")
async def edit_device_page(device_id: int, password: str, request: Request, db: Session = Depends(get_db)):
    """显示编辑设备页面"""
    # 验证管理员密码
    if not verify_admin_password(password):
        error_message = "密码错误，无权限执行此操作。"
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
    
    # 获取设备信息
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        error_message = "设备不存在。"
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
    
    return templates.TemplateResponse("edit_device.html", {
        "request": request,
        "device": device
    })

# 删除设备
@app.delete("/devices/{device_id}")
async def delete_device(device_id: int, request: Request, db: Session = Depends(get_db)):
    """删除设备"""
    try:
        # 获取请求体中的密码
        body = await request.json()
        password = body.get("password")
        
        # 验证管理员密码
        if not verify_admin_password(password):
            raise HTTPException(status_code=403, detail="密码错误，无权限执行此操作。")
        
        # 获取要删除的设备
        device = db.query(Device).filter(Device.id == device_id).first()
        if not device:
            raise HTTPException(status_code=404, detail="设备不存在。")
        
        device_name = device.name
        
        # 删除相关的连接记录
        db.query(Connection).filter(
            (Connection.source_device_id == device_id) | 
            (Connection.target_device_id == device_id)
        ).delete()
        
        # 删除设备
        db.delete(device)
        db.commit()
        
        return {"message": f"设备 {device_name} 删除成功。"}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"删除设备失败：{str(e)}")

@app.post("/devices")
async def create_device(
    asset_id: str = Form(...),
    name: str = Form(...),
    station: str = Form(...),
    model: str = Form(None),
    device_type: str = Form(None),
    location: str = Form(None),
    power_rating: str = Form(None),
    vendor: str = Form(None),
    commission_date: str = Form(None),
    remark: str = Form(None),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    # 验证管理员密码
    if not verify_admin_password(password):
        error_message = "密码错误，无权限执行此操作。"
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
    
    # 增加资产编号唯一性校验
    existing_device = db.query(Device).filter(Device.asset_id == asset_id).first()
    if existing_device:
        # 如果存在，则重定向回主页并显示错误信息
        error_message = f"创建失败：资产编号 '{asset_id}' 已存在。"
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)

    new_device = Device(
        asset_id=asset_id,
        name=name,
        station=station,
        location=location,
        power_rating=power_rating,
        vendor=vendor,
        commission_date=commission_date,
        remark=remark
    )
    db.add(new_device)
    db.commit()
    return RedirectResponse(url="/", status_code=303)

def _adjust_color_brightness(hex_color: str, factor: float) -> str:
    """调整颜色亮度"""
    try:
        # 移除 # 符号
        hex_color = hex_color.lstrip('#')
        
        # 转换为 RGB
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        
        # 调整亮度
        r = min(255, max(0, int(r * factor)))
        g = min(255, max(0, int(g * factor)))
        b = min(255, max(0, int(b * factor)))
        
        # 转换回十六进制
        return f"#{r:02x}{g:02x}{b:02x}"
    except:
        return hex_color  # 如果转换失败，返回原色


def _get_device_lifecycle_status(device: Device, db: Session) -> str:
    """计算设备的生命周期状态 - 复用已有的完整实现逻辑"""
    try:
        from datetime import datetime
        import re
        
        # 查找对应的生命周期规则
        rule = db.query(LifecycleRule).filter(
            LifecycleRule.device_type == device.device_type,
            LifecycleRule.is_active == "true"
        ).first()
        
        if not rule:
            return "未配置规则"
        
        # 解析投产日期
        if not device.commission_date:
            return "投产日期未知"
            
        try:
            # 计算设备年龄（年）
            age_years = (datetime.now().date() - device.commission_date).days / 365.25
            
            # 根据规则判断状态
            if age_years < rule.suggested_scrap_age * 0.7:
                return "在用"
            elif age_years < rule.suggested_scrap_age:
                return "即将报废"
            else:
                return "已超期"
        except Exception as e:
            print(f"计算设备年龄失败: {str(e)}")
            return "计算失败"
    except Exception as e:
        print(f"计算生命周期状态失败: {str(e)}")
        return "未知"





def _get_device_lifecycle_status(device: Device, db: Session) -> str:
    """计算设备的生命周期状态 - 复用已有的完整实现逻辑"""
    try:
        from datetime import datetime
        import re
        
        # 查找对应的生命周期规则
        rule = db.query(LifecycleRule).filter(
            LifecycleRule.device_type == device.device_type,
            LifecycleRule.is_active == "true"
        ).first()
        
        # 如果没有找到规则，返回未知
        if not rule:
            return "未知"
            
        # 计算设备年龄
        if not device.commission_date:
            return "未知"
            
        age_years = (datetime.now().date() - device.commission_date).days / 365.25
        
        # 根据规则判断状态
        if age_years < rule.suggested_scrap_age * 0.7:
            return "在用"
        elif age_years < rule.suggested_scrap_age:
            return "即将报废"
        else:
            return "已超期"
            
    except Exception as e:
        print(f"计算生命周期状态失败: {str(e)}")
        return "未知"


# 新增API路径：/api/power-chain/{device_id} - 与/graph_data/{device_id}功能相同，保持向后兼容


# 新增API路径：/api/power-chain/{device_id} - 与/graph_data/{device_id}功能相同，保持向后兼容
@app.get("/api/power-chain/{device_id}")
async def get_graph_data(
    device_id: int,
    level: str = Query("device", regex="^(device|port)$", description="显示级别：device=设备级，port=端口级"),
    layout_type: str = Query("standard", regex="^(standard|bus)$", description="布局类型：standard=标准布局，bus=总线式布局"),
    connection_type: str = Query(None, description="连接类型筛选条件"),
    station: str = Query(None, description="站点筛选条件"),
    device_type: str = Query(None, description="设备类型筛选条件"),
    show_critical_only: bool = Query(False, description="是否只显示关键设备"),
    only_selected_device: bool = Query(False, description="是否只显示选中设备"),
    group_size: int = Query(1, description="总线式布局时每组端口数量"),
    db: Session = Depends(get_db)
):
    """获取拓扑图数据，按照用户需求格式化"""
    logger.info(f"/graph_data called: device_id={device_id}, level={level}, layout_type={layout_type}, connection_type={connection_type}, station={station}, device_type={device_type}, show_critical_only={show_critical_only}, only_selected_device={only_selected_device}, group_size={group_size}")
    
    # 查找选中的设备
    selected_device = db.query(Device).filter(Device.id == device_id).first()
    if not selected_device:
        raise HTTPException(status_code=404, detail="Device not found")
    
    nodes = []
    edges = []
    processed_device_ids = set()
    
    # 添加选中设备节点
    lifecycle_status = _get_device_lifecycle_status(selected_device, db)
    
    # 修复设备名称显示格式：按照设计规范只显示设备名称
    device_label = selected_device.name
    
    # 记录设备名称格式化日志
    topology_error_tracker.log_error(
        category=ErrorCategory.NODE_RENDERING,
        level=ErrorLevel.INFO,
        message=f"设备节点标签格式化完成",
        context={
            "device_id": selected_device.id,
            "device_name": selected_device.name,
            "station": selected_device.station,
            "formatted_label": device_label,
            "original_format": selected_device.name
        }
    )
    
    node_data = {
        "id": selected_device.id,
        "label": device_label,
        "title": f"""资产编号: {selected_device.asset_id}\n名称: {selected_device.name}\n设备类型: {selected_device.device_type or 'N/A'}\n站点: {selected_device.station or 'N/A'}\n型号: {selected_device.model or 'N/A'}\n位置: {selected_device.location or 'N/A'}\n额定容量: {selected_device.power_rating or 'N/A'}\n生产厂家: {selected_device.vendor or 'N/A'}\n投产时间: {selected_device.commission_date or 'N/A'}\n生命周期状态: {lifecycle_status}""",
        "level": 0,  # 选中设备在第一层
        "device_type": selected_device.device_type,
        "station": selected_device.station
    }
    nodes.append(node_data)
    processed_device_ids.add(selected_device.id)

    # 获取与选中设备直接连接的所有连接
    direct_connections = list(selected_device.target_connections) + list(selected_device.source_connections)
    
    # 处理每个直接连接
    for conn in direct_connections:
        # 应用连接筛选条件
        if not _should_include_connection(conn, connection_type):
            continue
            
        # 跳过无效连接（连接类型为None或空的连接）
        if not conn.connection_type or conn.connection_type == "None":
            continue
            
        # 确定连接的对端设备
        if conn.source_device_id == selected_device.id:
            # 选中设备是源设备，对端是目标设备
            connected_device = conn.target_device
            is_selected_source = True
        else:
            # 选中设备是目标设备，对端是源设备
            connected_device = conn.source_device
            is_selected_source = False
            
        # 检查对端设备是否存在且符合筛选条件
        if not connected_device or not _should_include_device(connected_device, station, device_type, show_critical_only):
            continue
            
        # 避免重复添加设备节点
        if connected_device.id not in processed_device_ids:
            if level == "port":
                # 端口级显示：为连接设备创建端口节点
                if layout_type == "bus":
                    # 总线式布局
                    if not only_selected_device:
                        # 双端视图：为对端设备也创建总线与端口节点
                        bus_data = _create_bus_topology_nodes(connected_device, db, group_size=group_size)
                        for node in bus_data['bus_nodes']:
                            nodes.append(node)
                        for node in bus_data['port_nodes']:
                            nodes.append(node)
                        edges.extend(bus_data['bus_port_edges'])
                    else:
                        # 仅A端视图：不为对端设备创建节点，由选中设备侧的“对端合并端口”承担展示
                        pass
                else:
                    # 标准布局：端口节点已在前面统一创建，此处不再重复创建
                    pass
            else:
                # 设备级显示：为连接设备创建节点
                lifecycle_status = _get_device_lifecycle_status(connected_device, db)
                
                # 修复连接设备名称显示格式：只显示设备名称
                connected_device_label = connected_device.name
                
                node_data = {
                    "id": connected_device.id,
                    "label": connected_device_label,
                    "title": f"""资产编号: {connected_device.asset_id}\n名称: {connected_device.name}\n设备类型: {connected_device.device_type or 'N/A'}\n站点: {connected_device.station or 'N/A'}\n型号: {connected_device.model or 'N/A'}\n位置: {connected_device.location or 'N/A'}\n额定容量: {connected_device.power_rating or 'N/A'}\n生产厂家: {connected_device.vendor or 'N/A'}\n投产时间: {connected_device.commission_date or 'N/A'}\n生命周期状态: {lifecycle_status}""",
                    "level": 1,  # 连接设备在第二层
                    "device_type": connected_device.device_type,
                    "station": connected_device.station
                }
                nodes.append(node_data)
            processed_device_ids.add(connected_device.id)
            
        # 创建连接边，根据上下游字段确定箭头方向
        if level == "port":
            # 端口级连接
            if layout_type == "bus":
                # 总线式布局：创建端口到端口的连接
                direction = "upstream" if is_selected_source else "downstream"
                bus_port_edges = _create_bus_port_edges(conn, direction)
                edges.extend(bus_port_edges)
            else:
                # 标准布局：创建标准端口连接
                direction = "upstream" if is_selected_source else "downstream"
                port_edges = _create_port_edges(conn, direction)
                edges.extend(port_edges)
        else:
            # 设备级连接：根据上下游字段确定箭头方向
            if conn.upstream_downstream == "上游":
                # 上游表示电流从source流向target
                edge_data = {
                    "from": conn.source_device_id,
                    "to": conn.target_device_id,
                    "arrows": "to",
                    "label": conn.connection_type or conn.cable_type or "",
                    "connection_type": conn.connection_type,
                    "cable_type": conn.cable_type,
                    "cable_model": conn.cable_model,
                    "remark": conn.remark,
                    "connection_id": conn.id
                }
            elif conn.upstream_downstream == "下游":
                # 下游表示电流从target流向source
                edge_data = {
                    "from": conn.target_device_id,
                    "to": conn.source_device_id,
                    "arrows": "to",
                    "label": conn.connection_type or conn.cable_type or "",
                    "connection_type": conn.connection_type,
                    "cable_type": conn.cable_type,
                    "cable_model": conn.cable_model,
                    "remark": conn.remark,
                    "connection_id": conn.id
                }
            else:
                # 没有明确的上下游关系，使用默认方向（从source到target）
                edge_data = {
                    "from": conn.source_device_id,
                    "to": conn.target_device_id,
                    "arrows": "to",
                    "label": conn.connection_type or conn.cable_type or "",
                    "connection_type": conn.connection_type,
                    "cable_type": conn.cable_type,
                    "cable_model": conn.cable_model,
                    "remark": conn.remark,
                    "connection_id": conn.id
                }
            edges.append(edge_data)
                
    # 构建返回数据
    response_data = {"nodes": nodes, "edges": edges, "level": level}
    
    # 如果是总线式布局，添加额外的元数据
    if level == "port" and layout_type == "bus":
        response_data["layout_type"] = "bus"
        response_data["metadata"] = {
            "bus_count": len([n for n in nodes if n.get('type') == 'bus']),
            "port_count": len([n for n in nodes if n.get('type') == 'port']),
            "connection_count": len([e for e in edges if e.get('type') == 'port_connection'])
        }
    else:
        response_data["layout_type"] = "standard"
    
    logger.info(f"/graph_data response: nodes={len(nodes)}, edges={len(edges)}, layout_type={response_data['layout_type']}")
    return JSONResponse(content=response_data)


def _get_device_lifecycle_status(device: Device, db: Session) -> str:
    """计算设备的生命周期状态 - 复用已有的完整实现逻辑"""
    try:
        from datetime import datetime
        import re
        
        # 查找对应的生命周期规则
        rule = db.query(LifecycleRule).filter(
            LifecycleRule.device_type == device.device_type,
            LifecycleRule.is_active == "true"
        ).first()
        
        if not rule:
            return "未配置规则"
        
        # 解析投产日期
        if not device.commission_date:
            return "投产日期未填写"
        
        commission_date = None
        date_str = device.commission_date.strip()
        current_date = datetime.now()
        
        # 处理特殊格式：YYYYMM (如 202312)
        if re.match(r'^\d{6}$', date_str):
            try:
                year = int(date_str[:4])
                month = int(date_str[4:6])
                commission_date = datetime(year, month, 1)
            except ValueError:
                pass
        
        # 如果特殊格式解析失败，尝试标准格式
        if not commission_date:
            date_formats = [
                "%Y-%m-%d",
                "%Y/%m/%d", 
                "%Y.%m.%d",
                "%Y-%m",
                "%Y/%m",
                "%Y.%m",
                "%Y"
            ]
            
            for fmt in date_formats:
                try:
                    if fmt == "%Y":
                        # 只有年份的情况，默认为该年的1月1日
                        commission_date = datetime.strptime(device.commission_date, fmt).replace(month=1, day=1)
                    elif fmt in ["%Y-%m", "%Y/%m", "%Y.%m"]:
                        # 只有年月的情况，默认为该月的1日
                        commission_date = datetime.strptime(device.commission_date, fmt).replace(day=1)
                    else:
                        commission_date = datetime.strptime(device.commission_date, fmt)
                    break
                except ValueError:
                    continue
        
        if not commission_date:
            return "投产日期格式无法识别"
        
        # 计算服役时间和剩余时间
        days_in_service = (current_date - commission_date).days
        lifecycle_days = rule.lifecycle_years * 365
        remaining_days = lifecycle_days - days_in_service
        warning_days = rule.warning_months * 30
        
        # 确定生命周期状态
        if remaining_days < 0:
            return f"已超期 {abs(remaining_days)} 天"
        elif remaining_days <= warning_days:
            return f"临近超限，剩余 {remaining_days} 天"
        else:
            return f"正常，剩余 {remaining_days} 天"
            
    except Exception as e:
        return "计算错误"


def _should_include_device(device: Device, station: Optional[str], device_type: Optional[str], show_critical_only: bool) -> bool:
    """判断设备是否应该包含在拓扑图中"""
    # 基础数据验证：过滤掉名称无效的设备
    if not device.name or device.name.strip() == "" or device.name.lower() in ["nan", "null", "none"]:
        return False
    
    # 站点筛选
    if station and device.station != station:
        return False
    
    # 设备类型筛选
    if device_type and device.device_type != device_type:
        return False
    
    # 关键设备筛选（这里可以根据业务需求定义关键设备的判断逻辑）
    if show_critical_only:
        # 示例：将发电机组、UPS、变压器等视为关键设备
        critical_types = ["发电机组", "UPS", "变压器", "高压配电柜", "低压配电柜"]
        if device.device_type not in critical_types:
            return False
    
    return True


def _should_include_connection(connection: Connection, connection_type: Optional[str]) -> bool:
    """判断连接是否应该包含在拓扑图中"""
    # 连接类型筛选
    if connection_type and connection.connection_type != connection_type:
        return False
    
    return True


def _create_port_nodes(device: Device, db: Session, level: int = 1) -> list:
    """为设备创建端口级节点 - 标准模式
    只显示选中设备的端口和直接连接的对端设备端口
    """
    port_nodes = []
    
    # 获取设备的所有连接，提取端口信息
    connections = db.query(Connection).filter(
        or_(Connection.source_device_id == device.id, Connection.target_device_id == device.id)
    ).all()
    
    # 收集选中设备的端口信息
    selected_device_ports = set()
    connected_device_ports = set()
    
    for conn in connections:
        # 选中设备的端口
        if conn.source_device_id == device.id:
            if conn.source_fuse_number:
                selected_device_ports.add(("熔丝", conn.source_fuse_number, conn))
            if conn.source_breaker_number:
                selected_device_ports.add(("空开", conn.source_breaker_number, conn))
        if conn.target_device_id == device.id:
            if conn.target_fuse_number:
                selected_device_ports.add(("熔丝", conn.target_fuse_number, conn))
            if conn.target_breaker_number:
                selected_device_ports.add(("空开", conn.target_breaker_number, conn))
        
        # 对端设备的端口（只显示直接连接的）
        if conn.source_device_id == device.id:
            # 当前设备是源设备，收集目标设备端口
            target_device = conn.target_device
            if target_device:
                if conn.target_fuse_number:
                    connected_device_ports.add((target_device, "熔丝", conn.target_fuse_number, conn))
                if conn.target_breaker_number:
                    connected_device_ports.add((target_device, "空开", conn.target_breaker_number, conn))
        elif conn.target_device_id == device.id:
            # 当前设备是目标设备，收集源设备端口
            source_device = conn.source_device
            if source_device:
                if conn.source_fuse_number:
                    connected_device_ports.add((source_device, "熔丝", conn.source_fuse_number, conn))
                if conn.source_breaker_number:
                    connected_device_ports.add((source_device, "空开", conn.source_breaker_number, conn))
    
    # 创建选中设备的端口节点
    for port_type, port_number, conn in selected_device_ports:
        port_nodes.append({
            "id": f"{device.id}_{port_type}_{port_number}",
            "label": f"{port_type}-{port_number}",  # 简化端口名称，移除设备名称
            "title": f"""<b>设备:</b> {device.name}<br>
                         <b>端口:</b> {port_type}-{port_number}<br>
                         <b>设备类型:</b> {device.device_type or 'N/A'}""",
            "level": level,  # 端口节点在 level 层
            "device_id": device.id,
            "device_name": device.name,  # 新增：设备名称，便于前端三行显示
            "station": device.station,   # 新增：站点信息
            "port_name": f"{port_type}-{port_number}",  # 新增：端口名（简化）
            "port_type": port_type,
            "port_number": port_number,
            "node_type": "selected_device_port",  # 标记为选中设备端口
            "color": {"background": "#E6F3FF", "border": "#4169E1"}  # 端口使用蓝色
        })
    
    # 创建对端设备的端口节点
    for connected_device, port_type, port_number, conn in connected_device_ports:
        port_nodes.append({
            "id": f"{connected_device.id}_{port_type}_{port_number}",
            "label": f"{connected_device.name}·{port_type}-{port_number}",  # 对端标签采用“设备名·端口名”
            "title": f"""<b>设备:</b> {connected_device.name}<br>
                         <b>端口:</b> {port_type}-{port_number}<br>
                         <b>设备类型:</b> {connected_device.device_type or 'N/A'}<br>
                         <b>连接到:</b> {device.name}""",
            "level": level + 1,  # 对端端口节点在 level + 1 层
            "device_id": connected_device.id,
            "device_name": connected_device.name,  # 新增
            "station": connected_device.station,    # 新增
            "port_name": f"{port_type}-{port_number}",  # 新增
            "port_type": port_type,
            "port_number": port_number,
            "node_type": "connected_device_port"  # 标记为对端设备端口
        })
    
    return port_nodes


def _create_port_edges(connection: Connection, direction: str) -> list:
    """为连接创建端口级边 - 标准模式
    根据upstream_downstream字段确定箭头方向
    """
    edges = []
    
    # 创建端口间的连接
    source_ports = []
    target_ports = []
    
    # 使用新的端口ID格式（中文）
    if connection.source_fuse_number:
        source_ports.append(f"{connection.source_device_id}_熔丝_{connection.source_fuse_number}")
    if connection.source_breaker_number:
        source_ports.append(f"{connection.source_device_id}_空开_{connection.source_breaker_number}")
    if connection.target_fuse_number:
        target_ports.append(f"{connection.target_device_id}_熔丝_{connection.target_fuse_number}")
    if connection.target_breaker_number:
        target_ports.append(f"{connection.target_device_id}_空开_{connection.target_breaker_number}")
    
    # 根据upstream_downstream字段确定箭头方向
    if connection.upstream_downstream == "上游":
        # 上游表示电流从source流向target
        arrow_direction = "to"
        from_ports = source_ports
        to_ports = target_ports
    elif connection.upstream_downstream == "下游":
        # 下游表示电流从target流向source
        arrow_direction = "to"
        from_ports = target_ports
        to_ports = source_ports
    else:
        # 没有明确方向，使用默认方向（source到target）
        arrow_direction = "to"
        from_ports = source_ports
        to_ports = target_ports
    
    # 创建端口间的连接边
    for from_port in from_ports:
        for to_port in to_ports:
            edges.append({
                "from": from_port,
                "to": to_port,
                "arrows": arrow_direction,
                "label": connection.connection_type or connection.cable_model or "",
                "connection_type": connection.connection_type,
                "cable_model": connection.cable_model,
                "connection_id": connection.id,
                "remark": connection.remark
            })
    
    return edges







@app.get("/api/port-selection/{device_id}")
async def get_port_selection_options(device_id: int, db: Session = Depends(get_db)):
    """获取端口选择选项"""
    try:
        from port_topology_service import PortTopologyService
        service = PortTopologyService(db)
        data = service.get_port_selection_options(device_id)
        return {"success": True, "data": data}
    except Exception as e:
        topology_error_tracker.log_error(
            category=ErrorCategory.API_ERROR,
            level=ErrorLevel.ERROR,
            message=f"端口选择选项API调用失败: {str(e)}",
            context={"device_id": device_id},
            exception=e
        )
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/graph", response_class=HTMLResponse)
async def get_topology_graph_page(request: Request, db: Session = Depends(get_db)):
    """拓扑图页面 - 显示所有设备供用户选择"""
    devices = db.query(Device).all()
    return templates.TemplateResponse("graph.html", {"request": request, "devices": devices})


@app.get("/graph/{device_id}", response_class=HTMLResponse)
async def get_power_chain_graph(request: Request, device_id: int, db: Session = Depends(get_db)):
    """特定设备的拓扑图页面 - 兼容旧版本URL"""
    devices = db.query(Device).all()
    return templates.TemplateResponse("graph.html", {"request": request, "devices": devices, "selected_device_id": device_id})


# --- 设备生命周期规则管理 API ---

@app.get("/api/lifecycle-rules")
async def get_lifecycle_rules(db: Session = Depends(get_db)):
    """
    获取所有生命周期规则
    """
    try:
        rules = db.query(LifecycleRule).all()
        return JSONResponse(content={
            "success": True,
            "data": [{
                "id": rule.id,
                "device_type": rule.device_type,
                "lifecycle_years": rule.lifecycle_years,
                "warning_months": rule.warning_months,
                "description": rule.description,
                "is_active": rule.is_active,
                "created_at": rule.created_at,
                "updated_at": rule.updated_at
            } for rule in rules]
        })
    except Exception as e:
        print(f"获取生命周期规则失败: {e}")
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)


@app.post("/api/lifecycle-rules")
async def create_lifecycle_rule(
    device_type: str = Form(...),
    lifecycle_years: int = Form(...),
    warning_months: int = Form(6),
    description: str = Form(""),
    password: str = Form(...),  # 添加密码参数
    db: Session = Depends(get_db)
):
    """
    创建生命周期规则
    """
    try:
        # 验证管理员密码
        if not verify_admin_password(password):
            return JSONResponse(content={"success": False, "message": "密码错误"}, status_code=401)
        
        from datetime import datetime
        
        # 检查设备类型是否已存在规则
        existing_rule = db.query(LifecycleRule).filter(LifecycleRule.device_type == device_type).first()
        if existing_rule:
            return JSONResponse(content={
                "success": False, 
                "message": f"设备类型 '{device_type}' 的生命周期规则已存在"
            }, status_code=400)
        
        # 创建新规则
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_rule = LifecycleRule(
            device_type=device_type,
            lifecycle_years=lifecycle_years,
            warning_months=warning_months,
            description=description,
            is_active="true",
            created_at=current_time,
            updated_at=current_time
        )
        
        db.add(new_rule)
        db.commit()
        db.refresh(new_rule)
        
        return JSONResponse(content={
            "success": True,
            "message": "生命周期规则创建成功",
            "data": {
                "id": new_rule.id,
                "device_type": new_rule.device_type,
                "lifecycle_years": new_rule.lifecycle_years,
                "warning_months": new_rule.warning_months
            }
        })
        
    except Exception as e:
        db.rollback()




        print(f"创建生命周期规则失败: {e}")
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)


@app.put("/api/lifecycle-rules/{rule_id}")
async def update_lifecycle_rule(
    rule_id: int,
    device_type: str = Form(...),
    lifecycle_years: int = Form(...),
    warning_months: int = Form(6),
    description: str = Form(""),
    is_active: str = Form("true"),
    password: str = Form(...),  # 添加密码参数
    db: Session = Depends(get_db)
):
    """
    更新生命周期规则
    """
    try:
        # 验证管理员密码
        if not verify_admin_password(password):
            return JSONResponse(content={"success": False, "message": "密码错误"}, status_code=401)
        
        from datetime import datetime
        
        rule = db.query(LifecycleRule).filter(LifecycleRule.id == rule_id).first()
        if not rule:
            return JSONResponse(content={"success": False, "message": "规则不存在"}, status_code=404)
        
        # 检查设备类型是否与其他规则冲突
        existing_rule = db.query(LifecycleRule).filter(
            LifecycleRule.device_type == device_type,
            LifecycleRule.id != rule_id
        ).first()
        if existing_rule:
            return JSONResponse(content={
                "success": False, 
                "message": f"设备类型 '{device_type}' 的生命周期规则已存在"
            }, status_code=400)
        
        # 更新规则
        rule.device_type = device_type
        rule.lifecycle_years = lifecycle_years
        rule.warning_months = warning_months
        rule.description = description
        rule.is_active = is_active
        rule.updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        db.commit()
        
        return JSONResponse(content={
            "success": True,
            "message": "生命周期规则更新成功"
        })
        
    except Exception as e:
        db.rollback()
        print(f"更新生命周期规则失败: {e}")
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)


@app.delete("/api/lifecycle-rules/{rule_id}")
async def delete_lifecycle_rule(rule_id: int, password: str = Form(...), db: Session = Depends(get_db)):
    """
    删除生命周期规则
    """
    try:
        # 验证管理员密码
        if not verify_admin_password(password):
            return JSONResponse(content={"success": False, "message": "密码错误"}, status_code=401)
        
        rule = db.query(LifecycleRule).filter(LifecycleRule.id == rule_id).first()
        if not rule:
            return JSONResponse(content={"success": False, "message": "规则不存在"}, status_code=404)
        
        db.delete(rule)
        db.commit()
        
        return JSONResponse(content={
            "success": True,
            "message": "生命周期规则删除成功"
        })
        
    except Exception as e:
        db.rollback()
        print(f"删除生命周期规则失败: {e}")
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)


@app.get("/api/devices")
async def get_devices_api(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(50, ge=1, le=200, description="每页数量"),
    db: Session = Depends(get_db)
):
    """
    获取设备列表API接口
    """
    try:
        # 构建查询
        query = db.query(Device)
        
        # 计算总数
        total = query.count()
        
        # 应用分页
        offset = (page - 1) * page_size
        devices = query.offset(offset).limit(page_size).all()
        
        # 构建响应数据
        result = []
        for device in devices:
            result.append({
                "id": device.id,
                "asset_id": device.asset_id,
                "name": device.name,
                "station": device.station,
                "model": device.model,
                "device_type": device.device_type,
                "location": device.location,
                "power_rating": device.power_rating,
                "vendor": device.vendor,
                "commission_date": device.commission_date.isoformat() if device.commission_date and hasattr(device.commission_date, 'isoformat') else device.commission_date,
                "remark": device.remark
            })
        
        return JSONResponse(content={
            "success": True,
            "data": result,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total": total,
                "total_pages": (total + page_size - 1) // page_size
            }
        })
        
    except Exception as e:
        print(f"获取设备列表失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取设备列表失败: {str(e)}")


@app.get("/api/topology/filter-options")
async def get_filter_options(db: Session = Depends(get_db)):
    """
    获取拓扑图筛选选项
    返回设备类型、连接类型、局站等筛选选项
    """
    try:
        # 获取所有局站
        stations = db.query(Device.station).filter(Device.station.isnot(None)).filter(Device.station != '').distinct().all()
        station_list = [station[0] for station in stations if station[0]]
        station_list.sort()
        
        # 获取所有连接类型
        connection_types = db.query(Connection.connection_type).filter(Connection.connection_type.isnot(None)).filter(Connection.connection_type != '').distinct().all()
        connection_type_list = [conn_type[0] for conn_type in connection_types if conn_type[0]]
        connection_type_list.sort()
        
        return JSONResponse(content={
            "success": True,
            "data": {
                "device_types": STANDARD_DEVICE_TYPES,  # 使用新的标准设备类型列表
                "connection_types": connection_type_list,
                "stations": station_list
            }
        })
        
    except Exception as e:
        print(f"获取筛选选项失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取筛选选项失败: {str(e)}")


@app.get("/api/devices/lifecycle-status")
async def get_devices_lifecycle_status(
    status_filter: Optional[str] = None,  # normal, warning, expired, all
    db: Session = Depends(get_db)
):
    """
    获取设备生命周期状态
    status_filter: normal(正常), warning(临近超限), expired(已超期), all(全部)
    """
    try:
        from datetime import datetime, timedelta
        import re
        
        # 获取所有设备和规则
        devices = db.query(Device).all()
        rules = {rule.device_type: rule for rule in db.query(LifecycleRule).filter(LifecycleRule.is_active == "true").all()}
        
        result_devices = []
        current_date = datetime.now()
        
        for device in devices:
            # 查找对应的生命周期规则
            rule = rules.get(device.device_type)
            if not rule:
                # 没有规则的设备标记为未知状态
                device_info = {
                    "id": device.id,
                    "asset_id": device.asset_id,
                    "name": device.name,
                    "station": device.station,
                    "model": device.model,
                    "vendor": device.vendor,
                    "commission_date": device.commission_date,
                    "lifecycle_status": "unknown",
                    "lifecycle_status_text": "未配置规则",
                    "days_in_service": None,
                    "remaining_days": None,
                    "rule_years": None
                }
                if not status_filter or status_filter == "all":
                    result_devices.append(device_info)
                continue
            
            # 解析投产日期
            if not device.commission_date:
                device_info = {
                    "id": device.id,
                    "asset_id": device.asset_id,
                    "name": device.name,
                    "station": device.station,
                    "model": device.model,
                    "vendor": device.vendor,
                    "commission_date": device.commission_date,
                    "lifecycle_status": "unknown",
                    "lifecycle_status_text": "投产日期未填写",
                    "days_in_service": None,
                    "remaining_days": None,
                    "rule_years": rule.lifecycle_years
                }
                if not status_filter or status_filter == "all":
                    result_devices.append(device_info)
                continue
            
            # 尝试解析多种日期格式
            commission_date = None
            date_str = device.commission_date.strip()
            
            # 处理特殊格式：YYYYMM (如 202312)
            if re.match(r'^\d{6}$', date_str):
                try:
                    year = int(date_str[:4])
                    month = int(date_str[4:6])
                    commission_date = datetime(year, month, 1)
                except ValueError:
                    pass
            
            # 如果特殊格式解析失败，尝试标准格式
            if not commission_date:
                date_formats = [
                    "%Y-%m-%d",
                    "%Y/%m/%d", 
                    "%Y.%m.%d",
                    "%Y-%m",
                    "%Y/%m",
                    "%Y.%m",
                    "%Y"
                ]
                
                for fmt in date_formats:
                    try:
                        if fmt == "%Y":
                            # 只有年份的情况，默认为该年的1月1日
                            commission_date = datetime.strptime(device.commission_date, fmt).replace(month=1, day=1)
                        elif fmt in ["%Y-%m", "%Y/%m", "%Y.%m"]:
                            # 只有年月的情况，默认为该月的1日
                            commission_date = datetime.strptime(device.commission_date, fmt).replace(day=1)
                        else:
                            commission_date = datetime.strptime(device.commission_date, fmt)
                        break
                    except ValueError:
                        continue
            
            if not commission_date:
                device_info = {
                    "id": device.id,
                    "asset_id": device.asset_id,
                    "name": device.name,
                    "station": device.station,
                    "model": device.model,
                    "vendor": device.vendor,
                    "commission_date": device.commission_date,
                    "lifecycle_status": "unknown",
                    "lifecycle_status_text": "投产日期格式无法识别",
                    "days_in_service": None,
                    "remaining_days": None,
                    "rule_years": rule.lifecycle_years
                }
                if not status_filter or status_filter == "all":
                    result_devices.append(device_info)
                continue
            
            # 计算服役时间和剩余时间
            days_in_service = (current_date - commission_date).days
            lifecycle_days = rule.lifecycle_years * 365
            remaining_days = lifecycle_days - days_in_service
            warning_days = rule.warning_months * 30
            
            # 确定生命周期状态
            if remaining_days < 0:
                lifecycle_status = "expired"
                lifecycle_status_text = f"已超期 {abs(remaining_days)} 天"
            elif remaining_days <= warning_days:
                lifecycle_status = "warning"
                lifecycle_status_text = f"临近超限，剩余 {remaining_days} 天"
            else:
                lifecycle_status = "normal"
                lifecycle_status_text = f"正常，剩余 {remaining_days} 天"
            
            device_info = {
                "id": device.id,
                "asset_id": device.asset_id,
                "name": device.name,
                "station": device.station,
                "model": device.model,
                "vendor": device.vendor,
                "commission_date": device.commission_date,
                "lifecycle_status": lifecycle_status,
                "lifecycle_status_text": lifecycle_status_text,
                "days_in_service": days_in_service,
                "remaining_days": remaining_days,
                "rule_years": rule.lifecycle_years
            }
            
            # 根据筛选条件添加设备
            if not status_filter or status_filter == "all" or status_filter == lifecycle_status:
                result_devices.append(device_info)
        
        # 统计信息
        total_count = len(result_devices)
        normal_count = len([d for d in result_devices if d["lifecycle_status"] == "normal"])
        warning_count = len([d for d in result_devices if d["lifecycle_status"] == "warning"])
        expired_count = len([d for d in result_devices if d["lifecycle_status"] == "expired"])
        unknown_count = len([d for d in result_devices if d["lifecycle_status"] == "unknown"])
        
        return JSONResponse(content={
            "success": True,
            "data": result_devices,
            "statistics": {
                "total": total_count,
                "normal": normal_count,
                "warning": warning_count,
                "expired": expired_count,
                "unknown": unknown_count
            }
        })
        
    except Exception as e:
        print(f"获取设备生命周期状态失败: {e}")
        traceback.print_exc()
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)


@app.get("/test-route")
async def test_route():
    """
    测试路由
    """
    print("=== 测试路由被调用 ===")
    return {"message": "测试路由正常工作", "timestamp": "updated"}

@app.get("/debug-routes")
async def debug_routes():
    """
    调试路由 - 显示所有已注册的路由
    """
    routes = []
    for route in app.routes:
        if hasattr(route, 'path') and hasattr(route, 'methods'):
            routes.append({
                "path": route.path,
                "methods": list(route.methods) if route.methods else [],
                "name": getattr(route, 'name', 'unknown')
            })
    return {"registered_routes": routes, "total_count": len(routes)}

@app.get("/debug-lifecycle")
async def debug_lifecycle():
    """
    调试生命周期路由
    """
    print("=== 调试生命周期路由被调用 ===")
    return {"message": "调试路由正常工作", "status": "ok"}

@app.post("/api/verify-password")
async def verify_password(request: Request):
    """
    验证管理员密码
    """
    try:
        data = await request.json()
        password = data.get("password", "")
        
        if verify_admin_password(password):
            return {"success": True, "message": "密码验证成功"}
        else:
            return {"success": False, "message": "密码错误"}
    except Exception as e:
        print(f"Error verifying password: {e}")
        return {"success": False, "message": "验证失败"}

@app.get("/lifecycle-management", response_class=HTMLResponse)
async def lifecycle_management_page(request: Request):
    """
    生命周期管理页面
    """
    print("=== 访问生命周期管理页面 ===")
    print(f"请求URL: {request.url}")
    print(f"请求方法: {request.method}")
    try:
        print("正在渲染模板...")
        response = templates.TemplateResponse("lifecycle_management.html", {"request": request})
        print("模板渲染成功")
        return response
    except Exception as e:
        print(f"生命周期管理页面错误: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/connections", response_class=HTMLResponse)
async def connections_page(request: Request):
    """
    连接管理页面
    """
    print("=== 访问连接管理页面 ===")
    print(f"请求URL: {request.url}")
    print(f"请求方法: {request.method}")
    try:
        print("正在渲染模板...")
        response = templates.TemplateResponse("connections.html", {"request": request})
        print("模板渲染成功")
        return response
    except Exception as e:
        print(f"连接管理页面错误: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/analytics", response_class=HTMLResponse)
async def analytics_page(request: Request):
    """
    统计分析页面
    """
    print("=== 访问统计分析页面 ===")
    print(f"请求URL: {request.url}")
    print(f"请求方法: {request.method}")
    try:
        print("正在渲染统计分析模板...")
        response = templates.TemplateResponse("analytics.html", {"request": request})
        print("统计分析模板渲染成功")
        return response
    except Exception as e:
        print(f"统计分析页面错误: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/export")
async def export_devices(
    password: str = Form(...),
    export_range: str = Form("all"),
    station_filter: str = Form(""),
    name_filter: str = Form(""),
    device_type_filter: str = Form(""),
    vendor_filter: str = Form(""),
    lifecycle_filter: str = Form(""),
    db: Session = Depends(get_db)
):
    """
    导出设备数据为Excel文件
    支持全量导出和筛选导出，需要管理员密码验证
    """
    try:
        # 验证管理员密码
        if not verify_admin_password(password):
            raise HTTPException(status_code=401, detail="密码错误，无权限导出数据")
        
        # 根据导出范围查询设备数据
        query = db.query(Device)
        
        # 如果是筛选导出，应用筛选条件
        if export_range == "filtered":
            if station_filter:
                query = query.filter(Device.station.ilike(f"%{station_filter}%"))
            if name_filter:
                query = query.filter(Device.name.ilike(f"%{name_filter}%"))
            if device_type_filter:
                query = query.filter(Device.device_type.ilike(f"%{device_type_filter}%"))
            if vendor_filter:
                query = query.filter(Device.vendor.ilike(f"%{vendor_filter}%"))
            if lifecycle_filter:
                # 这里需要根据生命周期状态筛选，暂时跳过复杂的生命周期逻辑
                pass
        
        devices = query.all()
        
        if not devices:
            raise HTTPException(status_code=404, detail="没有找到设备数据")
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "设备列表"
        
        # 定义表头
        headers = [
            "ID", "资产编号", "设备名称", "局站", "设备类型", "设备型号", 
            "所在位置", "额定容量", "设备生产厂家", "投产日期", "备注"
        ]
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 写入设备数据
        for row, device in enumerate(devices, 2):
            data = [
                device.id,
                device.asset_id,
                device.name,
                device.station,
                device.device_type,
                device.model,
                device.location,
                device.power_rating,
                device.vendor,
                device.commission_date.strftime("%Y-%m-%d") if device.commission_date else "",
                device.remark
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # 设置斑马纹效果
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 限制最大宽度
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 冻结首行
        ws.freeze_panes = "A2"
        
        # 添加筛选器
        ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}1"
        
        # 生成文件名（包含时间戳）
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if export_range == "filtered":
            filename = f"设备列表_筛选导出_{timestamp}.xlsx"
        else:
            filename = f"设备列表_全量导出_{timestamp}.xlsx"
        
        # 将Excel文件保存到内存
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # 设置响应头
        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        
        # 返回文件流
        return StreamingResponse(
            io.BytesIO(excel_buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"导出设备数据错误: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


# --- 连接管理 Pydantic 模型 ---

class ConnectionCreate(BaseModel):
    """创建连接的请求模型"""
    source_device_id: int
    target_device_id: int
    connection_type: Optional[str] = None
    cable_model: Optional[str] = None
    source_fuse_number: Optional[str] = None
    source_fuse_spec: Optional[str] = None
    source_breaker_number: Optional[str] = None
    source_breaker_spec: Optional[str] = None
    target_fuse_number: Optional[str] = None
    target_fuse_spec: Optional[str] = None
    target_breaker_number: Optional[str] = None
    target_breaker_spec: Optional[str] = None
    hierarchy_relation: Optional[str] = None
    upstream_downstream: Optional[str] = None
    parallel_count: Optional[int] = 1
    rated_current: Optional[float] = None
    cable_length: Optional[float] = None
    source_device_photo: Optional[str] = None
    target_device_photo: Optional[str] = None
    remark: Optional[str] = None
    installation_date: Optional[date] = None

class ConnectionUpdate(BaseModel):
    """更新连接的请求模型"""
    source_device_id: Optional[int] = None
    target_device_id: Optional[int] = None
    connection_type: Optional[str] = None
    cable_model: Optional[str] = None
    source_fuse_number: Optional[str] = None
    source_fuse_spec: Optional[str] = None
    source_breaker_number: Optional[str] = None
    source_breaker_spec: Optional[str] = None
    target_fuse_number: Optional[str] = None
    target_fuse_spec: Optional[str] = None
    target_breaker_number: Optional[str] = None
    target_breaker_spec: Optional[str] = None
    hierarchy_relation: Optional[str] = None
    upstream_downstream: Optional[str] = None
    parallel_count: Optional[int] = None
    rated_current: Optional[float] = None
    cable_length: Optional[float] = None
    source_device_photo: Optional[str] = None
    target_device_photo: Optional[str] = None
    remark: Optional[str] = None
    installation_date: Optional[date] = None

class ConnectionResponse(BaseModel):
    """连接响应模型"""
    id: int
    source_device_id: int
    target_device_id: int
    source_device_name: str
    target_device_name: str
    source_port: Optional[str]  # 源端口名称（带前缀）
    target_port: Optional[str]  # 目标端口名称（带前缀）
    connection_type: Optional[str]
    cable_model: Optional[str]
    source_fuse_number: Optional[str]
    source_fuse_spec: Optional[str]
    source_breaker_number: Optional[str]
    source_breaker_spec: Optional[str]
    target_fuse_number: Optional[str]
    target_fuse_spec: Optional[str]
    target_breaker_number: Optional[str]
    target_breaker_spec: Optional[str]
    hierarchy_relation: Optional[str]
    upstream_downstream: Optional[str]
    parallel_count: Optional[int]
    rated_current: Optional[float]
    cable_length: Optional[float]
    source_device_photo: Optional[str]
    target_device_photo: Optional[str]
    remark: Optional[str]
    installation_date: Optional[date]
    created_at: datetime
    updated_at: Optional[datetime]
    
    class Config:
        # 启用ORM模式，允许从SQLAlchemy模型创建
        from_attributes = True
        # 自定义JSON编码器处理日期时间对象
        json_encoders = {
            datetime: lambda v: v.isoformat() if v else None,
            date: lambda v: v.isoformat() if v else None
        }


# --- 连接管理 RESTful API 接口 ---

def get_unique_connections_count(db: Session) -> int:
    """
    获取去重后的连接数量
    通过识别双向连接并去重来获得真实的连接数量
    """
    # 获取所有有效连接
    connections = db.query(Connection).filter(Connection.connection_type.isnot(None)).all()
    
    # 使用集合存储唯一连接
    unique_connections = set()
    
    for conn in connections:
        # 创建连接的唯一标识
        # 对于双向连接，使用较小的设备ID作为第一个参数，确保A->B和B->A生成相同的标识
        device_pair = tuple(sorted([conn.source_device_id, conn.target_device_id]))
        
        # 结合端口信息创建更精确的连接标识
        source_port = conn.source_fuse_number or conn.source_breaker_number or ""
        target_port = conn.target_fuse_number or conn.target_breaker_number or ""
        
        # 为双向连接创建统一的标识
        if conn.source_device_id == device_pair[0]:
            connection_key = (device_pair[0], device_pair[1], source_port, target_port, conn.connection_type)
        else:
            connection_key = (device_pair[0], device_pair[1], target_port, source_port, conn.connection_type)
        
        unique_connections.add(connection_key)
    
    return len(unique_connections)


def get_connected_ports_count(db: Session) -> int:
    """
    直接统计所有有连接的端口数量
    这种方法能够准确处理内部设备互连和外部设备连接
    """
    connections = db.query(Connection).filter(Connection.connection_type.isnot(None)).all()
    connected_ports = set()
    
    for conn in connections:
        # 添加源端口
        if conn.source_fuse_number:
            port_id = f"{conn.source_device_id}_fuse_{conn.source_fuse_number}"
            connected_ports.add(port_id)
        if conn.source_breaker_number:
            port_id = f"{conn.source_device_id}_breaker_{conn.source_breaker_number}"
            connected_ports.add(port_id)
            
        # 添加目标端口
        if conn.target_fuse_number:
            port_id = f"{conn.target_device_id}_fuse_{conn.target_fuse_number}"
            connected_ports.add(port_id)
        if conn.target_breaker_number:
            port_id = f"{conn.target_device_id}_breaker_{conn.target_breaker_number}"
            connected_ports.add(port_id)
    
    return len(connected_ports)


@app.get("/api/connections/statistics")
async def get_connections_statistics(db: Session = Depends(get_db)):
    """
    获取连接统计信息
    """
    try:
        # 使用去重算法获取真实的连接数量
        total_connections = get_unique_connections_count(db)
        
        # 使用PortStatisticsService统一的统计逻辑，确保数据一致性
        port_service = PortStatisticsService(db)
        port_summary = port_service._get_device_port_summary()
        
        # 从统一的端口统计服务获取数据
        total_ports = port_summary.get('total_ports', 0)
        connected_ports_count = port_summary.get('connected_ports', 0)
        idle_ports = port_summary.get('idle_ports', 0)
        
        # 获取设备总数
        total_devices = db.query(Device).count()
        
        # 按连接类型统计
        connection_type_stats = db.query(
            Connection.connection_type,
            func.count(Connection.id).label('count')
        ).group_by(Connection.connection_type).all()
        
        # 将混合的中英文连接类型统计合并为标准格式
        cable_count = 0
        busbar_count = 0
        bus_count = 0
        
        for item in connection_type_stats:
            conn_type = item[0] or ""
            count = item[1]
            
            # 电缆类型（cable 或 电缆）
            if conn_type.lower() in ['cable', '电缆']:
                cable_count += count
            # 铜排类型（busbar 或 铜排）
            elif conn_type.lower() in ['busbar', '铜排']:
                busbar_count += count
            # 母线类型（bus、busway 或 母线）
            elif conn_type.lower() in ['bus', 'busway', '母线']:
                bus_count += count
        
        # 按设备类型统计（源设备）
        device_type_stats = db.query(
            Device.device_type,
            func.count(Connection.id).label('count')
        ).join(Connection, Device.id == Connection.source_device_id)\
         .group_by(Device.device_type).all()
        
        # 最近30天新增连接数
        thirty_days_ago = datetime.now() - timedelta(days=30)
        recent_connections = db.query(Connection)\
            .filter(Connection.created_at >= thirty_days_ago).count()
        
        return JSONResponse(content={
            "success": True,
            "data": {
                "total_devices": total_devices,
                "total_ports": total_ports,
                "connected_ports": connected_ports_count,
                "idle_ports": idle_ports,
                "total_connections": total_connections,
                "cable": cable_count,
                "busbar": busbar_count,
                "bus": bus_count,
                "recent_connections": recent_connections,
                "connection_types": [
                    {"type": item[0] or "未分类", "count": item[1]} 
                    for item in connection_type_stats
                ],
                "device_types": [
                    {"type": item[0] or "未分类", "count": item[1]} 
                    for item in device_type_stats
                ]
            }
        })
        
    except Exception as e:
        print(f"获取连接统计失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取连接统计失败: {str(e)}")


@app.get("/api/ports/statistics")
async def get_port_statistics(db: Session = Depends(get_db)):
    """
    获取端口统计信息
    """
    try:
        # 创建端口统计服务实例
        port_service = PortStatisticsService(db)
        
        # 获取端口统计数据
        statistics = port_service.get_port_statistics()
        
        return JSONResponse(content={
            "success": True,
            "data": statistics
        })
        
    except Exception as e:
        print(f"获取端口统计失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取端口统计失败: {str(e)}")


@app.get("/api/devices/{device_id}/ports")
async def get_device_port_details(device_id: int, db: Session = Depends(get_db)):
    """
    获取指定设备的端口详情
    """
    try:
        # 创建端口统计服务实例
        port_service = PortStatisticsService(db)
        
        # 获取设备端口详情
        port_details = port_service.get_device_port_details(device_id)
        
        return JSONResponse(content={
            "success": True,
            "data": port_details
        })
        
    except Exception as e:
        print(f"获取设备端口详情失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取设备端口详情失败: {str(e)}")


# ==================== 统计分析API端点 ====================

@app.get("/api/analytics/utilization-rates")
async def get_utilization_rates(db: Session = Depends(get_db)):
    """
    获取使用率分析数据
    包括端口总体使用率、按设备类型统计、按站点统计等
    """
    try:
        # 创建统计分析服务实例
        analytics_service = AnalyticsService(db)
        
        # 获取使用率分析数据
        utilization_data = analytics_service.get_utilization_rates()
        
        return JSONResponse(content={
            "success": True,
            "data": utilization_data
        })
        
    except Exception as e:
        print(f"获取使用率分析数据失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取使用率分析数据失败: {str(e)}")


@app.get("/api/analytics/idle-rates")
async def get_idle_rates(db: Session = Depends(get_db)):
    """
    获取空闲率分析数据
    包括端口总体空闲率、按设备类型统计、按站点统计、空闲率预警等
    """
    try:
        # 创建统计分析服务实例
        analytics_service = AnalyticsService(db)
        
        # 获取空闲率分析数据
        idle_data = analytics_service.get_idle_rates()
        
        return JSONResponse(content={
            "success": True,
            "data": idle_data
        })
        
    except Exception as e:
        print(f"获取空闲率分析数据失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取空闲率分析数据失败: {str(e)}")





@app.get("/api/analytics/summary-dashboard")
async def get_summary_dashboard(db: Session = Depends(get_db)):
    """
    获取仪表板汇总数据
    包括所有关键指标的汇总信息，用于统计分析仪表板显示
    """
    try:
        # 创建统计分析服务实例
        analytics_service = AnalyticsService(db)
        
        # 获取仪表板汇总数据
        dashboard_data = analytics_service.get_summary_dashboard()
        
        return JSONResponse(content={
            "success": True,
            "data": dashboard_data
        })
        
    except Exception as e:
        print(f"获取仪表板汇总数据失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取仪表板汇总数据失败: {str(e)}")


# 辅助函数：根据熔丝/空开编号为端口名称添加前缀
def build_port_name_with_prefix(fuse_number, breaker_number, original_port=None):
    """根据熔丝编号或空开编号为端口名称添加前缀"""
    fuse_num = str(fuse_number).strip() if fuse_number and str(fuse_number).strip() not in ['', 'nan', 'None'] else ''
    breaker_num = str(breaker_number).strip() if breaker_number and str(breaker_number).strip() not in ['', 'nan', 'None'] else ''
    
    # 优先使用熔丝编号
    if fuse_num:
        return f"熔丝_{fuse_num}"
    elif breaker_num:
        return f"空开_{breaker_num}"
    else:
        # 如果都没有，返回原始端口名称或空字符串
        return original_port if original_port else ''


@app.get("/api/connections")
async def get_connections(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(100, ge=1, le=5000, description="每页数量"),
    source_device_id: Optional[int] = Query(None, description="源设备ID"),
    target_device_id: Optional[int] = Query(None, description="目标设备ID"),
    connection_type: Optional[str] = Query(None, description="连接类型"),
    device_name: Optional[str] = Query(None, description="设备名称（模糊查询，匹配源设备或目标设备）"),
    db: Session = Depends(get_db)
):
    """
    获取连接列表
    支持分页和筛选功能
    """
    try:
        # 构建查询
        # 创建Device表的别名用于目标设备
        target_device = aliased(Device)
        query = db.query(Connection, Device.name.label('source_device_name'), target_device.name.label('target_device_name'))\
                  .join(Device, Connection.source_device_id == Device.id)\
                  .join(target_device, Connection.target_device_id == target_device.id)
        
        # 应用筛选条件
        if source_device_id:
            query = query.filter(Connection.source_device_id == source_device_id)
        if target_device_id:
            query = query.filter(Connection.target_device_id == target_device_id)
        if connection_type:
            if connection_type == "空闲":
                # 筛选空闲端口：连接类型为空且A端设备有熔丝或空开数据
                query = query.filter(
                    and_(
                        Connection.connection_type.is_(None),
                        or_(
                            Connection.source_fuse_number.isnot(None),
                            Connection.source_breaker_number.isnot(None)
                        )
                    )
                )
            elif connection_type == "已使用总量":
                # 筛选已使用总量：显示所有有连接类型的记录（非空闲）
                query = query.filter(Connection.connection_type.isnot(None))
            else:
                query = query.filter(Connection.connection_type.ilike(f"%{connection_type}%"))
        else:
            # 如果没有指定连接类型筛选，默认显示所有记录（包括空闲端口）
            # 但要确保A端设备有端口数据（熔丝或空开）
            query = query.filter(
                or_(
                    Connection.source_fuse_number.isnot(None),
                    Connection.source_breaker_number.isnot(None)
                )
            )
        # 按设备名称模糊查询（匹配源设备或目标设备）
        if device_name:
            query = query.filter(
                or_(
                    Device.name.ilike(f"%{device_name}%"),  # 匹配源设备名称
                    target_device.name.ilike(f"%{device_name}%")  # 匹配目标设备名称
                )
            )
        
        # 计算总数
        total = query.count()
        
        # 应用分页
        offset = (page - 1) * page_size
        results = query.offset(offset).limit(page_size).all()
        
        # 构建响应数据 - 手动序列化日期字段以避免JSON序列化错误
        result = []
        for conn, source_name, target_name in results:
            # 手动处理日期字段的序列化
            installation_date_str = conn.installation_date.isoformat() if conn.installation_date else None
            created_at_str = conn.created_at.isoformat() if conn.created_at else None
            updated_at_str = conn.updated_at.isoformat() if conn.updated_at else None
            
            # 构建带前缀的端口名称
            source_port_with_prefix = build_port_name_with_prefix(
                conn.source_fuse_number, 
                conn.source_breaker_number, 
                conn.source_port
            )
            target_port_with_prefix = build_port_name_with_prefix(
                conn.target_fuse_number, 
                conn.target_breaker_number, 
                conn.target_port
            )
            
            result.append({
                "id": conn.id,
                "source_device_id": conn.source_device_id,
                "target_device_id": conn.target_device_id,
                "source_device_name": source_name,
                "target_device_name": target_name,
                "connection_type": conn.connection_type,
                "cable_model": conn.cable_model,
                "source_port": source_port_with_prefix,  # 使用带前缀的端口名称
                "target_port": target_port_with_prefix,  # 使用带前缀的端口名称
                "source_fuse_number": conn.source_fuse_number,
                "source_fuse_spec": conn.source_fuse_spec,
                "source_breaker_number": conn.source_breaker_number,
                "source_breaker_spec": conn.source_breaker_spec,
                "target_fuse_number": conn.target_fuse_number,
                "target_fuse_spec": conn.target_fuse_spec,
                "target_breaker_number": conn.target_breaker_number,
                "target_breaker_spec": conn.target_breaker_spec,
                "hierarchy_relation": conn.hierarchy_relation,
                "upstream_downstream": conn.upstream_downstream,
                "parallel_count": conn.parallel_count,
                "rated_current": conn.rated_current,
                "a_rated_current": conn.a_rated_current,  # 添加A端额定电流字段
                "b_rated_current": conn.b_rated_current,  # 添加B端额定电流字段
                "cable_length": conn.cable_length,
                "source_device_photo": conn.source_device_photo,
                "target_device_photo": conn.target_device_photo,
                "remark": conn.remark,
                "installation_date": installation_date_str,
                "created_at": created_at_str,
                "updated_at": updated_at_str
            })

        return {
            "success": True,
            "data": result,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total": total,
                "total_pages": (total + page_size - 1) // page_size
            }
        }
        
    except Exception as e:
        print(f"获取连接列表失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取连接列表失败: {str(e)}")


@app.post("/api/connections")
async def create_connection(
    source_device_id: int = Form(...),
    target_device_id: int = Form(...),
    connection_type: Optional[str] = Form(None),
    cable_model: Optional[str] = Form(None),
    source_port: Optional[str] = Form(None),
    target_port: Optional[str] = Form(None),
    source_fuse_number: Optional[str] = Form(None),
    source_fuse_spec: Optional[str] = Form(None),
    source_breaker_number: Optional[str] = Form(None),
    source_breaker_spec: Optional[str] = Form(None),
    target_fuse_number: Optional[str] = Form(None),
    target_fuse_spec: Optional[str] = Form(None),
    target_breaker_number: Optional[str] = Form(None),
    target_breaker_spec: Optional[str] = Form(None),
    hierarchy_relation: Optional[str] = Form(None),
    upstream_downstream: Optional[str] = Form(None),
    parallel_count: Optional[int] = Form(1),
    rated_current: Optional[float] = Form(None),
    cable_length: Optional[float] = Form(None),
    source_device_photo: Optional[str] = Form(None),
    target_device_photo: Optional[str] = Form(None),
    remark: Optional[str] = Form(None),
    installation_date: Optional[str] = Form(None),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """
    创建新连接
    需要管理员密码验证
    """
    try:
        # 验证管理员密码
        if not verify_admin_password(password):
            raise HTTPException(status_code=401, detail="密码错误")
        
        # 处理日期字段 - 支持yyyymm格式
        parsed_installation_date = None
        if installation_date:
            try:
                # 支持yyyymm格式，如202412
                if len(installation_date) == 6 and installation_date.isdigit():
                    year = int(installation_date[:4])
                    month = int(installation_date[4:6])
                    parsed_installation_date = datetime(year, month, 1).date()
                else:
                    raise ValueError("日期格式不正确")
            except ValueError:
                raise HTTPException(status_code=400, detail="安装日期格式错误，请使用YYYYMM格式（如：202412）")
        
        # 验证源设备和目标设备是否存在
        source_device = db.query(Device).filter(Device.id == source_device_id).first()
        if not source_device:
            raise HTTPException(status_code=404, detail=f"源设备ID {source_device_id} 不存在")
        
        target_device = db.query(Device).filter(Device.id == target_device_id).first()
        if not target_device:
            raise HTTPException(status_code=404, detail=f"目标设备ID {target_device_id} 不存在")
        
        # 检查是否已存在相同的连接
        existing_connection = db.query(Connection).filter(
            Connection.source_device_id == source_device_id,
            Connection.target_device_id == target_device_id
        ).first()
        
        if existing_connection:
            raise HTTPException(status_code=400, detail="该连接已存在")
        
        # 创建新连接
        new_connection = Connection(
            source_device_id=source_device_id,
            target_device_id=target_device_id,
            source_port=source_port,
            target_port=target_port,
            connection_type=connection_type,
            cable_model=cable_model,
            source_fuse_number=source_fuse_number,
            source_fuse_spec=source_fuse_spec,
            source_breaker_number=source_breaker_number,
            source_breaker_spec=source_breaker_spec,
            target_fuse_number=target_fuse_number,
            target_fuse_spec=target_fuse_spec,
            target_breaker_number=target_breaker_number,
            target_breaker_spec=target_breaker_spec,
            hierarchy_relation=hierarchy_relation,
            upstream_downstream=upstream_downstream,
            parallel_count=parallel_count,
            rated_current=rated_current,
            cable_length=cable_length,
            source_device_photo=source_device_photo,
            target_device_photo=target_device_photo,
            remark=remark,
            installation_date=parsed_installation_date,
            created_at=datetime.now()
        )
        
        db.add(new_connection)
        db.commit()
        db.refresh(new_connection)
        
        # 构建响应
        response = ConnectionResponse(
            id=new_connection.id,
            source_device_id=new_connection.source_device_id,
            target_device_id=new_connection.target_device_id,
            source_device_name=source_device.name,
            target_device_name=target_device.name,
            connection_type=new_connection.connection_type,
            cable_model=new_connection.cable_model,
            source_fuse_number=new_connection.source_fuse_number,
            source_fuse_spec=new_connection.source_fuse_spec,
            source_breaker_number=new_connection.source_breaker_number,
            source_breaker_spec=new_connection.source_breaker_spec,
            target_fuse_number=new_connection.target_fuse_number,
            target_fuse_spec=new_connection.target_fuse_spec,
            target_breaker_number=new_connection.target_breaker_number,
            target_breaker_spec=new_connection.target_breaker_spec,
            hierarchy_relation=new_connection.hierarchy_relation,
            upstream_downstream=new_connection.upstream_downstream,
            parallel_count=new_connection.parallel_count,
            rated_current=new_connection.rated_current,
            cable_length=new_connection.cable_length,
            source_device_photo=new_connection.source_device_photo,
            target_device_photo=new_connection.target_device_photo,
            remark=new_connection.remark,
            installation_date=new_connection.installation_date,
            created_at=new_connection.created_at,
            updated_at=new_connection.updated_at
        )
        
        # 手动处理日期字段序列化
        response_data = {
            "id": new_connection.id,
            "source_device_id": new_connection.source_device_id,
            "target_device_id": new_connection.target_device_id,
            "source_device_name": source_device.name,
            "target_device_name": target_device.name,
            "connection_type": new_connection.connection_type,
            "cable_model": new_connection.cable_model,
            "source_fuse_number": new_connection.source_fuse_number,
            "source_fuse_spec": new_connection.source_fuse_spec,
            "source_breaker_number": new_connection.source_breaker_number,
            "source_breaker_spec": new_connection.source_breaker_spec,
            "target_fuse_number": new_connection.target_fuse_number,
            "target_fuse_spec": new_connection.target_fuse_spec,
            "target_breaker_number": new_connection.target_breaker_number,
            "target_breaker_spec": new_connection.target_breaker_spec,
            "hierarchy_relation": new_connection.hierarchy_relation,
            "upstream_downstream": new_connection.upstream_downstream,
            "parallel_count": new_connection.parallel_count,
            "rated_current": new_connection.rated_current,
            "cable_length": new_connection.cable_length,
            "source_device_photo": new_connection.source_device_photo,
            "target_device_photo": new_connection.target_device_photo,
            "remark": new_connection.remark,
            "installation_date": new_connection.installation_date.isoformat() if new_connection.installation_date else None,
            "created_at": new_connection.created_at.isoformat() if new_connection.created_at else None,
            "updated_at": new_connection.updated_at.isoformat() if new_connection.updated_at else None
        }
        
        return JSONResponse(content={
            "success": True,
            "message": "连接创建成功",
            "data": response_data
        })
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"创建连接失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"创建连接失败: {str(e)}")


@app.put("/api/connections/{connection_id}")
async def update_connection(
    connection_id: int,
    source_device_id: Optional[int] = Form(None),
    target_device_id: Optional[int] = Form(None),
    source_port: Optional[str] = Form(None),
    target_port: Optional[str] = Form(None),
    connection_type: Optional[str] = Form(None),
    cable_model: Optional[str] = Form(None),
    source_fuse_number: Optional[str] = Form(None),
    source_fuse_spec: Optional[str] = Form(None),
    source_breaker_number: Optional[str] = Form(None),
    source_breaker_spec: Optional[str] = Form(None),
    target_fuse_number: Optional[str] = Form(None),
    target_fuse_spec: Optional[str] = Form(None),
    target_breaker_number: Optional[str] = Form(None),
    target_breaker_spec: Optional[str] = Form(None),
    hierarchy_relation: Optional[str] = Form(None),
    upstream_downstream: Optional[str] = Form(None),
    parallel_count: Optional[int] = Form(None),
    rated_current: Optional[float] = Form(None),
    cable_length: Optional[float] = Form(None),
    source_device_photo: Optional[str] = Form(None),
    target_device_photo: Optional[str] = Form(None),
    remark: Optional[str] = Form(None),
    installation_date: Optional[str] = Form(None),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """
    更新连接信息
    需要管理员密码验证
    """
    try:
        # 验证管理员密码
        if not verify_admin_password(password):
            raise HTTPException(status_code=401, detail="密码错误")
        
        # 查找要更新的连接
        existing_connection = db.query(Connection).filter(Connection.id == connection_id).first()
        if not existing_connection:
            raise HTTPException(status_code=404, detail="连接不存在")
        
        # 如果要更新设备ID，验证设备是否存在
        if source_device_id is not None:
            source_device = db.query(Device).filter(Device.id == source_device_id).first()
            if not source_device:
                raise HTTPException(status_code=404, detail=f"源设备ID {source_device_id} 不存在")
            existing_connection.source_device_id = source_device_id
        
        if target_device_id is not None:
            target_device = db.query(Device).filter(Device.id == target_device_id).first()
            if not target_device:
                raise HTTPException(status_code=404, detail=f"目标设备ID {target_device_id} 不存在")
            existing_connection.target_device_id = target_device_id
        
        # 更新端口字段
        if source_port is not None:
            existing_connection.source_port = source_port
        if target_port is not None:
            existing_connection.target_port = target_port
        
        # 更新其他字段
        if connection_type is not None:
            existing_connection.connection_type = connection_type
        if cable_model is not None:
            existing_connection.cable_model = cable_model
        if source_fuse_number is not None:
            existing_connection.source_fuse_number = source_fuse_number
        if source_fuse_spec is not None:
            existing_connection.source_fuse_spec = source_fuse_spec
        if source_breaker_number is not None:
            existing_connection.source_breaker_number = source_breaker_number
        if source_breaker_spec is not None:
            existing_connection.source_breaker_spec = source_breaker_spec
        if target_fuse_number is not None:
            existing_connection.target_fuse_number = target_fuse_number
        if target_fuse_spec is not None:
            existing_connection.target_fuse_spec = target_fuse_spec
        if target_breaker_number is not None:
            existing_connection.target_breaker_number = target_breaker_number
        if target_breaker_spec is not None:
            existing_connection.target_breaker_spec = target_breaker_spec
        if hierarchy_relation is not None:
            existing_connection.hierarchy_relation = hierarchy_relation
        if upstream_downstream is not None:
            existing_connection.upstream_downstream = upstream_downstream
        if parallel_count is not None:
            existing_connection.parallel_count = parallel_count
        if rated_current is not None:
            existing_connection.rated_current = rated_current
        if cable_length is not None:
            existing_connection.cable_length = cable_length
        if source_device_photo is not None:
            existing_connection.source_device_photo = source_device_photo
        if target_device_photo is not None:
            existing_connection.target_device_photo = target_device_photo
        if remark is not None:
            existing_connection.remark = remark
        if installation_date is not None:
            try:
                existing_connection.installation_date = datetime.strptime(installation_date, '%Y-%m-%d').date()
            except ValueError:
                existing_connection.installation_date = None
        
        existing_connection.updated_at = datetime.now()
        
        db.commit()
        db.refresh(existing_connection)
        
        # 构建响应数据
        response_data = {
            "id": existing_connection.id,
            "source_device_id": existing_connection.source_device_id,
            "target_device_id": existing_connection.target_device_id,
            "source_device_name": existing_connection.source_device.name,
            "target_device_name": existing_connection.target_device.name,
            "source_port": existing_connection.source_port,
            "target_port": existing_connection.target_port,
            "connection_type": existing_connection.connection_type,
            "cable_model": existing_connection.cable_model,
            "source_fuse_number": existing_connection.source_fuse_number,
            "source_fuse_spec": existing_connection.source_fuse_spec,
            "source_breaker_number": existing_connection.source_breaker_number,
            "source_breaker_spec": existing_connection.source_breaker_spec,
            "target_fuse_number": existing_connection.target_fuse_number,
            "target_fuse_spec": existing_connection.target_fuse_spec,
            "target_breaker_number": existing_connection.target_breaker_number,
            "target_breaker_spec": existing_connection.target_breaker_spec,
            "hierarchy_relation": existing_connection.hierarchy_relation,
            "upstream_downstream": existing_connection.upstream_downstream,
            "parallel_count": existing_connection.parallel_count,
            "rated_current": existing_connection.rated_current,
            "cable_length": existing_connection.cable_length,
            "source_device_photo": existing_connection.source_device_photo,
            "target_device_photo": existing_connection.target_device_photo,
            "remark": existing_connection.remark,
            "installation_date": existing_connection.installation_date.isoformat() if existing_connection.installation_date else None,
            "created_at": existing_connection.created_at.isoformat() if existing_connection.created_at else None,
            "updated_at": existing_connection.updated_at.isoformat() if existing_connection.updated_at else None
        }
        
        return JSONResponse(content={
            "success": True,
            "message": "连接更新成功",
            "data": response_data
        })
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"更新连接失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"更新连接失败: {str(e)}")


@app.delete("/api/connections/{connection_id}")
async def delete_connection(
    connection_id: int,
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """
    删除连接
    需要管理员密码验证
    """
    try:
        # 验证管理员密码
        if not verify_admin_password(password):
            raise HTTPException(status_code=401, detail="密码错误")
        
        # 查找要删除的连接
        connection = db.query(Connection).filter(Connection.id == connection_id).first()
        if not connection:
            raise HTTPException(status_code=404, detail="连接不存在")
        
        # 删除连接
        db.delete(connection)
        db.commit()
        
        return JSONResponse(content={
            "success": True,
            "message": "连接删除成功"
        })
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"删除连接失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"删除连接失败: {str(e)}")


@app.get("/api/connections/{connection_id}", response_model=ConnectionResponse)
async def get_connection(
    connection_id: int,
    db: Session = Depends(get_db)
):
    """
    获取单个连接详情
    """
    try:
        connection = db.query(Connection).filter(Connection.id == connection_id).first()
        if not connection:
            raise HTTPException(status_code=404, detail="连接不存在")
        
        # 手动处理日期字段的序列化
        installation_date_str = connection.installation_date.isoformat() if connection.installation_date else None
        created_at_str = connection.created_at.isoformat() if connection.created_at else None
        updated_at_str = connection.updated_at.isoformat() if connection.updated_at else None
        
        response_data = {
            "id": connection.id,
            "source_device_id": connection.source_device_id,
            "target_device_id": connection.target_device_id,
            "source_device_name": connection.source_device.name,
            "target_device_name": connection.target_device.name,
            "source_port": build_port_name_with_prefix(
                connection.source_fuse_number, 
                connection.source_breaker_number
            ),
            "target_port": build_port_name_with_prefix(
                connection.target_fuse_number, 
                connection.target_breaker_number
            ),
            "connection_type": connection.connection_type,
            "cable_model": connection.cable_model,
            "source_fuse_number": connection.source_fuse_number,
            "source_fuse_spec": connection.source_fuse_spec,
            "source_breaker_number": connection.source_breaker_number,
            "source_breaker_spec": connection.source_breaker_spec,
            "target_fuse_number": connection.target_fuse_number,
            "target_fuse_spec": connection.target_fuse_spec,
            "target_breaker_number": connection.target_breaker_number,
            "target_breaker_spec": connection.target_breaker_spec,
            "hierarchy_relation": connection.hierarchy_relation,
            "upstream_downstream": connection.upstream_downstream,
            "parallel_count": connection.parallel_count,
            "rated_current": connection.rated_current,
            "a_rated_current": connection.a_rated_current,  # A端额定电流
            "b_rated_current": connection.b_rated_current,  # B端额定电流
            "cable_length": connection.cable_length,
            "source_device_photo": connection.source_device_photo,
            "target_device_photo": connection.target_device_photo,
            "remark": connection.remark,
            "installation_date": installation_date_str,
            "created_at": created_at_str,
            "updated_at": updated_at_str
        }
        
        return JSONResponse(content={
            "success": True,
            "data": response_data
        })
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"获取连接详情失败: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"获取连接详情失败: {str(e)}")


# --- 总线式端口拓扑图实现函数 ---

def _create_bus_topology_nodes(device: Device, db: Session, group_size: int = 12) -> dict:
    """为设备创建总线式拓扑节点（总线节点 + 端口节点 + 总线-端口连接）"""
    bus_nodes = []
    port_nodes = []
    bus_port_edges = []
    connected_device_ports = []  # 存储对端设备的端口节点
    
    try:
        # 1. 获取设备的所有端口信息
        device_ports = _extract_device_ports(device, db)
        
        # 2. 按电流方向分组端口
        port_groups = _group_ports_by_direction(device, device_ports)
        
        # 3. 为每个方向创建总线节点和端口节点（支持按端口数量分组）
        for direction, ports in port_groups.items():
            if ports:  # 只有当该方向有端口时才创建总线
                # 按端口数量分组，每group_size个端口一条总线
                port_chunks = _split_ports_into_chunks(ports, max_ports_per_bus=group_size)
                
                for chunk_index, port_chunk in enumerate(port_chunks):
                    # 创建总线节点（如果有多条总线，添加编号）
                    bus_suffix = f"_{chunk_index + 1}" if len(port_chunks) > 1 else ""
                    bus_node = _create_bus_node(device, direction + bus_suffix, port_chunk)
                    bus_nodes.append(bus_node)
                    
                    # 创建端口节点并连接到总线
                    for port in port_chunk:
                        port_node = _create_port_node_for_bus(device, port, direction, db)
                        port_nodes.append(port_node)
                        
                        # 创建总线到端口的连接
                        bus_port_edge = _create_bus_to_port_edge(bus_node['id'], port_node['id'])
                        bus_port_edges.append(bus_port_edge)
                        
                        # 为有连接的端口创建对端设备的简化端口节点
                        if port.get('connected_device_id') and port.get('connection_id'):
                            connected_port = _create_connected_device_port_node(device, port, db)
                            if connected_port:
                                connected_device_ports.append(connected_port)
        
        # 将对端设备端口节点添加到结果中（去重处理）
        # 使用字典去重，避免同一个对端端口被重复添加
        unique_connected_ports = {}
        for connected_port in connected_device_ports:
            port_id = connected_port['id']
            if port_id not in unique_connected_ports:
                unique_connected_ports[port_id] = connected_port
        
        port_nodes.extend(unique_connected_ports.values())
        
        return {
            'bus_nodes': bus_nodes,
            'port_nodes': port_nodes,
            'bus_port_edges': bus_port_edges
        }
        
    except Exception as e:
        print(f"创建总线式拓扑节点失败: {str(e)}")
        # 如果总线式创建失败，回退到标准端口节点
        standard_ports = _create_port_nodes(device, db)
        return {
            'bus_nodes': [],
            'port_nodes': standard_ports,
            'bus_port_edges': []
        }


def _extract_device_ports(device: Device, db: Session) -> list:
    """提取设备的所有端口信息，基于实际电流方向正确分类"""
    ports = []
    
    # 从作为目标设备的连接中提取端口
    for conn in device.target_connections:
        # 根据upstream_downstream字段判断实际电流方向
        # 对于选中设备作为target_device的连接：
        # - 如果upstream_downstream为"上游"，说明电流从source流向target，选中设备接收电力（输入端口）
        # - 如果upstream_downstream为"下游"，说明电流从target流向source，选中设备输出电力（输出端口）
        actual_direction = _determine_actual_port_direction(
            device.id, conn.source_device_id, conn.target_device_id, 
            conn.upstream_downstream, 'target'
        )
        
        if conn.target_fuse_number:
            ports.append({
                'name': conn.target_fuse_number,
                'type': '熔断器',
                'spec': conn.target_fuse_spec,
                'connection_id': conn.id,
                'direction': actual_direction,
                'connected_device_id': conn.source_device_id,
                'upstream_downstream': conn.upstream_downstream
            })
        if conn.target_breaker_number:
            ports.append({
                'name': conn.target_breaker_number,
                'type': '断路器',
                'spec': conn.target_breaker_spec,
                'connection_id': conn.id,
                'direction': actual_direction,
                'connected_device_id': conn.source_device_id,
                'upstream_downstream': conn.upstream_downstream
            })
    
    # 从作为源设备的连接中提取端口
    for conn in device.source_connections:
        # 对于选中设备作为source_device的连接：
        # - 如果upstream_downstream为"上游"，说明电流从source流向target，选中设备输出电力（输出端口）
        # - 如果upstream_downstream为"下游"，说明电流从target流向source，选中设备接收电力（输入端口）
        actual_direction = _determine_actual_port_direction(
            device.id, conn.source_device_id, conn.target_device_id, 
            conn.upstream_downstream, 'source'
        )
        
        if conn.source_fuse_number:
            ports.append({
                'name': conn.source_fuse_number,
                'type': '熔断器',
                'spec': conn.source_fuse_spec,
                'connection_id': conn.id,
                'direction': actual_direction,
                'connected_device_id': conn.target_device_id,
                'upstream_downstream': conn.upstream_downstream
            })
        if conn.source_breaker_number:
            ports.append({
                'name': conn.source_breaker_number,
                'type': '断路器',
                'spec': conn.source_breaker_spec,
                'connection_id': conn.id,
                'direction': actual_direction,
                'connected_device_id': conn.target_device_id,
                'upstream_downstream': conn.upstream_downstream
            })
    
    return ports


def _group_ports_by_direction(device: Device, ports: list) -> dict:
    """按实际电流方向分组端口（基于连接关系和upstream_downstream字段）"""
    groups = {
        'input': [],
        'output': [],
        'bidirectional': []
    }
    
    for port in ports:
        # 使用基于实际连接关系的方向判断，而不是基于设备类型和端口名称的推测
        direction = port.get('direction', 'bidirectional')
        
        # 如果方向仍然不明确，则使用设备类型和端口名称作为备用判断
        if direction == 'bidirectional':
            direction = _determine_port_direction(device.device_type, port['name'], 'bidirectional')
        
        groups[direction].append(port)
    
    return groups


def _determine_actual_port_direction(selected_device_id: int, source_device_id: int, target_device_id: int, upstream_downstream: str, device_role: str) -> str:
    """基于upstream_downstream字段和设备角色确定端口的实际电流方向
        selected_device_id: 用户选中查看的设备ID
        source_device_id: 连接中的源设备ID
        target_device_id: 连接中的目标设备ID
        upstream_downstream: 连接的上下游关系（"上游"或"下游"）
        device_role: 选中设备在此连接中的角色（"source"或"target"）
    
    Returns:
        str: 端口方向（"input"、"output"或"bidirectional"）
    """
    if not upstream_downstream:
        return 'bidirectional'
    
    # 根据upstream_downstream字段和设备角色判断电流方向
    if upstream_downstream == "上游":
        # 上游关系：电流从source设备流向target设备
        if device_role == 'source':
            # 选中设备是源设备，输出电力
            return 'output'
        else:  # device_role == 'target'
            # 选中设备是目标设备，接收电力
            return 'input'
    elif upstream_downstream == "下游":
        # 下游关系：电流从target设备流向source设备
        if device_role == 'source':
            # 选中设备是源设备，但电流方向相反，所以接收电力
            return 'input'
        else:  # device_role == 'target'
            # 选中设备是目标设备，但电流方向相反，所以输出电力
            return 'output'
    else:
        # 未知的upstream_downstream值，使用备用判断
        return 'bidirectional'

def _determine_port_direction(device_type: str, port_name: str, default_direction: str) -> str:
    """判断端口的电流方向"""
    if not device_type or not port_name:
        return default_direction
    
    device_type = device_type.strip()
    port_name = port_name.strip().upper()
    
    # 基于设备类型的端口方向规则
    device_rules = {
        '发电机组': {
        },
        'UPS': {
            'input': ['输入', 'INPUT', 'AC_IN', 'BYPASS', '旁路', '进线'],
            'output': ['输出', 'OUTPUT', 'AC_OUT', '出线']
        },
        '变压器': {
            'input': ['一次', 'PRIMARY', '高压', 'HV', '进线'],
            'output': ['二次', 'SECONDARY', '低压', 'LV', '出线']
        },
        '高压配电柜': {
            'input': ['进线', 'INPUT', '母线进线', '主进线'],
            'output': ['出线', 'OUTPUT', '馈线', '分支']
        },
        '低压配电柜': {
            'input': ['进线', 'INPUT', '母线进线', '主进线'],
            'output': ['出线', 'OUTPUT', '馈线', '分支']
        },
        'ATS柜': {
            'input': ['进线', 'INPUT', '常用', '备用', 'N', 'E'],
            'output': ['出线', 'OUTPUT', '负载']
        }
    }
    
    # 检查设备类型规则
    rules = device_rules.get(device_type, {})
    
    for direction, keywords in rules.items():
        if any(keyword in port_name for keyword in keywords):
            return direction
    
    # 通用关键词检查
    if any(keyword in port_name for keyword in ['进线', 'INPUT', 'IN', '输入', '一次', 'PRIMARY']):
        return 'input'
    elif any(keyword in port_name for keyword in ['出线', 'OUTPUT', 'OUT', '输出', '二次', 'SECONDARY', '馈线']):
        return 'output'
    
    # 如果无法判断，使用默认方向
    return default_direction


def _create_bus_node(device: Device, direction: str, ports: list) -> dict:
    """创建总线节点"""
    direction_labels = {
        'input': '输入侧',
        'output': '输出侧',
        'bidirectional': '双向'
    }
    
    direction_colors = {
        'input': {'background': '#E8F5E8', 'border': '#4CAF50'},
        'output': {'background': '#FFF3E0', 'border': '#FF9800'},
        'bidirectional': {'background': '#E3F2FD', 'border': '#2196F3'}
    }
    
    port_names = [port['name'] for port in ports]
    
    return {
        'id': f"bus_{device.id}_{direction}",
        'type': 'bus',
        'label': f"{direction_labels.get(direction, direction)}总线",
        'title': f"{device.name} {direction_labels.get(direction, direction)}侧端口总线\n包含端口: {', '.join(port_names[:5])}{'...' if len(port_names) > 5 else ''}",
        'device_id': device.id,
        'device_name': device.name,
        'device_type': device.device_type,
        'direction': direction,
        'port_count': len(ports),
        'port_list': port_names,
        'shape': 'box',
        'color': {
            'background': direction_colors.get(direction, {}).get('background', '#E3F2FD'),
            'border': direction_colors.get(direction, {}).get('border', '#2196F3'),
            'highlight': {
                'background': '#BBDEFB',
                'border': '#1976D2'
            }
        },
        'font': {
            'size': 14,
            'color': '#1565C0',
            'face': 'Arial'
        },
        'margin': 10,
        'widthConstraint': {'minimum': 120, 'maximum': 200},
        'heightConstraint': {'minimum': 40, 'maximum': 80}
    }


def _split_ports_into_chunks(ports: list, max_ports_per_bus: int = 10) -> list:
    """
    将端口列表按指定数量分组，每组创建一条总线
    
    Args:
        ports: 端口列表
        max_ports_per_bus: 每条总线最大端口数，默认10个
    
    Returns:
        list: 分组后的端口列表，每个元素是一个端口组
    """
    if not ports:
        return []
    
    # 将端口列表按指定数量分组
    chunks = []
    for i in range(0, len(ports), max_ports_per_bus):
        chunk = ports[i:i + max_ports_per_bus]
        chunks.append(chunk)
    
    return chunks


def _create_connected_device_node(device: Device, port: dict, db: Session) -> dict:
    """为总线式布局创建对端设备节点"""
    if not port.get('connected_device_id'):
        return None
        
    connected_device = db.query(Device).filter(Device.id == port['connected_device_id']).first()
    if not connected_device:
        return None
    
    # 设备类型颜色映射
    device_colors = {
        '发电机组': '#4CAF50',
        'UPS': '#2196F3', 
        '变压器': '#FF9800',
        '配电柜': '#9C27B0',
        '开关柜': '#795548',
        '电池组': '#607D8B'
    }
    
    base_color = device_colors.get(connected_device.device_type, '#757575')
    
    return {
        'id': f"connected_device_{connected_device.id}_from_{device.id}_{port['name']}",
        'type': 'connected_device',
        'label': connected_device.name,
        'title': f"对端设备: {connected_device.name}\n类型: {connected_device.device_type}\n站点: {connected_device.station}\n型号: {connected_device.model or 'N/A'}",
        'device_id': connected_device.id,
        'device_name': connected_device.name,
        'device_type': connected_device.device_type,
        'station': connected_device.station,
        'model': connected_device.model,
        'shape': 'box',
        'size': 30,
        'color': {
            'background': base_color,
            'border': '#424242',
            'highlight': {
                'background': _adjust_color_brightness(base_color, 1.2),
                'border': '#212121'
            }
        },
        'font': {
            'size': 12,
            'color': '#FFFFFF'
        }
    }

def _create_connected_device_port_node(device: Device, port: dict, db: Session) -> dict:
    """为总线式布局创建对端设备的简化端口节点"""
    if not port.get('connected_device_id') or not port.get('connection_id'):
        return None
    
    try:
        # 获取对端设备信息
        connected_device = db.query(Device).filter(Device.id == port['connected_device_id']).first()
        if not connected_device:
            return None
        
        # 获取连接信息
        connection = db.query(Connection).filter(Connection.id == port['connection_id']).first()
        if not connection:
            return None
        
        # 确定对端端口信息
        connected_port_name = f"连接{connection.id}"  # 改进默认名称
        connected_port_type = "未知"
        
        if connection.source_device_id == device.id:
            # 当前设备是源设备，对端是目标设备的端口
            if connection.target_fuse_number:
                connected_port_name = f"熔丝-{connection.target_fuse_number}"
                connected_port_type = "熔丝"
            elif connection.target_breaker_number:
                connected_port_name = f"空开-{connection.target_breaker_number}"
                connected_port_type = "空开"
            else:
                # 当目标端口信息缺失时，使用更有意义的标识
                connected_port_name = f"入线{connection.id}"
        else:
            # 当前设备是目标设备，对端是源设备的端口
            if connection.source_fuse_number:
                connected_port_name = f"熔丝-{connection.source_fuse_number}"
                connected_port_type = "熔丝"
            elif connection.source_breaker_number:
                connected_port_name = f"空开-{connection.source_breaker_number}"
                connected_port_type = "空开"
            else:
                # 当源端口信息缺失时，使用更有意义的标识
                connected_port_name = f"出线{connection.id}"
        
        # 端口颜色配置
        port_colors = {
            '熔丝': '#FF9800',  # 橙色
            '空开': '#4CAF50',  # 绿色
            '接触器': '#2196F3',  # 蓝色
            '开关': '#9C27B0'     # 紫色
        }
        
        base_color = port_colors.get(connected_port_type, '#757575')
        
        # 使用连接ID确保节点ID的唯一性，避免重复ID问题
        return {
            'id': f"connected_port_{connected_device.id}_{connection.id}_{connected_port_name.replace('-', '_')}",
            'type': 'connected_port',
            'label': f"{connected_device.name}·{connected_port_name}",
            'title': f"{connected_device.name}·{connected_port_name}\n设备类型: {connected_device.device_type or 'N/A'}\n连接到: {device.name}·{port['name']}",
            'device_id': connected_device.id,
            'device_name': connected_device.name,
            'device_type': connected_device.device_type,
            'station': connected_device.station,
            'port_name': connected_port_name,
            'port_type': connected_port_type,
            'connection_id': connection.id,
            'source_device_id': device.id,
            'source_port_name': port['name'],
            'shape': 'circle',
            'size': 20,
            'color': {
                'background': base_color,
                'border': '#424242',
                'highlight': {
                    'background': _adjust_color_brightness(base_color, 1.2),
                    'border': '#212121'
                }
            },
            'font': {
                'size': 9,
                'color': '#212121'
            },
            # 设置位置，使其显示在图的右侧
            'x': 300,  # 相对于查询设备的右侧位置
            'y': 0     # 垂直居中
        }
        
    except Exception as e:
        print(f"创建对端设备端口节点失败: {str(e)}")
        return None


def _create_port_node_for_bus(device: Device, port: dict, direction: str, db: Session) -> dict:
    """为总线式布局创建端口节点，包含对端设备信息"""
    port_colors = {
        '熔丝': '#FF9800',  # 橙色
        '空开': '#4CAF50',  # 绿色
        '接触器': '#2196F3',  # 蓝色
        '开关': '#9C27B0'     # 紫色
    }
    
    # 将端口类型中的英文转换为中文
    port_type_chinese = port['type']
    if 'fuse' in port['type'].lower() or '熔断器' in port['type']:
        port_type_chinese = '熔丝'
    elif 'breaker' in port['type'].lower() or '断路器' in port['type']:
        port_type_chinese = '空开'
    
    base_color = port_colors.get(port_type_chinese, '#757575')
    
    # 获取对端设备信息
    connected_device_info = "空闲"
    connected_port_info = "空闲"
    if port.get('connected_device_id'):
        connected_device = db.query(Device).filter(Device.id == port['connected_device_id']).first()
        if connected_device:
            connected_device_info = f"{connected_device.name} ({connected_device.device_type})"
            
            # 获取对端端口信息
            if port.get('connection_id'):
                connection = db.query(Connection).filter(Connection.id == port['connection_id']).first()
                if connection:
                    # 根据当前设备在连接中的角色确定对端端口
                    if connection.source_device_id == device.id:
                        # 当前设备是源设备，对端是目标设备的端口
                        if connection.target_fuse_number:
                            connected_port_info = f"熔丝-{connection.target_fuse_number}"
                        elif connection.target_breaker_number:
                            connected_port_info = f"空开-{connection.target_breaker_number}"
                    else:
                        # 当前设备是目标设备，对端是源设备的端口
                        if connection.source_fuse_number:
                            connected_port_info = f"熔丝-{connection.source_fuse_number}"
                        elif connection.source_breaker_number:
                            connected_port_info = f"空开-{connection.source_breaker_number}"
    
    # 使用用户提交的原始端口名，不做任何简化或翻译
    original_port_name = port['name']

    # 使用连接ID确保节点ID的唯一性，避免重复ID问题
    connection_suffix = f"_{port.get('connection_id', 'no_conn')}"
    
    return {
        'id': f"port_{device.id}_{original_port_name.replace(' ', '_')}{connection_suffix}",
        'type': 'port',
        'label': original_port_name,
        'title': f"{device.name} - {original_port_name}\n类型: {port_type_chinese}\n规格: {port.get('spec', 'N/A')}\n方向: {direction}\n连接到: {connected_device_info}\n对端端口: {connected_port_info}",
        'device_id': device.id,
        'device_name': device.name,
        'device_type': device.device_type,
        'station': device.station,
        'port_name': original_port_name,
        'port_type': port_type_chinese,
        'port_spec': port.get('spec'),
        'direction': direction,
        'parent_bus': f"bus_{device.id}_{direction}",
        'connection_id': port.get('connection_id'),
        'connected_device_id': port.get('connected_device_id'),
        'connected_device_info': connected_device_info,
        'connected_port_info': connected_port_info,
        'shape': 'circle',
        'size': 25,
        'color': {
            'background': base_color,
            'border': '#424242',
            'highlight': {
        'background': _adjust_color_brightness(base_color, 1.2),
                'border': '#212121'
            }
        },
        'font': {
            'size': 10,
            'color': '#212121'
        }
    }


def _create_bus_to_port_edge(bus_id: str, port_id: str) -> dict:
    """创建总线到端口的连接边"""
    return {
        'id': f"bus_port_{bus_id}_{port_id}",
        'type': 'bus_connection',
        'from': bus_id,
        'to': port_id,
        'arrows': 'none',
        'color': {
            'color': '#90A4AE',
            'width': 2,
            'opacity': 0.6
        },
        'dashes': [5, 5],
        'smooth': {
            'enabled': True,
            'type': 'curvedCW',
            'roundness': 0.2
        },
        'length': 50
    }


def _create_bus_port_edges(connection, direction: str) -> list:
    """为总线式布局创建端口到端口的连接边"""
    edges = []
    
    try:
        if direction == "upstream":
            # 上游连接：从源设备端口到目标设备端口
            source_device = connection.source_device
            target_device = connection.target_device
            
            # 源端口（输出端口）- 使用与_create_port_node_for_bus一致的ID格式
            source_ports = []
            if connection.source_fuse_number:
                source_ports.append(f"port_{source_device.id}_熔丝-{connection.source_fuse_number}_{connection.id}")
            if connection.source_breaker_number:
                source_ports.append(f"port_{source_device.id}_空开-{connection.source_breaker_number}_{connection.id}")
            
            # 目标端口（输入端口）- 使用与_create_port_node_for_bus一致的ID格式
            target_ports = []
            if connection.target_fuse_number:
                target_ports.append(f"port_{target_device.id}_熔丝-{connection.target_fuse_number}_{connection.id}")
            if connection.target_breaker_number:
                target_ports.append(f"port_{target_device.id}_空开-{connection.target_breaker_number}_{connection.id}")
            
            # 创建端口间连接
            for source_port in source_ports:
                for target_port in target_ports:
                    edge = {
                        'id': f"port_conn_{connection.id}_{source_port}_{target_port}",
                        'type': 'port_connection',
                        'from': source_port,
                        'to': target_port,
                        'arrows': 'to',
                        'label': connection.connection_type or connection.cable_type or '',
                        'connection_id': connection.id,
                        'connection_type': connection.connection_type,
                        'cable_type': connection.cable_type,
                        'cable_model': connection.cable_model,
                        'remark': connection.remark,
                        'color': {
                            'color': _get_connection_color(connection.connection_type),
                            'width': _get_connection_width_from_connection(connection),
                            'highlight': _get_connection_highlight_color(connection.connection_type)
                        },
                        'smooth': {
                            'enabled': True,
                            'type': 'dynamic'
                        }
                    }
                    edges.append(edge)
        
        elif direction == "downstream":
            # 下游连接：从当前设备端口到目标设备端口
            source_device = connection.source_device
            target_device = connection.target_device
            
            # 源端口（输出端口）- 使用与_create_port_node_for_bus一致的ID格式
            source_ports = []
            if connection.source_fuse_number:
                source_ports.append(f"port_{connection.id}_{source_device.id}_熔丝-{connection.source_fuse_number}")
            if connection.source_breaker_number:
                source_ports.append(f"port_{connection.id}_{source_device.id}_空开-{connection.source_breaker_number}")
            
            # 目标端口（输入端口）- 使用与_create_port_node_for_bus一致的ID格式
            target_ports = []
            if connection.target_fuse_number:
                target_ports.append(f"port_{connection.id}_{target_device.id}_熔丝-{connection.target_fuse_number}")
            if connection.target_breaker_number:
                target_ports.append(f"port_{connection.id}_{target_device.id}_空开-{connection.target_breaker_number}")
            
            # 创建端口间连接
            for source_port in source_ports:
                for target_port in target_ports:
                    edge = {
                        'id': f"port_conn_{connection.id}_{source_port}_{target_port}",
                        'type': 'port_connection',
                        'from': source_port,
                        'to': target_port,
                        'arrows': 'to',
                        'label': connection.connection_type or connection.cable_type or '',
                        'connection_id': connection.id,
                        'connection_type': connection.connection_type,
                        'cable_type': connection.cable_type,
                        'cable_model': connection.cable_model,
                        'remark': connection.remark,
                        'color': {
                            'color': _get_connection_color(connection.connection_type),
                            'width': _get_connection_width_from_connection(connection),
                            'highlight': _get_connection_highlight_color(connection.connection_type)
                        },
                        'smooth': {
                            'enabled': True,
                            'type': 'dynamic'
                        }
                    }
                    edges.append(edge)
    
    except Exception as e:
        print(f"创建总线端口连接失败: {str(e)}")
    
    return edges


def _get_connection_color(connection_type: str) -> str:
    """根据连接类型获取连线颜色"""
    colors = {
        '电力连接': '#F44336',    # 红色
        '控制连接': '#2196F3',    # 蓝色
        '通信连接': '#4CAF50',    # 绿色
        '接地连接': '#795548',    # 棕色
        '电缆连接': '#FF5722',    # 深橙色
        '母线连接': '#9C27B0'     # 紫色
    }
    return colors.get(connection_type, '#424242')


def _get_connection_width_from_connection(connection) -> int:
    """从连接对象推导电压等级并获取连线宽度"""
    try:
        # 尝试从额定电流推导电压等级
        rated_current = None
        
        # 优先使用数值型的rated_current字段
        if hasattr(connection, 'rated_current') and connection.rated_current:
            rated_current = float(connection.rated_current)
        # 其次尝试从A端额定电流解析
        elif hasattr(connection, 'a_rated_current') and connection.a_rated_current:
            # 提取数字部分，如"63A" -> 63
            import re
            match = re.search(r'(\d+(?:\.\d+)?)', str(connection.a_rated_current))
            if match:
                rated_current = float(match.group(1))
        # 最后尝试从B端额定电流解析
        elif hasattr(connection, 'b_rated_current') and connection.b_rated_current:
            import re
            match = re.search(r'(\d+(?:\.\d+)?)', str(connection.b_rated_current))
            if match:
                rated_current = float(match.group(1))
        
        # 根据额定电流推导电压等级和线宽
        if rated_current:
            if rated_current >= 1000:      # 大电流，可能是高压
                return 4
            elif rated_current >= 100:     # 中等电流，可能是中压
                return 3
            else:                          # 小电流，低压
                return 2
        
        # 如果无法从电流推导，尝试从连接类型推导
        if hasattr(connection, 'connection_type') and connection.connection_type:
            connection_type = connection.connection_type.lower()
            if 'high' in connection_type or '高压' in connection_type:
                return 4
            elif 'medium' in connection_type or '中压' in connection_type:
                return 3
        
        # 默认返回低压线宽
        return 2
        
    except Exception as e:
        print(f"推导连线宽度失败: {str(e)}")
        return 2


def _get_connection_width(voltage_level) -> int:
    """根据电压等级获取连线宽度（保留原函数以兼容其他调用）"""
    if voltage_level is None:
        return 2
    
    try:
        voltage = float(voltage_level)
        if voltage >= 10000:      # 高压
            return 4
        elif voltage >= 1000:     # 中压
            return 3
        else:                     # 低压
            return 2
    except (ValueError, TypeError):
        return 2


def _get_connection_highlight_color(connection_type: str) -> str:
    """获取连接高亮颜色"""
    base_color = _get_connection_color(connection_type)
    return _adjust_color_brightness(base_color, 1.3)


def _adjust_color_brightness(hex_color: str, factor: float) -> str:
    """调整颜色亮度"""
    try:
        # 移除 # 符号
        hex_color = hex_color.lstrip('#')
        
        # 转换为 RGB
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        
        # 调整亮度
        r = min(255, max(0, int(r * factor)))
        g = min(255, max(0, int(g * factor)))
        b = min(255, max(0, int(b * factor)))
        
        # 转换回十六进制
        return f"#{r:02x}{g:02x}{b:02x}"
    except:
        return hex_color  # 如果转换失败，返回原色


# --- 应用启动 ---
if __name__ == "__main__":
    import uvicorn
    print(f"\n🌐 服务器启动地址: http://localhost:{PORT} 或 http://0.0.0.0:{PORT}")
    print(f"📊 管理界面: http://localhost:{PORT}")
    print(f"🔗 连接管理: http://localhost:{PORT}/connections")
    print(f"⚙️  生命周期管理: http://localhost:{PORT}/lifecycle-management")
    print(f"\n注意：应用程序实际运行在端口 {PORT}")
    uvicorn.run(app, host="0.0.0.0", port=PORT, reload=False)

    uvicorn.run(app, host="0.0.0.0", port=PORT, reload=False)