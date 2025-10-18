import os
import logging

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

from fastapi import FastAPI, Request, Depends, Form, UploadFile, File, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse, FileResponse
from fastapi.encoders import jsonable_encoder
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session, aliased
from sqlalchemy import func, or_
import pandas as pd
from typing import List, Optional
from urllib.parse import quote
import io
import traceback # 瀵煎叆 traceback 鐢ㄤ簬鎵撳嵃璇︾粏鐨勯敊璇爢鏍?
from datetime import datetime, timedelta, date
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pydantic import BaseModel
import re
from sqlalchemy import and_
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse, FileResponse
from fastapi.encoders import jsonable_encoder
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session, aliased
from sqlalchemy import func, or_
import pandas as pd
from typing import List, Optional
from urllib.parse import quote
from pathlib import Path
import io
import traceback # 瀵煎叆 traceback 鐢ㄤ簬鎵撳嵃璇︾粏鐨勯敊璇爢鏍?
from datetime import datetime, timedelta, date
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pydantic import BaseModel
import re
from sqlalchemy import and_

# Import configuration
from config import ADMIN_PASSWORD, HOST, PORT

# Fixed imports, using correct function names and models
from models import SessionLocal, Device, Connection, LifecycleRule, create_db_and_tables
from device_types import STANDARD_DEVICE_TYPES, validate_device_type, get_device_type_suggestions, STANDARD_DEVICE_TYPES


# Import error tracking system
from topology_error_tracker import topology_error_tracker, ErrorCategory, ErrorLevel
from port_topology_service import PortTopologyService


# --- Port Statistics Service Class ---

class PortStatisticsService:
    """绔彛缁熻鏈嶅姟绫伙紝鐢ㄤ簬澶勭悊璁惧绔彛浣跨敤鎯呭喌鐨勭粺璁″垎鏋?"""
    def __init__(self, db: Session):
        self.db = db
    
    def _get_device_port_summary(self) -> dict:
        """鑾峰彇璁惧绔彛鎬昏 - 閲囩敤闆嗗悎缁熻閫昏緫锛岀粺璁℃墍鏈夋湁杩炴帴鐨勭鍙?"""
        try:
            # Count total devices
            total_devices = self.db.query(Device).count()
            
            # Use sets to avoid duplicate counting of the same port
            all_ports = set()
            connected_ports = set()
            
            # Get all connection records
            connections = self.db.query(Connection).all()
            
            for conn in connections:
                # Count source ports (A side)
                if conn.source_fuse_number and conn.source_device_id:
                    port_key = f"device_{conn.source_device_id}_fuse_{conn.source_fuse_number}"
                    all_ports.add(port_key)
                    # Determine port status by checking if connection_type field is empty
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
                        
                if conn.source_breaker_number and conn.source_device_id:
                    port_key = f"device_{conn.source_device_id}_breaker_{conn.source_breaker_number}"
                    all_ports.add(port_key)
                    # 閫氳繃杩炴帴绫诲瀷瀛楁鏄惁涓虹┖鍒ゆ柇绔彛浣跨敤鐘舵€?
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
                
                # Count target ports (B side) - meets the design document requirement of "one connection occupies two ports"
                if conn.target_fuse_number and conn.target_device_id:
                    port_key = f"device_{conn.target_device_id}_fuse_{conn.target_fuse_number}"
                    all_ports.add(port_key)
                    # 閫氳繃杩炴帴绫诲瀷瀛楁鏄惁涓虹┖鍒ゆ柇绔彛浣跨敤鐘舵€?
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
                        
                if conn.target_breaker_number and conn.target_device_id:
                    port_key = f"device_{conn.target_device_id}_breaker_{conn.target_breaker_number}"
                    all_ports.add(port_key)
                    # 閫氳繃杩炴帴绫诲瀷瀛楁鏄惁涓虹┖鍒ゆ柇绔彛浣跨敤鐘舵€?
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
            
            total_ports = len(all_ports)
            connected_count = len(connected_ports)
            idle_ports = total_ports - connected_count
            utilization_rate = (connected_count / total_ports * 100) if total_ports > 0 else 0
            
            return {
                "total_devices": total_devices,
                "total_ports": total_ports,
                "connected_ports": connected_count,
                "idle_ports": idle_ports,
                "utilization_rate": round(utilization_rate, 2)
            }
        except Exception as e:
            print(f"鑾峰彇璁惧绔彛鎬昏鏃跺嚭閿? {e}")
            return {
                "total_devices": 0,
                "total_ports": 0,
                "connected_ports": 0,
                "idle_ports": 0,
                "utilization_rate": 0
            }
    
    def get_device_port_details(self, device_id: int) -> dict:
        """鑾峰彇鎸囧畾璁惧鐨勭鍙ｈ鎯?- 鍩轰簬杩炴帴琛ㄤ腑璇ヨ澶囩殑瀹為檯绔彛鏁版嵁"""
        try:
            # 鑾峰彇璁惧淇℃伅
            device = self.db.query(Device).filter(Device.id == device_id).first()
            if not device:
                raise HTTPException(status_code=404, detail="设备不存在")
            
            # 鑾峰彇璇ヨ澶囦綔涓篈绔澶囩殑鎵€鏈夎繛鎺ヨ褰曪紝浠庝腑鎻愬彇瀹為檯绔彛淇℃伅
            connections = self.db.query(Connection).filter(
                Connection.source_device_id == device_id
            ).all()
            
            # 鏀堕泦璇ヨ澶囩殑鎵€鏈夌鍙ｄ俊鎭紙鍩轰簬杩炴帴琛ㄤ腑鐨勫疄闄呮暟鎹級
            ports = []
            port_usage_map = {}
            
            for conn in connections:
                # 澶勭悊鐔斾笣绔彛
                if conn.source_fuse_number:
                    port_key = f"鐔斾笣-{conn.source_fuse_number}"
                    if port_key not in port_usage_map:
                        port_info = {
                            "port_name": port_key,
                            "port_type": "鐔斾笣",
                            "port_number": conn.source_fuse_number,
                            "specification": conn.source_fuse_spec or "未知规格",
                            "rating": self._extract_rating_from_spec(conn.source_fuse_spec or ""),
                            "status": "已连接" if conn.connection_type else "空闲",
                            "connected_device": conn.target_device.name if conn.target_device and conn.connection_type else None,
                            "connection_id": conn.id if conn.connection_type else None
                        }
                        ports.append(port_info)
                        port_usage_map[port_key] = port_info
                
                # 澶勭悊绌哄紑绔彛
                if conn.source_breaker_number:
                    port_key = f"绌哄紑-{conn.source_breaker_number}"
                    if port_key not in port_usage_map:
                        port_info = {
                            "port_name": port_key,
                            "port_type": "绌哄紑",
                            "port_number": conn.source_breaker_number,
                            "specification": conn.source_breaker_spec or "未知规格",
                            "rating": self._extract_rating_from_spec(conn.source_breaker_spec or ""),
                            "status": "已连接" if conn.connection_type else "空闲",
                            "connected_device": conn.target_device.name if conn.target_device and conn.connection_type else None,
                            "connection_id": conn.id if conn.connection_type else None
                        }
                        ports.append(port_info)
                        port_usage_map[port_key] = port_info
            
            # If no ports are found, return an empty list (indicating the device has no port configuration or connection records)
            if not ports:
                return {
                    "device_info": {
                        "id": device.id,
                        "name": device.name,
                        "device_type": device.device_type or "鏈煡",
                        "station": device.station or "鏈煡",
                        "location": device.location or "鏈煡"
                    },
                    "port_summary": {
                        "total_ports": 0,
                        "connected_ports": 0,
                        "idle_ports": 0,
                        "utilization_rate": 0.0
                    },
                    "ports": []
                }
            
            # Statistics
            total_ports = len(ports)
            connected_ports = len([p for p in ports if p["status"] == "已连接"])
            idle_ports = total_ports - connected_ports
            utilization_rate = (connected_ports / total_ports * 100) if total_ports > 0 else 0
            
            return {
                "device_info": {
                    "id": device.id,
                    "name": device.name,
                    "device_type": device.device_type or "鏈煡",
                    "station": device.station or "鏈煡",
                    "location": device.location or "鏈煡"
                },
                "port_summary": {
                    "total_ports": total_ports,
                    "connected_ports": connected_ports,
                    "idle_ports": idle_ports,
                    "utilization_rate": round(utilization_rate, 2)
                },
                "ports": sorted(ports, key=lambda x: (x["port_type"], x["port_number"]))
            }
        except HTTPException:
            raise
        except Exception as e:
            print(f"鑾峰彇璁惧绔彛璇︽儏鏃跺嚭閿? {e}")
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"鑾峰彇璁惧绔彛璇︽儏澶辫触: {str(e)}")
    
    def _extract_rating_from_spec(self, spec: str) -> str:
        """浠庤鏍煎瓧绗︿覆涓彁鍙栭瀹氬€?"""
        if not spec:
            return "鏈煡"
        
        # 灏濊瘯鎻愬彇鏁板瓧鍜屽崟浣嶏紙濡傦細63A, 100A, 2.5mm虏绛夛級
        import re
        pattern = r'(\d+(?:\.\d+)?)\s*([A-Za-z虏]+)'
        match = re.search(pattern, spec)
        if match:
            return f"{match.group(1)}{match.group(2)}"
        else:
            return "鏈煡"

# --- Port topology service class ---
class PortTopologyService:
    """Port topology service class for generating port topology data"""
    def __init__(self, db: Session):
        self.db = db
    
    def get_port_service_data(self, device_id: int, mode: str = "all") -> dict:
        """
        鑾峰彇绔彛鎷撴墤鍥炬暟鎹?- 鏈嶅姟灞傛柟娉?
        
        Args:
            device_id: 璁惧ID
            mode: 鏄剧ず妯″紡锛屽彲閫?'all' 鎴?'used'
        
        Returns:
            鍖呭惈鑺傜偣鍜岃竟鐨勬嫇鎵戝浘鏁版嵁
        """
        try:
            # Get the center device
            logger.info(f"鏌ヨ涓績璁惧锛岃澶嘔D: {device_id}")
            center_device = self.db.query(Device).filter(Device.id == device_id).first()
            if not center_device:
                logger.error(f"设备不存在，设备ID: {device_id}")
                raise HTTPException(status_code=404, detail="设备不存在")
            
            # Get all connections related to this device
            logger.info(f"鏌ヨ涓庤澶嘔D {device_id} 鐩稿叧鐨勬墍鏈夎繛鎺?)
            connections = self.db.query(Connection).filter(
                (Connection.source_device_id == device_id) |
                (Connection.target_device_id == device_id)
            ).all()
            
            if not connections:
                logger.info(f"璁惧ID {device_id} 娌℃湁鐩稿叧杩炴帴")
                return {
                    "nodes": [],
                    "edges": []
                }
            
            # Build nodes and edges
            nodes = []
            edges = []
            
            # Add center device node
            center_node = {
                "id": f"device_{device_id}",
                "label": center_device.name,
                "type": "device",
                "x": 0,
                "y": 0,
                "size": 20,
                "color": "#848484"
            }
            nodes.append(center_node)
            
            # Process connections
            left_ports = []
            right_ports = []
            port_count = 0
            
            for conn in connections:
                # Check if display conditions are met
                if mode == "used" and not conn.connection_type:
                    continue  # 璺宠繃绌洪棽绔彛
                
                # Process source port (local)
                if conn.source_device_id == device_id:
                    source_port = conn.source_fuse_number or conn.source_breaker_number
                    if source_port:
                        port_count += 1
                        port_id = f"port_{device_id}_{source_port}"
                        port_label = f"{source_port}"
                        
                        # Distribute ports left/right based on count
                        if port_count % 2 == 1:
                            port_x = -100
                            port_y = 50 + (port_count // 2) * 80
                            port_side = "left"
                        else:
                            port_x = 100
                            port_y = 50 + ((port_count - 1) // 2) * 80
                            port_side = "right"
                        
                        port_node = {
                            "id": port_id,
                            "label": port_label,
                            "type": "port",
                            "x": port_x,
                            "y": port_y,
                            "size": 10,
                            "color": "#848484"
                        }
                        nodes.append(port_node)
                        
                        # 娣诲姞杩炴帴绾?
                        edge_id = f"edge_{device_id}_{source_port}"
                        edge_data = {
                            "id": edge_id,
                            "from": f"device_{device_id}",
                            "to": port_id,
                            "width": 2,
                            "color": "#848484",
                            "cable_model": conn.cable_model,
                            "remark": conn.remark,
                            "arrows": {"to": {"enabled": True, "type": "arrow"}}
                        }
                        edges.append(edge_data)
                        
                        # 澶勭悊瀵圭璁惧
                        if conn.target_device:
                            target_device = conn.target_device
                            target_port = conn.target_fuse_number or conn.target_breaker_number
                            
                            # 鍒涘缓瀵圭璁惧鑺傜偣
                            target_node_id = f"target_{conn.target_device_id}"
                            target_node = {
                                "id": target_node_id,
                                "label": f"{target_device.name} {target_port}",
                                "type": "device",
                                "x": port_x + (-150 if port_side == "left" else 150),
                                "y": port_y,
                                "size": 15,
                                "color": "#848484"
                            }
                            nodes.append(target_node)
                            
                            # 娣诲姞杩炴帴绾垮埌瀵圭璁惧
                            edge_target_id = f"edge_{device_id}_{source_port}_target"
                            edge_target_data = {
                                "id": edge_target_id,
                                "from": port_id,
                                "to": target_node_id,
                                "width": 2,
                                "color": "#848484",
                                "cable_model": conn.cable_model,
                                "remark": conn.remark,
                                "arrows": {"to": {"enabled": True, "type": "arrow"}}
                            }
                            edges.append(edge_target_data)
                
                # 澶勭悊鐩爣绔彛锛堟湰绔級
                elif conn.target_device_id == device_id:
                    target_port = conn.target_fuse_number or conn.target_breaker_number
                    if target_port:
                        port_count += 1
                        port_id = f"port_{device_id}_{target_port}"
                        port_label = f"{target_port}"
                        
                        # 鏍规嵁绔彛鏁伴噺鍐冲畾宸﹀彸鍒嗗竷
                        if port_count % 2 == 1:
                            port_x = -100
                            port_y = 50 + (port_count // 2) * 80
                            port_side = "left"
                        else:
                            port_x = 100
                            port_y = 50 + ((port_count - 1) // 2) * 80
                            port_side = "right"
                        
                        port_node = {
                            "id": port_id,
                            "label": port_label,
                            "type": "port",
                            "x": port_x,
                            "y": port_y,
                            "size": 10,
                            "color": "#848484"
                        }
                        nodes.append(port_node)
                        
                        # 娣诲姞杩炴帴绾?
                        edge_id = f"edge_{device_id}_{target_port}"
                        edge_data = {
                            "id": edge_id,
                            "from": f"device_{device_id}",
                            "to": port_id,
                            "width": 2,
                            "color": "#848484",
                            "cable_model": conn.cable_model,
                            "remark": conn.remark,
                            "arrows": {"to": {"enabled": True, "type": "arrow"}}
                        }
                        edges.append(edge_data)
                        
                        # 澶勭悊瀵圭璁惧
                        if conn.source_device:
                            source_device = conn.source_device
                            source_port = conn.source_fuse_number or conn.source_breaker_number
                            
                            # 鍒涘缓瀵圭璁惧鑺傜偣
                            source_node_id = f"source_{conn.source_device_id}"
                            source_node = {
                                "id": source_node_id,
                                "label": f"{source_device.name} {source_port}",
                                "type": "device",
                                "x": port_x + (-150 if port_side == "left" else 150),
                                "y": port_y,
                                "size": 15,
                                "color": "#848484"
                            }
                            nodes.append(source_node)
                            
                            # 娣诲姞杩炴帴绾垮埌瀵圭璁惧
                            edge_source_id = f"edge_{device_id}_{target_port}_source"
                            edge_source_data = {
                                "id": edge_source_id,
                                "from": port_id,
                                "to": source_node_id,
                                "width": 2,
                                "color": "#848484",
                                "cable_model": conn.cable_model,
                                "remark": conn.remark,
                                "arrows": {"to": {"enabled": True, "type": "arrow"}}
                            }
                            edges.append(edge_source_data)
            
            logger.info(f"鎴愬姛鐢熸垚绔彛鎷撴墤鍥炬暟鎹紝鑺傜偣鏁? {len(nodes)}, 杈规暟: {len(edges)}")
            return {
                "nodes": nodes,
                "edges": edges
            }
        except Exception as e:
            logger.error(f"鑾峰彇绔彛鎷撴墤鍥炬暟鎹椂鍑洪敊: {str(e)}")
            traceback.print_exc()
            # 杩斿洖鏍囧噯鏍煎紡鐨勭┖鏁版嵁锛岄伩鍏嶅墠绔В鏋愬け璐?
            return {
                "nodes": [],
                "edges": []
            }

# Create FastAPI application instance
app = FastAPI(title="DC Asset Manager", description="Power Resource Management System")

# 绔欑偣鍥炬爣璺敱锛岄伩鍏嶆祻瑙堝櫒璇锋眰 /favicon.ico 404
_STATIC_DIR = Path(__file__).parent / "static"
_FAVICON = _STATIC_DIR / "icons.svg"

@app.get("/favicon.ico")
async def favicon():
    try:
        if _FAVICON.exists():
            return FileResponse(str(_FAVICON), media_type="image/svg+xml")
    except Exception:
        pass
    return RedirectResponse(url="/static/icons.svg")

# Database session dependency
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Health check endpoints
@app.get("/health")
@app.get("/api/health")
async def health_check(db: Session = Depends(get_db)):
    """鍋ュ悍妫€鏌ワ細楠岃瘉鏈嶅姟涓庢暟鎹簱鍙敤鎬?"""
    try:
        # 绠€鍗曟暟鎹簱杩為€氭€ф鏌?
        device_count = db.query(func.count(Device.id)).scalar() or 0
        return {
            "status": "ok",
            "db": "ok",
            "device_count": int(device_count),
            "port": PORT
        }
    except Exception as e:
        topology_error_tracker.log_error(
            category=ErrorCategory.DATABASE_ERROR,
            level=ErrorLevel.ERROR,
            message=f"鍋ュ悍妫€鏌ュけ璐? {str(e)}",
            exception=e
        )
        raise HTTPException(status_code=500, detail="healthcheck failed")

# Register port topology API endpoint
@app.get("/api/port-topology/{device_id}")
async def get_port_topology_data(device_id: int, mode: str = "detailed", db: Session = Depends(get_db)):
    """
    获取端口级拓扑图数据（路由层）：委托 PortTopologyService 构建数据
    """
    try:
        service = PortTopologyService(db)
        result = service.get_port_topology_data(device_id=device_id, mode=mode)
        return JSONResponse(content=result)
    except ValueError as e:
        # 设备不存在等业务错误
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        topology_error_tracker.log_error(
            category=ErrorCategory.API_ERROR,
            level=ErrorLevel.ERROR,
            message=f"端口拓扑接口异常: {str(e)}",
            context={"device_id": device_id, "mode": mode},
            exception=e
        )
        raise HTTPException(status_code=500, detail="Port topology API failed")

# 绔彛鎷撴墤鍥炬暟鎹敓鎴愬嚱鏁?
def generate_port_topology_data(device_id: int):
    """鐢熸垚绔彛鎷撴墤鍥炬暟鎹?"""
    try:
        with SessionLocal() as db:
            device = db.query(Device).filter(Device.id == device_id).first()
            if not device:
                return {"nodes": [], "edges": []}
            
            nodes = []
            edges = []
            
            # 鑾峰彇璁惧鐨勬墍鏈夌鍙ｈ繛鎺?
            connections = db.query(Connection).filter(
                or_(Connection.source_device_id == device_id,
                    Connection.target_device_id == device_id)
            ).all()
            
            # 涓烘瘡涓鍙ｅ垱寤鸿妭鐐?
            for conn in connections:
                if conn.source_device_id == device_id:
                    port_info = f"{conn.source_port}"
                    connected_device = db.query(Device).filter(Device.id == conn.target_device_id).first()
                    connected_port = conn.target_port
                else:
                    port_info = f"{conn.target_port}"
                    connected_device = db.query(Device).filter(Device.id == conn.source_device_id).first()
                    connected_port = conn.source_port
                
                # 鍒涘缓绔彛鑺傜偣
                port_node = {
                    "id": f"port_{conn.id}",
                    "label": port_info,
                    "type": "port",
                    "x": 0,
                    "y": len(nodes) * 100,
                    "size": 15,
                    "color": "#3b82f6",
                    "shape": "circle"
                }
                nodes.append(port_node)
                
                # 鍒涘缓杩炴帴鐨勮澶囪妭鐐?
                if connected_device:
                    device_node = {
                        "id": f"device_{connected_device.id}",
                        "label": f"{connected_device.name}\
{connected_port}",
                        "type": "connected_device",
                        "x": 200,
                        "y": len(nodes) * 100,
                        "size": 20,
                        "color": "#f59e0b",
                        "shape": "box"
                    }
                    nodes.append(device_node)
                    
                    # 鍒涘缓杩炴帴杈?
                    edge = {
                        "id": f"edge_{conn.id}",
                        "from": f"port_{conn.id}",
                        "to": f"device_{connected_device.id}",
                        "width": 2,
                        "color": "#6b7280",
                        "label": conn.cable_model or """
                    }
                    edges.append(edge)
            
            return {"nodes": nodes, "edges": edges}
            
    except Exception as e:
        topology_error_tracker.log_error(
            category=ErrorCategory.API_ERROR,
            level=ErrorLevel.ERROR,
            message=f"绔彛鎷撴墤鍥炬暟鎹敓鎴愬け璐? {str(e)}",
            context={"device_id": device_id},
            exception=e
        )
        return {"nodes": [], "edges": []}
        
        # 3. 鎸夊乏鍙冲垎鍖哄師鍒欏竷灞€绔彛
        port_names = list(port_info.keys())
        total_ports = len(port_names)
        left_ports = port_names[:total_ports//2 + total_ports%2]  # 宸︿晶绔彛锛堝鏁版椂澶氫竴涓級
        right_ports = port_names[total_ports//2 + total_ports%2:]  # 鍙充晶绔彛
        
        # 4. 鍒涘缓绔彛鑺傜偣鍜岃繛鎺?
        for i, port_name in enumerate(left_ports):
            # 宸︿晶绔彛
            port_node_id = f"port_left_{i}"
            port_x = -200
            port_y = -100 + i * 80
            
            port_node = {
                "id": port_node_id,
                "label": port_name,
                "type": "port",
                "x": port_x,
                "y": port_y,
                "size": 15,
                "color": "#10b981",
                "shape": "ellipse",
                "font": {"size": 12}
            }
            nodes.append(port_node)
            
            # 杩炴帴鍒颁腑蹇冭澶?
            edge_to_center = {
                "id": f"edge_center_{port_node_id}",
                "from": f"device_{device_id}",
                "to": port_node_id,
                "width": 2,
                "color": "#6b7280"
            }
            edges.append(edge_to_center)
            
            # 澶勭悊璇ョ鍙ｇ殑鎵€鏈夎繛鎺?
            for conn_info in port_info[port_name]:
                conn = conn_info["connection"]
                remote_device = conn_info["remote_device"]
                remote_port = conn_info["remote_port"]
                
                if remote_device:
                    # 鍒涘缓瀵圭璁惧鑺傜偣
                    remote_node_id = f"remote_{remote_device.id}_{remote_port}"
                    remote_node = {
                        "id": remote_node_id,
                        "label": f"{remote_device.name}\
{remote_port}",
                        "type": "remote_device",
                        "x": port_x - 150,
                        "y": port_y,
                        "size": 20,
                        "color": "#f59e0b",
                        "shape": "box",
                        "font": {"size": 10}
                    }
                    nodes.append(remote_node)
                    
                    # 杩炴帴绾块鑹叉牴鎹繛鎺ョ被鍨?
                    cable_color = "#fbbf24" if conn.connection_type == "浜ゆ祦" else "#ef4444" if conn.connection_type == "鐩存祦" else "#6b7280"
                    
                    # 绠ご鏂瑰悜鏍规嵁涓婁笅娓稿叧绯?
                    arrows = {}
                    if conn.upstream_downstream == "涓婃父":
                        arrows = {"to": {"enabled": True}} if conn_info["is_source"] else {"from": {"enabled": True}}
                    elif conn.upstream_downstream == "涓嬫父":
                        arrows = {"from": {"enabled": True}} if conn_info["is_source"] else {"to": {"enabled": True}}
                    
                    edge_to_remote = {
                        "id": f"edge_{port_node_id}_{remote_node_id}",
                        "from": port_node_id,
                        "to": remote_node_id,
                        "width": 3,
                        "color": cable_color,
                        "arrows": arrows,
                        "label": conn.cable_model or "",
                        "title": f"鐢电紗鍨嬪彿: {conn.cable_model or 'N/A'}\
澶囨敞: {conn.remark or 'N/A'}"
                    }
                    edges.append(edge_to_remote)
        
        # 鍙充晶绔彛澶勭悊
        for i, port_name in enumerate(right_ports):
            port_node_id = f"port_right_{i}"
            port_x = 200
            port_y = -100 + i * 80
            
            port_node = {
                "id": port_node_id,
                "label": port_name,
                "type": "port",
                "x": port_x,
                "y": port_y,
                "size": 15,
                "color": "#10b981",
                "shape": "ellipse",
                "font": {"size": 12}
            }
            nodes.append(port_node)
            
            # 杩炴帴鍒颁腑蹇冭澶?
            edge_to_center = {
                "id": f"edge_center_{port_node_id}",
                "from": f"device_{device_id}",
                "to": port_node_id,
                "width": 2,
                "color": "#6b7280"
            }
            edges.append(edge_to_center)
            
            # 澶勭悊璇ョ鍙ｇ殑鎵€鏈夎繛鎺?
            for conn_info in port_info[port_name]:
                conn = conn_info["connection"]
                remote_device = conn_info["remote_device"]
                remote_port = conn_info["remote_port"]
                
                if remote_device:
                    # 鍒涘缓瀵圭璁惧鑺傜偣
                    remote_node_id = f"remote_{remote_device.id}_{remote_port}"
                    remote_node = {
                        "id": remote_node_id,
                        "label": f"{remote_device.name}\
{remote_port}",
                        "type": "remote_device",
                        "x": port_x + 150,
                        "y": port_y,
                        "size": 20,
                        "color": "#f59e0b",
                        "shape": "box",
                        "font": {"size": 10}
                    }
                    nodes.append(remote_node)
                    
                    # 杩炴帴绾块鑹叉牴鎹繛鎺ョ被鍨?
                    cable_color = "#fbbf24" if conn.connection_type == "浜ゆ祦" else "#ef4444" if conn.connection_type == "鐩存祦" else "#6b7280"
                    
                    # 绠ご鏂瑰悜鏍规嵁涓婁笅娓稿叧绯?
                    arrows = {}
                    if conn.upstream_downstream == "涓婃父":
                        arrows = {"to": {"enabled": True}} if conn_info["is_source"] else {"from": {"enabled": True}}
                    elif conn.upstream_downstream == "涓嬫父":
                        arrows = {"from": {"enabled": True}} if conn_info["is_source"] else {"to": {"enabled": True}}
                    
                    edge_to_remote = {
                        "id": f"edge_{port_node_id}_{remote_node_id}",
                        "from": port_node_id,
                        "to": remote_node_id,
                        "width": 3,
                        "color": cable_color,
                        "arrows": arrows,
                        "label": conn.cable_model or "",
                        "title": f"鐢电紗鍨嬪彿: {conn.cable_model or 'N/A'}\
澶囨敞: {conn.remark or 'N/A'}"
                    }
                    edges.append(edge_to_remote)
        
        topology_error_tracker.log_error(
            category=ErrorCategory.API_SUCCESS,
            level=ErrorLevel.INFO,
            message=f"绔彛鎷撴墤鍥炬暟鎹敓鎴愭垚鍔?,
            context={
                "device_id": device_id,
                "device_name": device.name,
                "total_ports": total_ports,
                "nodes_count": len(nodes),
                "edges_count": len(edges)
            }
        )
        
        return {"nodes": nodes, "edges": edges}
        
    except Exception as e:
        topology_error_tracker.log_error(
            category=ErrorCategory.API_ERROR,
            level=ErrorLevel.ERROR,
            message=f"绔彛鎷撴墤鍥続PI璋冪敤澶辫触: {str(e)}",
            context={"device_id": device_id},
            exception=e
        )
        return {"nodes": [], "edges": []}

# 绔彛鎷撴墤鍥炬暟鎹敓鎴愬嚱鏁?
def generate_port_topology_data(device_id: int):
    """鐢熸垚绔彛鎷撴墤鍥炬暟鎹?"""
    try:
        with SessionLocal() as db:
            device = db.query(Device).filter(Device.id == device_id).first()
            if not device:
                return {"nodes": [], "edges": []}
            
            nodes = []
            edges = []
            
            # 鑾峰彇璁惧鐨勬墍鏈夌鍙ｈ繛鎺?
            connections = db.query(Connection).filter(
                or_(Connection.source_device_id == device_id,
                    Connection.target_device_id == device_id)
            ).all()
            
            # 涓烘瘡涓鍙ｅ垱寤鸿妭鐐?
            for conn in connections:
                if conn.source_device_id == device_id:
                    port_info = f"{conn.source_port}"
                    connected_device = db.query(Device).filter(Device.id == conn.target_device_id).first()
                    connected_port = conn.target_port
                else:
                    port_info = f"{conn.target_port}"
                    connected_device = db.query(Device).filter(Device.id == conn.source_device_id).first()
                    connected_port = conn.source_port
                
                # 鍒涘缓绔彛鑺傜偣
                port_node = {
                    "id": f"port_{conn.id}",
                    "label": port_info,
                    "type": "port",
                    "x": 0,
                    "y": len(nodes) * 100,
                    "size": 15,
                    "color": "#3b82f6",
                    "shape": "circle"
                }
                nodes.append(port_node)
                
                # 鍒涘缓杩炴帴鐨勮澶囪妭鐐?
                if connected_device:
                    device_node = {
                        "id": f"device_{connected_device.id}",
                        "label": f"{connected_device.name}\
{connected_port}",
                        "type": "connected_device",
                        "x": 200,
                        "y": len(nodes) * 100,
                        "size": 20,
                        "color": "#f59e0b",
                        "shape": "box"
                    }
                    nodes.append(device_node)
                    
                    # 鍒涘缓杩炴帴杈?
                    edge = {
                        "id": f"edge_{conn.id}",
                        "from": f"port_{conn.id}",
                        "to": f"device_{connected_device.id}",
                        "width": 2,
                        "color": "#6b7280",
                        "label": conn.cable_model or """
                    }
                    edges.append(edge)
            
            return {"nodes": nodes, "edges": edges}
            
    except Exception as e:
        topology_error_tracker.log_error(
            category=ErrorCategory.API_ERROR,
            level=ErrorLevel.ERROR,
            message=f"绔彛鎷撴墤鍥炬暟鎹敓鎴愬け璐? {str(e)}",
            context={"device_id": device_id},
            exception=e
        )
        return {"nodes": [], "edges": []}

        # 娣诲姞杩炴帴杈?
        for conn in topology_data.get("connections", []):
            edge = {
                "id": conn["id"],
                "from": conn["from_port_id"],
                "to": conn["to_port_id"],
                "label": conn.get("connection_type", ""),
                "title": "杩炴帴绫诲瀷: " + str(conn.get('connection_type', 'Unknown')) + ", 甯﹀: " + str(conn.get('bandwidth', 'N/A')),
                "color": get_connection_edge_color(conn.get("status", "active")),
                "width": 2,
                "connectionData": conn
            }
            edges.append(edge)
        
        return {
            "nodes": nodes,
            "edges": edges,
            "metadata": topology_data.get("metadata", {}),
            "device": topology_data.get("device", {})
        }
    except HTTPException as e:
        logger.error(f"HTTP寮傚父: {e.detail}")
        raise e
    except Exception as e:
        logger.error(f"鑾峰彇绔彛鎷撴墤鍥炬暟鎹椂鍑洪敊: {str(e)}")
        traceback.print_exc()
        # 杩斿洖鏍囧噯鏍煎紡鐨勭┖鏁版嵁锛岄伩鍏嶅墠绔В鏋愬け璐?
        return JSONResponse(content={"nodes": [], "edges": []}, status_code=500)

        station_utilization.sort(key=lambda x: x["idle_rate"], reverse=True)
        return station_utilization
    
    def _check_idle_rate_alerts(self) -> list:
        """Check idle rate alerts"""
        alerts = []
        
        # 妫€鏌ユ€讳綋绌洪棽鐜?
        overall_idle = self._calculate_overall_idle_rate()
        if overall_idle["idle_rate"] < 10:  # 绌洪棽鐜囦綆浜?0%棰勮
            alerts.append({
                "type": "overall",
                "level": "warning",
                "message": f"绯荤粺鎬讳綋绌洪棽鐜囦粎涓?{overall_idle['idle_rate']}%锛岃祫婧愮揣寮?,
                "idle_rate": overall_idle["idle_rate"]
            })
        
        # 妫€鏌ヨ澶囩被鍨嬬┖闂茬巼
        device_type_idle = self._calculate_device_type_idle_rate()
        for item in device_type_idle:
            if item["idle_rate"] < 5:  # 璁惧绫诲瀷绌洪棽鐜囦綆浜?%棰勮
                alerts.append({
                    "type": "device_type",
                    "level": "critical",
                    "message": f"{item['device_type']} 绫诲瀷璁惧绌洪棽鐜囦粎涓?{item['idle_rate']}%锛屾€ラ渶鎵╁",
                    "device_type": item["device_type"],
                    "idle_rate": item["idle_rate"]
                })
        
        # 妫€鏌ョ珯鐐圭┖闂茬巼
        station_idle = self._calculate_station_idle_rate()
        for item in station_idle:
            if item["idle_rate"] < 5:  # 绔欑偣绌洪棽鐜囦綆浜?%棰勮
                alerts.append({
                    "type": "station",
                    "level": "critical",
                    "message": f"{item['station']} 绔欑偣绌洪棽鐜囦粎涓?{item['idle_rate']}%锛屾€ラ渶鎵╁",
                    "station": item["station"],
                    "idle_rate": item["idle_rate"]
                })
        
        return alerts
    

    
    def _calculate_port_capacity_distribution(self) -> dict:
        """Calculate port capacity distribution"""
        try:
            # 缁熻涓嶅悓瑙勬牸绔彛鐨勪娇鐢ㄥ垎甯?
            connections = self.db.query(Connection).all()
            
            fuse_specs = {}
            breaker_specs = {}
            
            for conn in connections:
                # 缁熻鐔旀柇鍣ㄨ鏍煎垎甯?
                if conn.source_fuse_spec:
                    spec = conn.source_fuse_spec
                    if spec not in fuse_specs:
                        fuse_specs[spec] = {"total": 0, "connected": 0}
                    fuse_specs[spec]["total"] += 1
                    if conn.connection_type and conn.connection_type.strip():
                        fuse_specs[spec]["connected"] += 1
                
                if conn.target_fuse_spec:
                    spec = conn.target_fuse_spec
                    if spec not in fuse_specs:
                        fuse_specs[spec] = {"total": 0, "connected": 0}
                    fuse_specs[spec]["total"] += 1
                    if conn.connection_type and conn.connection_type.strip():
                        fuse_specs[spec]["connected"] += 1
                
                # 缁熻绌哄紑瑙勬牸鍒嗗竷
                if conn.source_breaker_spec:
                    spec = conn.source_breaker_spec
                    if spec not in breaker_specs:
                        breaker_specs[spec] = {"total": 0, "connected": 0}
                    breaker_specs[spec]["total"] += 1
                    if conn.connection_type and conn.connection_type.strip():
                        breaker_specs[spec]["connected"] += 1
                
                if conn.target_breaker_spec:
                    spec = conn.target_breaker_spec
                    if spec not in breaker_specs:
                        breaker_specs[spec] = {"total": 0, "connected": 0}
                    breaker_specs[spec]["total"] += 1
                    if conn.connection_type and conn.connection_type.strip():
                        breaker_specs[spec]["connected"] += 1
            
            # 璁＄畻鍚勮鏍肩殑浣跨敤鐜?
            for spec_data in fuse_specs.values():
                spec_data["utilization_rate"] = round(
                    (spec_data["connected"] / spec_data["total"] * 100) if spec_data["total"] > 0 else 0, 2
                )
            
            for spec_data in breaker_specs.values():
                spec_data["utilization_rate"] = round(
                    (spec_data["connected"] / spec_data["total"] * 100) if spec_data["total"] > 0 else 0, 2
                )
            
            return {
                "fuse_specifications": fuse_specs,
                "breaker_specifications": breaker_specs
            }
            
        except Exception as e:
            print(f"璁＄畻绔彛瀹归噺鍒嗗竷鏃跺嚭閿? {e}")
            return {
                "fuse_specifications": {},
                "breaker_specifications": {}
            }
    
    def _calculate_load_balance_analysis(self) -> dict:
        """璁＄畻璐熻浇鍧囪　鍒嗘瀽"""
        try:
            # 鑾峰彇鎵€鏈夎澶囩殑浣跨敤鐜?
            devices = self.db.query(Device).all()
            device_utilizations = []
            
            for device in devices:
                utilization_rate = self._get_device_utilization_rate(device.id)
                device_utilizations.append({
                    "device_id": device.id,
                    "device_name": device.name,
                    "device_type": device.device_type or "鏈煡",
                    "station": device.station or "鏈煡",
                    "utilization_rate": utilization_rate
                })
            
            if not device_utilizations:
                return {
                    "balance_score": 0,
                    "average_utilization": 0,
                    "utilization_variance": 0,
                    "overloaded_devices": [],
                    "underutilized_devices": []
                }
            
            # 璁＄畻骞冲潎浣跨敤鐜囧拰鏂瑰樊
            utilization_rates = [d["utilization_rate"] for d in device_utilizations]
            average_utilization = sum(utilization_rates) / len(utilization_rates)
            variance = sum((rate - average_utilization) ** 2 for rate in utilization_rates) / len(utilization_rates)
            
            # 璐熻浇鍧囪　璇勫垎锛堟柟宸秺灏忥紝鍧囪　鎬ц秺濂斤級
            balance_score = max(0, 100 - variance)  # 绠€鍖栫殑璇勫垎绠楁硶
            
            # 璇嗗埆杩囪浇鍜屼綆鍒╃敤鐜囪澶?
            overloaded_devices = [d for d in device_utilizations if d["utilization_rate"] > 90]
            underutilized_devices = [d for d in device_utilizations if d["utilization_rate"] < 20]
            
            return {
                "balance_score": round(balance_score, 2),
                "average_utilization": round(average_utilization, 2),
                "utilization_variance": round(variance, 2),
                "overloaded_devices": overloaded_devices,
                "underutilized_devices": underutilized_devices
            }
            
        except Exception as e:
            print(f"璁＄畻璐熻浇鍧囪　鍒嗘瀽鏃跺嚭閿? {e}")
            return {
                "balance_score": 0,
                "average_utilization": 0,
                "utilization_variance": 0,
                "overloaded_devices": [],
                "underutilized_devices": []
            }
    
    def _get_top_utilized_devices(self, limit: int = 10) -> list:
        """鑾峰彇浣跨敤鐜囨渶楂樼殑璁惧"""
        try:
            devices = self.db.query(Device).all()
            device_utilizations = []
            
            for device in devices:
                utilization_rate = self._get_device_utilization_rate(device.id)
                device_utilizations.append({
                    "device_id": device.id,
                    "device_name": device.name,
                    "device_type": device.device_type or "鏈煡",
                    "station": device.station or "鏈煡",
                    "utilization_rate": utilization_rate
                })
            
            # 鎸変娇鐢ㄧ巼闄嶅簭鎺掑簭骞惰繑鍥炲墠N涓?
            device_utilizations.sort(key=lambda x: x["utilization_rate"], reverse=True)
            return device_utilizations[:limit]
            
        except Exception as e:
            print(f"鑾峰彇浣跨敤鐜囨渶楂樿澶囨椂鍑洪敊: {e}")
            return []
    
    def _get_device_utilization_rate(self, device_id: int) -> float:
        """鑾峰彇鍗曚釜璁惧鐨勪娇鐢ㄧ巼"""
        try:
            # 鑾峰彇璁惧鐨勬墍鏈夌鍙?
            connections = self.db.query(Connection).filter(
                (Connection.source_device_id == device_id) |
                (Connection.target_device_id == device_id)
            ).all()
            
            all_ports = set()
            connected_ports = set()
            
            for conn in connections:
                if conn.source_device_id == device_id:
                    if conn.source_fuse_number:
                        port_key = f"fuse_{conn.source_fuse_number}"
                        all_ports.add(port_key)
                        if conn.connection_type and conn.connection_type.strip():
                            connected_ports.add(port_key)
                    if conn.source_breaker_number:
                        port_key = f"breaker_{conn.source_breaker_number}"
                        all_ports.add(port_key)
                        if conn.connection_type and conn.connection_type.strip():
                            connected_ports.add(port_key)
                
                if conn.target_device_id == device_id:
                    if conn.target_fuse_number:
                        port_key = f"fuse_{conn.target_fuse_number}"
                        all_ports.add(port_key)
                        if conn.connection_type and conn.connection_type.strip():
                            connected_ports.add(port_key)
                    if conn.target_breaker_number:
                        port_key = f"breaker_{conn.target_breaker_number}"
                        all_ports.add(port_key)
                        if conn.connection_type and conn.connection_type.strip():
                            connected_ports.add(port_key)
            
            total_ports = len(all_ports)
            connected_count = len(connected_ports)
            
            return (connected_count / total_ports * 100) if total_ports > 0 else 0
            
        except Exception as e:
            print(f"鑾峰彇璁惧 {device_id} 浣跨敤鐜囨椂鍑洪敊: {e}")
            return 0
    

    
    def get_port_statistics(self) -> dict:
        """鑾峰彇鍏ㄥ眬绔彛缁熻淇℃伅"""
        try:
            # 1. 璁惧绔彛鎬昏
            device_port_summary = self._get_device_port_summary()
            
            # 2. 绔彛绫诲瀷缁熻
            port_type_statistics = self._get_port_type_statistics()
            
            # 3. 瀹归噺缁熻
            capacity_statistics = self._get_capacity_statistics()
            
            # 4. 璁惧绔彛璇︽儏
            device_port_details = self._get_device_port_details()
            
            return {
                "device_port_summary": device_port_summary,
                "port_type_statistics": port_type_statistics,
                "capacity_statistics": capacity_statistics,
                "device_port_details": device_port_details
            }
        except Exception as e:
            print(f"鑾峰彇绔彛缁熻淇℃伅鏃跺嚭閿? {e}")
            raise HTTPException(status_code=500, detail=f"鑾峰彇绔彛缁熻淇℃伅澶辫触: {str(e)}")
    
    def _get_device_port_summary(self) -> dict:
        """鑾峰彇璁惧绔彛鎬昏 - 閲囩敤闆嗗悎缁熻閫昏緫锛岀粺璁℃墍鏈夋湁杩炴帴鐨勭鍙?"""
        try:
            # 缁熻鎬昏澶囨暟
            total_devices = self.db.query(Device).count()
            
            # 浣跨敤闆嗗悎鏉ラ伩鍏嶉噸澶嶈绠楀悓涓€涓鍙?
            all_ports = set()
            connected_ports = set()
            
            # 鑾峰彇鎵€鏈夎繛鎺ヨ褰?
            connections = self.db.query(Connection).all()
            
            for conn in connections:
                # 缁熻婧愮鍙ｏ紙A绔級
                if conn.source_fuse_number and conn.source_device_id:
                    port_key = f"device_{conn.source_device_id}_fuse_{conn.source_fuse_number}"
                    all_ports.add(port_key)
                    # 閫氳繃杩炴帴绫诲瀷瀛楁鏄惁涓虹┖鍒ゆ柇绔彛浣跨敤鐘舵€?
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
                        
                if conn.source_breaker_number and conn.source_device_id:
                    port_key = f"device_{conn.source_device_id}_breaker_{conn.source_breaker_number}"
                    all_ports.add(port_key)
                    # 閫氳繃杩炴帴绫诲瀷瀛楁鏄惁涓虹┖鍒ゆ柇绔彛浣跨敤鐘舵€?
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
                
                # 缁熻鐩爣绔彛锛圔绔級- 绗﹀悎璁捐鏂囨。涓?涓€涓繛鎺ュ崰鐢ㄤ袱涓鍙?鐨勮姹?
                if conn.target_fuse_number and conn.target_device_id:
                    port_key = f"device_{conn.target_device_id}_fuse_{conn.target_fuse_number}"
                    all_ports.add(port_key)
                    # 閫氳繃杩炴帴绫诲瀷瀛楁鏄惁涓虹┖鍒ゆ柇绔彛浣跨敤鐘舵€?
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
                        
                if conn.target_breaker_number and conn.target_device_id:
                    port_key = f"device_{conn.target_device_id}_breaker_{conn.target_breaker_number}"
                    all_ports.add(port_key)
                    # 閫氳繃杩炴帴绫诲瀷瀛楁鏄惁涓虹┖鍒ゆ柇绔彛浣跨敤鐘舵€?
                    if conn.connection_type and conn.connection_type.strip():
                        connected_ports.add(port_key)
            
            total_ports = len(all_ports)
            connected_count = len(connected_ports)
            idle_ports = total_ports - connected_count
            utilization_rate = (connected_count / total_ports * 100) if total_ports > 0 else 0
            
            return {
                "total_devices": total_devices,
                "total_ports": total_ports,
                "connected_ports": connected_count,
                "idle_ports": idle_ports,
                "utilization_rate": round(utilization_rate, 2)
            }
        except Exception as e:
            print(f"鑾峰彇璁惧绔彛鎬昏鏃跺嚭閿? {e}")
            return {
                "total_devices": 0,
                "total_ports": 0,
                "connected_ports": 0,
                "idle_ports": 0,
                "utilization_rate": 0
            }
    
    def _get_port_type_statistics(self) -> dict:
        """鑾峰彇绔彛绫诲瀷缁熻 - 鍩轰簬A绔澶囩粺璁?"""
        try:
            # 鑾峰彇鎵€鏈夎繛鎺ヨ褰?
            connections = self.db.query(Connection).all()
            
            # 浣跨敤闆嗗悎鏉ラ伩鍏嶉噸澶嶈绠楃鍙?
            fuse_ports = set()
            breaker_ports = set()
            connected_fuse_ports = set()
            connected_breaker_ports = set()
            
            for conn in connections:
                # 鍙粺璁绔紙婧愮锛夎澶囩殑绔彛
                if conn.source_fuse_number and conn.source_device_id:
                    port_key = f"{conn.source_device_id}_fuse_{conn.source_fuse_number}"
                    fuse_ports.add(port_key)
                    # 閫氳繃杩炴帴绫诲瀷瀛楁鏄惁涓虹┖鍒ゆ柇绔彛浣跨敤鐘舵€?
                    if conn.connection_type and conn.connection_type.strip():
                        connected_fuse_ports.add(port_key)
                
                if conn.source_breaker_number and conn.source_device_id:
                    port_key = f"{conn.source_device_id}_breaker_{conn.source_breaker_number}"
                    breaker_ports.add(port_key)
                    # 閫氳繃杩炴帴绫诲瀷瀛楁鏄惁涓虹┖鍒ゆ柇绔彛浣跨敤鐘舵€?
                    if conn.connection_type and conn.connection_type.strip():
                        connected_breaker_ports.add(port_key)
            
            fuse_total = len(fuse_ports)
            fuse_connected = len(connected_fuse_ports)
            breaker_total = len(breaker_ports)
            breaker_connected = len(connected_breaker_ports)
            
            return {
                "fuse_ports": {
                    "total": fuse_total,
                    "connected": fuse_connected,
                    "idle": fuse_total - fuse_connected,
                    "utilization_rate": round((fuse_connected / fuse_total * 100) if fuse_total > 0 else 0, 2)
                },
                "breaker_ports": {
                    "total": breaker_total,
                    "connected": breaker_connected,
                    "idle": breaker_total - breaker_connected,
                    "utilization_rate": round((breaker_connected / breaker_total * 100) if breaker_total > 0 else 0, 2)
                }
            }
        except Exception as e:
            print(f"鑾峰彇绔彛绫诲瀷缁熻鏃跺嚭閿? {e}")
            return {
                "fuse_ports": {"total": 0, "connected": 0, "idle": 0, "utilization_rate": 0},
                "breaker_ports": {"total": 0, "connected": 0, "idle": 0, "utilization_rate": 0}
            }
    
    def _get_capacity_statistics(self) -> dict:
        """鑾峰彇瀹归噺缁熻"""
        try:
            # 鑾峰彇鎵€鏈夎繛鎺ョ殑瑙勬牸淇℃伅
            connections = self.db.query(Connection).all()
            
            capacity_stats = {}
            high_capacity_available = {"630A_above": 0, "400A_above": 0, "250A_above": 0}
            
            for conn in connections:
                # 澶勭悊鍚勭瑙勬牸瀛楁
                for spec_field in [conn.source_fuse_spec, conn.source_breaker_spec, 
                                 conn.target_fuse_spec, conn.target_breaker_spec]:
                    if spec_field:
                        rating = self._extract_rating_from_spec(spec_field)
                        if rating and rating != "鏈煡":
                            if rating not in capacity_stats:
                                capacity_stats[rating] = {"total": 0, "connected": 0, "idle": 0}
                            
                            capacity_stats[rating]["total"] += 1
                            
                            # 鍒ゆ柇鏄惁宸茶繛鎺?
                            if conn.source_device_id and conn.target_device_id:
                                capacity_stats[rating]["connected"] += 1
                            else:
                                capacity_stats[rating]["idle"] += 1
                                
                                # 缁熻澶у閲忓彲鐢ㄧ鍙?
                                try:
                                    rating_value = int(rating.replace('A', ''))
                                    if rating_value >= 630:
                                        high_capacity_available["630A_above"] += 1
                                    if rating_value >= 400:
                                        high_capacity_available["400A_above"] += 1
                                    if rating_value >= 250:
                                        high_capacity_available["250A_above"] += 1
                                except ValueError:
                                    pass  # 蹇界暐鏃犳硶杞崲鐨勫閲忓€?
            
            return {
                "by_rating": capacity_stats,
                "high_capacity_available": high_capacity_available
            }
        except Exception as e:
            print(f"鑾峰彇瀹归噺缁熻鏃跺嚭閿? {e}")
            return {
                "by_rating": {},
                "high_capacity_available": {"630A_above": 0, "400A_above": 0, "250A_above": 0}
            }
    
    def _get_device_port_details(self) -> list:
        """鑾峰彇璁惧绔彛璇︽儏 - 鍩轰簬A绔澶囩粺璁?"""
        try:
            # 鑾峰彇鎵€鏈夎澶囧強鍏剁鍙ｄ娇鐢ㄦ儏鍐?
            devices = self.db.query(Device).all()

            device_details = []
            
            for device in devices:
                # 鍙粺璁¤璁惧浣滀负A绔紙婧愮锛夌殑杩炴帴璁板綍
                device_connections = self.db.query(Connection).filter(
                    Connection.source_device_id == device.id
                ).all()

                # 浣跨敤闆嗗悎鏉ラ伩鍏嶉噸澶嶈绠楃鍙?
                all_ports = set()
                connected_ports_set = set()
                fuse_ports = set()
                breaker_ports = set()
                
                for conn in device_connections:
                    # 鍙粺璁¤璁惧浣滀负A绔殑绔彛
                    if conn.source_fuse_number:
                        port_key = f"fuse_{conn.source_fuse_number}"
                        all_ports.add(port_key)
                        fuse_ports.add(port_key)
                        # 閫氳繃杩炴帴绫诲瀷瀛楁鏄惁涓虹┖鍒ゆ柇绔彛浣跨敤鐘舵€?
                        if conn.connection_type and conn.connection_type.strip():
                            connected_ports_set.add(port_key)

                    if conn.source_breaker_number:
                        port_key = f"breaker_{conn.source_breaker_number}"
                        all_ports.add(port_key)
                        breaker_ports.add(port_key)
                        # 閫氳繃杩炴帴绫诲瀷瀛楁鏄惁涓虹┖鍒ゆ柇绔彛浣跨敤鐘舵€?
                        if conn.connection_type and conn.connection_type.strip():
                            connected_ports_set.add(port_key)

                
                total_ports = len(all_ports)
                connected_ports = len(connected_ports_set)
                
                idle_ports = total_ports - connected_ports
                utilization_rate = (connected_ports / total_ports * 100) if total_ports > 0 else 0
                
                device_details.append({
                    "device_id": device.id,
                    "device_name": device.name,
                    "device_type": device.device_type or "鏈煡",
                    "station": device.station or "鏈煡",
                    "total_ports": total_ports,
                    "connected_ports": connected_ports,
                    "idle_ports": idle_ports,
                    "utilization_rate": round(utilization_rate, 2),
                    "fuse_ports": len(fuse_ports),
                    "breaker_ports": len(breaker_ports)
                })
            
            # 鎸夊埄鐢ㄧ巼闄嶅簭鎺掑簭
            device_details.sort(key=lambda x: x["utilization_rate"], reverse=True)
            
            return device_details
        except Exception as e:
            print(f"鑾峰彇璁惧绔彛璇︽儏鏃跺嚭閿? {e}")
            return []
    
    def _extract_rating_from_spec(self, spec_string: str) -> str:
        """浠庤鏍煎瓧绗︿覆涓彁鍙栫數娴佺瓑绾?"""
        if not spec_string:
            return "鏈煡"
        
        try:
            # 鍖归厤鎷彿鍐呯殑鐢垫祦鍊硷紝濡?"NT4(500A)" -> "500A"
            match = re.search(r'\((\d+)A\)', spec_string)
            if match:
                return f"{match.group(1)}A"
            
            # 鍖归厤鐩存帴鐨勭數娴佸€硷紝濡?"500A" -> "500A"
            match = re.search(r'(\d+)A', spec_string)
            if match:
                return f"{match.group(1)}A"
            
            return "鏈煡"
        except Exception as e:
            print(f"鎻愬彇鐢垫祦绛夌骇鏃跺嚭閿? {e}")
            return "鏈煡"
    



def verify_admin_password(password: str) -> bool:
    """
    楠岃瘉绠＄悊鍛樺瘑鐮?
    Args:
        password: 鐢ㄦ埛杈撳叆鐨勫瘑鐮?
    Returns:
        bool: 瀵嗙爜鏄惁姝ｇ‘
    """
    return password == ADMIN_PASSWORD

# --- FastAPI 搴旂敤璁剧疆 ---


# 鎸傝浇闈欐€佹枃浠剁洰褰?
# 鎸傝浇闈欐€佹枃浠剁洰褰?
app.mount("/static", StaticFiles(directory="static"), name="static")
# 鎸傝浇涓存椂鐩綍浠ヤ緵鍔犺浇鏈湴fixture
app.mount("/temp", StaticFiles(directory="temp"), name="temp")
# 璁剧疆妯℃澘鐩綍
templates = Jinja2Templates(directory="templates")

# --- 鏁版嵁搴撲細璇濈鐞?---

def get_db():
    """
    鏁版嵁搴撲細璇濈鐞嗗嚱鏁?
    澧炲姞浜嗚缁嗙殑鏃ュ織璁板綍鏉ヨ窡韪暟鎹簱杩炴帴鐨勫垱寤哄拰鍏抽棴杩囩▼
    """
    print("\n--- 鍒涘缓鏁版嵁搴撲細璇?---")
    db = None
    try:
        db = SessionLocal()
        print(f"鏁版嵁搴撲細璇濆垱寤烘垚鍔? {id(db)}")
        yield db
    except Exception as e:
        print(f"鏁版嵁搴撲細璇濆垱寤哄け璐? {e}")
        if db:
            print("姝ｅ湪鍥炴粴鏁版嵁搴撲簨鍔?..")
            db.rollback()
        raise
    finally:
        if db:
            print(f"姝ｅ湪鍏抽棴鏁版嵁搴撲細璇? {id(db)}")
            db.close()
            print("鏁版嵁搴撲細璇濆凡鍏抽棴")
        print("--- 鏁版嵁搴撲細璇濈鐞嗙粨鏉?---\n")

# --- 搴旂敤鍚姩浜嬩欢 ---

@app.on_event("startup")
def on_startup():
    """
    搴旂敤鍚姩浜嬩欢澶勭悊鍑芥暟
    澧炲姞浜嗚缁嗙殑鏃ュ織璁板綍鏉ヨ窡韪簲鐢ㄥ惎鍔ㄨ繃绋?
    """
    print("\n" + "=" * 60)
    print("馃殌 鍔ㄥ姏璧勬簮璧勪骇绠＄悊绯荤粺鍚姩涓?..")
    print("=" * 60)
    
    try:
        # 妫€鏌ュ苟鍒涘缓鏁版嵁搴撶洰褰?
        db_dir = './database'
        if not os.path.exists(db_dir):
            print(f"馃搧 鍒涘缓鏁版嵁搴撶洰褰? {db_dir}")
            os.makedirs(db_dir)
        else:
            print(f"馃搧 鏁版嵁搴撶洰褰曞凡瀛樺湪: {db_dir}")
        
        # 鍒濆鍖栨暟鎹簱
        print("馃梽锔?姝ｅ湪鍒濆鍖栨暟鎹簱...")
        create_db_and_tables()
        
        print("鉁?搴旂敤鍚姩瀹屾垚锛?)
        print(f"馃寪 鏈嶅姟鍣ㄥ湴鍧€: http://localhost:{PORT}")
        print("=" * 60 + "\n")
        
    except Exception as e:
        print(f"\n鉂?搴旂敤鍚姩澶辫触!")
        print(f"閿欒绫诲瀷: {type(e).__name__}")
        print(f"閿欒淇℃伅: {e}")
        print("\n瀹屾暣閿欒鍫嗘爤:")
        traceback.print_exc()
        print("=" * 60)
        raise  # 閲嶆柊鎶涘嚭寮傚父锛屽仠姝㈠簲鐢ㄥ惎鍔?

# --- 璺敱鍜岃鍥惧嚱鏁?---

@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request, db: Session = Depends(get_db)):
    """
    棣栭〉璺敱 - 鏄剧ず鎵€鏈夎澶囧垪琛?
    澧炲姞浜嗚缁嗙殑鏃ュ織璁板綍鏉ヨ窡韪暟鎹幏鍙栬繃绋?
    """
    print("\n=== 棣栭〉鏁版嵁鑾峰彇寮€濮?===")
    
    try:
        # 鑾峰彇璁惧鏁版嵁
        print("姝ｅ湪浠庢暟鎹簱鏌ヨ璁惧鏁版嵁...")
        devices = db.query(Device).order_by(Device.id).all()
        device_count = len(devices)
        print(f"鏌ヨ鍒?{device_count} 涓澶?)
        
        # 鑾峰彇鐢熷懡鍛ㄦ湡瑙勫垯
        lifecycle_rules = db.query(LifecycleRule).filter(LifecycleRule.is_active == 'true').all()
        rules_dict = {rule.device_type: rule for rule in lifecycle_rules}
        print(f"鍔犺浇浜?{len(rules_dict)} 涓敓鍛藉懆鏈熻鍒?)
        
        # 涓烘瘡涓澶囪绠楃敓鍛藉懆鏈熺姸鎬?
        for device in devices:
            lifecycle_status = "unknown"
            lifecycle_status_text = "鏈厤缃鍒?
            
            if device.device_type and device.device_type in rules_dict:
                rule = rules_dict[device.device_type]
                if device.commission_date:
                    try:
                        # 瑙ｆ瀽鎶曚骇鏃ユ湡
                        commission_date = None
                        date_str = str(device.commission_date).strip()
                        
                        # 澶勭悊鐗规畩鏍煎紡锛歒YYYMM (濡?202312)
                        if re.match(r'^\d{6}$', date_str):
                            try:
                                year = int(date_str[:4])
                                month = int(date_str[4:6])
                                commission_date = datetime(year, month, 1)
                            except ValueError:
                                pass
                        
                        # 濡傛灉鐗规畩鏍煎紡瑙ｆ瀽澶辫触锛屽皾璇曟爣鍑嗘牸寮?
                        if not commission_date:
                            date_formats = [
                                "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
                                "%Y-%m", "%Y/%m", "%Y.%m",
                                "%Y骞?m鏈?d鏃?, "%Y骞?m鏈?
                            ]
                            
                            for fmt in date_formats:
                                try:
                                    commission_date = datetime.strptime(date_str, fmt)
                                    break
                                except ValueError:
                                    continue
                        
                        if commission_date:
                            # 璁＄畻鏈嶅焦鏃堕棿
                            today = datetime.now()
                            service_years = (today - commission_date).days / 365.25
                            
                            # 鍒ゆ柇鐘舵€?
                            if service_years >= rule.lifecycle_years:
                                lifecycle_status = "expired"
                                lifecycle_status_text = "宸茶秴鏈?
                            elif service_years >= (rule.lifecycle_years - rule.warning_months / 12):
                                lifecycle_status = "warning"
                                lifecycle_status_text = "涓磋繎瓒呴檺"
                            else:
                                lifecycle_status = "normal"
                                lifecycle_status_text = "姝ｅ父"
                        else:
                            lifecycle_status = "unknown"
                            lifecycle_status_text = "鎶曚骇鏃ユ湡鏍煎紡鏃犳硶璇嗗埆"
                    except Exception as e:
                        lifecycle_status = "unknown"
                        lifecycle_status_text = "鎶曚骇鏃ユ湡鏍煎紡鏃犳硶璇嗗埆"
                else:
                    lifecycle_status = "unknown"
                    lifecycle_status_text = "鎶曚骇鏃ユ湡鏈～鍐?
            
            # 灏嗙姸鎬佷俊鎭坊鍔犲埌璁惧瀵硅薄
            device.lifecycle_status = lifecycle_status
            device.lifecycle_status_text = lifecycle_status_text
        
        # 鏄剧ず鍓嶅嚑涓澶囩殑淇℃伅鐢ㄤ簬璋冭瘯
        if device_count > 0:
            print("\n鍓?涓澶囦俊鎭?")
            for i, device in enumerate(devices[:3]):
                print(f"  璁惧{i+1}: ID={device.id}, 璧勪骇缂栧彿={device.asset_id}, 鍚嶇О={device.name}, 鐢熷懡鍛ㄦ湡鐘舵€?{device.lifecycle_status}")
        else:
            print("璀﹀憡: 鏁版嵁搴撲腑娌℃湁璁惧鏁版嵁锛?)
        
        # 鑾峰彇杩炴帴鏁版嵁鐢ㄤ簬缁熻
        connections = db.query(Connection).all()
        connection_count = len(connections)
        print(f"鏁版嵁搴撲腑鍏辨湁 {connection_count} 涓繛鎺?)
        
        # 鑾峰彇鎵€鏈変笉閲嶅鐨勫眬绔欏垪琛紝鐢ㄤ簬绛涢€変笅鎷夋
        print("姝ｅ湪鑾峰彇灞€绔欏垪琛?..")
        stations = db.query(Device.station).filter(Device.station.isnot(None)).filter(Device.station != '').distinct().all()
        station_list = [station[0] for station in stations if station[0]]  # 鎻愬彇灞€绔欏悕绉板苟杩囨护绌哄€?
        station_list.sort()  # 鎸夊瓧姣嶉『搴忔帓搴?
        print(f"鎵惧埌 {len(station_list)} 涓笉鍚岀殑灞€绔? {station_list}")
        
        # 浣跨敤棰勫畾涔夌殑鏍囧噯璁惧绫诲瀷鍒楄〃
        print("姝ｅ湪鍔犺浇鏍囧噯璁惧绫诲瀷鍒楄〃...")
        device_type_list = sorted(STANDARD_DEVICE_TYPES)
        print(f"鍔犺浇浜?{len(device_type_list)} 涓爣鍑嗚澶囩被鍨? {device_type_list}")
        
        # 鑾峰彇鎵€鏈変笉閲嶅鐨勫巶瀹跺垪琛紝鐢ㄤ簬绛涢€変笅鎷夋
        print("姝ｅ湪鑾峰彇鍘傚鍒楄〃...")
        vendors = db.query(Device.vendor).filter(Device.vendor.isnot(None)).filter(Device.vendor != '').distinct().all()
        vendor_list = [vendor[0] for vendor in vendors if vendor[0]]  # 鎻愬彇鍘傚鍚嶇О骞惰繃婊ょ┖鍊?
        vendor_list.sort()  # 鎸夊瓧姣嶉『搴忔帓搴?
        print(f"鎵惧埌 {len(vendor_list)} 涓笉鍚岀殑鍘傚: {vendor_list}")
        
        # 妫€鏌ユ槸鍚︽湁涓婁紶閿欒淇℃伅
        upload_error = request.query_params.get("error")
        if upload_error:
            print(f"妫€娴嬪埌涓婁紶閿欒淇℃伅: {upload_error}")
        else:
            print("娌℃湁涓婁紶閿欒淇℃伅")
        
        # 妫€鏌ユ槸鍚︽湁鎴愬姛淇℃伅
        success_message = request.query_params.get("success")
        if success_message:
            print(f"妫€娴嬪埌鎴愬姛淇℃伅: {success_message}")
        else:
            print("娌℃湁鎴愬姛淇℃伅")
        
        print("=== 棣栭〉鏁版嵁鑾峰彇瀹屾垚 ===")
        
        return templates.TemplateResponse("index.html", {
            "request": request, 
            "devices": devices, 
            "stations": station_list,
            "device_types": device_type_list,
            "vendors": vendor_list,
            "upload_error": upload_error,
            "success_message": success_message
        })
        
    except Exception as e:
        print(f"\n!!! 棣栭〉鏁版嵁鑾峰彇澶辫触 !!!")
        print(f"閿欒绫诲瀷: {type(e).__name__}")
        print(f"閿欒淇℃伅: {e}")
        print("\n瀹屾暣閿欒鍫嗘爤:")
        traceback.print_exc()
        print("=" * 50)
        
        # 杩斿洖閿欒椤甸潰鎴栫┖璁惧鍒楄〃
        return templates.TemplateResponse("index.html", {
            "request": request, 
            "devices": [], 
            "stations": [],
            "device_types": [],
            "vendors": [],
            "upload_error": f"鑾峰彇璁惧鏁版嵁鏃跺嚭閿? {e}"
        })

@app.post("/upload")
async def upload_excel(file: UploadFile = File(...), password: str = Form(...), db: Session = Depends(get_db)):
    """
    澶勭悊 Excel 鏂囦欢涓婁紶銆?
    濡傛灉澶辫触锛屽垯閲嶅畾鍚戝洖涓婚〉骞堕檮甯﹁缁嗛敊璇俊鎭€?
    澧炲姞浜嗚缁嗙殑鏃ュ織璁板綍鏉ヨ窡韪鐞嗚繃绋嬨€?
    """
    print("\n=== 寮€濮嬪鐞嗕笂浼犵殑Excel鏂囦欢 ===")
    print(f"涓婁紶鏂囦欢鍚? {file.filename}")
    print(f"鏂囦欢绫诲瀷: {file.content_type}")
    
    # 楠岃瘉绠＄悊鍛樺瘑鐮?
    if not verify_admin_password(password):
        error_message = "瀵嗙爜閿欒锛屾棤鏉冮檺鎵ц姝ゆ搷浣溿€?
        print(f"鏉冮檺楠岃瘉澶辫触: {error_message}")
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
    
    print("绠＄悊鍛樺瘑鐮侀獙璇侀€氳繃")
    
    try:
        # 姝ラ 1: 澧為噺鏇存柊妯″紡 - 淇濈暀鎵嬪伐娣诲姞鐨勮澶囷紝鍙洿鏂癊xcel涓殑璁惧
        print("\n姝ラ 1: 閲囩敤澧為噺鏇存柊妯″紡锛屼繚鐣欑幇鏈夋墜宸ユ坊鍔犵殑璁惧...")
        
        # 璁板綍褰撳墠鏁版嵁閲?
        current_connections_count = db.query(Connection).count()
        current_devices_count = db.query(Device).count()
        print(f"褰撳墠鏁版嵁搴撶姸鎬? {current_connections_count} 涓繛鎺? {current_devices_count} 涓澶?)
        print("姝ラ 1: 瀹屾垚銆傚皢閲囩敤澧為噺鏇存柊妯″紡澶勭悊Excel鏁版嵁銆?)

        contents = await file.read()
        print(f"鏂囦欢澶у皬: {len(contents)} 瀛楄妭")
        buffer = io.BytesIO(contents)
        
        # 姝ラ 2: 璇诲彇Excel鏂囦欢
        print("\n姝ラ 2: 浣跨敤 pandas 璇诲彇Excel鏂囦欢...")
        # 閫氳繃 dtype 鍙傛暟鎸囧畾鍒椾互瀛楃涓插舰寮忚鍙栵紝閬垮厤鑷姩杞崲鏍煎紡
        # 閲嶈锛氬亣璁?涓婄骇璁惧"鍒楃幇鍦ㄥ寘鍚殑鏄埗璁惧鐨勮祫浜х紪鍙?
        df = pd.read_excel(buffer, dtype={
            '璧勪骇缂栧彿': str,
            '璁惧鎶曚骇鏃堕棿': str,
            '涓婄骇璁惧': str 
        })
        df = df.where(pd.notna(df), None) # 灏?NaN 鏇挎崲涓?None
        print(f"姝ラ 2: 瀹屾垚銆傝鍙栧埌 {len(df)} 琛屾暟鎹€?)
        print(f"Excel 鏂囦欢鍒楀悕: {df.columns.tolist()}")
        
        # 楠岃瘉蹇呰鐨勫垪鏄惁瀛樺湪
        required_columns = ['璧勪骇缂栧彿', '璁惧鍚嶇О']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            error_msg = f"Excel鏂囦欢缂哄皯蹇呰鐨勫垪: {missing_columns}"
            print(f"閿欒: {error_msg}")
            return RedirectResponse(url=f"/?error={quote(error_msg)}", status_code=303)
        
        # 鏄剧ず鍓嶅嚑琛屾暟鎹牱鏈敤浜庤皟璇?
        print("\n鍓?琛屾暟鎹牱鏈?")
        for i in range(min(3, len(df))):
            print(f"绗瑊i+1}琛? 璧勪骇缂栧彿={df.iloc[i].get('璧勪骇缂栧彿')}, 璁惧鍚嶇О={df.iloc[i].get('璁惧鍚嶇О')}")

        devices_map = {} # 杩欎釜鏄犲皠灏嗕互 璧勪骇缂栧彿 涓洪敭
        devices_created_count = 0
        devices_updated_count = 0
        skipped_rows = []

        # 姝ラ 3: 澧為噺鏇存柊璁惧锛堝垱寤烘垨鏇存柊锛?
        print("\n姝ラ 3: 寮€濮嬬涓€閬嶅鐞?- 澧為噺鏇存柊璁惧锛堝垱寤烘柊璁惧鎴栨洿鏂扮幇鏈夎澶囷級...")
        for index, row in df.iterrows():
            # 鏂板锛氳幏鍙栧苟鏍￠獙璧勪骇缂栧彿
            asset_id = row.get("璧勪骇缂栧彿")
            if isinstance(asset_id, str):
                asset_id = asset_id.strip()

            if not asset_id or asset_id == 'nan' or asset_id.lower() == 'none':
                skip_reason = f"璧勪骇缂栧彿涓虹┖鎴栨棤鏁? '{asset_id}'"
                print(f"  - 绗?{index+2} 琛岋細璺宠繃锛寋skip_reason}")
                skipped_rows.append((index+2, skip_reason))
                continue
            
            device_name = row.get("璁惧鍚嶇О")
            if isinstance(device_name, str):
                device_name = device_name.strip()

            if not device_name or device_name == 'nan' or device_name.lower() == 'none':
                skip_reason = f"璁惧鍚嶇О涓虹┖鎴栨棤鏁? '{device_name}'"
                print(f"  - 绗?{index+2} 琛岋細璺宠繃锛寋skip_reason}")
                skipped_rows.append((index+2, skip_reason))
                continue
            
            # 妫€鏌ヨ祫浜х紪鍙锋槸鍚﹀凡鍦ㄦ湰娆′笂浼犱腑閲嶅
            if asset_id in devices_map:
                skip_reason = f"璧勪骇缂栧彿 '{asset_id}' 鍦‥xcel鏂囦欢涓噸澶?
                print(f"  - 绗?{index+2} 琛岋細璺宠繃锛寋skip_reason}")
                skipped_rows.append((index+2, skip_reason))
                continue

            try:
                # 妫€鏌ユ暟鎹簱涓槸鍚﹀凡瀛樺湪璇ヨ祫浜х紪鍙风殑璁惧
                existing_device = db.query(Device).filter(Device.asset_id == asset_id).first()
                
                # 鑾峰彇灞€绔欎俊鎭?
                station = row.get("灞€绔?)
                if isinstance(station, str):
                    station = station.strip()
                if not station or station == 'nan' or station.lower() == 'none':
                    skip_reason = f"灞€绔欎俊鎭负绌烘垨鏃犳晥: '{station}'"
                    print(f"  - 绗?{index+2} 琛岋細璺宠繃锛寋skip_reason}")
                    skipped_rows.append((index+2, skip_reason))
                    continue
                
                # 鑾峰彇骞堕獙璇佽澶囩被鍨?
                device_type = row.get("璁惧绫诲瀷")
                if isinstance(device_type, str):
                    device_type = device_type.strip()
                
                # 楠岃瘉璁惧绫诲瀷鏄惁鍦ㄦ爣鍑嗗垪琛ㄤ腑
                if device_type and device_type != 'nan' and device_type.lower() != 'none':
                    if not validate_device_type(device_type):
                        # 鎻愪緵寤鸿鐨勮澶囩被鍨?
                        suggestions = get_device_type_suggestions(device_type)
                        if suggestions:
                            suggestion_text = f"锛屽缓璁娇鐢? {', '.join(suggestions[:3])}"
                        else:
                            suggestion_text = """
                        skip_reason = f"璁惧绫诲瀷 '{device_type}' 涓嶅湪鏍囧噯鍒楄〃涓瓄suggestion_text}"
                        print(f"  - 绗?{index+2} 琛岋細璺宠繃锛寋skip_reason}")
                        skipped_rows.append((index+2, skip_reason))
                        continue
                else:
                    # 濡傛灉璁惧绫诲瀷涓虹┖锛岃缃负"寰呯‘璁?
                    device_type = "寰呯‘璁?
                
                if existing_device:
                    # 鏇存柊鐜版湁璁惧
                    existing_device.name = device_name
                    existing_device.station = station
                    existing_device.model = row.get("璁惧鍨嬪彿")
                    existing_device.device_type = device_type  # 浣跨敤楠岃瘉鍚庣殑璁惧绫诲瀷
                    existing_device.location = row.get("鏈烘埧鍐呯┖闂翠綅缃?)
                    existing_device.power_rating = row.get("璁惧棰濆畾瀹归噺")
                    existing_device.vendor = row.get("璁惧鐢熶骇鍘傚")
                    existing_device.commission_date = row.get("璁惧鎶曚骇鏃堕棿")
                    existing_device.remark = row.get("澶囨敞")
                    
                    # 娉ㄦ剰锛氫互涓嬫満鎴跨浉鍏冲瓧娈佃蹇界暐锛堟牴鎹敤鎴疯姹傦級锛?
                    # - 鏈烘埧鍚嶇О
                    # - 璧勬簮绯荤粺鏈烘埧鍚嶇О  
                    # - 璧勬簮绯荤粺鏈烘埧缂栫爜
                    # - 鏈烘埧绛夌骇
                    
                    devices_map[asset_id] = existing_device
                    devices_updated_count += 1
                    print(f"  - 绗?{index+2} 琛岋細鍑嗗鏇存柊鐜版湁璁惧 '{device_name}' (璧勪骇缂栧彿: {asset_id}, 灞€绔? {station})")
                else:
                    # 鍒涘缓鏂拌澶?
                    device = Device(
                        asset_id=asset_id,
                        name=device_name,
                        station=station,
                        model=row.get("璁惧鍨嬪彿"),
                        device_type=device_type,  # 浣跨敤楠岃瘉鍚庣殑璁惧绫诲瀷
                        location=row.get("鏈烘埧鍐呯┖闂翠綅缃?),
                        power_rating=row.get("璁惧棰濆畾瀹归噺"),
                        vendor=row.get("璁惧鐢熶骇鍘傚"),
                        commission_date=row.get("璁惧鎶曚骇鏃堕棿"),
                        remark=row.get("澶囨敞")
                        # 娉ㄦ剰锛氫互涓嬫満鎴跨浉鍏冲瓧娈佃蹇界暐锛堟牴鎹敤鎴疯姹傦級锛?
                        # - 鏈烘埧鍚嶇О
                        # - 璧勬簮绯荤粺鏈烘埧鍚嶇О  
                        # - 璧勬簮绯荤粺鏈烘埧缂栫爜
                        # - 鏈烘埧绛夌骇
                    )
                    db.add(device)
                    devices_map[asset_id] = device
                    devices_created_count += 1
                    print(f"  - 绗?{index+2} 琛岋細鍑嗗鍒涘缓鏂拌澶?'{device_name}' (璧勪骇缂栧彿: {asset_id}, 灞€绔? {station})")
                    
            except Exception as device_error:
                skip_reason = f"澶勭悊璁惧澶辫触: {device_error}"
                print(f"  - 绗?{index+2} 琛岋細璺宠繃锛寋skip_reason}")
                skipped_rows.append((index+2, skip_reason))
                continue
        
        print(f"\n鍑嗗鎻愪氦璁惧鏇存敼鍒版暟鎹簱锛堟柊寤? {devices_created_count}, 鏇存柊: {devices_updated_count}锛?..")
        try:
            db.commit() # 鎻愪氦浜嬪姟浠ョ敓鎴愯澶嘔D
            print("璁惧鎻愪氦鎴愬姛锛?)
        except Exception as commit_error:
            print(f"璁惧鎻愪氦澶辫触: {commit_error}")
            db.rollback()
            raise commit_error
            
        # 楠岃瘉璁惧鏁伴噺
        actual_device_count = db.query(Device).count()
        print(f"姝ラ 3: 瀹屾垚銆傛柊寤?{devices_created_count} 涓澶囷紝鏇存柊 {devices_updated_count} 涓澶囷紝鏁版嵁搴撲腑鎬诲叡鏈?{actual_device_count} 涓澶囥€?)
        
        if skipped_rows:
            print(f"\n璺宠繃鐨勮鏁扮粺璁? {len(skipped_rows)} 琛?)
            for row_num, reason in skipped_rows[:5]:  # 鍙樉绀哄墠5涓?
                print(f"  绗瑊row_num}琛? {reason}")
            if len(skipped_rows) > 5:
                print(f"  ... 杩樻湁 {len(skipped_rows) - 5} 琛岃璺宠繃")

        # 鍒锋柊鏄犲皠锛岀‘淇濆璞″寘鍚暟鎹簱鐢熸垚鐨処D
        print("\n鍒锋柊璁惧瀵硅薄浠ヨ幏鍙栨暟鎹簱鐢熸垚鐨処D...")
        for asset_id_key in list(devices_map.keys()):
            try:
                db.refresh(devices_map[asset_id_key])
                print(f"  璁惧 {asset_id_key} ID: {devices_map[asset_id_key].id}")
            except Exception as refresh_error:
                print(f"  鍒锋柊璁惧 {asset_id_key} 澶辫触: {refresh_error}")

        # 姝ラ 4: 娓呯悊娑夊強Excel璁惧鐨勬棫杩炴帴
        print("\n姝ラ 4: 娓呯悊娑夊強Excel涓澶囩殑鏃ц繛鎺?..")
        excel_device_ids = [device.id for device in devices_map.values()]
        if excel_device_ids:
            # 鍒犻櫎娑夊強杩欎簺璁惧鐨勬墍鏈夎繛鎺ワ紙浣滀负婧愯澶囨垨鐩爣璁惧锛?
            old_connections_deleted = db.query(Connection).filter(
                (Connection.source_device_id.in_(excel_device_ids)) |
                (Connection.target_device_id.in_(excel_device_ids))
            ).delete(synchronize_session=False)
            db.commit()
            print(f"鍒犻櫎浜?{old_connections_deleted} 涓秹鍙奅xcel璁惧鐨勬棫杩炴帴")
        else:
            print("娌℃湁Excel璁惧锛岃烦杩囪繛鎺ユ竻鐞?)
            
        connections_created_count = 0
        connection_skipped_rows = []
        
        # 姝ラ 5: 鍒涘缓鏂拌繛鎺?
        print("\n姝ラ 5: 寮€濮嬬浜岄亶澶勭悊 - 鍒涘缓鏂拌繛鎺?..")
        for index, row in df.iterrows():
            # 浣跨敤璧勪骇缂栧彿鏉ユ煡鎵捐澶?
            source_asset_id = row.get("涓婄骇璁惧")
            target_asset_id = row.get("璧勪骇缂栧彿")

            if isinstance(source_asset_id, str):
                source_asset_id = source_asset_id.strip()
            if isinstance(target_asset_id, str):
                target_asset_id = target_asset_id.strip()
            
            # 妫€鏌ユ槸鍚︽湁涓婄骇璁惧淇℃伅
            if not source_asset_id or source_asset_id == 'nan' or source_asset_id.lower() == 'none':
                print(f"  - 绗?{index+2} 琛岋細璺宠繃杩炴帴鍒涘缓锛屾棤涓婄骇璁惧淇℃伅")
                continue
                
            # 纭繚婧愬拰鐩爣璁惧閮藉瓨鍦ㄤ簬鏄犲皠涓?
            if target_asset_id and source_asset_id:
                if source_asset_id not in devices_map:
                    skip_reason = f"涓婄骇璁惧 '{source_asset_id}' 涓嶅瓨鍦?
                    print(f"  - 绗?{index+2} 琛岋細璺宠繃杩炴帴锛寋skip_reason}")
                    connection_skipped_rows.append((index+2, skip_reason))
                    continue
                    
                if target_asset_id not in devices_map:
                    skip_reason = f"鐩爣璁惧 '{target_asset_id}' 涓嶅瓨鍦?
                    print(f"  - 绗?{index+2} 琛岋細璺宠繃杩炴帴锛寋skip_reason}")
                    connection_skipped_rows.append((index+2, skip_reason))
                    continue
                
                source_device = devices_map[source_asset_id]
                target_device = devices_map[target_asset_id]
                
                try:
                    connection = Connection(
                        source_device_id=source_device.id,
                        source_port=row.get("涓婄骇绔彛"),
                        target_device_id=target_device.id,
                        target_port=row.get("鏈绔彛"),
                        cable_type=row.get("绾跨紗绫诲瀷")
                    )
                    db.add(connection)
                    connections_created_count += 1
                    print(f"  - 绗?{index+2} 琛岋細鍑嗗鍒涘缓浠?'{source_device.name}' 鍒?'{target_device.name}' 鐨勮繛鎺?)
                except Exception as conn_error:
                    skip_reason = f"鍒涘缓杩炴帴瀵硅薄澶辫触: {conn_error}"
                    print(f"  - 绗?{index+2} 琛岋細璺宠繃杩炴帴锛寋skip_reason}")
                    connection_skipped_rows.append((index+2, skip_reason))
                    continue
        
        print(f"\n鍑嗗鎻愪氦 {connections_created_count} 涓繛鎺ュ埌鏁版嵁搴?..")
        try:
            db.commit()
            print("杩炴帴鎻愪氦鎴愬姛锛?)
        except Exception as commit_error:
            print(f"杩炴帴鎻愪氦澶辫触: {commit_error}")
            db.rollback()
            raise commit_error
            
        # 楠岃瘉杩炴帴鏄惁鐪熺殑琚垱寤?
        actual_connection_count = db.query(Connection).count()
        print(f"姝ラ 5: 瀹屾垚銆傞鏈熷垱寤?{connections_created_count} 涓繛鎺ワ紝瀹為檯鏁版嵁搴撲腑鏈?{actual_connection_count} 涓繛鎺ャ€?)
        
        if connection_skipped_rows:
            print(f"\n杩炴帴璺宠繃鐨勮鏁扮粺璁? {len(connection_skipped_rows)} 琛?)
            for row_num, reason in connection_skipped_rows[:5]:  # 鍙樉绀哄墠5涓?
                print(f"  绗瑊row_num}琛? {reason}")
            if len(connection_skipped_rows) > 5:
                print(f"  ... 杩樻湁 {len(connection_skipped_rows) - 5} 琛岃繛鎺ヨ璺宠繃")
        
        # 姝ラ 6: 澶勭悊Sheet2杩炴帴鏁版嵁
        sheet2_connections_count = 0
        sheet2_skipped_rows = []
        
        try:
            print("\n姝ラ 6: 寮€濮嬪鐞哠heet2杩炴帴鏁版嵁...")
            
            # 灏濊瘯璇诲彇Sheet2锛堣繛鎺ヨ〃锛?
            try:
                # 閲嶇疆buffer浣嶇疆鍒板紑澶达紝鍥犱负涔嬪墠璇诲彇Sheet1鏃跺凡缁忕Щ鍔ㄤ簡浣嶇疆
                buffer.seek(0)
                df_connections = pd.read_excel(buffer, sheet_name='杩炴帴')
                print(f"鎴愬姛璇诲彇Sheet2锛屽叡 {len(df_connections)} 琛岃繛鎺ユ暟鎹?)
            except Exception as sheet_error:
                print(f"鏃犳硶璇诲彇Sheet2锛堣繛鎺ヨ〃锛? {sheet_error}")
                print("璺宠繃Sheet2澶勭悊锛岀户缁畬鎴愬鍏?)
                df_connections = None
            
            if df_connections is not None and len(df_connections) > 0:
                # 杩炴帴绫诲瀷鏄犲皠 - 鎵╁睍鏄犲皠琛ㄤ互鍖呭惈鏇村鍙兘鐨勭┖鍊艰〃绀?
                CONNECTION_TYPE_MAPPING = {
                    # 鏍囧噯杩炴帴绫诲瀷
                    '鐢电紗': 'cable',
                    '閾滄帓': 'busbar', 
                    '姣嶇嚎': 'busway',
                    'cable': 'cable',
                    'busbar': 'busbar',
                    'busway': 'busway',
                    # 鐢垫皵杩炴帴绫诲瀷 - 鏍规嵁瀹為檯Excel鏁版嵁娣诲姞
                    '鐩存祦': 'DC',
                    '浜ゆ祦': 'AC',
                    'DC': 'DC',
                    'AC': 'AC',
                    'dc': 'DC',
                    'ac': 'AC',
                    # 绌哄€肩殑鍚勭琛ㄧず鏂瑰紡 - 缁熶竴鏄犲皠涓篘one琛ㄧず绌洪棽绔彛
                    '鏃?: None,
                    '绌?: None,
                    '绌洪棽': None,
                    '鏈繛鎺?: None,
                    'N/A': None,
                    'n/a': None,
                    'NA': None,
                    'na': None,
                    '鏃犺繛鎺?: None,
                    '寰呰繛鎺?: None,
                    '棰勭暀': None,
                    'None': None,
                    'null': None,
                    'NULL': None,
                    '': None,  # 绌哄瓧绗︿覆
                    ' ': None,  # 绌烘牸
                }
                
                # 杈呭姪鍑芥暟锛氳幏鍙栨垨鍒涘缓璁惧
                def get_or_create_device(device_name: str, default_station: str = "鏈煡绔欑偣"):
                    """鑾峰彇璁惧锛屽鏋滀笉瀛樺湪鍒欒嚜鍔ㄥ垱寤?"""
                    if not device_name:
                        return None
                    
                    device = db.query(Device).filter(Device.name == device_name).first()
                    if not device:
                        # 鑷姩鍒涘缓璁惧
                        device = Device(
                            name=device_name,
                            asset_id=f"AUTO_{len(device_name)}_{hash(device_name) % 10000:04d}",  # 鐢熸垚鍞竴璧勪骇缂栧彿
                            station=default_station,
                            device_type="寰呯‘璁?,
                            location="寰呯‘璁?,
                            remark="閫氳繃Excel瀵煎叆鏃惰嚜鍔ㄥ垱寤猴紝璇峰畬鍠勮澶囦俊鎭?
                        )
                        db.add(device)
                        db.flush()  # 鑾峰彇ID浣嗕笉鎻愪氦
                        print(f"  * 鑷姩鍒涘缓璁惧: {device_name} (ID: {device.id})")
                    return device
                
                # 缁熻淇℃伅
                created_devices = []
                warnings = []
                
                for index, row in df_connections.iterrows():
                    try:
                        # 鑾峰彇璁惧鍚嶇О
                        source_device_name = str(row.get('A绔澶囧悕绉?, '')).strip()
                        target_device_name = str(row.get('B绔澶囧悕绉?, '')).strip()
                        
                        # 澶勭悊绌鸿澶囧悕绉扮殑鎯呭喌
                        if not source_device_name and not target_device_name:
                            skip_reason = "A绔拰B绔澶囧悕绉伴兘涓虹┖"
                            print(f"  - 绗?{index+2} 琛岋細璺宠繃杩炴帴锛寋skip_reason}")
                            sheet2_skipped_rows.append((index+2, skip_reason))
                            continue
                        elif not source_device_name:
                            skip_reason = "A绔澶囧悕绉颁负绌?
                            print(f"  - 绗?{index+2} 琛岋細璺宠繃杩炴帴锛寋skip_reason}")
                            sheet2_skipped_rows.append((index+2, skip_reason))
                            continue
                        elif not target_device_name:
                            skip_reason = "B绔澶囧悕绉颁负绌?
                            print(f"  - 绗?{index+2} 琛岋細璺宠繃杩炴帴锛寋skip_reason}")
                            sheet2_skipped_rows.append((index+2, skip_reason))
                            continue
                        
                        # 鑾峰彇鎴栧垱寤鸿澶?
                        source_device = get_or_create_device(source_device_name)
                        target_device = get_or_create_device(target_device_name)
                        
                        if not source_device or not target_device:
                            skip_reason = "璁惧鍒涘缓澶辫触"
                            print(f"  - 绗?{index+2} 琛岋細璺宠繃杩炴帴锛寋skip_reason}")
                            sheet2_skipped_rows.append((index+2, skip_reason))
                            continue
                        
                        # 璁板綍鏂板垱寤虹殑璁惧
                        if source_device.remark and "閫氳繃Excel瀵煎叆鏃惰嚜鍔ㄥ垱寤? in source_device.remark:
                            if source_device_name not in created_devices:
                                created_devices.append(source_device_name)
                        if target_device.remark and "閫氳繃Excel瀵煎叆鏃惰嚜鍔ㄥ垱寤? in target_device.remark:
                            if target_device_name not in created_devices:
                                created_devices.append(target_device_name)
                        
                        # 澶勭悊绔彛閫昏緫
                        def build_port_info(fuse_number, fuse_spec, breaker_number, breaker_spec):
                            """鏋勫缓绔彛淇℃伅锛屼紭鍏堜娇鐢ㄧ啍涓濓紝鍏舵浣跨敤绌哄紑"""
                            fuse_num = str(fuse_number).strip() if pd.notna(fuse_number) else ''
                            fuse_sp = str(fuse_spec).strip() if pd.notna(fuse_spec) else ''
                            breaker_num = str(breaker_number).strip() if pd.notna(breaker_number) else ''
                            breaker_sp = str(breaker_spec).strip() if pd.notna(breaker_spec) else ''
                            
                            if fuse_num and fuse_num != 'nan':
                                return f"{fuse_num} ({fuse_sp})" if fuse_sp and fuse_sp != 'nan' else fuse_num
                            elif breaker_num and breaker_num != 'nan':
                                return f"{breaker_num} ({breaker_sp})" if breaker_sp and breaker_sp != 'nan' else breaker_num
                            else:
                                return None
                        
                        # 鏋勫缓A绔拰B绔鍙ｄ俊鎭?
                        source_port = build_port_info(
                            row.get('A绔啍涓濈紪鍙?), row.get('A绔啍涓濊鏍?),
                            row.get('A绔┖寮€缂栧彿'), row.get('A绔┖寮€瑙勬牸')
                        )
                        target_port = build_port_info(
                            row.get('B绔啍涓濈紪鍙?), row.get('B绔啍涓濊鏍?),
                            row.get('B绔┖寮€缂栧彿'), row.get('绌哄紑瑙勬牸')
                        )
                        
                        # 澶勭悊杩炴帴绫诲瀷 - 淇绌洪棽绔彛琚敊璇綊绫讳负鐢电紗鐨勯棶棰?
                        connection_type_raw = row.get('杩炴帴绫诲瀷锛堜氦娴?鐩存祦锛?)  # 淇锛氫娇鐢ㄥ疄闄匛xcel鍒楀悕
                        # 瑙勫垯锛氱┖鍊?-> None锛涢潪绌哄垯鎸夋槧灏勮〃杞崲锛屾棤娉曟槧灏勪篃缃?None 骞惰褰曡鍛?
                        connection_type = None
                        if pd.isna(connection_type_raw) or str(connection_type_raw).strip() == '':
                            connection_type = None
                        else:
                            raw = str(connection_type_raw).strip()
                            connection_type = CONNECTION_TYPE_MAPPING.get(raw, None)
                            if raw not in CONNECTION_TYPE_MAPPING:
                                print(f"  * 璀﹀憡锛氱 {index+2} 琛岃繛鎺ョ被鍨?'{raw}' 鏃犳硶璇嗗埆锛岃缃负绌洪棽绔彛")
                                warnings.append(f"绗?{index+2} 琛岋細杩炴帴绫诲瀷 '{raw}' 鏃犳硶璇嗗埆")

                        # 妫€鏌ユ槸鍚﹀凡瀛樺湪鐩稿悓杩炴帴锛堟寜璁惧涓庣鍙ｇ淮搴﹀幓閲嶏級
                        existing_connection = db.query(Connection).filter(
                            Connection.source_device_id == source_device.id,
                            Connection.target_device_id == target_device.id,
                            Connection.source_port == source_port,
                            Connection.target_port == target_port
                        ).first()

                        if existing_connection:
                            skip_reason = "杩炴帴宸插瓨鍦?
                            print(f"  - 绗?{index+2} 琛岋細璺宠繃杩炴帴锛寋skip_reason}")
                            sheet2_skipped_rows.append((index+2, skip_reason))
                            continue

                        # 鍒涘缓杩炴帴璁板綍锛堜繚鐣欑幇鏈夊瓧娈碉紝鏈缃?cable_type锛岄伩鍏嶈鐢級
                        connection = Connection(
                            source_device_id=source_device.id,
                            target_device_id=target_device.id,
                            source_port=source_port,
                            target_port=target_port,
                            # A绔俊鎭?
                            source_fuse_number=str(row.get('A绔啍涓濈紪鍙?, '')).strip() if pd.notna(row.get('A绔啍涓濈紪鍙?)) else None,
                            source_fuse_spec=str(row.get('A绔啍涓濊鏍?, '')).strip() if pd.notna(row.get('A绔啍涓濊鏍?)) else None,
                            source_breaker_number=str(row.get('A绔┖寮€缂栧彿', '')).strip() if pd.notna(row.get('A绔┖寮€缂栧彿')) else None,
                            source_breaker_spec=str(row.get('A绔┖寮€瑙勬牸', '')).strip() if pd.notna(row.get('A绔┖寮€瑙勬牸')) else None,
                            # B绔俊鎭?
                            target_fuse_number=str(row.get('B绔啍涓濈紪鍙?, '')).strip() if pd.notna(row.get('B绔啍涓濈紪鍙?)) else None,
                            target_fuse_spec=str(row.get('B绔啍涓濊鏍?, '')).strip() if pd.notna(row.get('B绔啍涓濊鏍?)) else None,
                            target_breaker_number=str(row.get('B绔┖寮€缂栧彿', '')).strip() if pd.notna(row.get('B绔┖寮€缂栧彿')) else None,
                            target_breaker_spec=str(row.get('绌哄紑瑙勬牸', '')).strip() if pd.notna(row.get('绌哄紑瑙勬牸')) else None,
                            target_device_location=str(row.get('B绔澶囦綅缃紙闈炲姩鍔涜澶囷級', '')).strip() if pd.notna(row.get('B绔澶囦綅缃紙闈炲姩鍔涜澶囷級')) else None,
                            # 棰濆畾鐢垫祦淇℃伅
                            a_rated_current=str(row.get('A绔瀹氱數娴?, '')).strip() if pd.notna(row.get('A绔瀹氱數娴?)) else None,
                            b_rated_current=str(row.get('B绔瀹氱數娴?, '')).strip() if pd.notna(row.get('B绔瀹氱數娴?)) else None,
                            # 杩炴帴淇℃伅
                            hierarchy_relation=str(row.get('涓婁笅绾?, '')).strip() if pd.notna(row.get('涓婁笅绾?)) else None,
                            upstream_downstream=str(row.get('涓婁笅娓?, '')).strip() if pd.notna(row.get('涓婁笅娓?)) else None,
                            connection_type=connection_type,
                            cable_model=str(row.get('鐢电紗鍨嬪彿', '')).strip() if pd.notna(row.get('鐢电紗鍨嬪彿')) else None,
                            # 闄勫姞淇℃伅
                            source_device_photo=str(row.get('A绔澶囩収鐗?, '')).strip() if pd.notna(row.get('A绔澶囩収鐗?)) else None,
                            target_device_photo=str(row.get('B绔澶囩収鐗?, '')).strip() if pd.notna(row.get('B绔澶囩収鐗?)) else None,
                            remark=str(row.get('澶囨敞', '')).strip() if pd.notna(row.get('澶囨敞')) else None,
                            # 瀹夎鏃ユ湡锛圗xcel涓病鏈夋瀛楁锛岃缃负None锛?
                            installation_date=None
                        )

                        db.add(connection)
                        sheet2_connections_count += 1
                        print(f"  - 绗?{index+2} 琛岋細鍑嗗鍒涘缓浠?'{source_device_name}' 鍒?'{target_device_name}' 鐨勮繛鎺?)
                        print(f"    婧愮鍙? {source_port}, 鐩爣绔彛: {target_port}, 杩炴帴绫诲瀷: {connection_type}")
                        
                    except Exception as conn_error:
                        skip_reason = f"澶勭悊杩炴帴澶辫触: {conn_error}"
                        print(f"  - 绗?{index+2} 琛岋細璺宠繃杩炴帴锛寋skip_reason}")
                        sheet2_skipped_rows.append((index+2, skip_reason))
                        continue
                
                # 鎻愪氦Sheet2杩炴帴
                if sheet2_connections_count > 0:
                    print(f"\n鍑嗗鎻愪氦 {sheet2_connections_count} 涓猄heet2杩炴帴鍒版暟鎹簱...")
                    try:
                        db.commit()
                        print("Sheet2杩炴帴鎻愪氦鎴愬姛锛?)
                    except Exception as commit_error:
                        print(f"Sheet2杩炴帴鎻愪氦澶辫触: {commit_error}")
                        db.rollback()
                        raise commit_error
                
                # 鐢熸垚璇︾粏鐨勫鍏ユ姤鍛?
                print(f"\n=== Sheet2杩炴帴瀵煎叆鎶ュ憡 ===")
                print(f"鎬昏繛鎺ユ暟: {len(df_connections)} 琛?)
                print(f"鎴愬姛瀵煎叆: {sheet2_connections_count} 涓繛鎺?)
                print(f"璺宠繃杩炴帴: {len(sheet2_skipped_rows)} 琛?)
                
                if created_devices:
                    print(f"\n鑷姩鍒涘缓鐨勮澶?({len(created_devices)} 涓?:")
                    for device_name in created_devices:
                        print(f"  + {device_name}")
                    print("\n娉ㄦ剰: 鑷姩鍒涘缓鐨勮澶囦俊鎭笉瀹屾暣锛岃鍦ㄨ澶囩鐞嗛〉闈㈠畬鍠勭浉鍏充俊鎭€?)
                
                if sheet2_skipped_rows:
                    print(f"\n璺宠繃鐨勮繛鎺ヨ鎯?")
                    skip_reasons = {}
                    for row_num, reason in sheet2_skipped_rows:
                        if reason not in skip_reasons:
                            skip_reasons[reason] = []
                        skip_reasons[reason].append(row_num)
                    
                    for reason, rows in skip_reasons.items():
                        print(f"  {reason}: {len(rows)} 琛?(绗瑊', '.join(map(str, rows[:3]))}琛寋'...' if len(rows) > 3 else ''})")
                
                # 璁＄畻瀵煎叆鎴愬姛鐜?
                success_rate = (sheet2_connections_count / len(df_connections)) * 100 if len(df_connections) > 0 else 0
                print(f"\n瀵煎叆鎴愬姛鐜? {success_rate:.1f}% ({sheet2_connections_count}/{len(df_connections)})")
            
            print(f"姝ラ 6: 瀹屾垚銆備粠Sheet2鍒涘缓浜?{sheet2_connections_count} 涓繛鎺?)
            
        except Exception as sheet2_error:
            print(f"澶勭悊Sheet2鏃跺嚭閿? {sheet2_error}")
            print("缁х画瀹屾垚瀵煎叆锛屽拷鐣heet2閿欒")
        
        # 鏈€缁堢粺璁?
        final_connection_count = db.query(Connection).count()
        total_connections_created = connections_created_count + sheet2_connections_count
        
        print("\n=== Excel鏂囦欢澧為噺鏇存柊澶勭悊鎴愬姛 ===")
        print(f"澶勭悊缁撴灉: 鏂板缓 {devices_created_count} 涓澶? 鏇存柊 {devices_updated_count} 涓澶?)
        print(f"杩炴帴鍒涘缓: Sheet1鍒涘缓 {connections_created_count} 涓? Sheet2鍒涘缓 {sheet2_connections_count} 涓? 鎬昏 {total_connections_created} 涓?)
        print(f"鏁版嵁搴撴渶缁堢姸鎬? {actual_device_count} 涓澶? {final_connection_count} 涓繛鎺?)

    except Exception as e:
        print(f"\n!!! 鍙戠敓寮傚父锛屽紑濮嬪洖婊氫簨鍔?!!!")
        try:
            db.rollback()
            print("浜嬪姟鍥炴粴鎴愬姛")
        except Exception as rollback_error:
            print(f"浜嬪姟鍥炴粴澶辫触: {rollback_error}")
            
        error_message = f"澶勭悊Excel鏂囦欢鏃跺嚭閿? {e}"
        print(f"\n=== Excel鏂囦欢澶勭悊澶辫触 ===")
        print(f"閿欒绫诲瀷: {type(e).__name__}")
        print(f"閿欒淇℃伅: {error_message}")
        print("\n瀹屾暣閿欒鍫嗘爤:")
        traceback.print_exc()
        print("=" * 50)
        
        # 妫€鏌ユ暟鎹簱鐘舵€?
        try:
            final_device_count = db.query(Device).count()
            final_connection_count = db.query(Connection).count()
            print(f"\n閿欒鍚庢暟鎹簱鐘舵€? {final_device_count} 涓澶? {final_connection_count} 涓繛鎺?)
        except Exception as db_check_error:
            print(f"鏃犳硶妫€鏌ユ暟鎹簱鐘舵€? {db_check_error}")
            
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)

    print(f"\n涓婁紶澶勭悊瀹屾垚锛岄噸瀹氬悜鍒伴椤?..")
    return RedirectResponse(url="/", status_code=303)

# 鏇存柊璁惧淇℃伅
@app.post("/devices/{device_id}")
async def update_device(
    device_id: int,
    asset_id: str = Form(...),
    name: str = Form(...),
    station: str = Form(...),
    model: str = Form(None),
    device_type: str = Form(None),
    location: str = Form(None),
    power_rating: str = Form(None),
    vendor: str = Form(None),
    commission_date: str = Form(None),
    remark: str = Form(None),
    db: Session = Depends(get_db)
):
    """鏇存柊璁惧淇℃伅锛堢紪杈戝姛鑳戒笉闇€瑕佸瘑鐮侀獙璇侊紝鍥犱负鍦ㄨ繘鍏ョ紪杈戦〉闈㈡椂宸查獙璇侊級"""
    try:
        # 鑾峰彇瑕佹洿鏂扮殑璁惧
        device = db.query(Device).filter(Device.id == device_id).first()
        if not device:
            error_message = "璁惧涓嶅瓨鍦ㄣ€?
            return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
        
        # 妫€鏌ヨ祫浜х紪鍙峰敮涓€鎬э紙鎺掗櫎褰撳墠璁惧锛?
        existing_device = db.query(Device).filter(
            Device.asset_id == asset_id,
            Device.id != device_id
        ).first()
        if existing_device:
            error_message = f"璧勪骇缂栧彿 {asset_id} 宸插瓨鍦紝璇蜂娇鐢ㄥ叾浠栫紪鍙枫€?
            return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
        
        # 鏇存柊璁惧淇℃伅
        device.asset_id = asset_id
        device.name = name
        device.station = station
        device.model = model if model else None
        device.device_type = device_type if device_type else None
        device.location = location if location else None
        device.power_rating = power_rating if power_rating else None
        device.vendor = vendor if vendor else None
        device.commission_date = commission_date if commission_date else None
        device.remark = remark if remark else None
        
        db.commit()
        
        success_message = f"璁惧 {name} 鏇存柊鎴愬姛銆?
        return RedirectResponse(url=f"/?success={quote(success_message)}", status_code=303)
        
    except Exception as e:
        db.rollback()
        error_message = f"鏇存柊璁惧澶辫触锛歿str(e)}"
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)

# 缂栬緫璁惧椤甸潰
@app.get("/edit/{device_id}")
async def edit_device_page(device_id: int, password: str, request: Request, db: Session = Depends(get_db)):
    """鏄剧ず缂栬緫璁惧椤甸潰"""
    # 楠岃瘉绠＄悊鍛樺瘑鐮?
    if not verify_admin_password(password):
        error_message = "瀵嗙爜閿欒锛屾棤鏉冮檺鎵ц姝ゆ搷浣溿€?
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
    
    # 鑾峰彇璁惧淇℃伅
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        error_message = "璁惧涓嶅瓨鍦ㄣ€?
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
    
    return templates.TemplateResponse("edit_device.html", {
        "request": request,
        "device": device
    })

# 鍒犻櫎璁惧
@app.delete("/devices/{device_id}")
async def delete_device(device_id: int, request: Request, db: Session = Depends(get_db)):
    """鍒犻櫎璁惧"""
    try:
        # 鑾峰彇璇锋眰浣撲腑鐨勫瘑鐮?
        body = await request.json()
        password = body.get("password")
        
        # 楠岃瘉绠＄悊鍛樺瘑鐮?
        if not verify_admin_password(password):
            raise HTTPException(status_code=403, detail="瀵嗙爜閿欒锛屾棤鏉冮檺鎵ц姝ゆ搷浣溿€?)
        
        # 鑾峰彇瑕佸垹闄ょ殑璁惧
        device = db.query(Device).filter(Device.id == device_id).first()
        if not device:
            raise HTTPException(status_code=404, detail="璁惧涓嶅瓨鍦ㄣ€?)
        
        device_name = device.name
        
        # 鍒犻櫎鐩稿叧鐨勮繛鎺ヨ褰?
        db.query(Connection).filter(
            (Connection.source_device_id == device_id) | 
            (Connection.target_device_id == device_id)
        ).delete()
        
        # 鍒犻櫎璁惧
        db.delete(device)
        db.commit()
        
        return {"message": f"璁惧 {device_name} 鍒犻櫎鎴愬姛銆?}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"鍒犻櫎璁惧澶辫触锛歿str(e)}")

@app.post("/devices")
async def create_device(
    asset_id: str = Form(...),
    name: str = Form(...),
    station: str = Form(...),
    model: str = Form(None),
    device_type: str = Form(None),
    location: str = Form(None),
    power_rating: str = Form(None),
    vendor: str = Form(None),
    commission_date: str = Form(None),
    remark: str = Form(None),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    # 楠岃瘉绠＄悊鍛樺瘑鐮?
    if not verify_admin_password(password):
        error_message = "瀵嗙爜閿欒锛屾棤鏉冮檺鎵ц姝ゆ搷浣溿€?
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)
    
    # 澧炲姞璧勪骇缂栧彿鍞竴鎬ф牎楠?
    existing_device = db.query(Device).filter(Device.asset_id == asset_id).first()
    if existing_device:
        # 濡傛灉瀛樺湪锛屽垯閲嶅畾鍚戝洖涓婚〉骞舵樉绀洪敊璇俊鎭?
        error_message = f"鍒涘缓澶辫触锛氳祫浜х紪鍙?'{asset_id}' 宸插瓨鍦ㄣ€?
        return RedirectResponse(url=f"/?error={quote(error_message)}", status_code=303)

    new_device = Device(
        asset_id=asset_id,
        name=name,
        station=station,
        location=location,
        power_rating=power_rating,
        vendor=vendor,
        commission_date=commission_date,
        remark=remark
    )
    db.add(new_device)
    db.commit()
    return RedirectResponse(url="/", status_code=303)

def _adjust_color_brightness(hex_color: str, factor: float) -> str:
    """璋冩暣棰滆壊浜害"""
    try:
        # 绉婚櫎 # 绗﹀彿
        hex_color = hex_color.lstrip('#')
        
        # 杞崲涓?RGB
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        
        # 璋冩暣浜害
        r = min(255, max(0, int(r * factor)))
        g = min(255, max(0, int(g * factor)))
        b = min(255, max(0, int(b * factor)))
        
        # 杞崲鍥炲崄鍏繘鍒?
        return f"#{r:02x}{g:02x}{b:02x}"
    except:
        return hex_color  # 濡傛灉杞崲澶辫触锛岃繑鍥炲師鑹?


def _get_device_lifecycle_status(device: Device, db: Session) -> str:
    """璁＄畻璁惧鐨勭敓鍛藉懆鏈熺姸鎬?- 澶嶇敤宸叉湁鐨勫畬鏁村疄鐜伴€昏緫"""
    try:
        from datetime import datetime
        import re
        
        # 鏌ユ壘瀵瑰簲鐨勭敓鍛藉懆鏈熻鍒?
        rule = db.query(LifecycleRule).filter(
            LifecycleRule.device_type == device.device_type,
            LifecycleRule.is_active == "true"
        ).first()
        
        if not rule:
            return "鏈厤缃鍒?
        
        # 瑙ｆ瀽鎶曚骇鏃ユ湡
        if not device.commission_date:
            return "鎶曚骇鏃ユ湡鏈煡"
            
        try:
            # 璁＄畻璁惧骞撮緞锛堝勾锛?
            age_years = (datetime.now().date() - device.commission_date).days / 365.25
            
            # 鏍规嵁瑙勫垯鍒ゆ柇鐘舵€?
            if age_years < rule.suggested_scrap_age * 0.7:
                return "鍦ㄧ敤"
            elif age_years < rule.suggested_scrap_age:
                return "鍗冲皢鎶ュ簾"
            else:
                return "宸茶秴鏈?
        except Exception as e:
            print(f"璁＄畻璁惧骞撮緞澶辫触: {str(e)}")
            return "璁＄畻澶辫触"
    except Exception as e:
        print(f"璁＄畻鐢熷懡鍛ㄦ湡鐘舵€佸け璐? {str(e)}")
        return "鏈煡"





def _get_device_lifecycle_status(device: Device, db: Session) -> str:
    """璁＄畻璁惧鐨勭敓鍛藉懆鏈熺姸鎬?- 澶嶇敤宸叉湁鐨勫畬鏁村疄鐜伴€昏緫"""
    try:
        from datetime import datetime
        import re
        
        # 鏌ユ壘瀵瑰簲鐨勭敓鍛藉懆鏈熻鍒?
        rule = db.query(LifecycleRule).filter(
            LifecycleRule.device_type == device.device_type,
            LifecycleRule.is_active == "true"
        ).first()
        
        # 濡傛灉娌℃湁鎵惧埌瑙勫垯锛岃繑鍥炴湭鐭?
        if not rule:
            return "鏈煡"
            
        # 璁＄畻璁惧骞撮緞
        if not device.commission_date:
            return "鏈煡"
            
        age_years = (datetime.now().date() - device.commission_date).days / 365.25
        
        # 鏍规嵁瑙勫垯鍒ゆ柇鐘舵€?
        if age_years < rule.suggested_scrap_age * 0.7:
            return "鍦ㄧ敤"
        elif age_years < rule.suggested_scrap_age:
            return "鍗冲皢鎶ュ簾"
        else:
            return "宸茶秴鏈?
            
    except Exception as e:
        print(f"璁＄畻鐢熷懡鍛ㄦ湡鐘舵€佸け璐? {str(e)}")
        return "鏈煡"


# 鏂板API璺緞锛?api/power-chain/{device_id} - 涓?graph_data/{device_id}鍔熻兘鐩稿悓锛屼繚鎸佸悜鍚庡吋瀹?


# 鏂板API璺緞锛?api/power-chain/{device_id} - 涓?graph_data/{device_id}鍔熻兘鐩稿悓锛屼繚鎸佸悜鍚庡吋瀹?
@app.get("/graph_data/{device_id}")
@app.get("/api/power-chain/{device_id}")
async def get_graph_data(
    device_id: int,
    level: str = Query("device", regex="^(device|port)$", description="鏄剧ず绾у埆锛歞evice=璁惧绾э紝port=绔彛绾?),
    layout_type: str = Query("standard", regex="^(standard|bus)$", description="甯冨眬绫诲瀷锛歴tandard=鏍囧噯甯冨眬锛宐us=鎬荤嚎寮忓竷灞€"),
    connection_type: str = Query(None, description="杩炴帴绫诲瀷绛涢€夋潯浠?),
    station: str = Query(None, description="绔欑偣绛涢€夋潯浠?),
    device_type: str = Query(None, description="璁惧绫诲瀷绛涢€夋潯浠?),
    show_critical_only: bool = Query(False, description="鏄惁鍙樉绀哄叧閿澶?),
    only_selected_device: bool = Query(False, description="鏄惁鍙樉绀洪€変腑璁惧"),
    group_size: int = Query(1, description="鎬荤嚎寮忓竷灞€鏃舵瘡缁勭鍙ｆ暟閲?),
    db: Session = Depends(get_db)
):
    """鑾峰彇鎷撴墤鍥炬暟鎹紝鎸夌収鐢ㄦ埛闇€姹傛牸寮忓寲"""
    logger.info(f"/graph_data called: device_id={device_id}, level={level}, layout_type={layout_type}, connection_type={connection_type}, station={station}, device_type={device_type}, show_critical_only={show_critical_only}, only_selected_device={only_selected_device}, group_size={group_size}")
    
    # 鏌ユ壘閫変腑鐨勮澶?
    selected_device = db.query(Device).filter(Device.id == device_id).first()
    if not selected_device:
        raise HTTPException(status_code=404, detail="Device not found")
    
    nodes = []
    edges = []
    processed_device_ids = set()
    
    # 娣诲姞閫変腑璁惧鑺傜偣
    lifecycle_status = _get_device_lifecycle_status(selected_device, db)
    
    # 淇璁惧鍚嶇О鏄剧ず鏍煎紡锛氭寜鐓ц璁¤鑼冨彧鏄剧ず璁惧鍚嶇О
    device_label = selected_device.name
    
    # 璁板綍璁惧鍚嶇О鏍煎紡鍖栨棩蹇?
    topology_error_tracker.log_error(
        category=ErrorCategory.NODE_RENDERING,
        level=ErrorLevel.INFO,
        message=f"璁惧鑺傜偣鏍囩鏍煎紡鍖栧畬鎴?,
        context={
            "device_id": selected_device.id,
            "device_name": selected_device.name,
            "station": selected_device.station,
            "formatted_label": device_label,
            "original_format": selected_device.name
        }
    )
    
    node_data = {
        "id": selected_device.id,
        "label": device_label,
        "title": f"""璧勪骇缂栧彿: {selected_device.asset_id}\n鍚嶇О: {selected_device.name}\n璁惧绫诲瀷: {selected_device.device_type or 'N/A'}\n绔欑偣: {selected_device.station or 'N/A'}\n鍨嬪彿: {selected_device.model or 'N/A'}\n浣嶇疆: {selected_device.location or 'N/A'}\n棰濆畾瀹归噺: {selected_device.power_rating or 'N/A'}\n鐢熶骇鍘傚: {selected_device.vendor or 'N/A'}\n鎶曚骇鏃堕棿: {selected_device.commission_date or 'N/A'}\n鐢熷懡鍛ㄦ湡鐘舵€? {lifecycle_status}""",
        "level": 0,  # 閫変腑璁惧鍦ㄧ涓€灞?
        "device_type": selected_device.device_type,
        "station": selected_device.station
    }
    nodes.append(node_data)
    processed_device_ids.add(selected_device.id)

    # 鑾峰彇涓庨€変腑璁惧鐩存帴杩炴帴鐨勬墍鏈夎繛鎺?
    direct_connections = list(selected_device.target_connections) + list(selected_device.source_connections)
    
    # 澶勭悊姣忎釜鐩存帴杩炴帴
    for conn in direct_connections:
        # 搴旂敤杩炴帴绛涢€夋潯浠?
        if not _should_include_connection(conn, connection_type):
            continue
            
        # 璺宠繃鏃犳晥杩炴帴锛堣繛鎺ョ被鍨嬩负None鎴栫┖鐨勮繛鎺ワ級
        if not conn.connection_type or conn.connection_type == "None":
            continue
            
        # 纭畾杩炴帴鐨勫绔澶?
        if conn.source_device_id == selected_device.id:
            # 閫変腑璁惧鏄簮璁惧锛屽绔槸鐩爣璁惧
            connected_device = conn.target_device
            is_selected_source = True
        else:
            # 閫変腑璁惧鏄洰鏍囪澶囷紝瀵圭鏄簮璁惧
            connected_device = conn.source_device
            is_selected_source = False
            
        # 妫€鏌ュ绔澶囨槸鍚﹀瓨鍦ㄤ笖绗﹀悎绛涢€夋潯浠?
        if not connected_device or not _should_include_device(connected_device, station, device_type, show_critical_only):
            continue
            
        # 閬垮厤閲嶅娣诲姞璁惧鑺傜偣
        if connected_device.id not in processed_device_ids:
            if level == "port":
                # 绔彛绾ф樉绀猴細涓鸿繛鎺ヨ澶囧垱寤虹鍙ｈ妭鐐?
                if layout_type == "bus":
                    # 鎬荤嚎寮忓竷灞€
                    if not only_selected_device:
                        # 鍙岀瑙嗗浘锛氫负瀵圭璁惧涔熷垱寤烘€荤嚎涓庣鍙ｈ妭鐐?
                        bus_data = _create_bus_topology_nodes(connected_device, db, group_size=group_size)
                        for node in bus_data['bus_nodes']:
                            nodes.append(node)
                        for node in bus_data['port_nodes']:
                            nodes.append(node)
                        edges.extend(bus_data['bus_port_edges'])
                    else:
                        # 浠匒绔鍥撅細涓嶄负瀵圭璁惧鍒涘缓鑺傜偣锛岀敱閫変腑璁惧渚х殑鈥滃绔悎骞剁鍙ｂ€濇壙鎷呭睍绀?
                        pass
                else:
                    # 鏍囧噯甯冨眬锛氱鍙ｈ妭鐐瑰凡鍦ㄥ墠闈㈢粺涓€鍒涘缓锛屾澶勪笉鍐嶉噸澶嶅垱寤?
                    pass
            else:
                # 璁惧绾ф樉绀猴細涓鸿繛鎺ヨ澶囧垱寤鸿妭鐐?
                lifecycle_status = _get_device_lifecycle_status(connected_device, db)
                
                # 淇杩炴帴璁惧鍚嶇О鏄剧ず鏍煎紡锛氬彧鏄剧ず璁惧鍚嶇О
                connected_device_label = connected_device.name
                
                node_data = {
                    "id": connected_device.id,
                    "label": connected_device_label,
                    "title": f"""璧勪骇缂栧彿: {connected_device.asset_id}\n鍚嶇О: {connected_device.name}\n璁惧绫诲瀷: {connected_device.device_type or 'N/A'}\n绔欑偣: {connected_device.station or 'N/A'}\n鍨嬪彿: {connected_device.model or 'N/A'}\n浣嶇疆: {connected_device.location or 'N/A'}\n棰濆畾瀹归噺: {connected_device.power_rating or 'N/A'}\n鐢熶骇鍘傚: {connected_device.vendor or 'N/A'}\n鎶曚骇鏃堕棿: {connected_device.commission_date or 'N/A'}\n鐢熷懡鍛ㄦ湡鐘舵€? {lifecycle_status}""",
                    "level": 1,  # 杩炴帴璁惧鍦ㄧ浜屽眰
                    "device_type": connected_device.device_type,
                    "station": connected_device.station
                }
                nodes.append(node_data)
            processed_device_ids.add(connected_device.id)
            
        # 鍒涘缓杩炴帴杈癸紝鏍规嵁涓婁笅娓稿瓧娈电‘瀹氱澶存柟鍚?
        if level == "port":
            # 绔彛绾ц繛鎺?
            if layout_type == "bus":
                # 鎬荤嚎寮忓竷灞€锛氬垱寤虹鍙ｅ埌绔彛鐨勮繛鎺?
                direction = "upstream" if is_selected_source else "downstream"
                bus_port_edges = _create_bus_port_edges(conn, direction)
                edges.extend(bus_port_edges)
            else:
                # 鏍囧噯甯冨眬锛氬垱寤烘爣鍑嗙鍙ｈ繛鎺?
                direction = "upstream" if is_selected_source else "downstream"
                port_edges = _create_port_edges(conn, direction)
                edges.extend(port_edges)
        else:
            # 璁惧绾ц繛鎺ワ細鏍规嵁涓婁笅娓稿瓧娈电‘瀹氱澶存柟鍚?
            if conn.upstream_downstream == "涓婃父":
                # 涓婃父琛ㄧず鐢垫祦浠巗ource娴佸悜target
                edge_data = {
                    "from": conn.source_device_id,
                    "to": conn.target_device_id,
                    "arrows": "to",
                    "label": conn.connection_type or conn.cable_type or "",
                    "connection_type": conn.connection_type,
                    "cable_type": conn.cable_type,
                    "cable_model": conn.cable_model,
                    "remark": conn.remark,
                    "connection_id": conn.id
                }
            elif conn.upstream_downstream == "涓嬫父":
                # 涓嬫父琛ㄧず鐢垫祦浠巘arget娴佸悜source
                edge_data = {
                    "from": conn.target_device_id,
                    "to": conn.source_device_id,
                    "arrows": "to",
                    "label": conn.connection_type or conn.cable_type or "",
                    "connection_type": conn.connection_type,
                    "cable_type": conn.cable_type,
                    "cable_model": conn.cable_model,
                    "remark": conn.remark,
                    "connection_id": conn.id
                }
            else:
                # 娌℃湁鏄庣‘鐨勪笂涓嬫父鍏崇郴锛屼娇鐢ㄩ粯璁ゆ柟鍚戯紙浠巗ource鍒皌arget锛?
                edge_data = {
                    "from": conn.source_device_id,
                    "to": conn.target_device_id,
                    "arrows": "to",
                    "label": conn.connection_type or conn.cable_type or "",
                    "connection_type": conn.connection_type,
                    "cable_type": conn.cable_type,
                    "cable_model": conn.cable_model,
                    "remark": conn.remark,
                    "connection_id": conn.id
                }
            # 杩囨护鏃犳晥杩炴帴涓庣己澶辩鐐癸紝缁熶竴绠ご鏂瑰悜
            _label = (edge_data.get("label") or "").strip()
            _from = edge_data.get("from")
            _to = edge_data.get("to")
            if _label and _label.lower() not in ("nan", "none", "null", "绌?, "鏈煡") and _from and _to:
                edge_data["arrows"] = "to"
                edges.append(edge_data)
            else:
                logger.debug(f"璺宠繃鏃犳晥璁惧绾ц繛鎺? label={_label}, from={_from}, to={_to}, connection_id={edge_data.get('connection_id')}")
                
    # 鏋勫缓杩斿洖鏁版嵁
    response_data = {"nodes": nodes, "edges": edges, "level": level}
    
    # 濡傛灉鏄€荤嚎寮忓竷灞€锛屾坊鍔犻澶栫殑鍏冩暟鎹?
    if level == "port" and layout_type == "bus":
        response_data["layout_type"] = "bus"
        response_data["metadata"] = {
            "bus_count": len([n for n in nodes if n.get('type') == 'bus']),
            "port_count": len([n for n in nodes if n.get('type') == 'port']),
            "connection_count": len([e for e in edges if e.get('type') == 'port_connection'])
        }
    else:
        response_data["layout_type"] = "standard"
    
    logger.info(f"/graph_data response: nodes={len(nodes)}, edges={len(edges)}, layout_type={response_data['layout_type']}")
    return JSONResponse(content=response_data)


def _get_device_lifecycle_status(device: Device, db: Session) -> str:
    """璁＄畻璁惧鐨勭敓鍛藉懆鏈熺姸鎬?- 澶嶇敤宸叉湁鐨勫畬鏁村疄鐜伴€昏緫"""
    try:
        from datetime import datetime
        import re
        
        # 鏌ユ壘瀵瑰簲鐨勭敓鍛藉懆鏈熻鍒?
        rule = db.query(LifecycleRule).filter(
            LifecycleRule.device_type == device.device_type,
            LifecycleRule.is_active == "true"
        ).first()
        
        if not rule:
            return "鏈厤缃鍒?
        
        # 瑙ｆ瀽鎶曚骇鏃ユ湡
        if not device.commission_date:
            return "鎶曚骇鏃ユ湡鏈～鍐?
        
        commission_date = None
        date_str = device.commission_date.strip()
        current_date = datetime.now()
        
        # 澶勭悊鐗规畩鏍煎紡锛歒YYYMM (濡?202312)
        if re.match(r'^\d{6}$', date_str):
            try:
                year = int(date_str[:4])
                month = int(date_str[4:6])
                commission_date = datetime(year, month, 1)
            except ValueError:
                pass
        
        # 濡傛灉鐗规畩鏍煎紡瑙ｆ瀽澶辫触锛屽皾璇曟爣鍑嗘牸寮?
        if not commission_date:
            date_formats = [
                "%Y-%m-%d",
                "%Y/%m/%d", 
                "%Y.%m.%d",
                "%Y-%m",
                "%Y/%m",
                "%Y.%m",
                "%Y"
            ]
            
            for fmt in date_formats:
                try:
                    if fmt == "%Y":
                        # 鍙湁骞翠唤鐨勬儏鍐碉紝榛樿涓鸿骞寸殑1鏈?鏃?
                        commission_date = datetime.strptime(device.commission_date, fmt).replace(month=1, day=1)
                    elif fmt in ["%Y-%m", "%Y/%m", "%Y.%m"]:
                        # 鍙湁骞存湀鐨勬儏鍐碉紝榛樿涓鸿鏈堢殑1鏃?
                        commission_date = datetime.strptime(device.commission_date, fmt).replace(day=1)
                    else:
                        commission_date = datetime.strptime(device.commission_date, fmt)
                    break
                except ValueError:
                    continue
        
        if not commission_date:
            return "鎶曚骇鏃ユ湡鏍煎紡鏃犳硶璇嗗埆"
        
        # 璁＄畻鏈嶅焦鏃堕棿鍜屽墿浣欐椂闂?
        days_in_service = (current_date - commission_date).days
        lifecycle_days = rule.lifecycle_years * 365
        remaining_days = lifecycle_days - days_in_service
        warning_days = rule.warning_months * 30
        
        # 纭畾鐢熷懡鍛ㄦ湡鐘舵€?
        if remaining_days < 0:
            return f"宸茶秴鏈?{abs(remaining_days)} 澶?
        elif remaining_days <= warning_days:
            return f"涓磋繎瓒呴檺锛屽墿浣?{remaining_days} 澶?
        else:
            return f"姝ｅ父锛屽墿浣?{remaining_days} 澶?
            
    except Exception as e:
        return "璁＄畻閿欒"


def _should_include_device(device: Device, station: Optional[str], device_type: Optional[str], show_critical_only: bool) -> bool:
    """鍒ゆ柇璁惧鏄惁搴旇鍖呭惈鍦ㄦ嫇鎵戝浘涓?"""
    # 鍩虹鏁版嵁楠岃瘉锛氳繃婊ゆ帀鍚嶇О鏃犳晥鐨勮澶?
    if not device.name or device.name.strip() == "" or device.name.lower() in ["nan", "null", "none"]:
        return False
    
    # 绔欑偣绛涢€?
    if station and device.station != station:
        return False
    
    # 璁惧绫诲瀷绛涢€?
    if device_type and device.device_type != device_type:
        return False
    
    # 鍏抽敭璁惧绛涢€夛紙杩欓噷鍙互鏍规嵁涓氬姟闇€姹傚畾涔夊叧閿澶囩殑鍒ゆ柇閫昏緫锛?
    if show_critical_only:
        # 绀轰緥锛氬皢鍙戠數鏈虹粍銆乁PS銆佸彉鍘嬪櫒绛夎涓哄叧閿澶?
        critical_types = ["鍙戠數鏈虹粍", "UPS", "鍙樺帇鍣?, "楂樺帇閰嶇數鏌?, "浣庡帇閰嶇數鏌?]
        if device.device_type not in critical_types:
            return False
    
    return True


def _should_include_connection(connection: Connection, connection_type: Optional[str]) -> bool:
    """鍒ゆ柇杩炴帴鏄惁搴旇鍖呭惈鍦ㄦ嫇鎵戝浘涓?"""
    # 杩炴帴绫诲瀷绛涢€?
    if connection_type and connection.connection_type != connection_type:
        return False
    
    return True


def _create_port_nodes(device: Device, db: Session, level: int = 1) -> list:
    """涓鸿澶囧垱寤虹鍙ｇ骇鑺傜偣 - 鏍囧噯妯″紡
    鍙樉绀洪€変腑璁惧鐨勭鍙ｅ拰鐩存帴杩炴帴鐨勫绔澶囩鍙?
    """
    port_nodes = []
    
    # 鑾峰彇璁惧鐨勬墍鏈夎繛鎺ワ紝鎻愬彇绔彛淇℃伅
    connections = db.query(Connection).filter(
        or_(Connection.source_device_id == device.id, Connection.target_device_id == device.id)
    ).all()
    
    # 鏀堕泦閫変腑璁惧鐨勭鍙ｄ俊鎭?
    selected_device_ports = set()
    connected_device_ports = set()
    
    for conn in connections:
        # 閫変腑璁惧鐨勭鍙?
        if conn.source_device_id == device.id:
            if conn.source_fuse_number:
                selected_device_ports.add(("鐔斾笣", conn.source_fuse_number, conn))
            if conn.source_breaker_number:
                selected_device_ports.add(("绌哄紑", conn.source_breaker_number, conn))
        if conn.target_device_id == device.id:
            if conn.target_fuse_number:
                selected_device_ports.add(("鐔斾笣", conn.target_fuse_number, conn))
            if conn.target_breaker_number:
                selected_device_ports.add(("绌哄紑", conn.target_breaker_number, conn))
        
        # 瀵圭璁惧鐨勭鍙ｏ紙鍙樉绀虹洿鎺ヨ繛鎺ョ殑锛?
        if conn.source_device_id == device.id:
            # 褰撳墠璁惧鏄簮璁惧锛屾敹闆嗙洰鏍囪澶囩鍙?
            target_device = conn.target_device
            if target_device:
                if conn.target_fuse_number:
                    connected_device_ports.add((target_device, "鐔斾笣", conn.target_fuse_number, conn))
                if conn.target_breaker_number:
                    connected_device_ports.add((target_device, "绌哄紑", conn.target_breaker_number, conn))
        elif conn.target_device_id == device.id:
            # 褰撳墠璁惧鏄洰鏍囪澶囷紝鏀堕泦婧愯澶囩鍙?
            source_device = conn.source_device
            if source_device:
                if conn.source_fuse_number:
                    connected_device_ports.add((source_device, "鐔斾笣", conn.source_fuse_number, conn))
                if conn.source_breaker_number:
                    connected_device_ports.add((source_device, "绌哄紑", conn.source_breaker_number, conn))
    
    # 鍒涘缓閫変腑璁惧鐨勭鍙ｈ妭鐐?
    for port_type, port_number, conn in selected_device_ports:
        port_nodes.append({
            "id": f"{device.id}_{port_type}_{port_number}",
            "label": f"{port_type}-{port_number}",  # 绠€鍖栫鍙ｅ悕绉帮紝绉婚櫎璁惧鍚嶇О
            "title": f"""<b>璁惧:</b> {device.name}<br>
                         <b>绔彛:</b> {port_type}-{port_number}<br>
                         <b>璁惧绫诲瀷:</b> {device.device_type or 'N/A'}""",
            "level": level,  # 绔彛鑺傜偣鍦?level 灞?
            "device_id": device.id,
            "device_name": device.name,  # 鏂板锛氳澶囧悕绉帮紝渚夸簬鍓嶇涓夎鏄剧ず
            "station": device.station,   # 鏂板锛氱珯鐐逛俊鎭?
            "port_name": f"{port_type}-{port_number}",  # 鏂板锛氱鍙ｅ悕锛堢畝鍖栵級
            "port_type": port_type,
            "port_number": port_number,
            "node_type": "selected_device_port",  # 鏍囪涓洪€変腑璁惧绔彛
            "color": {"background": "#E6F3FF", "border": "#4169E1"}  # 绔彛浣跨敤钃濊壊
        })
    
    # 鍒涘缓瀵圭璁惧鐨勭鍙ｈ妭鐐?
    for connected_device, port_type, port_number, conn in connected_device_ports:
        port_nodes.append({
            "id": f"{connected_device.id}_{port_type}_{port_number}",
            "label": f"{connected_device.name}路{port_type}-{port_number}",  # 瀵圭鏍囩閲囩敤鈥滆澶囧悕路绔彛鍚嶁€?
            "title": f"""<b>璁惧:</b> {connected_device.name}<br>
                         <b>绔彛:</b> {port_type}-{port_number}<br>
                         <b>璁惧绫诲瀷:</b> {connected_device.device_type or 'N/A'}<br>
                         <b>杩炴帴鍒?</b> {device.name}""",
            "level": level + 1,  # 瀵圭绔彛鑺傜偣鍦?level + 1 灞?
            "device_id": connected_device.id,
            "device_name": connected_device.name,  # 鏂板
            "station": connected_device.station,    # 鏂板
            "port_name": f"{port_type}-{port_number}",  # 鏂板
            "port_type": port_type,
            "port_number": port_number,
            "node_type": "connected_device_port"  # 鏍囪涓哄绔澶囩鍙?
        })
    
    return port_nodes


def _create_port_edges(connection: Connection, direction: str) -> list:
    """涓鸿繛鎺ュ垱寤虹鍙ｇ骇杈?- 鏍囧噯妯″紡
    鏍规嵁upstream_downstream瀛楁纭畾绠ご鏂瑰悜
    """
    edges = []
    
    # 鍒涘缓绔彛闂寸殑杩炴帴
    source_ports = []
    target_ports = []
    
    # 浣跨敤鏂扮殑绔彛ID鏍煎紡锛堜腑鏂囷級
    if connection.source_fuse_number:
        source_ports.append(f"{connection.source_device_id}_鐔斾笣_{connection.source_fuse_number}")
    if connection.source_breaker_number:
        source_ports.append(f"{connection.source_device_id}_绌哄紑_{connection.source_breaker_number}")
    if connection.target_fuse_number:
        target_ports.append(f"{connection.target_device_id}_鐔斾笣_{connection.target_fuse_number}")
    if connection.target_breaker_number:
        target_ports.append(f"{connection.target_device_id}_绌哄紑_{connection.target_breaker_number}")
    
    # 鏍规嵁upstream_downstream瀛楁纭畾绠ご鏂瑰悜
    if connection.upstream_downstream == "涓婃父":
        # 涓婃父琛ㄧず鐢垫祦浠巗ource娴佸悜target
        arrow_direction = "to"
        from_ports = source_ports
        to_ports = target_ports
    elif connection.upstream_downstream == "涓嬫父":
        # 涓嬫父琛ㄧず鐢垫祦浠巘arget娴佸悜source
        arrow_direction = "to"
        from_ports = target_ports
        to_ports = source_ports
    else:
        # 娌℃湁鏄庣‘鏂瑰悜锛屼娇鐢ㄩ粯璁ゆ柟鍚戯紙source鍒皌arget锛?
        arrow_direction = "to"
        from_ports = source_ports
        to_ports = target_ports
    
    # 鍒涘缓绔彛闂寸殑杩炴帴杈癸紙杩囨护鏃犳晥鏍囩鍜岀鐐癸級
    _edge_label = (connection.cable_model or connection.connection_type or "").strip()
    if not _edge_label or _edge_label.lower() in ("nan", "none", "null", "绌?, "鏈煡"):
        logger.debug(f"璺宠繃鏃犳晥绔彛绾ц繛鎺? label={_edge_label}, conn_id={connection.id}")
        return edges
    for from_port in from_ports:
        for to_port in to_ports:
            if from_port and to_port:
                edges.append({
                    "from": from_port,
                    "to": to_port,
                    "arrows": "to",
                    "label": _edge_label,
                    "connection_type": connection.connection_type,
                    "cable_model": connection.cable_model,
                    "connection_id": connection.id,
                    "remark": connection.remark
                })
    
    return edges







@app.get("/api/port-selection/{device_id}")
async def get_port_selection_options(device_id: int, db: Session = Depends(get_db)):
    """鑾峰彇绔彛閫夋嫨閫夐」"""
    try:
        from port_topology_service import PortTopologyService
        service = PortTopologyService(db)
        data = service.get_port_selection_options(device_id)
        return {"success": True, "data": data}
    except Exception as e:
        topology_error_tracker.log_error(
            category=ErrorCategory.API_ERROR,
            level=ErrorLevel.ERROR,
            message=f"绔彛閫夋嫨閫夐」API璋冪敤澶辫触: {str(e)}",
            context={"device_id": device_id},
            exception=e
        )
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/graph", response_class=HTMLResponse)
async def get_topology_graph_page(request: Request, db: Session = Depends(get_db)):
    """鎷撴墤鍥鹃〉闈?- 鏄剧ず鎵€鏈夎澶囦緵鐢ㄦ埛閫夋嫨"""
    devices = db.query(Device).all()
    return templates.TemplateResponse("graph.html", {"request": request, "devices": devices})


@app.get("/topology")
async def get_fixture_topology_page(request: Request):
    """
    涓洪伩鍏嶉〉闈㈡贩娣嗭紝/topology 缁熶竴閲嶅畾鍚戝埌涓诲叆鍙?/graph銆?
    鍘熸嫇鎵戝鑸紙topology.html锛変粛淇濈暀鍦ㄤ唬鐮佷腑浣滀负瀹為獙椤碉紝浣嗕笉鍐嶄綔涓洪粯璁ゅ叆鍙ｃ€?
    """
    return templates.TemplateResponse("topology.html", {"request": request})


@app.get("/graph/{device_id}", response_class=HTMLResponse)
async def get_power_chain_graph(request: Request, device_id: int, db: Session = Depends(get_db)):
    """鐗瑰畾璁惧鐨勬嫇鎵戝浘椤甸潰 - 鍏煎鏃х増鏈琔RL"""
    devices = db.query(Device).all()
    return templates.TemplateResponse("graph.html", {"request": request, "devices": devices, "selected_device_id": device_id})


# --- 璁惧鐢熷懡鍛ㄦ湡瑙勫垯绠＄悊 API ---

@app.get("/api/lifecycle-rules")
async def get_lifecycle_rules(db: Session = Depends(get_db)):
    """
    鑾峰彇鎵€鏈夌敓鍛藉懆鏈熻鍒?
    """
    try:
        rules = db.query(LifecycleRule).all()
        return JSONResponse(content={
            "success": True,
            "data": [{
                "id": rule.id,
                "device_type": rule.device_type,
                "lifecycle_years": rule.lifecycle_years,
                "warning_months": rule.warning_months,
                "description": rule.description,
                "is_active": rule.is_active,
                "created_at": rule.created_at,
                "updated_at": rule.updated_at
            } for rule in rules]
        })
    except Exception as e:
        print(f"鑾峰彇鐢熷懡鍛ㄦ湡瑙勫垯澶辫触: {e}")
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)


@app.post("/api/lifecycle-rules")
async def create_lifecycle_rule(
    device_type: str = Form(...),
    lifecycle_years: int = Form(...),
    warning_months: int = Form(6),
    description: str = Form(""),
    password: str = Form(...),  # 娣诲姞瀵嗙爜鍙傛暟
    db: Session = Depends(get_db)
):
    """
    鍒涘缓鐢熷懡鍛ㄦ湡瑙勫垯
    """
    try:
        # 楠岃瘉绠＄悊鍛樺瘑鐮?
        if not verify_admin_password(password):
            return JSONResponse(content={"success": False, "message": "瀵嗙爜閿欒"}, status_code=401)
        
        from datetime import datetime
        
        # 妫€鏌ヨ澶囩被鍨嬫槸鍚﹀凡瀛樺湪瑙勫垯
        existing_rule = db.query(LifecycleRule).filter(LifecycleRule.device_type == device_type).first()
        if existing_rule:
            return JSONResponse(content={
                "success": False, 
                "message": f"璁惧绫诲瀷 '{device_type}' 鐨勭敓鍛藉懆鏈熻鍒欏凡瀛樺湪"
            }, status_code=400)
        
        # 鍒涘缓鏂拌鍒?
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_rule = LifecycleRule(
            device_type=device_type,
            lifecycle_years=lifecycle_years,
            warning_months=warning_months,
            description=description,
            is_active="true",
            created_at=current_time,
            updated_at=current_time
        )
        
        db.add(new_rule)
        db.commit()
        db.refresh(new_rule)
        
        return JSONResponse(content={
            "success": True,
            "message": "鐢熷懡鍛ㄦ湡瑙勫垯鍒涘缓鎴愬姛",
            "data": {
                "id": new_rule.id,
                "device_type": new_rule.device_type,
                "lifecycle_years": new_rule.lifecycle_years,
                "warning_months": new_rule.warning_months
            }
        })
        
    except Exception as e:
        db.rollback()




        print(f"鍒涘缓鐢熷懡鍛ㄦ湡瑙勫垯澶辫触: {e}")
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)


@app.put("/api/lifecycle-rules/{rule_id}")
async def update_lifecycle_rule(
    rule_id: int,
    device_type: str = Form(...),
    lifecycle_years: int = Form(...),
    warning_months: int = Form(6),
    description: str = Form(""),
    is_active: str = Form("true"),
    password: str = Form(...),  # 娣诲姞瀵嗙爜鍙傛暟
    db: Session = Depends(get_db)
):
    """
    鏇存柊鐢熷懡鍛ㄦ湡瑙勫垯
    """
    try:
        # 楠岃瘉绠＄悊鍛樺瘑鐮?
        if not verify_admin_password(password):
            return JSONResponse(content={"success": False, "message": "瀵嗙爜閿欒"}, status_code=401)
        
        from datetime import datetime
        
        rule = db.query(LifecycleRule).filter(LifecycleRule.id == rule_id).first()
        if not rule:
            return JSONResponse(content={"success": False, "message": "瑙勫垯涓嶅瓨鍦?}, status_code=404)
        
        # 妫€鏌ヨ澶囩被鍨嬫槸鍚︿笌鍏朵粬瑙勫垯鍐茬獊
        existing_rule = db.query(LifecycleRule).filter(
            LifecycleRule.device_type == device_type,
            LifecycleRule.id != rule_id
        ).first()
        if existing_rule:
            return JSONResponse(content={
                "success": False, 
                "message": f"璁惧绫诲瀷 '{device_type}' 鐨勭敓鍛藉懆鏈熻鍒欏凡瀛樺湪"
            }, status_code=400)
        
        # 鏇存柊瑙勫垯
        rule.device_type = device_type
        rule.lifecycle_years = lifecycle_years
        rule.warning_months = warning_months
        rule.description = description
        rule.is_active = is_active
        rule.updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        db.commit()
        
        return JSONResponse(content={
            "success": True,
            "message": "鐢熷懡鍛ㄦ湡瑙勫垯鏇存柊鎴愬姛"
        })
        
    except Exception as e:
        db.rollback()
        print(f"鏇存柊鐢熷懡鍛ㄦ湡瑙勫垯澶辫触: {e}")
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)


@app.delete("/api/lifecycle-rules/{rule_id}")
async def delete_lifecycle_rule(rule_id: int, password: str = Form(...), db: Session = Depends(get_db)):
    """
    鍒犻櫎鐢熷懡鍛ㄦ湡瑙勫垯
    """
    try:
        # 楠岃瘉绠＄悊鍛樺瘑鐮?
        if not verify_admin_password(password):
            return JSONResponse(content={"success": False, "message": "瀵嗙爜閿欒"}, status_code=401)
        
        rule = db.query(LifecycleRule).filter(LifecycleRule.id == rule_id).first()
        if not rule:
            return JSONResponse(content={"success": False, "message": "瑙勫垯涓嶅瓨鍦?}, status_code=404)
        
        db.delete(rule)
        db.commit()
        
        return JSONResponse(content={
            "success": True,
            "message": "鐢熷懡鍛ㄦ湡瑙勫垯鍒犻櫎鎴愬姛"
        })
        
    except Exception as e:
        db.rollback()
        print(f"鍒犻櫎鐢熷懡鍛ㄦ湡瑙勫垯澶辫触: {e}")
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)


@app.get("/api/devices")
async def get_devices_api(
    page: int = Query(1, ge=1, description="椤电爜"),
    page_size: int = Query(50, ge=1, le=200, description="姣忛〉鏁伴噺"),
    db: Session = Depends(get_db)
):
    """
    鑾峰彇璁惧鍒楄〃API鎺ュ彛
    """
    try:
        # 鏋勫缓鏌ヨ
        query = db.query(Device)
        
        # 璁＄畻鎬绘暟
        total = query.count()
        
        # 搴旂敤鍒嗛〉
        offset = (page - 1) * page_size
        devices = query.offset(offset).limit(page_size).all()
        
        # 鏋勫缓鍝嶅簲鏁版嵁
        result = []
        for device in devices:
            result.append({
                "id": device.id,
                "asset_id": device.asset_id,
                "name": device.name,
                "station": device.station,
                "model": device.model,
                "device_type": device.device_type,
                "location": device.location,
                "power_rating": device.power_rating,
                "vendor": device.vendor,
                "commission_date": device.commission_date.isoformat() if device.commission_date and hasattr(device.commission_date, 'isoformat') else device.commission_date,
                "remark": device.remark
            })
        
        return JSONResponse(content={
            "success": True,
            "data": result,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total": total,
                "total_pages": (total + page_size - 1) // page_size
            }
        })
        
    except Exception as e:
        print(f"鑾峰彇璁惧鍒楄〃澶辫触: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"鑾峰彇璁惧鍒楄〃澶辫触: {str(e)}")


@app.get("/api/topology/filter-options")
async def get_filter_options(db: Session = Depends(get_db)):
    """
    鑾峰彇鎷撴墤鍥剧瓫閫夐€夐」
    杩斿洖璁惧绫诲瀷銆佽繛鎺ョ被鍨嬨€佸眬绔欑瓑绛涢€夐€夐」
    """
    try:
        # 鑾峰彇鎵€鏈夊眬绔?
        stations = db.query(Device.station).filter(Device.station.isnot(None)).filter(Device.station != '').distinct().all()
        station_list = [station[0] for station in stations if station[0]]
        station_list.sort()
        
        # 鑾峰彇鎵€鏈夎繛鎺ョ被鍨?
        connection_types = db.query(Connection.connection_type).filter(Connection.connection_type.isnot(None)).filter(Connection.connection_type != '').distinct().all()
        connection_type_list = [conn_type[0] for conn_type in connection_types if conn_type[0]]
        connection_type_list.sort()
        
        return JSONResponse(content={
            "success": True,
            "data": {
                "device_types": STANDARD_DEVICE_TYPES,  # 浣跨敤鏂扮殑鏍囧噯璁惧绫诲瀷鍒楄〃
                "connection_types": connection_type_list,
                "stations": station_list
            }
        })
        
    except Exception as e:
        print(f"鑾峰彇绛涢€夐€夐」澶辫触: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"鑾峰彇绛涢€夐€夐」澶辫触: {str(e)}")


@app.get("/api/devices/lifecycle-status")
async def get_devices_lifecycle_status(
    status_filter: Optional[str] = None,  # normal, warning, expired, all
    db: Session = Depends(get_db)
):
    """
    鑾峰彇璁惧鐢熷懡鍛ㄦ湡鐘舵€?
    status_filter: normal(姝ｅ父), warning(涓磋繎瓒呴檺), expired(宸茶秴鏈?, all(鍏ㄩ儴)
    """
    try:
        from datetime import datetime, timedelta
        import re
        
        # 鑾峰彇鎵€鏈夎澶囧拰瑙勫垯
        devices = db.query(Device).all()
        rules = {rule.device_type: rule for rule in db.query(LifecycleRule).filter(LifecycleRule.is_active == "true").all()}
        
        result_devices = []
        current_date = datetime.now()
        
        for device in devices:
            # 鏌ユ壘瀵瑰簲鐨勭敓鍛藉懆鏈熻鍒?
            rule = rules.get(device.device_type)
            if not rule:
                # 娌℃湁瑙勫垯鐨勮澶囨爣璁颁负鏈煡鐘舵€?
                device_info = {
                    "id": device.id,
                    "asset_id": device.asset_id,
                    "name": device.name,
                    "station": device.station,
                    "model": device.model,
                    "vendor": device.vendor,
                    "commission_date": device.commission_date,
                    "lifecycle_status": "unknown",
                    "lifecycle_status_text": "鏈厤缃鍒?,
                    "days_in_service": None,
                    "remaining_days": None,
                    "rule_years": None
                }
                if not status_filter or status_filter == "all":
                    result_devices.append(device_info)
                continue
            
            # 瑙ｆ瀽鎶曚骇鏃ユ湡
            if not device.commission_date:
                device_info = {
                    "id": device.id,
                    "asset_id": device.asset_id,
                    "name": device.name,
                    "station": device.station,
                    "model": device.model,
                    "vendor": device.vendor,
                    "commission_date": device.commission_date,
                    "lifecycle_status": "unknown",
                    "lifecycle_status_text": "鎶曚骇鏃ユ湡鏈～鍐?,
                    "days_in_service": None,
                    "remaining_days": None,
                    "rule_years": rule.lifecycle_years
                }
                if not status_filter or status_filter == "all":
                    result_devices.append(device_info)
                continue
            
            # 灏濊瘯瑙ｆ瀽澶氱鏃ユ湡鏍煎紡
            commission_date = None
            date_str = device.commission_date.strip()
            
            # 澶勭悊鐗规畩鏍煎紡锛歒YYYMM (濡?202312)
            if re.match(r'^\d{6}$', date_str):
                try:
                    year = int(date_str[:4])
                    month = int(date_str[4:6])
                    commission_date = datetime(year, month, 1)
                except ValueError:
                    pass
            
            # 濡傛灉鐗规畩鏍煎紡瑙ｆ瀽澶辫触锛屽皾璇曟爣鍑嗘牸寮?
            if not commission_date:
                date_formats = [
                    "%Y-%m-%d",
                    "%Y/%m/%d", 
                    "%Y.%m.%d",
                    "%Y-%m",
                    "%Y/%m",
                    "%Y.%m",
                    "%Y"
                ]
                
                for fmt in date_formats:
                    try:
                        if fmt == "%Y":
                            # 鍙湁骞翠唤鐨勬儏鍐碉紝榛樿涓鸿骞寸殑1鏈?鏃?
                            commission_date = datetime.strptime(device.commission_date, fmt).replace(month=1, day=1)
                        elif fmt in ["%Y-%m", "%Y/%m", "%Y.%m"]:
                            # 鍙湁骞存湀鐨勬儏鍐碉紝榛樿涓鸿鏈堢殑1鏃?
                            commission_date = datetime.strptime(device.commission_date, fmt).replace(day=1)
                        else:
                            commission_date = datetime.strptime(device.commission_date, fmt)
                        break
                    except ValueError:
                        continue
            
            if not commission_date:
                device_info = {
                    "id": device.id,
                    "asset_id": device.asset_id,
                    "name": device.name,
                    "station": device.station,
                    "model": device.model,
                    "vendor": device.vendor,
                    "commission_date": device.commission_date,
                    "lifecycle_status": "unknown",
                    "lifecycle_status_text": "鎶曚骇鏃ユ湡鏍煎紡鏃犳硶璇嗗埆",
                    "days_in_service": None,
                    "remaining_days": None,
                    "rule_years": rule.lifecycle_years
                }
                if not status_filter or status_filter == "all":
                    result_devices.append(device_info)
                continue
            
            # 璁＄畻鏈嶅焦鏃堕棿鍜屽墿浣欐椂闂?
            days_in_service = (current_date - commission_date).days
            lifecycle_days = rule.lifecycle_years * 365
            remaining_days = lifecycle_days - days_in_service
            warning_days = rule.warning_months * 30
            
            # 纭畾鐢熷懡鍛ㄦ湡鐘舵€?
            if remaining_days < 0:
                lifecycle_status = "expired"
                lifecycle_status_text = f"宸茶秴鏈?{abs(remaining_days)} 澶?
            elif remaining_days <= warning_days:
                lifecycle_status = "warning"
                lifecycle_status_text = f"涓磋繎瓒呴檺锛屽墿浣?{remaining_days} 澶?
            else:
                lifecycle_status = "normal"
                lifecycle_status_text = f"姝ｅ父锛屽墿浣?{remaining_days} 澶?
            
            device_info = {
                "id": device.id,
                "asset_id": device.asset_id,
                "name": device.name,
                "station": device.station,
                "model": device.model,
                "vendor": device.vendor,
                "commission_date": device.commission_date,
                "lifecycle_status": lifecycle_status,
                "lifecycle_status_text": lifecycle_status_text,
                "days_in_service": days_in_service,
                "remaining_days": remaining_days,
                "rule_years": rule.lifecycle_years
            }
            
            # 鏍规嵁绛涢€夋潯浠舵坊鍔犺澶?
            if not status_filter or status_filter == "all" or status_filter == lifecycle_status:
                result_devices.append(device_info)
        
        # 缁熻淇℃伅
        total_count = len(result_devices)
        normal_count = len([d for d in result_devices if d["lifecycle_status"] == "normal"])
        warning_count = len([d for d in result_devices if d["lifecycle_status"] == "warning"])
        expired_count = len([d for d in result_devices if d["lifecycle_status"] == "expired"])
        unknown_count = len([d for d in result_devices if d["lifecycle_status"] == "unknown"])
        
        return JSONResponse(content={
            "success": True,
            "data": result_devices,
            "statistics": {
                "total": total_count,
                "normal": normal_count,
                "warning": warning_count,
                "expired": expired_count,
                "unknown": unknown_count
            }
        })
        
    except Exception as e:
        print(f"鑾峰彇璁惧鐢熷懡鍛ㄦ湡鐘舵€佸け璐? {e}")
        traceback.print_exc()
        return JSONResponse(content={"success": False, "message": str(e)}, status_code=500)


@app.get("/test-route")
async def test_route():
    """
    娴嬭瘯璺敱
    """
    print("=== 娴嬭瘯璺敱琚皟鐢?===")
    return {"message": "娴嬭瘯璺敱姝ｅ父宸ヤ綔", "timestamp": "updated"}

@app.get("/debug-routes")
async def debug_routes():
    """
    璋冭瘯璺敱 - 鏄剧ず鎵€鏈夊凡娉ㄥ唽鐨勮矾鐢?
    """
    routes = []
    for route in app.routes:
        if hasattr(route, 'path') and hasattr(route, 'methods'):
            routes.append({
                "path": route.path,
                "methods": list(route.methods) if route.methods else [],
                "name": getattr(route, 'name', 'unknown')
            })
    return {"registered_routes": routes, "total_count": len(routes)}

@app.get("/debug-lifecycle")
async def debug_lifecycle():
    """
    璋冭瘯鐢熷懡鍛ㄦ湡璺敱
    """
    print("=== 璋冭瘯鐢熷懡鍛ㄦ湡璺敱琚皟鐢?===")
    return {"message": "璋冭瘯璺敱姝ｅ父宸ヤ綔", "status": "ok"}

@app.post("/api/verify-password")
async def verify_password(request: Request):
    """
    楠岃瘉绠＄悊鍛樺瘑鐮?
    """
    try:
        data = await request.json()
        password = data.get("password", "")
        
        if verify_admin_password(password):
            return {"success": True, "message": "瀵嗙爜楠岃瘉鎴愬姛"}
        else:
            return {"success": False, "message": "瀵嗙爜閿欒"}
    except Exception as e:
        print(f"Error verifying password: {e}")
        return {"success": False, "message": "楠岃瘉澶辫触"}

@app.get("/lifecycle-management", response_class=HTMLResponse)
async def lifecycle_management_page(request: Request):
    """
    鐢熷懡鍛ㄦ湡绠＄悊椤甸潰
    """
    print("=== 璁块棶鐢熷懡鍛ㄦ湡绠＄悊椤甸潰 ===")
    print(f"璇锋眰URL: {request.url}")
    print(f"璇锋眰鏂规硶: {request.method}")
    try:
        print("姝ｅ湪娓叉煋妯℃澘...")
        response = templates.TemplateResponse("lifecycle_management.html", {"request": request})
        print("妯℃澘娓叉煋鎴愬姛")
        return response
    except Exception as e:
        print(f"鐢熷懡鍛ㄦ湡绠＄悊椤甸潰閿欒: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/connections", response_class=HTMLResponse)
async def connections_page(request: Request):
    """
    杩炴帴绠＄悊椤甸潰
    """
    print("=== 璁块棶杩炴帴绠＄悊椤甸潰 ===")
    print(f"璇锋眰URL: {request.url}")
    print(f"璇锋眰鏂规硶: {request.method}")
    try:
        print("姝ｅ湪娓叉煋妯℃澘...")
        response = templates.TemplateResponse("connections.html", {"request": request})
        print("妯℃澘娓叉煋鎴愬姛")
        return response
    except Exception as e:
        print(f"杩炴帴绠＄悊椤甸潰閿欒: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/analytics", response_class=HTMLResponse)
async def analytics_page(request: Request):
    """
    缁熻鍒嗘瀽椤甸潰
    """
    print("=== 璁块棶缁熻鍒嗘瀽椤甸潰 ===")
    print(f"璇锋眰URL: {request.url}")
    print(f"璇锋眰鏂规硶: {request.method}")
    try:
        print("姝ｅ湪娓叉煋缁熻鍒嗘瀽妯℃澘...")
        response = templates.TemplateResponse("analytics.html", {"request": request})
        print("缁熻鍒嗘瀽妯℃澘娓叉煋鎴愬姛")
        return response
    except Exception as e:
        print(f"缁熻鍒嗘瀽椤甸潰閿欒: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/export")
async def export_devices(
    password: str = Form(...),
    export_range: str = Form("all"),
    station_filter: str = Form(""),
    name_filter: str = Form(""),
    device_type_filter: str = Form(""),
    vendor_filter: str = Form(""),
    lifecycle_filter: str = Form(""),
    db: Session = Depends(get_db)
):
    """
    瀵煎嚭璁惧鏁版嵁涓篍xcel鏂囦欢
    鏀寔鍏ㄩ噺瀵煎嚭鍜岀瓫閫夊鍑猴紝闇€瑕佺鐞嗗憳瀵嗙爜楠岃瘉
    """
    try:
        # 楠岃瘉绠＄悊鍛樺瘑鐮?
        if not verify_admin_password(password):
            raise HTTPException(status_code=401, detail="瀵嗙爜閿欒锛屾棤鏉冮檺瀵煎嚭鏁版嵁")
        
        # 鏍规嵁瀵煎嚭鑼冨洿鏌ヨ璁惧鏁版嵁
        query = db.query(Device)
        
        # 濡傛灉鏄瓫閫夊鍑猴紝搴旂敤绛涢€夋潯浠?
        if export_range == "filtered":
            if station_filter:
                query = query.filter(Device.station.ilike(f"%{station_filter}%"))
            if name_filter:
                query = query.filter(Device.name.ilike(f"%{name_filter}%"))
            if device_type_filter:
                query = query.filter(Device.device_type.ilike(f"%{device_type_filter}%"))
            if vendor_filter:
                query = query.filter(Device.vendor.ilike(f"%{vendor_filter}%"))
            if lifecycle_filter:
                # 杩欓噷闇€瑕佹牴鎹敓鍛藉懆鏈熺姸鎬佺瓫閫夛紝鏆傛椂璺宠繃澶嶆潅鐨勭敓鍛藉懆鏈熼€昏緫
                pass
        
        devices = query.all()
        
        if not devices:
            raise HTTPException(status_code=404, detail="娌℃湁鎵惧埌璁惧鏁版嵁")
        
        # 鍒涘缓Excel宸ヤ綔绨?
        wb = Workbook()
        ws = wb.active
        ws.title = "璁惧鍒楄〃"
        
        # 瀹氫箟琛ㄥご
        headers = [
            "ID", "璧勪骇缂栧彿", "璁惧鍚嶇О", "灞€绔?, "璁惧绫诲瀷", "璁惧鍨嬪彿", 
            "鎵€鍦ㄤ綅缃?, "棰濆畾瀹归噺", "璁惧鐢熶骇鍘傚", "鎶曚骇鏃ユ湡", "澶囨敞"
        ]
        
        # 璁剧疆琛ㄥご鏍峰紡
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # 鍐欏叆琛ㄥご
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 鍐欏叆璁惧鏁版嵁
        for row, device in enumerate(devices, 2):
            data = [
                device.id,
                device.asset_id,
                device.name,
                device.station,
                device.device_type,
                device.model,
                device.location,
                device.power_rating,
                device.vendor,
                device.commission_date.strftime("%Y-%m-%d") if device.commission_date else "",
                device.remark
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # 璁剧疆鏂戦┈绾规晥鏋?
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # 鑷姩璋冩暣鍒楀
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 闄愬埗鏈€澶у搴?
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 鍐荤粨棣栬
        ws.freeze_panes = "A2"
        
        # 娣诲姞绛涢€夊櫒
        ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}1"
        
        # 鐢熸垚鏂囦欢鍚嶏紙鍖呭惈鏃堕棿鎴筹級
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if export_range == "filtered":
            filename = f"璁惧鍒楄〃_绛涢€夊鍑篲{timestamp}.xlsx"
        else:
            filename = f"璁惧鍒楄〃_鍏ㄩ噺瀵煎嚭_{timestamp}.xlsx"
        
        # 灏咵xcel鏂囦欢淇濆瓨鍒板唴瀛?
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # 璁剧疆鍝嶅簲澶?
        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        
        # 杩斿洖鏂囦欢娴?
        return StreamingResponse(
            io.BytesIO(excel_buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"瀵煎嚭璁惧鏁版嵁閿欒: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"瀵煎嚭澶辫触: {str(e)}")


# --- 杩炴帴绠＄悊 Pydantic 妯″瀷 ---

class ConnectionCreate(BaseModel):
    """鍒涘缓杩炴帴鐨勮姹傛ā鍨?"""
    source_device_id: int
    target_device_id: int
    connection_type: Optional[str] = None
    cable_model: Optional[str] = None
    source_fuse_number: Optional[str] = None
    source_fuse_spec: Optional[str] = None
    source_breaker_number: Optional[str] = None
    source_breaker_spec: Optional[str] = None
    target_fuse_number: Optional[str] = None
    target_fuse_spec: Optional[str] = None
    target_breaker_number: Optional[str] = None
    target_breaker_spec: Optional[str] = None
    hierarchy_relation: Optional[str] = None
    upstream_downstream: Optional[str] = None
    parallel_count: Optional[int] = 1
    rated_current: Optional[float] = None
    cable_length: Optional[float] = None
    source_device_photo: Optional[str] = None
    target_device_photo: Optional[str] = None
    remark: Optional[str] = None
    installation_date: Optional[date] = None

class ConnectionUpdate(BaseModel):
    """鏇存柊杩炴帴鐨勮姹傛ā鍨?"""
    source_device_id: Optional[int] = None
    target_device_id: Optional[int] = None
    connection_type: Optional[str] = None
    cable_model: Optional[str] = None
    source_fuse_number: Optional[str] = None
    source_fuse_spec: Optional[str] = None
    source_breaker_number: Optional[str] = None
    source_breaker_spec: Optional[str] = None
    target_fuse_number: Optional[str] = None
    target_fuse_spec: Optional[str] = None
    target_breaker_number: Optional[str] = None
    target_breaker_spec: Optional[str] = None
    hierarchy_relation: Optional[str] = None
    upstream_downstream: Optional[str] = None
    parallel_count: Optional[int] = None
    rated_current: Optional[float] = None
    cable_length: Optional[float] = None
    source_device_photo: Optional[str] = None
    target_device_photo: Optional[str] = None
    remark: Optional[str] = None
    installation_date: Optional[date] = None

class ConnectionResponse(BaseModel):
    """杩炴帴鍝嶅簲妯″瀷"""
    id: int
    source_device_id: int
    target_device_id: int
    source_device_name: str
    target_device_name: str
    source_port: Optional[str]  # 婧愮鍙ｅ悕绉帮紙甯﹀墠缂€锛?
    target_port: Optional[str]  # 鐩爣绔彛鍚嶇О锛堝甫鍓嶇紑锛?
    connection_type: Optional[str]
    cable_model: Optional[str]
    source_fuse_number: Optional[str]
    source_fuse_spec: Optional[str]
    source_breaker_number: Optional[str]
    source_breaker_spec: Optional[str]
    target_fuse_number: Optional[str]
    target_fuse_spec: Optional[str]
    target_breaker_number: Optional[str]
    target_breaker_spec: Optional[str]
    hierarchy_relation: Optional[str]
    upstream_downstream: Optional[str]
    parallel_count: Optional[int]
    rated_current: Optional[float]
    cable_length: Optional[float]
    source_device_photo: Optional[str]
    target_device_photo: Optional[str]
    remark: Optional[str]
    installation_date: Optional[date]
    created_at: datetime
    updated_at: Optional[datetime]
    
    class Config:
        # 鍚敤ORM妯″紡锛屽厑璁镐粠SQLAlchemy妯″瀷鍒涘缓
        from_attributes = True
        # 鑷畾涔塉SON缂栫爜鍣ㄥ鐞嗘棩鏈熸椂闂村璞?
        json_encoders = {
            datetime: lambda v: v.isoformat() if v else None,
            date: lambda v: v.isoformat() if v else None
        }


# --- 杩炴帴绠＄悊 RESTful API 鎺ュ彛 ---

def get_unique_connections_count(db: Session) -> int:
    """
    鑾峰彇鍘婚噸鍚庣殑杩炴帴鏁伴噺
    閫氳繃璇嗗埆鍙屽悜杩炴帴骞跺幓閲嶆潵鑾峰緱鐪熷疄鐨勮繛鎺ユ暟閲?
    """
    # 鑾峰彇鎵€鏈夋湁鏁堣繛鎺?
    connections = db.query(Connection).filter(Connection.connection_type.isnot(None)).all()
    
    # 浣跨敤闆嗗悎瀛樺偍鍞竴杩炴帴
    unique_connections = set()
    
    for conn in connections:
        # 鍒涘缓杩炴帴鐨勫敮涓€鏍囪瘑
        # 瀵逛簬鍙屽悜杩炴帴锛屼娇鐢ㄨ緝灏忕殑璁惧ID浣滀负绗竴涓弬鏁帮紝纭繚A->B鍜孊->A鐢熸垚鐩稿悓鐨勬爣璇?
        device_pair = tuple(sorted([conn.source_device_id, conn.target_device_id]))
        
        # 缁撳悎绔彛淇℃伅鍒涘缓鏇寸簿纭殑杩炴帴鏍囪瘑
        source_port = conn.source_fuse_number or conn.source_breaker_number or """
        target_port = conn.target_fuse_number or conn.target_breaker_number or """
        # 涓哄弻鍚戣繛鎺ュ垱寤虹粺涓€鐨勬爣璇?
        if conn.source_device_id == device_pair[0]:
            connection_key = (device_pair[0], device_pair[1], source_port, target_port, conn.connection_type)
        else:
            connection_key = (device_pair[0], device_pair[1], target_port, source_port, conn.connection_type)
        
        unique_connections.add(connection_key)
    
    return len(unique_connections)


def get_connected_ports_count(db: Session) -> int:
    """
    鐩存帴缁熻鎵€鏈夋湁杩炴帴鐨勭鍙ｆ暟閲?
    杩欑鏂规硶鑳藉鍑嗙‘澶勭悊鍐呴儴璁惧浜掕繛鍜屽閮ㄨ澶囪繛鎺?
    """
    connections = db.query(Connection).filter(Connection.connection_type.isnot(None)).all()
    connected_ports = set()
    
    for conn in connections:
        # 娣诲姞婧愮鍙?
        if conn.source_fuse_number:
            port_id = f"{conn.source_device_id}_fuse_{conn.source_fuse_number}"
            connected_ports.add(port_id)
        if conn.source_breaker_number:
            port_id = f"{conn.source_device_id}_breaker_{conn.source_breaker_number}"
            connected_ports.add(port_id)
            
        # 娣诲姞鐩爣绔彛
        if conn.target_fuse_number:
            port_id = f"{conn.target_device_id}_fuse_{conn.target_fuse_number}"
            connected_ports.add(port_id)
        if conn.target_breaker_number:
            port_id = f"{conn.target_device_id}_breaker_{conn.target_breaker_number}"
            connected_ports.add(port_id)
    
    return len(connected_ports)


@app.get("/api/connections/statistics")
async def get_connections_statistics(db: Session = Depends(get_db)):
    """
    鑾峰彇杩炴帴缁熻淇℃伅
    """
    try:
        # 浣跨敤鍘婚噸绠楁硶鑾峰彇鐪熷疄鐨勮繛鎺ユ暟閲?
        total_connections = get_unique_connections_count(db)
        
        # 浣跨敤PortStatisticsService缁熶竴鐨勭粺璁￠€昏緫锛岀‘淇濇暟鎹竴鑷存€?
        port_service = PortStatisticsService(db)
        port_summary = port_service._get_device_port_summary()
        
        # 浠庣粺涓€鐨勭鍙ｇ粺璁℃湇鍔¤幏鍙栨暟鎹?
        total_ports = port_summary.get('total_ports', 0)
        connected_ports_count = port_summary.get('connected_ports', 0)
        idle_ports = port_summary.get('idle_ports', 0)
        
        # 鑾峰彇璁惧鎬绘暟
        total_devices = db.query(Device).count()
        
        # 鎸夎繛鎺ョ被鍨嬬粺璁?
        connection_type_stats = db.query(
            Connection.connection_type,
            func.count(Connection.id).label('count')
        ).group_by(Connection.connection_type).all()
        
        # 灏嗘贩鍚堢殑涓嫳鏂囪繛鎺ョ被鍨嬬粺璁″悎骞朵负鏍囧噯鏍煎紡
        cable_count = 0
        busbar_count = 0
        bus_count = 0
        
        for item in connection_type_stats:
            conn_type = item[0] or """
            count = item[1]
            
            # 鐢电紗绫诲瀷锛坈able 鎴?鐢电紗锛?
            if conn_type.lower() in ['cable', '鐢电紗']:
                cable_count += count
            # 閾滄帓绫诲瀷锛坆usbar 鎴?閾滄帓锛?
            elif conn_type.lower() in ['busbar', '閾滄帓']:
                busbar_count += count
            # 姣嶇嚎绫诲瀷锛坆us銆乥usway 鎴?姣嶇嚎锛?
            elif conn_type.lower() in ['bus', 'busway', '姣嶇嚎']:
                bus_count += count
        
        # 鎸夎澶囩被鍨嬬粺璁★紙婧愯澶囷級
        device_type_stats = db.query(
            Device.device_type,
            func.count(Connection.id).label('count')
        ).join(Connection, Device.id == Connection.source_device_id)\
         .group_by(Device.device_type).all()
        
        # 鏈€杩?0澶╂柊澧炶繛鎺ユ暟
        thirty_days_ago = datetime.now() - timedelta(days=30)
        recent_connections = db.query(Connection)\
            .filter(Connection.created_at >= thirty_days_ago).count()
        
        return JSONResponse(content={
            "success": True,
            "data": {
                "total_devices": total_devices,
                "total_ports": total_ports,
                "connected_ports": connected_ports_count,
                "idle_ports": idle_ports,
                "total_connections": total_connections,
                "cable": cable_count,
                "busbar": busbar_count,
                "bus": bus_count,
                "recent_connections": recent_connections,
                "connection_types": [
                    {"type": item[0] or "鏈垎绫?, "count": item[1]} 
                    for item in connection_type_stats
                ],
                "device_types": [
                    {"type": item[0] or "鏈垎绫?, "count": item[1]} 
                    for item in device_type_stats
                ]
            }
        })
        
    except Exception as e:
        print(f"鑾峰彇杩炴帴缁熻澶辫触: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"鑾峰彇杩炴帴缁熻澶辫触: {str(e)}")


@app.get("/api/ports/statistics")
async def get_port_statistics(db: Session = Depends(get_db)):
    """
    鑾峰彇绔彛缁熻淇℃伅
    """
    try:
        # 鍒涘缓绔彛缁熻鏈嶅姟瀹炰緥
        port_service = PortStatisticsService(db)
        
        # 鑾峰彇绔彛缁熻鏁版嵁
        statistics = port_service.get_port_statistics()
        
        return JSONResponse(content={
            "success": True,
            "data": statistics
        })
        
    except Exception as e:
        print(f"鑾峰彇绔彛缁熻澶辫触: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"鑾峰彇绔彛缁熻澶辫触: {str(e)}")


@app.get("/api/devices/{device_id}/ports")
async def get_device_port_details(device_id: int, db: Session = Depends(get_db)):
    """
    鑾峰彇鎸囧畾璁惧鐨勭鍙ｈ鎯?
    """
    try:
        # 鍒涘缓绔彛缁熻鏈嶅姟瀹炰緥
        port_service = PortStatisticsService(db)
        
        # 鑾峰彇璁惧绔彛璇︽儏
        port_details = port_service.get_device_port_details(device_id)
        
        return JSONResponse(content={
            "success": True,
            "data": port_details
        })
        
    except Exception as e:
        print(f"鑾峰彇璁惧绔彛璇︽儏澶辫触: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"鑾峰彇璁惧绔彛璇︽儏澶辫触: {str(e)}")


# ==================== 缁熻鍒嗘瀽API绔偣 ====================

@app.get("/api/analytics/utilization-rates")
async def get_utilization_rates(db: Session = Depends(get_db)):
    """
    鑾峰彇浣跨敤鐜囧垎鏋愭暟鎹?
    鍖呮嫭绔彛鎬讳綋浣跨敤鐜囥€佹寜璁惧绫诲瀷缁熻銆佹寜绔欑偣缁熻绛?
    """
    try:
        # 鍒涘缓缁熻鍒嗘瀽鏈嶅姟瀹炰緥
        analytics_service = AnalyticsService(db)
        
        # 鑾峰彇浣跨敤鐜囧垎鏋愭暟鎹?
        utilization_data = analytics_service.get_utilization_rates()
        
        return JSONResponse(content={
            "success": True,
            "data": utilization_data
        })
        
    except Exception as e:
        print(f"鑾峰彇浣跨敤鐜囧垎鏋愭暟鎹け璐? {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"鑾峰彇浣跨敤鐜囧垎鏋愭暟鎹け璐? {str(e)}")


@app.get("/api/analytics/idle-rates")
async def get_idle_rates(db: Session = Depends(get_db)):
    """
    鑾峰彇绌洪棽鐜囧垎鏋愭暟鎹?
    鍖呮嫭绔彛鎬讳綋绌洪棽鐜囥€佹寜璁惧绫诲瀷缁熻銆佹寜绔欑偣缁熻銆佺┖闂茬巼棰勮绛?
    """
    try:
        # 鍒涘缓缁熻鍒嗘瀽鏈嶅姟瀹炰緥
        analytics_service = AnalyticsService(db)
        
        # 鑾峰彇绌洪棽鐜囧垎鏋愭暟鎹?
        idle_data = analytics_service.get_idle_rates()
        
        return JSONResponse(content={
            "success": True,
            "data": idle_data
        })
        
    except Exception as e:
        print(f"鑾峰彇绌洪棽鐜囧垎鏋愭暟鎹け璐? {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"鑾峰彇绌洪棽鐜囧垎鏋愭暟鎹け璐? {str(e)}")





@app.get("/api/analytics/summary-dashboard")
async def get_summary_dashboard(db: Session = Depends(get_db)):
    """
    鑾峰彇浠〃鏉挎眹鎬绘暟鎹?
    鍖呮嫭鎵€鏈夊叧閿寚鏍囩殑姹囨€讳俊鎭紝鐢ㄤ簬缁熻鍒嗘瀽浠〃鏉挎樉绀?
    """
    try:
        # 鍒涘缓缁熻鍒嗘瀽鏈嶅姟瀹炰緥
        analytics_service = AnalyticsService(db)
        
        # 鑾峰彇浠〃鏉挎眹鎬绘暟鎹?
        dashboard_data = analytics_service.get_summary_dashboard()
        
        return JSONResponse(content={
            "success": True,
            "data": dashboard_data
        })
        
    except Exception as e:
        print(f"鑾峰彇浠〃鏉挎眹鎬绘暟鎹け璐? {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"鑾峰彇浠〃鏉挎眹鎬绘暟鎹け璐? {str(e)}")


# 杈呭姪鍑芥暟锛氭牴鎹啍涓?绌哄紑缂栧彿涓虹鍙ｅ悕绉版坊鍔犲墠缂€
def build_port_name_with_prefix(fuse_number, breaker_number, original_port=None):
    """鏍规嵁鐔斾笣缂栧彿鎴栫┖寮€缂栧彿涓虹鍙ｅ悕绉版坊鍔犲墠缂€"""
    fuse_num = str(fuse_number).strip() if fuse_number and str(fuse_number).strip() not in ['', 'nan', 'None'] else ''
    breaker_num = str(breaker_number).strip() if breaker_number and str(breaker_number).strip() not in ['', 'nan', 'None'] else ''
    
    # 浼樺厛浣跨敤鐔斾笣缂栧彿
    if fuse_num:
        return f"鐔斾笣_{fuse_num}"
    elif breaker_num:
        return f"绌哄紑_{breaker_num}"
    else:
        # 濡傛灉閮芥病鏈夛紝杩斿洖鍘熷绔彛鍚嶇О鎴栫┖瀛楃涓?
        return original_port if original_port else ''


@app.get("/api/connections")
async def get_connections(
    page: int = Query(1, ge=1, description="椤电爜"),
    page_size: int = Query(100, ge=1, le=5000, description="姣忛〉鏁伴噺"),
    source_device_id: Optional[int] = Query(None, description="婧愯澶嘔D"),
    target_device_id: Optional[int] = Query(None, description="鐩爣璁惧ID"),
    connection_type: Optional[str] = Query(None, description="杩炴帴绫诲瀷"),
    device_name: Optional[str] = Query(None, description="璁惧鍚嶇О锛堟ā绯婃煡璇紝鍖归厤婧愯澶囨垨鐩爣璁惧锛?),
    db: Session = Depends(get_db)
):
    """
    鑾峰彇杩炴帴鍒楄〃
    鏀寔鍒嗛〉鍜岀瓫閫夊姛鑳?
    """
    try:
        # 鏋勫缓鏌ヨ
        # 鍒涘缓Device琛ㄧ殑鍒悕鐢ㄤ簬鐩爣璁惧
        target_device = aliased(Device)
        query = db.query(Connection, Device.name.label('source_device_name'), target_device.name.label('target_device_name'))\
                  .join(Device, Connection.source_device_id == Device.id)\
                  .join(target_device, Connection.target_device_id == target_device.id)
        
        # 搴旂敤绛涢€夋潯浠?
        if source_device_id:
            query = query.filter(Connection.source_device_id == source_device_id)
        if target_device_id:
            query = query.filter(Connection.target_device_id == target_device_id)
        if connection_type:
            if connection_type == "绌洪棽":
                # 绛涢€夌┖闂茬鍙ｏ細杩炴帴绫诲瀷涓虹┖涓擜绔澶囨湁鐔斾笣鎴栫┖寮€鏁版嵁
                query = query.filter(
                    and_(
                        Connection.connection_type.is_(None),
                        or_(
                            Connection.source_fuse_number.isnot(None),
                            Connection.source_breaker_number.isnot(None)
                        )
                    )
                )
            elif connection_type == "宸蹭娇鐢ㄦ€婚噺":
                # 绛涢€夊凡浣跨敤鎬婚噺锛氭樉绀烘墍鏈夋湁杩炴帴绫诲瀷鐨勮褰曪紙闈炵┖闂诧級
                query = query.filter(Connection.connection_type.isnot(None))
            else:
                query = query.filter(Connection.connection_type.ilike(f"%{connection_type}%"))
        else:
            # 濡傛灉娌℃湁鎸囧畾杩炴帴绫诲瀷绛涢€夛紝榛樿鏄剧ず鎵€鏈夎褰曪紙鍖呮嫭绌洪棽绔彛锛?
            # 浣嗚纭繚A绔澶囨湁绔彛鏁版嵁锛堢啍涓濇垨绌哄紑锛?
            query = query.filter(
                or_(
                    Connection.source_fuse_number.isnot(None),
                    Connection.source_breaker_number.isnot(None)
                )
            )
        # 鎸夎澶囧悕绉版ā绯婃煡璇紙鍖归厤婧愯澶囨垨鐩爣璁惧锛?
        if device_name:
            query = query.filter(
                or_(
                    Device.name.ilike(f"%{device_name}%"),  # 鍖归厤婧愯澶囧悕绉?
                    target_device.name.ilike(f"%{device_name}%")  # 鍖归厤鐩爣璁惧鍚嶇О
                )
            )
        
        # 璁＄畻鎬绘暟
        total = query.count()
        
        # 搴旂敤鍒嗛〉
        offset = (page - 1) * page_size
        results = query.offset(offset).limit(page_size).all()
        
        # 鏋勫缓鍝嶅簲鏁版嵁 - 鎵嬪姩搴忓垪鍖栨棩鏈熷瓧娈典互閬垮厤JSON搴忓垪鍖栭敊璇?
        result = []
        for conn, source_name, target_name in results:
            # 鎵嬪姩澶勭悊鏃ユ湡瀛楁鐨勫簭鍒楀寲
            installation_date_str = conn.installation_date.isoformat() if conn.installation_date else None
            created_at_str = conn.created_at.isoformat() if conn.created_at else None
            updated_at_str = conn.updated_at.isoformat() if conn.updated_at else None
            
            # 鏋勫缓甯﹀墠缂€鐨勭鍙ｅ悕绉?
            source_port_with_prefix = build_port_name_with_prefix(
                conn.source_fuse_number, 
                conn.source_breaker_number, 
                conn.source_port
            )
            target_port_with_prefix = build_port_name_with_prefix(
                conn.target_fuse_number, 
                conn.target_breaker_number, 
                conn.target_port
            )
            
            result.append({
                "id": conn.id,
                "source_device_id": conn.source_device_id,
                "target_device_id": conn.target_device_id,
                "source_device_name": source_name,
                "target_device_name": target_name,
                "connection_type": conn.connection_type,
                "cable_model": conn.cable_model,
                "source_port": source_port_with_prefix,  # 浣跨敤甯﹀墠缂€鐨勭鍙ｅ悕绉?
                "target_port": target_port_with_prefix,  # 浣跨敤甯﹀墠缂€鐨勭鍙ｅ悕绉?
                "source_fuse_number": conn.source_fuse_number,
                "source_fuse_spec": conn.source_fuse_spec,
                "source_breaker_number": conn.source_breaker_number,
                "source_breaker_spec": conn.source_breaker_spec,
                "target_fuse_number": conn.target_fuse_number,
                "target_fuse_spec": conn.target_fuse_spec,
                "target_breaker_number": conn.target_breaker_number,
                "target_breaker_spec": conn.target_breaker_spec,
                "hierarchy_relation": conn.hierarchy_relation,
                "upstream_downstream": conn.upstream_downstream,
                "parallel_count": conn.parallel_count,
                "rated_current": conn.rated_current,
                "a_rated_current": conn.a_rated_current,  # 娣诲姞A绔瀹氱數娴佸瓧娈?
                "b_rated_current": conn.b_rated_current,  # 娣诲姞B绔瀹氱數娴佸瓧娈?
                "cable_length": conn.cable_length,
                "source_device_photo": conn.source_device_photo,
                "target_device_photo": conn.target_device_photo,
                "remark": conn.remark,
                "installation_date": installation_date_str,
                "created_at": created_at_str,
                "updated_at": updated_at_str
            })

        return {
            "success": True,
            "data": result,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total": total,
                "total_pages": (total + page_size - 1) // page_size
            }
        }
        
    except Exception as e:
        print(f"鑾峰彇杩炴帴鍒楄〃澶辫触: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"鑾峰彇杩炴帴鍒楄〃澶辫触: {str(e)}")


@app.post("/api/connections")
async def create_connection(
    source_device_id: int = Form(...),
    target_device_id: int = Form(...),
    connection_type: Optional[str] = Form(None),
    cable_model: Optional[str] = Form(None),
    source_port: Optional[str] = Form(None),
    target_port: Optional[str] = Form(None),
    source_fuse_number: Optional[str] = Form(None),
    source_fuse_spec: Optional[str] = Form(None),
    source_breaker_number: Optional[str] = Form(None),
    source_breaker_spec: Optional[str] = Form(None),
    target_fuse_number: Optional[str] = Form(None),
    target_fuse_spec: Optional[str] = Form(None),
    target_breaker_number: Optional[str] = Form(None),
    target_breaker_spec: Optional[str] = Form(None),
    hierarchy_relation: Optional[str] = Form(None),
    upstream_downstream: Optional[str] = Form(None),
    parallel_count: Optional[int] = Form(1),
    rated_current: Optional[float] = Form(None),
    cable_length: Optional[float] = Form(None),
    source_device_photo: Optional[str] = Form(None),
    target_device_photo: Optional[str] = Form(None),
    remark: Optional[str] = Form(None),
    installation_date: Optional[str] = Form(None),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """
    鍒涘缓鏂拌繛鎺?
    闇€瑕佺鐞嗗憳瀵嗙爜楠岃瘉
    """
    try:
        # 楠岃瘉绠＄悊鍛樺瘑鐮?
        if not verify_admin_password(password):
            raise HTTPException(status_code=401, detail="瀵嗙爜閿欒")
        
        # 澶勭悊鏃ユ湡瀛楁 - 鏀寔yyyymm鏍煎紡
        parsed_installation_date = None
        if installation_date:
            try:
                # 鏀寔yyyymm鏍煎紡锛屽202412
                if len(installation_date) == 6 and installation_date.isdigit():
                    year = int(installation_date[:4])
                    month = int(installation_date[4:6])
                    parsed_installation_date = datetime(year, month, 1).date()
                else:
                    raise ValueError("鏃ユ湡鏍煎紡涓嶆纭?)
            except ValueError:
                raise HTTPException(status_code=400, detail="瀹夎鏃ユ湡鏍煎紡閿欒锛岃浣跨敤YYYYMM鏍煎紡锛堝锛?02412锛?)
        
        # 楠岃瘉婧愯澶囧拰鐩爣璁惧鏄惁瀛樺湪
        source_device = db.query(Device).filter(Device.id == source_device_id).first()
        if not source_device:
            raise HTTPException(status_code=404, detail=f"婧愯澶嘔D {source_device_id} 涓嶅瓨鍦?)
        
        target_device = db.query(Device).filter(Device.id == target_device_id).first()
        if not target_device:
            raise HTTPException(status_code=404, detail=f"鐩爣璁惧ID {target_device_id} 涓嶅瓨鍦?)
        
        # 妫€鏌ユ槸鍚﹀凡瀛樺湪鐩稿悓鐨勮繛鎺?
        existing_connection = db.query(Connection).filter(
            Connection.source_device_id == source_device_id,
            Connection.target_device_id == target_device_id
        ).first()
        
        if existing_connection:
            raise HTTPException(status_code=400, detail="璇ヨ繛鎺ュ凡瀛樺湪")
        
        # 鍒涘缓鏂拌繛鎺?
        new_connection = Connection(
            source_device_id=source_device_id,
            target_device_id=target_device_id,
            source_port=source_port,
            target_port=target_port,
            connection_type=connection_type,
            cable_model=cable_model,
            source_fuse_number=source_fuse_number,
            source_fuse_spec=source_fuse_spec,
            source_breaker_number=source_breaker_number,
            source_breaker_spec=source_breaker_spec,
            target_fuse_number=target_fuse_number,
            target_fuse_spec=target_fuse_spec,
            target_breaker_number=target_breaker_number,
            target_breaker_spec=target_breaker_spec,
            hierarchy_relation=hierarchy_relation,
            upstream_downstream=upstream_downstream,
            parallel_count=parallel_count,
            rated_current=rated_current,
            cable_length=cable_length,
            source_device_photo=source_device_photo,
            target_device_photo=target_device_photo,
            remark=remark,
            installation_date=parsed_installation_date,
            created_at=datetime.now()
        )
        
        db.add(new_connection)
        db.commit()
        db.refresh(new_connection)
        
        # 鏋勫缓鍝嶅簲
        response = ConnectionResponse(
            id=new_connection.id,
            source_device_id=new_connection.source_device_id,
            target_device_id=new_connection.target_device_id,
            source_device_name=source_device.name,
            target_device_name=target_device.name,
            connection_type=new_connection.connection_type,
            cable_model=new_connection.cable_model,
            source_fuse_number=new_connection.source_fuse_number,
            source_fuse_spec=new_connection.source_fuse_spec,
            source_breaker_number=new_connection.source_breaker_number,
            source_breaker_spec=new_connection.source_breaker_spec,
            target_fuse_number=new_connection.target_fuse_number,
            target_fuse_spec=new_connection.target_fuse_spec,
            target_breaker_number=new_connection.target_breaker_number,
            target_breaker_spec=new_connection.target_breaker_spec,
            hierarchy_relation=new_connection.hierarchy_relation,
            upstream_downstream=new_connection.upstream_downstream,
            parallel_count=new_connection.parallel_count,
            rated_current=new_connection.rated_current,
            cable_length=new_connection.cable_length,
            source_device_photo=new_connection.source_device_photo,
            target_device_photo=new_connection.target_device_photo,
            remark=new_connection.remark,
            installation_date=new_connection.installation_date,
            created_at=new_connection.created_at,
            updated_at=new_connection.updated_at
        )
        
        # 鎵嬪姩澶勭悊鏃ユ湡瀛楁搴忓垪鍖?
        response_data = {
            "id": new_connection.id,
            "source_device_id": new_connection.source_device_id,
            "target_device_id": new_connection.target_device_id,
            "source_device_name": source_device.name,
            "target_device_name": target_device.name,
            "connection_type": new_connection.connection_type,
            "cable_model": new_connection.cable_model,
            "source_fuse_number": new_connection.source_fuse_number,
            "source_fuse_spec": new_connection.source_fuse_spec,
            "source_breaker_number": new_connection.source_breaker_number,
            "source_breaker_spec": new_connection.source_breaker_spec,
            "target_fuse_number": new_connection.target_fuse_number,
            "target_fuse_spec": new_connection.target_fuse_spec,
            "target_breaker_number": new_connection.target_breaker_number,
            "target_breaker_spec": new_connection.target_breaker_spec,
            "hierarchy_relation": new_connection.hierarchy_relation,
            "upstream_downstream": new_connection.upstream_downstream,
            "parallel_count": new_connection.parallel_count,
            "rated_current": new_connection.rated_current,
            "cable_length": new_connection.cable_length,
            "source_device_photo": new_connection.source_device_photo,
            "target_device_photo": new_connection.target_device_photo,
            "remark": new_connection.remark,
            "installation_date": new_connection.installation_date.isoformat() if new_connection.installation_date else None,
            "created_at": new_connection.created_at.isoformat() if new_connection.created_at else None,
            "updated_at": new_connection.updated_at.isoformat() if new_connection.updated_at else None
        }
        
        return JSONResponse(content={
            "success": True,
            "message": "杩炴帴鍒涘缓鎴愬姛",
            "data": response_data
        })
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"鍒涘缓杩炴帴澶辫触: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"鍒涘缓杩炴帴澶辫触: {str(e)}")


@app.put("/api/connections/{connection_id}")
async def update_connection(
    connection_id: int,
    source_device_id: Optional[int] = Form(None),
    target_device_id: Optional[int] = Form(None),
    source_port: Optional[str] = Form(None),
    target_port: Optional[str] = Form(None),
    connection_type: Optional[str] = Form(None),
    cable_model: Optional[str] = Form(None),
    source_fuse_number: Optional[str] = Form(None),
    source_fuse_spec: Optional[str] = Form(None),
    source_breaker_number: Optional[str] = Form(None),
    source_breaker_spec: Optional[str] = Form(None),
    target_fuse_number: Optional[str] = Form(None),
    target_fuse_spec: Optional[str] = Form(None),
    target_breaker_number: Optional[str] = Form(None),
    target_breaker_spec: Optional[str] = Form(None),
    hierarchy_relation: Optional[str] = Form(None),
    upstream_downstream: Optional[str] = Form(None),
    parallel_count: Optional[int] = Form(None),
    rated_current: Optional[float] = Form(None),
    cable_length: Optional[float] = Form(None),
    source_device_photo: Optional[str] = Form(None),
    target_device_photo: Optional[str] = Form(None),
    remark: Optional[str] = Form(None),
    installation_date: Optional[str] = Form(None),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """
    鏇存柊杩炴帴淇℃伅
    闇€瑕佺鐞嗗憳瀵嗙爜楠岃瘉
    """
    try:
        # 楠岃瘉绠＄悊鍛樺瘑鐮?
        if not verify_admin_password(password):
            raise HTTPException(status_code=401, detail="瀵嗙爜閿欒")
        
        # 鏌ユ壘瑕佹洿鏂扮殑杩炴帴
        existing_connection = db.query(Connection).filter(Connection.id == connection_id).first()
        if not existing_connection:
            raise HTTPException(status_code=404, detail="杩炴帴涓嶅瓨鍦?)
        
        # 濡傛灉瑕佹洿鏂拌澶嘔D锛岄獙璇佽澶囨槸鍚﹀瓨鍦?
        if source_device_id is not None:
            source_device = db.query(Device).filter(Device.id == source_device_id).first()
            if not source_device:
                raise HTTPException(status_code=404, detail=f"婧愯澶嘔D {source_device_id} 涓嶅瓨鍦?)
            existing_connection.source_device_id = source_device_id
        
        if target_device_id is not None:
            target_device = db.query(Device).filter(Device.id == target_device_id).first()
            if not target_device:
                raise HTTPException(status_code=404, detail=f"鐩爣璁惧ID {target_device_id} 涓嶅瓨鍦?)
            existing_connection.target_device_id = target_device_id
        
        # 鏇存柊绔彛瀛楁
        if source_port is not None:
            existing_connection.source_port = source_port
        if target_port is not None:
            existing_connection.target_port = target_port
        
        # 鏇存柊鍏朵粬瀛楁
        if connection_type is not None:
            existing_connection.connection_type = connection_type
        if cable_model is not None:
            existing_connection.cable_model = cable_model
        if source_fuse_number is not None:
            existing_connection.source_fuse_number = source_fuse_number
        if source_fuse_spec is not None:
            existing_connection.source_fuse_spec = source_fuse_spec
        if source_breaker_number is not None:
            existing_connection.source_breaker_number = source_breaker_number
        if source_breaker_spec is not None:
            existing_connection.source_breaker_spec = source_breaker_spec
        if target_fuse_number is not None:
            existing_connection.target_fuse_number = target_fuse_number
        if target_fuse_spec is not None:
            existing_connection.target_fuse_spec = target_fuse_spec
        if target_breaker_number is not None:
            existing_connection.target_breaker_number = target_breaker_number
        if target_breaker_spec is not None:
            existing_connection.target_breaker_spec = target_breaker_spec
        if hierarchy_relation is not None:
            existing_connection.hierarchy_relation = hierarchy_relation
        if upstream_downstream is not None:
            existing_connection.upstream_downstream = upstream_downstream
        if parallel_count is not None:
            existing_connection.parallel_count = parallel_count
        if rated_current is not None:
            existing_connection.rated_current = rated_current
        if cable_length is not None:
            existing_connection.cable_length = cable_length
        if source_device_photo is not None:
            existing_connection.source_device_photo = source_device_photo
        if target_device_photo is not None:
            existing_connection.target_device_photo = target_device_photo
        if remark is not None:
            existing_connection.remark = remark
        if installation_date is not None:
            try:
                existing_connection.installation_date = datetime.strptime(installation_date, '%Y-%m-%d').date()
            except ValueError:
                existing_connection.installation_date = None
        
        existing_connection.updated_at = datetime.now()
        
        db.commit()
        db.refresh(existing_connection)
        
        # 鏋勫缓鍝嶅簲鏁版嵁
        response_data = {
            "id": existing_connection.id,
            "source_device_id": existing_connection.source_device_id,
            "target_device_id": existing_connection.target_device_id,
            "source_device_name": existing_connection.source_device.name,
            "target_device_name": existing_connection.target_device.name,
            "source_port": existing_connection.source_port,
            "target_port": existing_connection.target_port,
            "connection_type": existing_connection.connection_type,
            "cable_model": existing_connection.cable_model,
            "source_fuse_number": existing_connection.source_fuse_number,
            "source_fuse_spec": existing_connection.source_fuse_spec,
            "source_breaker_number": existing_connection.source_breaker_number,
            "source_breaker_spec": existing_connection.source_breaker_spec,
            "target_fuse_number": existing_connection.target_fuse_number,
            "target_fuse_spec": existing_connection.target_fuse_spec,
            "target_breaker_number": existing_connection.target_breaker_number,
            "target_breaker_spec": existing_connection.target_breaker_spec,
            "hierarchy_relation": existing_connection.hierarchy_relation,
            "upstream_downstream": existing_connection.upstream_downstream,
            "parallel_count": existing_connection.parallel_count,
            "rated_current": existing_connection.rated_current,
            "cable_length": existing_connection.cable_length,
            "source_device_photo": existing_connection.source_device_photo,
            "target_device_photo": existing_connection.target_device_photo,
            "remark": existing_connection.remark,
            "installation_date": existing_connection.installation_date.isoformat() if existing_connection.installation_date else None,
            "created_at": existing_connection.created_at.isoformat() if existing_connection.created_at else None,
            "updated_at": existing_connection.updated_at.isoformat() if existing_connection.updated_at else None
        }
        
        return JSONResponse(content={
            "success": True,
            "message": "杩炴帴鏇存柊鎴愬姛",
            "data": response_data
        })
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"鏇存柊杩炴帴澶辫触: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"鏇存柊杩炴帴澶辫触: {str(e)}")


@app.delete("/api/connections/{connection_id}")
async def delete_connection(
    connection_id: int,
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    """
    鍒犻櫎杩炴帴
    闇€瑕佺鐞嗗憳瀵嗙爜楠岃瘉
    """
    try:
        # 楠岃瘉绠＄悊鍛樺瘑鐮?
        if not verify_admin_password(password):
            raise HTTPException(status_code=401, detail="瀵嗙爜閿欒")
        
        # 鏌ユ壘瑕佸垹闄ょ殑杩炴帴
        connection = db.query(Connection).filter(Connection.id == connection_id).first()
        if not connection:
            raise HTTPException(status_code=404, detail="杩炴帴涓嶅瓨鍦?)
        
        # 鍒犻櫎杩炴帴
        db.delete(connection)
        db.commit()
        
        return JSONResponse(content={
            "success": True,
            "message": "杩炴帴鍒犻櫎鎴愬姛"
        })
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"鍒犻櫎杩炴帴澶辫触: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"鍒犻櫎杩炴帴澶辫触: {str(e)}")


@app.get("/api/connections/{connection_id}", response_model=ConnectionResponse)
async def get_connection(
    connection_id: int,
    db: Session = Depends(get_db)
):
    """
    鑾峰彇鍗曚釜杩炴帴璇︽儏
    """
    try:
        connection = db.query(Connection).filter(Connection.id == connection_id).first()
        if not connection:
            raise HTTPException(status_code=404, detail="杩炴帴涓嶅瓨鍦?)
        
        # 鎵嬪姩澶勭悊鏃ユ湡瀛楁鐨勫簭鍒楀寲
        installation_date_str = connection.installation_date.isoformat() if connection.installation_date else None
        created_at_str = connection.created_at.isoformat() if connection.created_at else None
        updated_at_str = connection.updated_at.isoformat() if connection.updated_at else None
        
        response_data = {
            "id": connection.id,
            "source_device_id": connection.source_device_id,
            "target_device_id": connection.target_device_id,
            "source_device_name": connection.source_device.name,
            "target_device_name": connection.target_device.name,
            "source_port": build_port_name_with_prefix(
                connection.source_fuse_number, 
                connection.source_breaker_number
            ),
            "target_port": build_port_name_with_prefix(
                connection.target_fuse_number, 
                connection.target_breaker_number
            ),
            "connection_type": connection.connection_type,
            "cable_model": connection.cable_model,
            "source_fuse_number": connection.source_fuse_number,
            "source_fuse_spec": connection.source_fuse_spec,
            "source_breaker_number": connection.source_breaker_number,
            "source_breaker_spec": connection.source_breaker_spec,
            "target_fuse_number": connection.target_fuse_number,
            "target_fuse_spec": connection.target_fuse_spec,
            "target_breaker_number": connection.target_breaker_number,
            "target_breaker_spec": connection.target_breaker_spec,
            "hierarchy_relation": connection.hierarchy_relation,
            "upstream_downstream": connection.upstream_downstream,
            "parallel_count": connection.parallel_count,
            "rated_current": connection.rated_current,
            "a_rated_current": connection.a_rated_current,  # A绔瀹氱數娴?
            "b_rated_current": connection.b_rated_current,  # B绔瀹氱數娴?
            "cable_length": connection.cable_length,
            "source_device_photo": connection.source_device_photo,
            "target_device_photo": connection.target_device_photo,
            "remark": connection.remark,
            "installation_date": installation_date_str,
            "created_at": created_at_str,
            "updated_at": updated_at_str
        }
        
        return JSONResponse(content={
            "success": True,
            "data": response_data
        })
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"鑾峰彇杩炴帴璇︽儏澶辫触: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"鑾峰彇杩炴帴璇︽儏澶辫触: {str(e)}")


# --- 鎬荤嚎寮忕鍙ｆ嫇鎵戝浘瀹炵幇鍑芥暟 ---

def _create_bus_topology_nodes(device: Device, db: Session, group_size: int = 12) -> dict:
    """涓鸿澶囧垱寤烘€荤嚎寮忔嫇鎵戣妭鐐癸紙鎬荤嚎鑺傜偣 + 绔彛鑺傜偣 + 鎬荤嚎-绔彛杩炴帴锛?"""
    bus_nodes = []
    port_nodes = []
    bus_port_edges = []
    connected_device_ports = []  # 瀛樺偍瀵圭璁惧鐨勭鍙ｈ妭鐐?
    
    try:
        # 1. 鑾峰彇璁惧鐨勬墍鏈夌鍙ｄ俊鎭?
        device_ports = _extract_device_ports(device, db)
        
        # 2. 鎸夌數娴佹柟鍚戝垎缁勭鍙?
        port_groups = _group_ports_by_direction(device, device_ports)
        
        # 3. 涓烘瘡涓柟鍚戝垱寤烘€荤嚎鑺傜偣鍜岀鍙ｈ妭鐐癸紙鏀寔鎸夌鍙ｆ暟閲忓垎缁勶級
        for direction, ports in port_groups.items():
            if ports:  # 鍙湁褰撹鏂瑰悜鏈夌鍙ｆ椂鎵嶅垱寤烘€荤嚎
                # 鎸夌鍙ｆ暟閲忓垎缁勶紝姣廹roup_size涓鍙ｄ竴鏉℃€荤嚎
                port_chunks = _split_ports_into_chunks(ports, max_ports_per_bus=group_size)
                
                for chunk_index, port_chunk in enumerate(port_chunks):
                    # 鍒涘缓鎬荤嚎鑺傜偣锛堝鏋滄湁澶氭潯鎬荤嚎锛屾坊鍔犵紪鍙凤級
                    bus_suffix = f"_{chunk_index + 1}" if len(port_chunks) > 1 else """
                    bus_node = _create_bus_node(device, direction + bus_suffix, port_chunk)
                    bus_nodes.append(bus_node)
                    
                    # 鍒涘缓绔彛鑺傜偣骞惰繛鎺ュ埌鎬荤嚎
                    for port in port_chunk:
                        port_node = _create_port_node_for_bus(device, port, direction, db)
                        port_nodes.append(port_node)
                        
                        # 鍒涘缓鎬荤嚎鍒扮鍙ｇ殑杩炴帴
                        bus_port_edge = _create_bus_to_port_edge(bus_node['id'], port_node['id'])
                        bus_port_edges.append(bus_port_edge)
                        
                        # 涓烘湁杩炴帴鐨勭鍙ｅ垱寤哄绔澶囩殑绠€鍖栫鍙ｈ妭鐐?
                        if port.get('connected_device_id') and port.get('connection_id'):
                            connected_port = _create_connected_device_port_node(device, port, db)
                            if connected_port:
                                connected_device_ports.append(connected_port)
        
        # 灏嗗绔澶囩鍙ｈ妭鐐规坊鍔犲埌缁撴灉涓紙鍘婚噸澶勭悊锛?
        # 浣跨敤瀛楀吀鍘婚噸锛岄伩鍏嶅悓涓€涓绔鍙ｈ閲嶅娣诲姞
        unique_connected_ports = {}
        for connected_port in connected_device_ports:
            port_id = connected_port['id']
            if port_id not in unique_connected_ports:
                unique_connected_ports[port_id] = connected_port
        
        port_nodes.extend(unique_connected_ports.values())
        
        return {
            'bus_nodes': bus_nodes,
            'port_nodes': port_nodes,
            'bus_port_edges': bus_port_edges
        }
        
    except Exception as e:
        print(f"鍒涘缓鎬荤嚎寮忔嫇鎵戣妭鐐瑰け璐? {str(e)}")
        # 濡傛灉鎬荤嚎寮忓垱寤哄け璐ワ紝鍥為€€鍒版爣鍑嗙鍙ｈ妭鐐?
        standard_ports = _create_port_nodes(device, db)
        return {
            'bus_nodes': [],
            'port_nodes': standard_ports,
            'bus_port_edges': []
        }


def _extract_device_ports(device: Device, db: Session) -> list:
    """鎻愬彇璁惧鐨勬墍鏈夌鍙ｄ俊鎭紝鍩轰簬瀹為檯鐢垫祦鏂瑰悜姝ｇ‘鍒嗙被"""
    ports = []
    
    # 浠庝綔涓虹洰鏍囪澶囩殑杩炴帴涓彁鍙栫鍙?
    for conn in device.target_connections:
        # 鏍规嵁upstream_downstream瀛楁鍒ゆ柇瀹為檯鐢垫祦鏂瑰悜
        # 瀵逛簬閫変腑璁惧浣滀负target_device鐨勮繛鎺ワ細
        # - 濡傛灉upstream_downstream涓?涓婃父"锛岃鏄庣數娴佷粠source娴佸悜target锛岄€変腑璁惧鎺ユ敹鐢靛姏锛堣緭鍏ョ鍙ｏ級
        # - 濡傛灉upstream_downstream涓?涓嬫父"锛岃鏄庣數娴佷粠target娴佸悜source锛岄€変腑璁惧杈撳嚭鐢靛姏锛堣緭鍑虹鍙ｏ級
        actual_direction = _determine_actual_port_direction(
            device.id, conn.source_device_id, conn.target_device_id, 
            conn.upstream_downstream, 'target'
        )
        
        if conn.target_fuse_number:
            ports.append({
                'name': conn.target_fuse_number,
                'type': '鐔旀柇鍣?,
                'spec': conn.target_fuse_spec,
                'connection_id': conn.id,
                'direction': actual_direction,
                'connected_device_id': conn.source_device_id,
                'upstream_downstream': conn.upstream_downstream
            })
        if conn.target_breaker_number:
            ports.append({
                'name': conn.target_breaker_number,
                'type': '鏂矾鍣?,
                'spec': conn.target_breaker_spec,
                'connection_id': conn.id,
                'direction': actual_direction,
                'connected_device_id': conn.source_device_id,
                'upstream_downstream': conn.upstream_downstream
            })
    
    # 浠庝綔涓烘簮璁惧鐨勮繛鎺ヤ腑鎻愬彇绔彛
    for conn in device.source_connections:
        # 瀵逛簬閫変腑璁惧浣滀负source_device鐨勮繛鎺ワ細
        # - 濡傛灉upstream_downstream涓?涓婃父"锛岃鏄庣數娴佷粠source娴佸悜target锛岄€変腑璁惧杈撳嚭鐢靛姏锛堣緭鍑虹鍙ｏ級
        # - 濡傛灉upstream_downstream涓?涓嬫父"锛岃鏄庣數娴佷粠target娴佸悜source锛岄€変腑璁惧鎺ユ敹鐢靛姏锛堣緭鍏ョ鍙ｏ級
        actual_direction = _determine_actual_port_direction(
            device.id, conn.source_device_id, conn.target_device_id, 
            conn.upstream_downstream, 'source'
        )
        
        if conn.source_fuse_number:
            ports.append({
                'name': conn.source_fuse_number,
                'type': '鐔旀柇鍣?,
                'spec': conn.source_fuse_spec,
                'connection_id': conn.id,
                'direction': actual_direction,
                'connected_device_id': conn.target_device_id,
                'upstream_downstream': conn.upstream_downstream
            })
        if conn.source_breaker_number:
            ports.append({
                'name': conn.source_breaker_number,
                'type': '鏂矾鍣?,
                'spec': conn.source_breaker_spec,
                'connection_id': conn.id,
                'direction': actual_direction,
                'connected_device_id': conn.target_device_id,
                'upstream_downstream': conn.upstream_downstream
            })
    
    return ports


def _group_ports_by_direction(device: Device, ports: list) -> dict:
    """鎸夊疄闄呯數娴佹柟鍚戝垎缁勭鍙ｏ紙鍩轰簬杩炴帴鍏崇郴鍜寀pstream_downstream瀛楁锛?"""
    groups = {
        'input': [],
        'output': [],
        'bidirectional': []
    }
    
    for port in ports:
        # 浣跨敤鍩轰簬瀹為檯杩炴帴鍏崇郴鐨勬柟鍚戝垽鏂紝鑰屼笉鏄熀浜庤澶囩被鍨嬪拰绔彛鍚嶇О鐨勬帹娴?
        direction = port.get('direction', 'bidirectional')
        
        # 濡傛灉鏂瑰悜浠嶇劧涓嶆槑纭紝鍒欎娇鐢ㄨ澶囩被鍨嬪拰绔彛鍚嶇О浣滀负澶囩敤鍒ゆ柇
        if direction == 'bidirectional':
            direction = _determine_port_direction(device.device_type, port['name'], 'bidirectional')
        
        groups[direction].append(port)
    
    return groups


def _determine_actual_port_direction(selected_device_id: int, source_device_id: int, target_device_id: int, upstream_downstream: str, device_role: str) -> str:
    """鍩轰簬upstream_downstream瀛楁鍜岃澶囪鑹茬‘瀹氱鍙ｇ殑瀹為檯鐢垫祦鏂瑰悜
        selected_device_id: 鐢ㄦ埛閫変腑鏌ョ湅鐨勮澶嘔D
        source_device_id: 杩炴帴涓殑婧愯澶嘔D
        target_device_id: 杩炴帴涓殑鐩爣璁惧ID
        upstream_downstream: 杩炴帴鐨勪笂涓嬫父鍏崇郴锛?涓婃父"鎴?涓嬫父"锛?
        device_role: 閫変腑璁惧鍦ㄦ杩炴帴涓殑瑙掕壊锛?source"鎴?target"锛?
    
    Returns:
        str: 绔彛鏂瑰悜锛?input"銆?output"鎴?bidirectional"锛?
    """
    if not upstream_downstream:
        return 'bidirectional'
    
    # 鏍规嵁upstream_downstream瀛楁鍜岃澶囪鑹插垽鏂數娴佹柟鍚?
    if upstream_downstream == "涓婃父":
        # 涓婃父鍏崇郴锛氱數娴佷粠source璁惧娴佸悜target璁惧
        if device_role == 'source':
            # 閫変腑璁惧鏄簮璁惧锛岃緭鍑虹數鍔?
            return 'output'
        else:  # device_role == 'target'
            # 閫変腑璁惧鏄洰鏍囪澶囷紝鎺ユ敹鐢靛姏
            return 'input'
    elif upstream_downstream == "涓嬫父":
        # 涓嬫父鍏崇郴锛氱數娴佷粠target璁惧娴佸悜source璁惧
        if device_role == 'source':
            # 閫変腑璁惧鏄簮璁惧锛屼絾鐢垫祦鏂瑰悜鐩稿弽锛屾墍浠ユ帴鏀剁數鍔?
            return 'input'
        else:  # device_role == 'target'
            # 閫変腑璁惧鏄洰鏍囪澶囷紝浣嗙數娴佹柟鍚戠浉鍙嶏紝鎵€浠ヨ緭鍑虹數鍔?
            return 'output'
    else:
        # 鏈煡鐨剈pstream_downstream鍊硷紝浣跨敤澶囩敤鍒ゆ柇
        return 'bidirectional'

def _determine_port_direction(device_type: str, port_name: str, default_direction: str) -> str:
    """鍒ゆ柇绔彛鐨勭數娴佹柟鍚?"""
    if not device_type or not port_name:
        return default_direction
    
    device_type = device_type.strip()
    port_name = port_name.strip().upper()
    
    # 鍩轰簬璁惧绫诲瀷鐨勭鍙ｆ柟鍚戣鍒?
    device_rules = {
        '鍙戠數鏈虹粍': {
        },
        'UPS': {
            'input': ['杈撳叆', 'INPUT', 'AC_IN', 'BYPASS', '鏃佽矾', '杩涚嚎'],
            'output': ['杈撳嚭', 'OUTPUT', 'AC_OUT', '鍑虹嚎']
        },
        '鍙樺帇鍣?: {
            'input': ['涓€娆?, 'PRIMARY', '楂樺帇', 'HV', '杩涚嚎'],
            'output': ['浜屾', 'SECONDARY', '浣庡帇', 'LV', '鍑虹嚎']
        },
        '楂樺帇閰嶇數鏌?: {
            'input': ['杩涚嚎', 'INPUT', '姣嶇嚎杩涚嚎', '涓昏繘绾?],
            'output': ['鍑虹嚎', 'OUTPUT', '棣堢嚎', '鍒嗘敮']
        },
        '浣庡帇閰嶇數鏌?: {
            'input': ['杩涚嚎', 'INPUT', '姣嶇嚎杩涚嚎', '涓昏繘绾?],
            'output': ['鍑虹嚎', 'OUTPUT', '棣堢嚎', '鍒嗘敮']
        },
        'ATS鏌?: {
            'input': ['杩涚嚎', 'INPUT', '甯哥敤', '澶囩敤', 'N', 'E'],
            'output': ['鍑虹嚎', 'OUTPUT', '璐熻浇']
        }
    }
    
    # 妫€鏌ヨ澶囩被鍨嬭鍒?
    rules = device_rules.get(device_type, {})
    
    for direction, keywords in rules.items():
        if any(keyword in port_name for keyword in keywords):
            return direction
    
    # 閫氱敤鍏抽敭璇嶆鏌?
    if any(keyword in port_name for keyword in ['杩涚嚎', 'INPUT', 'IN', '杈撳叆', '涓€娆?, 'PRIMARY']):
        return 'input'
    elif any(keyword in port_name for keyword in ['鍑虹嚎', 'OUTPUT', 'OUT', '杈撳嚭', '浜屾', 'SECONDARY', '棣堢嚎']):
        return 'output'
    
    # 濡傛灉鏃犳硶鍒ゆ柇锛屼娇鐢ㄩ粯璁ゆ柟鍚?
    return default_direction


def _create_bus_node(device: Device, direction: str, ports: list) -> dict:
    """鍒涘缓鎬荤嚎鑺傜偣"""
    direction_labels = {
        'input': '杈撳叆渚?,
        'output': '杈撳嚭渚?,
        'bidirectional': '鍙屽悜'
    }
    
    direction_colors = {
        'input': {'background': '#E8F5E8', 'border': '#4CAF50'},
        'output': {'background': '#FFF3E0', 'border': '#FF9800'},
        'bidirectional': {'background': '#E3F2FD', 'border': '#2196F3'}
    }
    
    port_names = [port['name'] for port in ports]
    
    return {
        'id': f"bus_{device.id}_{direction}",
        'type': 'bus',
        'label': f"{direction_labels.get(direction, direction)}鎬荤嚎",
        'title': f"{device.name} {direction_labels.get(direction, direction)}渚х鍙ｆ€荤嚎\n鍖呭惈绔彛: {', '.join(port_names[:5])}{'...' if len(port_names) > 5 else ''}",
        'device_id': device.id,
        'device_name': device.name,
        'device_type': device.device_type,
        'direction': direction,
        'port_count': len(ports),
        'port_list': port_names,
        'shape': 'box',
        'color': {
            'background': direction_colors.get(direction, {}).get('background', '#E3F2FD'),
            'border': direction_colors.get(direction, {}).get('border', '#2196F3'),
            'highlight': {
                'background': '#BBDEFB',
                'border': '#1976D2'
            }
        },
        'font': {
            'size': 14,
            'color': '#1565C0',
            'face': 'Arial'
        },
        'margin': 10,
        'widthConstraint': {'minimum': 120, 'maximum': 200},
        'heightConstraint': {'minimum': 40, 'maximum': 80}
    }


def _split_ports_into_chunks(ports: list, max_ports_per_bus: int = 10) -> list:
    """
    灏嗙鍙ｅ垪琛ㄦ寜鎸囧畾鏁伴噺鍒嗙粍锛屾瘡缁勫垱寤轰竴鏉℃€荤嚎
    
    Args:
        ports: 绔彛鍒楄〃
        max_ports_per_bus: 姣忔潯鎬荤嚎鏈€澶х鍙ｆ暟锛岄粯璁?0涓?
    
    Returns:
        list: 鍒嗙粍鍚庣殑绔彛鍒楄〃锛屾瘡涓厓绱犳槸涓€涓鍙ｇ粍
    """
    if not ports:
        return []
    
    # 灏嗙鍙ｅ垪琛ㄦ寜鎸囧畾鏁伴噺鍒嗙粍
    chunks = []
    for i in range(0, len(ports), max_ports_per_bus):
        chunk = ports[i:i + max_ports_per_bus]
        chunks.append(chunk)
    
    return chunks


def _create_connected_device_node(device: Device, port: dict, db: Session) -> dict:
    """涓烘€荤嚎寮忓竷灞€鍒涘缓瀵圭璁惧鑺傜偣"""
    if not port.get('connected_device_id'):
        return None
        
    connected_device = db.query(Device).filter(Device.id == port['connected_device_id']).first()
    if not connected_device:
        return None
    
    # 璁惧绫诲瀷棰滆壊鏄犲皠
    device_colors = {
        '鍙戠數鏈虹粍': '#4CAF50',
        'UPS': '#2196F3', 
        '鍙樺帇鍣?: '#FF9800',
        '閰嶇數鏌?: '#9C27B0',
        '寮€鍏虫煖': '#795548',
        '鐢垫睜缁?: '#607D8B'
    }
    
    base_color = device_colors.get(connected_device.device_type, '#757575')
    
    return {
        'id': f"connected_device_{connected_device.id}_from_{device.id}_{port['name']}",
        'type': 'connected_device',
        'label': connected_device.name,
        'title': f"瀵圭璁惧: {connected_device.name}\n绫诲瀷: {connected_device.device_type}\n绔欑偣: {connected_device.station}\n鍨嬪彿: {connected_device.model or 'N/A'}",
        'device_id': connected_device.id,
        'device_name': connected_device.name,
        'device_type': connected_device.device_type,
        'station': connected_device.station,
        'model': connected_device.model,
        'shape': 'box',
        'size': 30,
        'color': {
            'background': base_color,
            'border': '#424242',
            'highlight': {
                'background': _adjust_color_brightness(base_color, 1.2),
                'border': '#212121'
            }
        },
        'font': {
            'size': 12,
            'color': '#FFFFFF'
        }
    }

def _create_connected_device_port_node(device: Device, port: dict, db: Session) -> dict:
    """涓烘€荤嚎寮忓竷灞€鍒涘缓瀵圭璁惧鐨勭畝鍖栫鍙ｈ妭鐐?"""
    if not port.get('connected_device_id') or not port.get('connection_id'):
        return None
    
    try:
        # 鑾峰彇瀵圭璁惧淇℃伅
        connected_device = db.query(Device).filter(Device.id == port['connected_device_id']).first()
        if not connected_device:
            return None
        
        # 鑾峰彇杩炴帴淇℃伅
        connection = db.query(Connection).filter(Connection.id == port['connection_id']).first()
        if not connection:
            return None
        
        # 纭畾瀵圭绔彛淇℃伅
        connected_port_name = f"杩炴帴{connection.id}"  # 鏀硅繘榛樿鍚嶇О
        connected_port_type = "鏈煡"
        
        if connection.source_device_id == device.id:
            # 褰撳墠璁惧鏄簮璁惧锛屽绔槸鐩爣璁惧鐨勭鍙?
            if connection.target_fuse_number:
                connected_port_name = f"鐔斾笣-{connection.target_fuse_number}"
                connected_port_type = "鐔斾笣"
            elif connection.target_breaker_number:
                connected_port_name = f"绌哄紑-{connection.target_breaker_number}"
                connected_port_type = "绌哄紑"
            else:
                # 褰撶洰鏍囩鍙ｄ俊鎭己澶辨椂锛屼娇鐢ㄦ洿鏈夋剰涔夌殑鏍囪瘑
                connected_port_name = f"鍏ョ嚎{connection.id}"
        else:
            # 褰撳墠璁惧鏄洰鏍囪澶囷紝瀵圭鏄簮璁惧鐨勭鍙?
            if connection.source_fuse_number:
                connected_port_name = f"鐔斾笣-{connection.source_fuse_number}"
                connected_port_type = "鐔斾笣"
            elif connection.source_breaker_number:
                connected_port_name = f"绌哄紑-{connection.source_breaker_number}"
                connected_port_type = "绌哄紑"
            else:
                # 褰撴簮绔彛淇℃伅缂哄け鏃讹紝浣跨敤鏇存湁鎰忎箟鐨勬爣璇?
                connected_port_name = f"鍑虹嚎{connection.id}"
        
        # 绔彛棰滆壊閰嶇疆
        port_colors = {
            '鐔斾笣': '#FF9800',  # 姗欒壊
            '绌哄紑': '#4CAF50',  # 缁胯壊
            '鎺ヨЕ鍣?: '#2196F3',  # 钃濊壊
            '寮€鍏?: '#9C27B0'     # 绱壊
        }
        
        base_color = port_colors.get(connected_port_type, '#757575')
        
        # 浣跨敤杩炴帴ID纭繚鑺傜偣ID鐨勫敮涓€鎬э紝閬垮厤閲嶅ID闂
        return {
            'id': f"connected_port_{connected_device.id}_{connection.id}_{connected_port_name.replace('-', '_')}",
            'type': 'connected_port',
            'label': f"{connected_device.name}路{connected_port_name}",
            'title': f"{connected_device.name}路{connected_port_name}\n璁惧绫诲瀷: {connected_device.device_type or 'N/A'}\n杩炴帴鍒? {device.name}路{port['name']}",
            'device_id': connected_device.id,
            'device_name': connected_device.name,
            'device_type': connected_device.device_type,
            'station': connected_device.station,
            'port_name': connected_port_name,
            'port_type': connected_port_type,
            'connection_id': connection.id,
            'source_device_id': device.id,
            'source_port_name': port['name'],
            'shape': 'circle',
            'size': 20,
            'color': {
                'background': base_color,
                'border': '#424242',
                'highlight': {
                    'background': _adjust_color_brightness(base_color, 1.2),
                    'border': '#212121'
                }
            },
            'font': {
                'size': 9,
                'color': '#212121'
            },
            # 璁剧疆浣嶇疆锛屼娇鍏舵樉绀哄湪鍥剧殑鍙充晶
            'x': 300,  # 鐩稿浜庢煡璇㈣澶囩殑鍙充晶浣嶇疆
            'y': 0     # 鍨傜洿灞呬腑
        }
        
    except Exception as e:
        print(f"鍒涘缓瀵圭璁惧绔彛鑺傜偣澶辫触: {str(e)}")
        return None


def _create_port_node_for_bus(device: Device, port: dict, direction: str, db: Session) -> dict:
    """涓烘€荤嚎寮忓竷灞€鍒涘缓绔彛鑺傜偣锛屽寘鍚绔澶囦俊鎭?"""
    port_colors = {
        '鐔斾笣': '#FF9800',  # 姗欒壊
        '绌哄紑': '#4CAF50',  # 缁胯壊
        '鎺ヨЕ鍣?: '#2196F3',  # 钃濊壊
        '寮€鍏?: '#9C27B0'     # 绱壊
    }
    
    # 灏嗙鍙ｇ被鍨嬩腑鐨勮嫳鏂囪浆鎹负涓枃
    port_type_chinese = port['type']
    if 'fuse' in port['type'].lower() or '鐔旀柇鍣? in port['type']:
        port_type_chinese = '鐔斾笣'
    elif 'breaker' in port['type'].lower() or '鏂矾鍣? in port['type']:
        port_type_chinese = '绌哄紑'
    
    base_color = port_colors.get(port_type_chinese, '#757575')
    
    # 鑾峰彇瀵圭璁惧淇℃伅
    connected_device_info = "绌洪棽"
    connected_port_info = "绌洪棽"
    if port.get('connected_device_id'):
        connected_device = db.query(Device).filter(Device.id == port['connected_device_id']).first()
        if connected_device:
            connected_device_info = f"{connected_device.name} ({connected_device.device_type})"
            
            # 鑾峰彇瀵圭绔彛淇℃伅
            if port.get('connection_id'):
                connection = db.query(Connection).filter(Connection.id == port['connection_id']).first()
                if connection:
                    # 鏍规嵁褰撳墠璁惧鍦ㄨ繛鎺ヤ腑鐨勮鑹茬‘瀹氬绔鍙?
                    if connection.source_device_id == device.id:
                        # 褰撳墠璁惧鏄簮璁惧锛屽绔槸鐩爣璁惧鐨勭鍙?
                        if connection.target_fuse_number:
                            connected_port_info = f"鐔斾笣-{connection.target_fuse_number}"
                        elif connection.target_breaker_number:
                            connected_port_info = f"绌哄紑-{connection.target_breaker_number}"
                    else:
                        # 褰撳墠璁惧鏄洰鏍囪澶囷紝瀵圭鏄簮璁惧鐨勭鍙?
                        if connection.source_fuse_number:
                            connected_port_info = f"鐔斾笣-{connection.source_fuse_number}"
                        elif connection.source_breaker_number:
                            connected_port_info = f"绌哄紑-{connection.source_breaker_number}"
    
    # 浣跨敤鐢ㄦ埛鎻愪氦鐨勫師濮嬬鍙ｅ悕锛屼笉鍋氫换浣曠畝鍖栨垨缈昏瘧
    original_port_name = port['name']

    # 浣跨敤杩炴帴ID纭繚鑺傜偣ID鐨勫敮涓€鎬э紝閬垮厤閲嶅ID闂
    connection_suffix = f"_{port.get('connection_id', 'no_conn')}"
    
    return {
        'id': f"port_{device.id}_{original_port_name.replace(' ', '_')}{connection_suffix}",
        'type': 'port',
        'label': original_port_name,
        'title': f"{device.name} - {original_port_name}\n绫诲瀷: {port_type_chinese}\n瑙勬牸: {port.get('spec', 'N/A')}\n鏂瑰悜: {direction}\n杩炴帴鍒? {connected_device_info}\n瀵圭绔彛: {connected_port_info}",
        'device_id': device.id,
        'device_name': device.name,
        'device_type': device.device_type,
        'station': device.station,
        'port_name': original_port_name,
        'port_type': port_type_chinese,
        'port_spec': port.get('spec'),
        'direction': direction,
        'parent_bus': f"bus_{device.id}_{direction}",
        'connection_id': port.get('connection_id'),
        'connected_device_id': port.get('connected_device_id'),
        'connected_device_info': connected_device_info,
        'connected_port_info': connected_port_info,
        'shape': 'circle',
        'size': 25,
        'color': {
            'background': base_color,
            'border': '#424242',
            'highlight': {
        'background': _adjust_color_brightness(base_color, 1.2),
                'border': '#212121'
            }
        },
        'font': {
            'size': 10,
            'color': '#212121'
        }
    }


def _create_bus_to_port_edge(bus_id: str, port_id: str) -> dict:
    """鍒涘缓鎬荤嚎鍒扮鍙ｇ殑杩炴帴杈?"""
    return {
        'id': f"bus_port_{bus_id}_{port_id}",
        'type': 'bus_connection',
        'from': bus_id,
        'to': port_id,
        'arrows': 'none',
        'color': {
            'color': '#90A4AE',
            'width': 2,
            'opacity': 0.6
        },
        'dashes': [5, 5],
        'smooth': {
            'enabled': True,
            'type': 'curvedCW',
            'roundness': 0.2
        },
        'length': 50
    }


def _create_bus_port_edges(connection, direction: str) -> list:
    """涓烘€荤嚎寮忓竷灞€鍒涘缓绔彛鍒扮鍙ｇ殑杩炴帴杈?"""
    edges = []
    
    try:
        if direction == "upstream":
            # 涓婃父杩炴帴锛氫粠婧愯澶囩鍙ｅ埌鐩爣璁惧绔彛
            source_device = connection.source_device
            target_device = connection.target_device
            
            # 婧愮鍙ｏ紙杈撳嚭绔彛锛? 浣跨敤涓巁create_port_node_for_bus涓€鑷寸殑ID鏍煎紡
            source_ports = []
            if connection.source_fuse_number:
                source_ports.append(f"port_{source_device.id}_鐔斾笣-{connection.source_fuse_number}_{connection.id}")
            if connection.source_breaker_number:
                source_ports.append(f"port_{source_device.id}_绌哄紑-{connection.source_breaker_number}_{connection.id}")
            
            # 鐩爣绔彛锛堣緭鍏ョ鍙ｏ級- 浣跨敤涓巁create_port_node_for_bus涓€鑷寸殑ID鏍煎紡
            target_ports = []
            if connection.target_fuse_number:
                target_ports.append(f"port_{target_device.id}_鐔斾笣-{connection.target_fuse_number}_{connection.id}")
            if connection.target_breaker_number:
                target_ports.append(f"port_{target_device.id}_绌哄紑-{connection.target_breaker_number}_{connection.id}")
            
            # 鍒涘缓绔彛闂磋繛鎺?
            for source_port in source_ports:
                for target_port in target_ports:
                    edge = {
                        'id': f"port_conn_{connection.id}_{source_port}_{target_port}",
                        'type': 'port_connection',
                        'from': source_port,
                        'to': target_port,
                        'arrows': 'to',
                        'label': connection.connection_type or connection.cable_type or '',
                        'connection_id': connection.id,
                        'connection_type': connection.connection_type,
                        'cable_type': connection.cable_type,
                        'cable_model': connection.cable_model,
                        'remark': connection.remark,
                        'color': {
                            'color': _get_connection_color(connection.connection_type),
                            'width': _get_connection_width_from_connection(connection),
                            'highlight': _get_connection_highlight_color(connection.connection_type)
                        },
                        'smooth': {
                            'enabled': True,
                            'type': 'dynamic'
                        }
                    }
                    edges.append(edge)
        
        elif direction == "downstream":
            # 涓嬫父杩炴帴锛氫粠褰撳墠璁惧绔彛鍒扮洰鏍囪澶囩鍙?
            source_device = connection.source_device
            target_device = connection.target_device
            
            # 婧愮鍙ｏ紙杈撳嚭绔彛锛? 浣跨敤涓巁create_port_node_for_bus涓€鑷寸殑ID鏍煎紡
            source_ports = []
            if connection.source_fuse_number:
                source_ports.append(f"port_{connection.id}_{source_device.id}_鐔斾笣-{connection.source_fuse_number}")
            if connection.source_breaker_number:
                source_ports.append(f"port_{connection.id}_{source_device.id}_绌哄紑-{connection.source_breaker_number}")
            
            # 鐩爣绔彛锛堣緭鍏ョ鍙ｏ級- 浣跨敤涓巁create_port_node_for_bus涓€鑷寸殑ID鏍煎紡
            target_ports = []
            if connection.target_fuse_number:
                target_ports.append(f"port_{connection.id}_{target_device.id}_鐔斾笣-{connection.target_fuse_number}")
            if connection.target_breaker_number:
                target_ports.append(f"port_{connection.id}_{target_device.id}_绌哄紑-{connection.target_breaker_number}")
            
            # 鍒涘缓绔彛闂磋繛鎺?
            for source_port in source_ports:
                for target_port in target_ports:
                    edge = {
                        'id': f"port_conn_{connection.id}_{source_port}_{target_port}",
                        'type': 'port_connection',
                        'from': source_port,
                        'to': target_port,
                        'arrows': 'to',
                        'label': connection.connection_type or connection.cable_type or '',
                        'connection_id': connection.id,
                        'connection_type': connection.connection_type,
                        'cable_type': connection.cable_type,
                        'cable_model': connection.cable_model,
                        'remark': connection.remark,
                        'color': {
                            'color': _get_connection_color(connection.connection_type),
                            'width': _get_connection_width_from_connection(connection),
                            'highlight': _get_connection_highlight_color(connection.connection_type)
                        },
                        'smooth': {
                            'enabled': True,
                            'type': 'dynamic'
                        }
                    }
                    edges.append(edge)
    
    except Exception as e:
        print(f"鍒涘缓鎬荤嚎绔彛杩炴帴澶辫触: {str(e)}")
    
    return edges


def _get_connection_color(connection_type: str) -> str:
    """鏍规嵁杩炴帴绫诲瀷鑾峰彇杩炵嚎棰滆壊"""
    colors = {
        '鐢靛姏杩炴帴': '#F44336',    # 绾㈣壊
        '鎺у埗杩炴帴': '#2196F3',    # 钃濊壊
        '閫氫俊杩炴帴': '#4CAF50',    # 缁胯壊
        '鎺ュ湴杩炴帴': '#795548',    # 妫曡壊
        '鐢电紗杩炴帴': '#FF5722',    # 娣辨鑹?
        '姣嶇嚎杩炴帴': '#9C27B0'     # 绱壊
    }
    return colors.get(connection_type, '#424242')


def _get_connection_width_from_connection(connection) -> int:
    """浠庤繛鎺ュ璞℃帹瀵肩數鍘嬬瓑绾у苟鑾峰彇杩炵嚎瀹藉害"""
    try:
        # 灏濊瘯浠庨瀹氱數娴佹帹瀵肩數鍘嬬瓑绾?
        rated_current = None
        
        # 浼樺厛浣跨敤鏁板€煎瀷鐨剅ated_current瀛楁
        if hasattr(connection, 'rated_current') and connection.rated_current:
            rated_current = float(connection.rated_current)
        # 鍏舵灏濊瘯浠嶢绔瀹氱數娴佽В鏋?
        elif hasattr(connection, 'a_rated_current') and connection.a_rated_current:
            # 鎻愬彇鏁板瓧閮ㄥ垎锛屽"63A" -> 63
            import re
            match = re.search(r'(\d+(?:\.\d+)?)', str(connection.a_rated_current))
            if match:
                rated_current = float(match.group(1))
        # 鏈€鍚庡皾璇曚粠B绔瀹氱數娴佽В鏋?
        elif hasattr(connection, 'b_rated_current') and connection.b_rated_current:
            import re
            match = re.search(r'(\d+(?:\.\d+)?)', str(connection.b_rated_current))
            if match:
                rated_current = float(match.group(1))
        
        # 鏍规嵁棰濆畾鐢垫祦鎺ㄥ鐢靛帇绛夌骇鍜岀嚎瀹?
        if rated_current:
            if rated_current >= 1000:      # 澶х數娴侊紝鍙兘鏄珮鍘?
                return 4
            elif rated_current >= 100:     # 涓瓑鐢垫祦锛屽彲鑳芥槸涓帇
                return 3
            else:                          # 灏忕數娴侊紝浣庡帇
                return 2
        
        # 濡傛灉鏃犳硶浠庣數娴佹帹瀵硷紝灏濊瘯浠庤繛鎺ョ被鍨嬫帹瀵?
        if hasattr(connection, 'connection_type') and connection.connection_type:
            connection_type = connection.connection_type.lower()
            if 'high' in connection_type or '楂樺帇' in connection_type:
                return 4
            elif 'medium' in connection_type or '涓帇' in connection_type:
                return 3
        
        # 榛樿杩斿洖浣庡帇绾垮
        return 2
        
    except Exception as e:
        print(f"鎺ㄥ杩炵嚎瀹藉害澶辫触: {str(e)}")
        return 2


def _get_connection_width(voltage_level) -> int:
    """鏍规嵁鐢靛帇绛夌骇鑾峰彇杩炵嚎瀹藉害锛堜繚鐣欏師鍑芥暟浠ュ吋瀹瑰叾浠栬皟鐢級"""
    if voltage_level is None:
        return 2
    
    try:
        voltage = float(voltage_level)
        if voltage >= 10000:      # 楂樺帇
            return 4
        elif voltage >= 1000:     # 涓帇
            return 3
        else:                     # 浣庡帇
            return 2
    except (ValueError, TypeError):
        return 2


def _get_connection_highlight_color(connection_type: str) -> str:
    """鑾峰彇杩炴帴楂樹寒棰滆壊"""
    base_color = _get_connection_color(connection_type)
    return _adjust_color_brightness(base_color, 1.3)


def _adjust_color_brightness(hex_color: str, factor: float) -> str:
    """璋冩暣棰滆壊浜害"""
    try:
        # 绉婚櫎 # 绗﹀彿
        hex_color = hex_color.lstrip('#')
        
        # 杞崲涓?RGB
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        
        # 璋冩暣浜害
        r = min(255, max(0, int(r * factor)))
        g = min(255, max(0, int(g * factor)))
        b = min(255, max(0, int(b * factor)))
        
        # 杞崲鍥炲崄鍏繘鍒?
        return f"#{r:02x}{g:02x}{b:02x}"
    except:
        return hex_color  # 濡傛灉杞崲澶辫触锛岃繑鍥炲師鑹?


# --- 搴旂敤鍚姩鍏ュ彛缁熶竴鍒?run.py ---
# 涓洪伩鍏嶉噸澶嶅惎鍔ㄦ湇鍔★紝main.py 涓嶅啀鐩存帴杩愯 uvicorn銆?
# 璇蜂娇鐢?`python run.py` 鍚姩搴旂敤锛屾垨鍦ㄩ儴缃茬幆澧冪敱杩涚▼绠＄悊鍣ㄥ惎鍔ㄣ€?
@app.get("/api/devices/search")
async def search_devices_api(
    q: Optional[str] = Query(None, description="Fuzzy search by name/asset_id/model/vendor/station"),
    station: Optional[str] = Query(None, description="Station exact filter"),
    device_type: Optional[str] = Query(None, description="Device type exact filter"),
    vendor: Optional[str] = Query(None, description="Vendor exact filter"),
    limit: int = Query(20, ge=1, le=200, description="Max results"),
    db: Session = Depends(get_db)
):
    """
    Devices search API for device selector.
    - Supports fuzzy querying by name/asset_id/model/vendor/station.
    - Supports exact filters by station/device_type/vendor.
    Returns fields: id/asset_id/name/station/device_type/vendor/model.
    """
    try:
        query = db.query(Device)

        # 锟斤拷确锟斤拷锟斤拷
        if station:
            query = query.filter(Device.station == station)
        if device_type:
            query = query.filter(Device.device_type == device_type)
        if vendor:
            query = query.filter(Device.vendor == vendor)

        # 模锟斤拷锟斤拷询
        if q:
            q_lower = q.strip().lower()
            like_pattern = f"%{q_lower}%"
            query = query.filter(or_(
                func.lower(Device.name).like(like_pattern),
                func.lower(Device.asset_id).like(like_pattern),
                func.lower(Device.model).like(like_pattern),
                func.lower(Device.vendor).like(like_pattern),
                func.lower(Device.station).like(like_pattern)
            ))

        # 锟斤拷锟狡凤拷锟斤拷锟斤拷锟斤拷锟斤拷锟斤拷锟斤拷
        devices = query.order_by(Device.station.asc(), Device.name.asc()).limit(limit).all()

        result = [
            {
                "id": d.id,
                "asset_id": d.asset_id,
                "name": d.name,
                "station": d.station,
                "device_type": d.device_type,
                "vendor": d.vendor,
                "model": d.model,
            }
            for d in devices
        ]

        return JSONResponse(content={
            "success": True,
            "data": result
        })
    except Exception as e:
        print(f"Device search failed: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Device search failed: {str(e)}")


@app.get("/api/port-topology")
async def get_port_topology_data_alias(
    device_id: int = Query(..., description="Device ID"),
    mode: str = Query("detailed", description="Mode: detailed/compact"),
    db: Session = Depends(get_db)
):
    """
    Alias for port topology data API (query variant).
    Equivalent to /api/port-topology/{device_id}, for unified frontend usage.
    """
    return await get_port_topology_data(device_id=device_id, mode=mode, db=db)